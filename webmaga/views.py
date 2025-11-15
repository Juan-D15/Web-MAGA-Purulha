from django.http import JsonResponse, HttpResponse
from django.db import connection
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.db.models import Count, Q, Sum, Avg, Max, Min, F, Prefetch
from django.db.models.functions import TruncMonth, TruncYear, Extract
from django.core.files.storage import FileSystemStorage
from django.core.mail import send_mail
from django.conf import settings
from django.db import transaction, IntegrityError
from django.utils import timezone
from datetime import datetime
from django.contrib.auth import authenticate, login as auth_login
import os
import json
from datetime import datetime, timedelta
import uuid
import hashlib
import random
import unicodedata
import re
import logging
from .models import (
    Region, Comunidad, Actividad, TipoActividad,
    Beneficiario, BeneficiarioIndividual, BeneficiarioFamilia, BeneficiarioInstitucion,
    TipoBeneficiario, Usuario, TipoComunidad, ActividadPersonal,
    Colaborador, Puesto,
    ActividadBeneficiario, ActividadComunidad, ActividadPortada, TarjetaDato, Evidencia, ActividadCambio,
    EventoCambioColaborador, ActividadArchivo, EventosGaleria, CambioEvidencia, EventosEvidenciasCambios,
    RegionGaleria, RegionArchivo, ComunidadGaleria, ComunidadArchivo, ComunidadAutoridad,
    PasswordResetCode, UsuarioFotoPerfil, SesionOffline
)
from .decorators import (
    solo_administrador,
    permiso_gestionar_eventos,
    permiso_gestionar_eventos_api,
    permiso_generar_reportes,
    usuario_autenticado,
    get_usuario_maga,
    permiso_admin_o_personal_api,
    usuario_puede_gestionar_evento,
)
from .ratelimit_decorators import (
    api_ratelimit_read,
    api_ratelimit_write,
    api_ratelimit_auth,
    api_ratelimit_upload,
    api_ratelimit_strict,
    api_ratelimit_bulk,
    api_ratelimit_login_smart,
)
from .views_utils import (
    aplicar_modificaciones_beneficiarios,
    eliminar_portada_evento,
    guardar_portada_evento,
    obtener_cambios_evento,
    obtener_comunidades_evento,
    obtener_detalle_beneficiario,
    obtener_portada_evento,
    obtener_tarjetas_datos,
    obtener_url_portada_o_evidencia,
    normalizar_dpi,
    buscar_conflicto_dpi,
)

logger = logging.getLogger(__name__)


# =====================================================
# Funciones utilitarias internas
# =====================================================

# (Las funciones se importan desde views_utils para mantener
# compatibilidad con los m√≥dulos que contin√∫an usando
# `from webmaga import views`.)


# =====================================================
# VISTAS DE P√ÅGINAS HTML
# =====================================================

from .views_pages import (
    configgeneral,
    preguntas_frecuentes,
    index,
    comunidades,
    regiones,
    proyectos,
    gestioneseventos,
    generarreportes,
    gestionusuarios,
    mapa_completo,
    login_view,
    logout_view,
    reportes_index,
    perfilusuario,
)


# =====================================================
# VISTAS API JSON (para AJAX)
# =====================================================

def _parse_request_data(request):
    """Obtiene datos de la petici√≥n soportando JSON y formularios."""
    if request.content_type and 'application/json' in request.content_type:
        try:
            return json.loads(request.body.decode('utf-8') if request.body else '{}')
        except json.JSONDecodeError:
            return {}
    if request.method == 'POST':
        return request.POST.dict()
    return {}


def _generar_codigo_verificacion():
    return f"{random.randint(0, 999999):06d}"


def _normalizar_ruta_media(url):
    """Convierte una URL/relativa de MEDIA en una ruta absoluta del sistema de archivos."""
    if not url:
        return None

    relative_path = str(url).strip()

    if not relative_path:
        return None

    # Si ya apunta directamente dentro de MEDIA_ROOT, normalizar y devolver
    media_root_normalizado = os.path.normpath(settings.MEDIA_ROOT)
    ruta_normalizada = os.path.normpath(relative_path)
    if ruta_normalizada.startswith(media_root_normalizado):
        return ruta_normalizada

    # Normalizar separadores para comparar prefijos de MEDIA_URL
    relative_path = relative_path.replace('\\', '/')

    media_url = getattr(settings, 'MEDIA_URL', '') or ''
    posibles_prefijos = [
        media_url,
        media_url.lstrip('/'),
        '/media/',
        'media/',
    ]
    for prefijo in posibles_prefijos:
        prefijo_normalizado = prefijo.replace('\\', '/')
        if prefijo_normalizado and relative_path.startswith(prefijo_normalizado):
            relative_path = relative_path[len(prefijo_normalizado):]
            break

    # Eliminar posibles separadores iniciales restantes
    relative_path = relative_path.lstrip('/\\')
    if not relative_path:
        return None

    return os.path.normpath(os.path.join(str(settings.MEDIA_ROOT), relative_path))


def _eliminar_archivo_media(url):
    """Elimina un archivo f√≠sico ubicado dentro de MEDIA_ROOT."""
    file_path = _normalizar_ruta_media(url)
    if not file_path:
        return False

    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            return True
        except Exception as error:
            print(f"‚ö†Ô∏è No se pudo eliminar el archivo '{file_path}': {error}")
    return False


def _eliminar_queryset_con_archivos(queryset):
    """Elimina los registros de un queryset removiendo previamente los archivos f√≠sicos asociados."""
    objetos = list(queryset)
    if not objetos:
        return 0

    ids = []
    for obj in objetos:
        ids.append(obj.pk)
        _eliminar_archivo_media(getattr(obj, 'url_almacenamiento', None))

    queryset.model.objects.filter(pk__in=ids).delete()
    return len(ids)


def parse_fecha_agregacion(valor):
    """
    Convierte una cadena (YYYY-MM-DD o ISO8601) en datetime consciente de zona horaria.
    Retorna None si el valor no es v√°lido.
    """
    if not valor:
        return None

    if isinstance(valor, str):
        valor = valor.strip()
        if not valor:
            return None

    try:
        if isinstance(valor, datetime):
            dt = valor
        elif isinstance(valor, str):
            try:
                dt = datetime.fromisoformat(valor)
            except ValueError:
                try:
                    dt = datetime.strptime(valor, '%Y-%m-%d')
                except ValueError:
                    return None
        else:
            return None

        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        else:
            dt = dt.astimezone(timezone.get_current_timezone())

        return dt
    except Exception:
        return None


@require_http_methods(["POST"])
def api_enviar_codigo_recuperacion(request):
    """Genera y env√≠a un c√≥digo de recuperaci√≥n de contrase√±a por correo."""
    data = _parse_request_data(request)
    email = (data.get('email') or '').strip().lower()

    if not email:
        return JsonResponse({'success': False, 'error': 'El correo es obligatorio.'}, status=400)

    try:
        usuario = Usuario.objects.get(email__iexact=email, activo=True)
    except Usuario.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Error al enviar el correo.'
        }, status=404)

    PasswordResetCode.objects.filter(usuario=usuario, usado=False).update(usado=True, usado_en=timezone.now())

    codigo = _generar_codigo_verificacion()
    expira_en = timezone.now() + timedelta(minutes=10)

    registro = PasswordResetCode.objects.create(
        usuario=usuario,
        codigo=codigo,
        expira_en=expira_en
    )

    destinatario = usuario.nombre or usuario.username or usuario.email
    remitente = settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER or 'no-reply@maga-purulha.local'

    asunto = 'C√≥digo de verificaci√≥n - MAGA Purulh√°'
    mensaje_texto = (
        f"Hola {destinatario},\n\n"
        "Recibimos una solicitud para restablecer la contrase√±a de tu cuenta en la plataforma MAGA Purulh√°.\n\n"
        f"Tu c√≥digo de verificaci√≥n es: {codigo}\n"
        "Introduce este c√≥digo en la pantalla de recuperaci√≥n para continuar.\n"
        "El c√≥digo tiene una vigencia de 10 minutos y solo puede usarse una vez.\n\n"
        "Si t√∫ no solicitaste este cambio, ignora este mensaje y tu contrase√±a seguir√° siendo la misma.\n\n"
        "Saludos,\n"
        "Equipo MAGA Purulh√°"
    )

    mensaje_html = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; margin:0; padding:24px; background:#f5f7fb; color:#1f2933;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="max-width:520px; margin:auto; background:#ffffff; border-radius:16px; box-shadow:0 18px 45px rgba(15,23,42,0.12); overflow:hidden;">
        <tr>
          <td style="background:linear-gradient(135deg,#0f4c81,#155e98); padding:28px 32px;">
            <h1 style="margin:0; font-size:24px; font-weight:700; color:#ffffff;">C√≥digo de verificaci√≥n</h1>
            <p style="margin:8px 0 0; font-size:14px; color:rgba(255,255,255,0.85);">
              Ministerio de Agricultura, Ganader√≠a y Alimentaci√≥n ‚Äî Purulh√°
            </p>
          </td>
        </tr>
        <tr>
          <td style="padding:32px;">
            <p style="margin:0 0 12px; font-size:16px;">Hola <strong>{destinatario}</strong>,</p>
            <p style="margin:0 0 24px; font-size:14px; line-height:1.6; color:#334155;">
              Recibimos una solicitud para restablecer la contrase√±a de tu cuenta. Introduce el siguiente c√≥digo de verificaci√≥n para continuar con el proceso:
            </p>
            <div style="text-align:center; margin:24px 0;">
              <span style="display:inline-block; padding:18px 28px; background:linear-gradient(135deg,#0f4c81,#155e98); color:#ffffff; font-size:28px; letter-spacing:6px; border-radius:12px; font-weight:700;">
                {codigo}
              </span>
            </div>
            <p style="margin:0 0 16px; font-size:14px; line-height:1.6; color:#334155;">
              ‚ö†Ô∏è Este c√≥digo es v√°lido por <strong>10 minutos</strong> y solo se puede utilizar una vez. Si no solicitaste este cambio, puedes ignorar este mensaje y tu contrase√±a seguir√° siendo la misma.
            </p>
            <div style="margin:24px 0 0; padding:16px 20px; background:#f1f5f9; border-radius:12px;">
              <p style="margin:0; font-size:13px; color:#475569;">¬øNecesitas ayuda? Comun√≠cate con el equipo de soporte del MAGA Purulh√°.</p>
            </div>
            <p style="margin:32px 0 0; font-size:13px; color:#94a3b8; text-transform:uppercase; letter-spacing:1.5px;">Equipo MAGA Purulh√°</p>
          </td>
        </tr>
      </table>
      <p style="margin:24px auto 0; max-width:520px; font-size:11px; color:#94a3b8; text-align:center;">Recibiste este correo porque se solicit√≥ un restablecimiento de contrase√±a en la plataforma MAGA Purulh√°.</p>
    </div>
    """

    try:
        send_mail(asunto, mensaje_texto, remitente, [usuario.email], fail_silently=False, html_message=mensaje_html)
    except Exception as exc:
        registro.marcar_usado()
        return JsonResponse({
            'success': False,
            'error': 'No se pudo enviar el correo de verificaci√≥n. Int√©ntalo nuevamente en unos minutos.',
            'detail': str(exc)
        }, status=500)

    return JsonResponse({
        'success': True,
        'message': 'Enviamos un c√≥digo de verificaci√≥n a tu correo. Revisa la bandeja de entrada o spam.',
        'expires_in_minutes': 10
    })


@require_http_methods(["POST"])
def api_verificar_codigo_recuperacion(request):
    """Valida el c√≥digo ingresado por el usuario."""
    data = _parse_request_data(request)
    email = (data.get('email') or '').strip().lower()
    codigo = (data.get('code') or data.get('codigo') or '').strip()

    if not email or not codigo:
        return JsonResponse({'success': False, 'error': 'Correo y c√≥digo son obligatorios.'}, status=400)

    if len(codigo) != 6 or not codigo.isdigit():
        return JsonResponse({'success': False, 'error': 'El c√≥digo debe tener 6 d√≠gitos.'}, status=400)

    try:
        usuario = Usuario.objects.get(email__iexact=email, activo=True)
    except Usuario.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Correo inv√°lido.'}, status=404)

    registro = PasswordResetCode.objects.filter(
        usuario=usuario,
        codigo=codigo
    ).order_by('-creado_en').first()

    if not registro:
        return JsonResponse({'success': False, 'error': 'El c√≥digo ingresado no es v√°lido.'}, status=400)

    if registro.usado:
        return JsonResponse({'success': False, 'error': 'El c√≥digo ya fue utilizado. Solicita uno nuevo.'}, status=400)

    if registro.expira_en < timezone.now():
        registro.usado = True
        registro.usado_en = timezone.now()
        registro.save(update_fields=['usado', 'usado_en'])
        return JsonResponse({'success': False, 'error': 'El c√≥digo expir√≥. Solicita uno nuevo.'}, status=400)

    if registro.intentos >= 5:
        registro.usado = True
        registro.usado_en = timezone.now()
        registro.save(update_fields=['usado', 'usado_en'])
        return JsonResponse({'success': False, 'error': 'Se super√≥ el l√≠mite de intentos. Solicita un nuevo c√≥digo.'}, status=400)

    registro.intentos += 1
    registro.verificado_en = timezone.now()
    registro.save(update_fields=['intentos', 'verificado_en'])

    return JsonResponse({
        'success': True,
        'message': 'C√≥digo verificado. Ahora puedes crear una nueva contrase√±a.',
        'reset_token': str(registro.token)
    })


@require_http_methods(["POST"])
def api_resetear_password(request):
    """Permite establecer una nueva contrase√±a usando un token v√°lido."""
    from django.contrib.auth.models import User
    from .authentication import hash_password_for_maga

    data = _parse_request_data(request)
    email = (data.get('email') or '').strip().lower()
    token = data.get('token') or data.get('reset_token')
    nueva_password = data.get('password') or data.get('new_password')

    if not email or not token or not nueva_password:
        return JsonResponse({'success': False, 'error': 'Correo, token y nueva contrase√±a son obligatorios.'}, status=400)

    if len(nueva_password) < 8 or not any(c.isupper() for c in nueva_password) or not any(c.islower() for c in nueva_password) or not any(c.isdigit() for c in nueva_password):
        return JsonResponse({
            'success': False,
            'error': 'La contrase√±a debe tener al menos 8 caracteres, incluir una may√∫scula, una min√∫scula y un n√∫mero.'
        }, status=400)

    try:
        usuario = Usuario.objects.get(email__iexact=email, activo=True)
    except Usuario.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Correo inv√°lido.'}, status=404)

    try:
        registro = PasswordResetCode.objects.get(usuario=usuario, token=token)
    except PasswordResetCode.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'El token proporcionado no es v√°lido.'}, status=400)

    if registro.usado:
        return JsonResponse({'success': False, 'error': 'El token ya fue utilizado. Solicita un nuevo c√≥digo.'}, status=400)

    if registro.expira_en < timezone.now():
        registro.usado = True
        registro.usado_en = timezone.now()
        registro.save(update_fields=['usado', 'usado_en'])
        return JsonResponse({'success': False, 'error': 'El token expir√≥. Solicita un nuevo c√≥digo.'}, status=400)

    if not registro.verificado_en:
        return JsonResponse({'success': False, 'error': 'Debes validar el c√≥digo antes de cambiar la contrase√±a.'}, status=400)

    nuevo_hash = hash_password_for_maga(nueva_password)
    usuario.password_hash = nuevo_hash
    usuario.actualizado_en = timezone.now()
    usuario.save(update_fields=['password_hash', 'actualizado_en'])

    try:
        user_django = User.objects.get(username=usuario.username)
        user_django.set_password(nueva_password)
        user_django.is_active = usuario.activo
        user_django.is_staff = usuario.rol == 'admin'
        user_django.is_superuser = usuario.rol == 'admin'
        user_django.save(update_fields=['password', 'is_active', 'is_staff', 'is_superuser'])
    except User.DoesNotExist:
        User.objects.create_user(
            username=usuario.username,
            email=usuario.email,
            password=nueva_password,
            is_staff=usuario.rol == 'admin',
            is_superuser=usuario.rol == 'admin',
            is_active=usuario.activo
        )

    registro.marcar_usado()

    remitente = settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER or 'no-reply@maga-purulha.local'
    asunto = '‚úÖ Contrase√±a actualizada - MAGA Purulh√°'
    mensaje_texto = (
        f"Hola {usuario.nombre or usuario.username or usuario.email},\n\n"
        "Te confirmamos que la contrase√±a de tu cuenta se actualiz√≥ correctamente.\n\n"
        "Si t√∫ no realizaste este cambio, contacta inmediatamente al administrador del sistema.\n\n"
        "Saludos,\n"
        "Equipo MAGA Purulh√°"
    )

    mensaje_html = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; margin:0; padding:24px; background:#f5f7fb; color:#1f2933;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="max-width:520px; margin:auto; background:#ffffff; border-radius:16px; box-shadow:0 18px 45px rgba(15,23,42,0.12); overflow:hidden;">
        <tr>
          <td style="background:linear-gradient(135deg,#0f4c81,#155e98); padding:28px 32px;">
            <h1 style="margin:0; font-size:24px; font-weight:700; color:#ffffff;">Contrase√±a actualizada</h1>
            <p style="margin:8px 0 0; font-size:14px; color:rgba(255,255,255,0.85);">
              Ministerio de Agricultura, Ganader√≠a y Alimentaci√≥n ‚Äî Purulh√°
            </p>
          </td>
        </tr>
        <tr>
          <td style="padding:32px;">
            <p style="margin:0 0 12px; font-size:16px;">Hola <strong>{usuario.nombre or usuario.username or usuario.email}</strong>,</p>
            <p style="margin:0 0 18px; font-size:14px; line-height:1.6; color:#334155;">
              Te confirmamos que la contrase√±a de tu cuenta se actualiz√≥ correctamente. Ya puedes iniciar sesi√≥n con tu nueva contrase√±a.
            </p>
            <div style="margin:24px 0 0; padding:16px 20px; background:#f1f5f9; border-radius:12px;">
              <p style="margin:0; font-size:13px; color:#475569;">Si t√∫ no realizaste este cambio, ponte en contacto inmediatamente con el administrador del sistema para proteger tu cuenta.</p>
            </div>
            <p style="margin:32px 0 0; font-size:13px; color:#94a3b8; text-transform:uppercase; letter-spacing:1.5px;">Equipo MAGA Purulh√°</p>
          </td>
        </tr>
      </table>
      <p style="margin:24px auto 0; max-width:520px; font-size:11px; color:#94a3b8; text-align:center;">Recibiste este correo como confirmaci√≥n de seguridad despu√©s de un cambio de contrase√±a en la plataforma MAGA Purulh√°.</p>
    </div>
    """

    try:
        send_mail(asunto, mensaje_texto, remitente, [usuario.email], fail_silently=True, html_message=mensaje_html)
    except Exception:
        pass

    return JsonResponse({'success': True, 'message': 'La contrase√±a se actualiz√≥ correctamente. Ya puedes iniciar sesi√≥n.'})


@login_required
@require_http_methods(["POST"])
def api_registrar_sesion_offline(request):
    """Registra o actualiza la autorizaci√≥n de sesi√≥n offline para un dispositivo."""
    usuario_maga = get_usuario_maga(request.user)
    if not usuario_maga:
        return JsonResponse({'success': False, 'error': 'Usuario MAGA no encontrado.'}, status=404)

    data = _parse_request_data(request)
    device_id = (data.get('device_id') or '').strip()
    credential_hash = (data.get('credential_hash') or '').strip()
    salt = (data.get('salt') or '').strip()
    expires_at_str = data.get('expires_at')
    permisos = data.get('permisos') or {}
    metadata = data.get('metadata') or {}

    if not device_id or not credential_hash or not salt:
        return JsonResponse({
            'success': False,
            'error': 'Datos incompletos para registrar la sesi√≥n offline.'
        }, status=400)

    expires_at = timezone.now() + timedelta(hours=72)
    if expires_at_str:
        try:
            parsed = datetime.fromisoformat(str(expires_at_str).replace('Z', '+00:00'))
            if timezone.is_naive(parsed):
                expires_at = timezone.make_aware(parsed, timezone.get_current_timezone())
            else:
                expires_at = parsed.astimezone(timezone.get_current_timezone())
        except ValueError:
            pass

    # Derivar hash que no expone directamente el hash enviado por el cliente
    token_material = f"{credential_hash}:{salt}:{device_id}:{usuario_maga.id}"
    token_hash = hashlib.sha256(token_material.encode('utf-8')).hexdigest()

    defaults = {
        'token_hash': token_hash,
        'permisos_offline': permisos,
        'datos_cache': metadata,
        'sesion_activa': True,
        'expira_en': expires_at,
        'ultima_actividad': timezone.now(),
        'ultima_sincronizacion': timezone.now(),
        'requiere_reautenticacion': False,
        'navegador': metadata.get('navegador'),
        'sistema_operativo': metadata.get('sistema_operativo'),
    }

    SesionOffline.objects.update_or_create(
        usuario=usuario_maga,
        dispositivo_id=device_id,
        defaults=defaults
    )

    return JsonResponse({
        'success': True,
        'device_id': device_id,
        'expires_at': expires_at.isoformat()
    })


@login_required
@require_http_methods(["POST"])
def api_enviar_asistencia_tecnica(request):
    """API: Env√≠a un correo de asistencia t√©cnica/comentario/problema"""
    data = _parse_request_data(request)
    
    nombre = (data.get('nombre') or '').strip()
    tipo = (data.get('tipo') or '').strip()
    mensaje = (data.get('mensaje') or '').strip()
    
    # Validaciones
    if not nombre:
        return JsonResponse({
            'success': False,
            'error': 'El nombre es requerido'
        }, status=400)
    
    if not tipo:
        return JsonResponse({
            'success': False,
            'error': 'El tipo es requerido'
        }, status=400)
    
    if not mensaje:
        return JsonResponse({
            'success': False,
            'error': 'El mensaje es requerido'
        }, status=400)
    
    # Obtener informaci√≥n del usuario
    usuario_maga = get_usuario_maga(request.user)
    usuario_email = request.user.email if hasattr(request.user, 'email') and request.user.email else 'No disponible'
    usuario_username = request.user.username if request.user.is_authenticated else 'No disponible'
    
    # Mapear tipos a espa√±ol
    tipo_map = {
        'comentario': 'Comentario',
        'problema': 'Reportar Problema',
        'sugerencia': 'Sugerencia',
        'otro': 'Otro'
    }
    tipo_display = tipo_map.get(tipo, tipo)
    
    # Configurar correo
    # El remitente debe ser el mismo correo configurado en EMAIL_HOST_USER
    remitente = settings.EMAIL_HOST_USER or 'recupmagabvpurulha@gmail.com'
    destinatario = getattr(settings, 'ASISTENCIA_CONTACT_EMAIL', 'recupmagabvpurulha@gmail.com')
    asunto = f'[Asistencia T√©cnica] {tipo_display} - {nombre}'
    
    # Mensaje en texto plano
    mensaje_texto = f"""Nuevo mensaje de Asistencia T√©cnica

Tipo: {tipo_display}
Nombre: {nombre}
Usuario: {usuario_username}
Email del usuario: {usuario_email}

Mensaje:
{mensaje}

---
Este correo fue enviado desde el sistema MAGA Purulh√°."""
    
    # Mensaje HTML (sin espacios iniciales)
    mensaje_html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Asistencia T√©cnica - MAGA Purulh√°</title>
</head>
<body style="margin:0; padding:0; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Helvetica Neue',Arial,sans-serif; background-color:#f8fafc;">
  <div style="max-width:600px; margin:0 auto; background-color:#ffffff;">
    <div style="padding:32px 24px; background:linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);">
      <h1 style="margin:0; font-size:24px; font-weight:600; color:#ffffff; text-align:center;">Asistencia T√©cnica</h1>
      <p style="margin:8px 0 0; font-size:14px; color:#e0e7ff; text-align:center;">Sistema MAGA Purulh√°</p>
    </div>
    
    <div style="padding:32px 24px;">
      <div style="margin-bottom:24px; padding:20px; background:#f1f5f9; border-radius:12px; border-left:4px solid #3b82f6;">
        <h2 style="margin:0 0 16px; font-size:18px; font-weight:600; color:#1e293b;">Nuevo mensaje recibido</h2>
        <div style="margin-bottom:12px;">
          <strong style="color:#475569; font-size:13px; text-transform:uppercase; letter-spacing:0.5px;">Tipo:</strong>
          <p style="margin:4px 0 0; font-size:15px; color:#1e293b; font-weight:500;">{tipo_display}</p>
        </div>
        <div style="margin-bottom:12px;">
          <strong style="color:#475569; font-size:13px; text-transform:uppercase; letter-spacing:0.5px;">Nombre:</strong>
          <p style="margin:4px 0 0; font-size:15px; color:#1e293b; font-weight:500;">{nombre}</p>
        </div>
        <div style="margin-bottom:12px;">
          <strong style="color:#475569; font-size:13px; text-transform:uppercase; letter-spacing:0.5px;">Usuario:</strong>
          <p style="margin:4px 0 0; font-size:15px; color:#1e293b;">{usuario_username}</p>
        </div>
        <div style="margin-bottom:12px;">
          <strong style="color:#475569; font-size:13px; text-transform:uppercase; letter-spacing:0.5px;">Email del usuario:</strong>
          <p style="margin:4px 0 0; font-size:15px; color:#1e293b;">{usuario_email}</p>
        </div>
      </div>
      
      <div style="margin-bottom:24px;">
        <strong style="display:block; margin-bottom:8px; color:#475569; font-size:13px; text-transform:uppercase; letter-spacing:0.5px;">Mensaje:</strong>
        <div style="padding:16px 20px; background:#f8fafc; border-radius:8px; border:1px solid #e2e8f0;">
          <p style="margin:0; font-size:15px; line-height:1.6; color:#334155; white-space:pre-wrap;">{mensaje}</p>
        </div>
      </div>
      
      <div style="margin-top:32px; padding:16px 20px; background:#f1f5f9; border-radius:12px;">
        <p style="margin:0; font-size:13px; color:#475569;">Este correo fue enviado autom√°ticamente desde el sistema MAGA Purulh√°.</p>
      </div>
    </div>
    
    <div style="padding:24px; background:#f8fafc; border-top:1px solid #e2e8f0; text-align:center;">
      <p style="margin:0; font-size:12px; color:#94a3b8;">¬©2025 Oficina Regional MAGA Purulh√° B.V.</p>
    </div>
  </div>
</body>
</html>"""
    
    import logging
    logger = logging.getLogger(__name__)
    
    # Verificar configuraci√≥n de correo antes de enviar
    logger.info(f'Configuraci√≥n de correo - Host: {settings.EMAIL_HOST}, Port: {settings.EMAIL_PORT}, User: {settings.EMAIL_HOST_USER}')
    
    if not settings.EMAIL_HOST_USER or not settings.EMAIL_HOST_PASSWORD:
        logger.warning('Configuraci√≥n de correo no encontrada. Usando valores por defecto.')
        return JsonResponse({
            'success': False,
            'error': 'Configuraci√≥n de correo no encontrada. Contacta al administrador.',
            'detail': 'EMAIL_HOST_USER o EMAIL_HOST_PASSWORD no est√°n configurados'
        }, status=500)
    
    # Enviar correo
    try:
        logger.info(f'Intentando enviar correo desde {remitente} a {destinatario}')
        result = send_mail(
            asunto,
            mensaje_texto,
            remitente,
            [destinatario],
            fail_silently=False,
            html_message=mensaje_html
        )
        
        logger.info(f'Correo de asistencia t√©cnica enviado. Resultado: {result}')
        
        if not result:
            return JsonResponse({
                'success': False,
                'error': 'El correo no se pudo enviar. Por favor, intenta nuevamente.',
                'detail': 'send_mail retorn√≥ False'
            }, status=500)
        
    except Exception as exc:
        logger.error(f'Error al enviar correo de asistencia t√©cnica: {str(exc)}', exc_info=True)
        
        # Mensaje de error m√°s detallado
        error_type = type(exc).__name__
        error_message = str(exc)
        
        # Mensaje de error para el usuario
        if 'Authentication' in error_type or '534' in error_message:
            user_error = 'Error de autenticaci√≥n. Verifica que la contrase√±a de aplicaci√≥n sea correcta.'
        elif 'Connection' in error_type or 'Connection refused' in error_message:
            user_error = 'Error de conexi√≥n. Verifica tu conexi√≥n a internet o la configuraci√≥n del servidor de correo.'
        else:
            user_error = 'No se pudo enviar el correo. Por favor, intenta nuevamente m√°s tarde.'
        
        return JsonResponse({
            'success': False,
            'error': user_error,
            'detail': error_message if settings.DEBUG else None,
            'error_type': error_type if settings.DEBUG else None
        }, status=500)
    
    return JsonResponse({
        'success': True,
        'message': '¬°Mensaje enviado exitosamente! Nos pondremos en contacto contigo pronto.'
    })


def api_usuario_actual(request):
    """API: Informaci√≥n del usuario actual y sus permisos"""
    if not request.user.is_authenticated:
        return JsonResponse({
            'autenticado': False,
            'error': 'Usuario no autenticado'
        }, status=401)
    
    usuario_maga = get_usuario_maga(request.user)
    
    if not usuario_maga:
        return JsonResponse({
            'autenticado': True,
            'error': 'Usuario MAGA no encontrado'
        }, status=404)
    
    # Obtener colaborador vinculado si existe
    colaborador_id = None
    colaborador_nombre = None
    if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
        colaborador_id = str(usuario_maga.colaborador.id)
        colaborador_nombre = usuario_maga.colaborador.nombre
    
    return JsonResponse({
        'autenticado': True,
        'userId': str(usuario_maga.id),
        'username': usuario_maga.username,
        'nombre': usuario_maga.nombre or '',
        'email': usuario_maga.email,
        'telefono': usuario_maga.telefono or '',
        'rol': usuario_maga.rol,
        'colaborador_id': colaborador_id,
        'collaboratorId': colaborador_id,
        'collaboratorName': colaborador_nombre,
        'isAdmin': usuario_maga.rol == 'admin',
        'permisos': {
            'es_admin': usuario_maga.rol == 'admin',
            'es_personal': usuario_maga.rol == 'personal',
            'puede_gestionar_eventos': usuario_maga.rol in ['admin', 'personal'],
            'puede_generar_reportes': True,
        }
    })


@require_http_methods(["POST"])
@permiso_admin_o_personal_api
def api_actualizar_perfil_usuario(request):
    """API: Actualizar perfil del usuario actual y sincronizar con colaborador vinculado"""
    usuario_maga = get_usuario_maga(request.user)
    
    if not usuario_maga:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    
    try:
        import json
        from django.db import transaction
        data = json.loads(request.body or '{}')
        
        username = data.get('username', '').strip()
        nombre = data.get('nombre', '').strip()
        email = data.get('email', '').strip()
        telefono = data.get('telefono', '').strip()
        
        # Validaciones
        if not username:
            return JsonResponse({
                'success': False,
                'error': 'El nombre de usuario es requerido'
            }, status=400)
        
        if not email:
            return JsonResponse({
                'success': False,
                'error': 'El email es requerido'
            }, status=400)
        
        # Validar email √∫nico (excepto el mismo usuario)
        if Usuario.objects.filter(email=email).exclude(id=usuario_maga.id).exists():
            return JsonResponse({
                'success': False,
                'error': 'El email ya est√° en uso'
            }, status=400)
        
        # Validar username √∫nico (excepto el mismo usuario)
        if Usuario.objects.filter(username=username).exclude(id=usuario_maga.id).exists():
            return JsonResponse({
                'success': False,
                'error': 'El nombre de usuario ya est√° en uso'
            }, status=400)
        
        # Validar tel√©fono si se proporciona (8 d√≠gitos)
        if telefono and (not telefono.isdigit() or len(telefono) != 8):
            return JsonResponse({
                'success': False,
                'error': 'El tel√©fono debe contener exactamente 8 d√≠gitos'
            }, status=400)
        
        # Obtener colaborador vinculado si existe
        colaborador = None
        if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
            colaborador = usuario_maga.colaborador
        
        # Validar email √∫nico en colaboradores si se est√° actualizando (excepto el colaborador vinculado)
        if colaborador:
            if Colaborador.objects.filter(correo=email).exclude(id=colaborador.id).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'El email ya est√° en uso por otro colaborador'
                }, status=400)
        
        # Actualizar en una transacci√≥n para mantener consistencia
        with transaction.atomic():
            # Actualizar usuario
            usuario_maga.username = username
            usuario_maga.nombre = nombre if nombre else None
            usuario_maga.email = email
            usuario_maga.telefono = telefono if telefono else None
            usuario_maga.save()
            
            # Si tiene colaborador vinculado, sincronizar campos comunes (nombre, email, telefono)
            # Estos campos siempre se sincronizan cuando se actualizan (incluso si se borran)
            if colaborador:
                # Sincronizar nombre: siempre usar el valor enviado (puede ser vac√≠o)
                colaborador.nombre = nombre if nombre else None
                # Sincronizar email: siempre usar el nuevo valor (es requerido)
                colaborador.correo = email
                # Sincronizar tel√©fono: siempre usar el valor enviado (puede ser vac√≠o)
                colaborador.telefono = telefono if telefono else None
                colaborador.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Perfil actualizado exitosamente' + (' (sincronizado con colaborador)' if colaborador else ''),
            'usuario': {
                'username': usuario_maga.username,
                'nombre': usuario_maga.nombre or '',
                'email': usuario_maga.email,
                'telefono': usuario_maga.telefono or ''
            },
            'sincronizado_colaborador': colaborador is not None
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar perfil: {str(e)}'
        }, status=500)


@require_http_methods(["GET", "POST"])
@permiso_admin_o_personal_api
def api_foto_perfil(request):
    """API: Obtener o subir/actualizar foto de perfil del usuario actual"""
    usuario_maga = get_usuario_maga(request.user)
    
    if not usuario_maga:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    
    if request.method == 'GET':
        # Obtener foto de perfil
        try:
            foto_perfil = UsuarioFotoPerfil.objects.get(usuario=usuario_maga)
            foto_url = foto_perfil.url_almacenamiento
            print(f"üì§ GET foto de perfil - URL en BD: {foto_url}")
            
            # Verificar si el archivo existe f√≠sicamente
            if foto_url:
                file_path = _normalizar_ruta_media(foto_url)
                if file_path:
                    file_exists = os.path.exists(file_path)
                    print(f"üîç Archivo existe f√≠sicamente: {file_exists}")
                    print(f"üìÇ Ruta del archivo: {file_path}")
                    if not file_exists:
                        print(f"‚ö†Ô∏è ADVERTENCIA: El archivo no existe en el sistema de archivos pero est√° en la BD")
            
            return JsonResponse({
                'success': True,
                'foto_url': foto_url,
                'archivo_nombre': foto_perfil.archivo_nombre
            })
        except UsuarioFotoPerfil.DoesNotExist:
            print(f"üì§ GET foto de perfil - No hay foto de perfil para este usuario")
            return JsonResponse({
                'success': True,
                'foto_url': None
            })
    
    elif request.method == 'POST':
        # Subir/actualizar foto de perfil
        try:
            print(f"üì§ Iniciando subida de foto de perfil para usuario: {usuario_maga.id}")
            foto = request.FILES.get('foto')
            if not foto:
                print("‚ùå No se proporcion√≥ ning√∫n archivo en request.FILES")
                print(f"üìã Keys en request.FILES: {list(request.FILES.keys())}")
                return JsonResponse({
                    'success': False,
                    'error': 'No se proporcion√≥ ning√∫n archivo'
                }, status=400)
            
            print(f"‚úÖ Archivo recibido: {foto.name}, tama√±o: {foto.size}, tipo: {foto.content_type}")
            
            # Validar tipo de archivo
            allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
            if foto.content_type not in allowed_types:
                return JsonResponse({
                    'success': False,
                    'error': 'Tipo de archivo no permitido. Solo se permiten im√°genes (JPEG, PNG, GIF, WEBP)'
                }, status=400)
            
            # COMPRIMIR IMAGEN autom√°ticamente si est√° habilitado
            if getattr(settings, 'COMPRESS_IMAGES', True):
                from .image_compression import compress_if_needed, get_compression_settings
                compression_settings = get_compression_settings('profile')
                foto = compress_if_needed(foto, **compression_settings)
                print(f"‚úÖ Imagen procesada (comprimida si era necesario): {foto.size} bytes")
            
            # Validar tama√±o (5MB m√°ximo) - despu√©s de comprimir
            max_size = 5 * 1024 * 1024  # 5MB
            if foto.size > max_size:
                return JsonResponse({
                    'success': False,
                    'error': 'El archivo es demasiado grande. El tama√±o m√°ximo es 5MB'
                }, status=400)
            
            # Crear directorio si no existe
            perfiles_dir = os.path.join(str(settings.MEDIA_ROOT), 'perfiles_img')
            print(f"üìÅ Directorio de perfiles: {perfiles_dir}")
            os.makedirs(perfiles_dir, exist_ok=True)
            print(f"‚úÖ Directorio creado/verificado: {perfiles_dir}")
            
            # Verificar permisos de escritura
            if not os.access(perfiles_dir, os.W_OK):
                print(f"‚ùå No hay permisos de escritura en {perfiles_dir}")
                return JsonResponse({
                    'success': False,
                    'error': f'No se tienen permisos de escritura en {perfiles_dir}'
                }, status=500)
            print(f"‚úÖ Permisos de escritura verificados")
            
            # Guardar archivo
            # Asegurar que location sea un string
            perfiles_dir_str = str(perfiles_dir)
            print(f"üîß Usando directorio (string): {perfiles_dir_str}")
            fs = FileSystemStorage(location=perfiles_dir_str)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
            file_extension = os.path.splitext(foto.name)[1]
            filename = f"{timestamp}_{usuario_maga.id}{file_extension}"
            print(f"üíæ Intentando guardar archivo: {filename} en {perfiles_dir}")
            try:
                saved_name = fs.save(filename, foto)
                print(f"‚úÖ Archivo guardado exitosamente: {saved_name}")
                print(f"üìÇ Ruta completa del archivo guardado: {os.path.join(perfiles_dir, saved_name)}")
                print(f"üîç Verificando si el archivo existe: {os.path.exists(os.path.join(perfiles_dir, saved_name))}")
            except Exception as save_error:
                import traceback
                print(f"‚ùå Error al guardar archivo: {str(save_error)}")
                print(f"üìã Traceback:\n{traceback.format_exc()}")
                raise
            file_url = f"/media/perfiles_img/{saved_name}"
            print(f"üîó URL generada: {file_url}")
            
            # Crear o actualizar registro en BD
            print(f"üíæ Creando/actualizando registro en BD...")
            foto_perfil, created = UsuarioFotoPerfil.objects.get_or_create(
                usuario=usuario_maga,
                defaults={
                    'archivo_nombre': foto.name,
                    'archivo_tipo': foto.content_type,
                    'archivo_tamanio': foto.size,
                    'url_almacenamiento': file_url
                }
            )
            print(f"{'‚úÖ Registro creado' if created else 'üîÑ Registro actualizado'}")
            
            if not created:
                # Si ya existe, eliminar archivo anterior
                old_url = foto_perfil.url_almacenamiento
                if old_url:
                    # Usar la funci√≥n de normalizaci√≥n existente o construir la ruta correctamente
                    old_path = _normalizar_ruta_media(old_url)
                    if old_path and os.path.exists(old_path):
                        try:
                            os.remove(old_path)
                        except:
                            pass
                
                # Actualizar registro
                foto_perfil.archivo_nombre = foto.name
                foto_perfil.archivo_tipo = foto.content_type
                foto_perfil.archivo_tamanio = foto.size
                foto_perfil.url_almacenamiento = file_url
                foto_perfil.save()
            
            print(f"‚úÖ Proceso completado exitosamente. URL final: {file_url}")
            return JsonResponse({
                'success': True,
                'message': 'Foto de perfil actualizada exitosamente',
                'foto_url': file_url
            })
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"‚ùå Error al subir foto de perfil: {str(e)}")
            print(f"üìã Traceback completo:\n{error_trace}")
            return JsonResponse({
                'success': False,
                'error': f'Error al subir la foto: {str(e)}'
            }, status=500)


@require_http_methods(["GET"])
@permiso_admin_o_personal_api
def api_usuario_estadisticas(request):
    """API: Obtener estad√≠sticas del usuario actual usando SOLO eventos_cambios_colaboradores"""
    usuario_maga = get_usuario_maga(request.user)
    
    if not usuario_maga:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    
    try:
        # Obtener colaborador vinculado al usuario
        colaborador = None
        if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
            colaborador = usuario_maga.colaborador
        else:
            # Si no tiene colaborador vinculado, no hay estad√≠sticas que mostrar
            return JsonResponse({
                'success': True,
                'total_eventos': 0,
                'total_avances': 0,
                'eventos': []
            })
        
        # Obtener SOLO cambios de eventos_cambios_colaboradores para este colaborador
        cambios_colaborador = EventoCambioColaborador.objects.filter(
            colaborador=colaborador
        ).select_related('actividad', 'actividad__tipo', 'actividad__comunidad', 'actividad__comunidad__region').order_by('fecha_cambio')
        
        # Inicializar contadores
        total_eventos = set()
        total_avances = 0
        eventos_dict = {}
        
        for cambio in cambios_colaborador:
            actividad = cambio.actividad
            if not actividad or actividad.eliminado_en:
                continue
            
            actividad_id = str(actividad.id)
            total_eventos.add(actividad_id)
            total_avances += 1
            
            if actividad_id not in eventos_dict:
                eventos_dict[actividad_id] = {
                    'evento_id': actividad_id,
                    'nombre': actividad.nombre,
                    'estado': actividad.estado,
                    'tipo': actividad.tipo.nombre if actividad.tipo else '-',
                    'comunidad': actividad.comunidad.nombre if actividad.comunidad else '-',
                    'total_avances': 0,
                    'fecha_primer_avance': None,
                    'fecha_ultimo_avance': None
                }
            
            eventos_dict[actividad_id]['total_avances'] += 1
            
            fecha_cambio = cambio.fecha_cambio
            if fecha_cambio:
                fecha_str = fecha_cambio.strftime('%Y-%m-%d')
                if not eventos_dict[actividad_id]['fecha_primer_avance'] or fecha_str < eventos_dict[actividad_id]['fecha_primer_avance']:
                    eventos_dict[actividad_id]['fecha_primer_avance'] = fecha_str
                if not eventos_dict[actividad_id]['fecha_ultimo_avance'] or fecha_str > eventos_dict[actividad_id]['fecha_ultimo_avance']:
                    eventos_dict[actividad_id]['fecha_ultimo_avance'] = fecha_str
        
        # Convertir eventos_dict a lista
        eventos_list = list(eventos_dict.values())
        
        return JsonResponse({
            'success': True,
            'total_eventos': len(total_eventos),
            'total_avances': total_avances,
            'eventos': eventos_list
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener estad√≠sticas: {str(e)}'
        }, status=500)


def api_regiones(request):
    """API: Listar todas las regiones con informaci√≥n b√°sica"""
    regiones_query = Region.objects.annotate(
        num_comunidades=Count('comunidades', filter=Q(comunidades__activo=True))
    ).order_by('codigo')
    
    regiones = []
    for region in regiones_query:
        # Obtener primera imagen de la galer√≠a si existe
        primera_imagen = None
        try:
            galeria = RegionGaleria.objects.filter(region=region).order_by('-creado_en').first()
            if galeria:
                primera_imagen = galeria.url_almacenamiento
        except:
            pass
        
        regiones.append({
            'id': str(region.id),
            'codigo': region.codigo,
            'nombre': region.nombre,
            'descripcion': region.descripcion,
            'comunidad_sede': region.comunidad_sede or '',
            'poblacion_aprox': region.poblacion_aprox,
            'num_comunidades': region.num_comunidades,
            'imagen_url': primera_imagen or 'https://images.unsplash.com/photo-1506905925346-21bda4d32df4?ixlib=rb-4.0.3&auto=format&fit=crop&w=400&q=80',
            'actualizado_en': region.actualizado_en.isoformat() if region.actualizado_en else None,
            'creado_en': region.creado_en.isoformat() if region.creado_en else None
        })
    
    response = JsonResponse(regiones, safe=False)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@api_ratelimit_read(rate='30/m')  # 30 peticiones por minuto
def api_regiones_recientes(request):
    """API: Obtener las √∫ltimas regiones actualizadas (OPTIMIZADO con Prefetch)
    Rate limit: 30 peticiones por minuto"""
    limite = int(request.GET.get('limite', 2))
    
    # Optimizaci√≥n: Usar Prefetch para evitar N+1 queries
    # IMPORTANTE: No usar slice [:1] dentro de Prefetch, causa error con filtros
    regiones_query = Region.objects.annotate(
        num_comunidades=Count('comunidades', filter=Q(comunidades__activo=True))
    ).prefetch_related(
        Prefetch(
            'galeria',
            queryset=RegionGaleria.objects.order_by('-creado_en')
        )
    ).order_by('-actualizado_en', '-creado_en')[:limite]
    
    regiones = []
    for region in regiones_query:
        primera_imagen = None
        try:
            # Tomar solo la primera imagen del prefetch
            galeria_list = list(region.galeria.all())
            if galeria_list:
                primera_imagen = galeria_list[0].url_almacenamiento
        except:
            pass
        
        regiones.append({
            'id': str(region.id),
            'codigo': region.codigo,
            'nombre': region.nombre,
            'descripcion': region.descripcion,
            'comunidad_sede': region.comunidad_sede or '',
            'poblacion_aprox': region.poblacion_aprox,
            'num_comunidades': region.num_comunidades,
            'imagen_url': primera_imagen or 'https://images.unsplash.com/photo-1506905925346-21bda4d32df4?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80',
            'actualizado_en': region.actualizado_en.isoformat() if region.actualizado_en else None
        })
    
    response = JsonResponse(regiones, safe=False)
    # Permitir cach√© de 5 minutos para regiones recientes
    response['Cache-Control'] = 'public, max-age=300'
    return response
def api_region_detalle(request, region_id):
    """API: Obtener detalle completo de una regi√≥n"""
    try:
        region = Region.objects.prefetch_related(
            'galeria',
            'archivos',
            'comunidades'
        ).annotate(
            num_comunidades=Count('comunidades', filter=Q(comunidades__activo=True)),
            num_actividades=Count('comunidades__actividades', filter=Q(comunidades__actividades__eliminado_en__isnull=True))
        ).get(id=region_id)
    except Region.DoesNotExist:
        return JsonResponse({'error': 'Regi√≥n no encontrada'}, status=404)
    
    # Obtener galer√≠a de im√°genes
    fotos = []
    for img in region.galeria.order_by('-creado_en').all():
        fotos.append({
            'id': str(img.id),
            'url': img.url_almacenamiento,
            'description': img.descripcion or 'Imagen de la regi√≥n'
        })
    
    # No agregar imagen por defecto si no hay im√°genes
    
    # Obtener archivos
    archivos = []
    for archivo in region.archivos.order_by('-creado_en').all():
        archivos.append({
            'id': str(archivo.id),
            'name': archivo.nombre_archivo,
            'description': archivo.descripcion or '',
            'type': archivo.archivo_tipo or 'pdf',
            'url': archivo.url_almacenamiento,
            'date': archivo.creado_en.isoformat() if archivo.creado_en else None
        })
    
    # Obtener comunidades
    comunidades_list = []
    for comunidad in region.comunidades.filter(activo=True):
        comunidades_list.append({
            'name': comunidad.nombre,
            'type': comunidad.tipo.get_nombre_display() if comunidad.tipo else ''
        })
    
    # Obtener proyectos activos de la regi√≥n
    proyectos = []
    actividades = (
        Actividad.objects.filter(eliminado_en__isnull=True)
        .filter(
            Q(comunidad__region=region) |
            Q(comunidades_relacionadas__region=region)
        )
        .select_related('tipo')
        .prefetch_related('comunidades_relacionadas')
        .order_by('-fecha')
        .distinct()[:10]
    )
    
    for actividad in actividades:
        proyectos.append({
            'name': actividad.nombre,
            'type': actividad.tipo.nombre if actividad.tipo else 'Actividad',
            'status': actividad.get_estado_display()
        })
    
    # Construir datos generales
    data = []
    if region.num_comunidades:
        data.append({
            'icon': 'üèòÔ∏è',
            'label': 'N√∫mero de Comunidades',
            'value': f'{region.num_comunidades} comunidades'
        })
    if region.poblacion_aprox:
        data.append({
            'icon': 'üë•',
            'label': 'Poblaci√≥n Aproximada',
            'value': f'{region.poblacion_aprox:,} habitantes'
        })
    if region.comunidad_sede:
        data.append({
            'icon': 'üèõÔ∏è',
            'label': 'Comunidad Sede',
            'value': region.comunidad_sede
        })
    
    response = JsonResponse({
        'id': str(region.id),
        'codigo': region.codigo,
        'nombre': region.nombre,
        'descripcion': region.descripcion or '',
        'comunidad_sede': region.comunidad_sede or '',
        'poblacion_aprox': region.poblacion_aprox,
        'num_comunidades': region.num_comunidades,
        'num_actividades': region.num_actividades,
        'latitud': float(region.latitud) if region.latitud else None,
        'longitud': float(region.longitud) if region.longitud else None,
        'photos': fotos,
        'data': data,
        'projects': proyectos,
        'communities': comunidades_list,
        'files': archivos,
        'location': f'Regi√≥n {region.codigo} - {region.nombre}',
        'actualizado_en': region.actualizado_en.isoformat() if region.actualizado_en else None,
        'creado_en': region.creado_en.isoformat() if region.creado_en else None
    })
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def _normalize_label(value):
    if not value:
        return ''
    normalized = unicodedata.normalize('NFKD', value)
    normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return normalized.strip().lower()


_COMUNIDAD_TITULOS_PREDEFINIDOS = {
    'poblacion',
    'poblaci√≥n',
    'coordenadas',
    'cocode',
    'telefono cocode',
    'tel√©fono cocode',
    'tipo de comunidad',
}
def _serialize_comunidad_detalle(comunidad, request=None):
    """Devuelve un diccionario serializado con la informaci√≥n completa de la comunidad."""
    galeria_prefetch = getattr(comunidad, 'galeria_api', None)
    if galeria_prefetch is None:
        galeria_prefetch = comunidad.galeria.order_by('-creado_en').all()

    fotos = [
        {
            'id': str(img.id),
            'url': img.url_almacenamiento,
            'description': img.descripcion or 'Imagen de la comunidad',
        }
        for img in galeria_prefetch
    ]

    cover_image = fotos[0]['url'] if fotos else DEFAULT_COMUNIDAD_IMAGE_LARGE

    archivos_prefetch = getattr(comunidad, 'archivos_api', None)
    if archivos_prefetch is None:
        archivos_prefetch = comunidad.archivos.order_by('-creado_en').all()

    archivos = [
        {
            'id': str(archivo.id),
            'name': archivo.nombre_archivo,
            'description': archivo.descripcion or '',
            'type': archivo.archivo_tipo or 'archivo',
            'url': archivo.url_almacenamiento,
            'date': archivo.creado_en.isoformat() if archivo.creado_en else None,
        }
        for archivo in archivos_prefetch
    ]

    autoridades_prefetch = getattr(comunidad, 'autoridades_api', None)
    if autoridades_prefetch is None:
        autoridades_prefetch = comunidad.autoridades.filter(activo=True).order_by('nombre')

    autoridades = [
        {
            'name': autoridad.nombre,
            'role': autoridad.rol,
            'phone': autoridad.telefono or '',
        }
        for autoridad in autoridades_prefetch
    ]

    actividades = (
        Actividad.objects.filter(eliminado_en__isnull=True)
        .filter(
            Q(comunidad=comunidad) |
            Q(comunidades_relacionadas__comunidad=comunidad)
        )
        .select_related('tipo')
        .prefetch_related('comunidades_relacionadas__comunidad')
        .order_by('-fecha')
        .distinct()[:10]
    )

    proyectos = [
        {
            'name': actividad.nombre,
            'type': actividad.tipo.nombre if actividad.tipo else 'Actividad',
            'status': actividad.get_estado_display(),
            'date': actividad.fecha.isoformat() if actividad.fecha else None,
        }
        for actividad in actividades
    ]

    poblacion_texto = f'{comunidad.poblacion:,} habitantes' if comunidad.poblacion is not None else ''

    if comunidad.latitud is not None and comunidad.longitud is not None:
        coordenadas_texto = f'{comunidad.latitud}, {comunidad.longitud}'
    else:
        coordenadas_texto = ''

    # Verificar si el usuario est√° autenticado
    usuario_autenticado = request and request.user.is_authenticated if request else False

    data_cards = [
        {
            'icon': 'üë•',
            'label': 'Poblaci√≥n',
            'value': poblacion_texto,
            'is_default': True,
            'has_value': bool(poblacion_texto),
        },
        {
            'icon': 'üß≠',
            'label': 'Coordenadas',
            'value': coordenadas_texto,
            'is_default': True,
            'has_value': bool(coordenadas_texto),
        },
        {
            'icon': 'üë§',
            'label': 'COCODE',
            'value': comunidad.cocode or '',
            'is_default': True,
            'has_value': bool(comunidad.cocode),
        },
    ]
    
    # Solo incluir la tarjeta de Tel√©fono COCODE si el usuario est√° autenticado
    if usuario_autenticado:
        data_cards.append({
            'icon': 'üìû',
            'label': 'Tel√©fono COCODE',
            'value': comunidad.telefono_cocode or '',
            'is_default': True,
            'has_value': bool(comunidad.telefono_cocode),
        })
    
    data_cards.append({
        'icon': 'üèòÔ∏è',
        'label': 'Tipo de Comunidad',
        'value': comunidad.tipo.get_nombre_display() if comunidad.tipo else '',
        'is_default': True,
        'has_value': bool(comunidad.tipo),
    })

    tarjetas_custom = []
    tarjetas_qs = (
        TarjetaDato.objects.filter(entidad_tipo='comunidad', entidad_id=comunidad.id)
        .order_by('orden', 'creado_en')
        .distinct()
    )

    ids_vistos = set()
    for tarjeta in tarjetas_qs:
        tarjeta_id = str(tarjeta.id)
        if tarjeta_id in ids_vistos:
            continue
        ids_vistos.add(tarjeta_id)

        titulo_norm = _normalize_label(tarjeta.titulo)
        if titulo_norm in _COMUNIDAD_TITULOS_PREDEFINIDOS:
            continue

        tarjetas_custom.append(
            {
                'id': tarjeta_id,
                'title': tarjeta.titulo,
                'value': tarjeta.valor or '',
                'icon': tarjeta.icono or '',
                'order': tarjeta.orden,
            }
        )

    tarjetas_custom.sort(key=lambda card: (card.get('order') or 0, card['title'].lower()))

    coordinates = None
    if comunidad.latitud is not None and comunidad.longitud is not None:
        coordinates = f'{comunidad.latitud}, {comunidad.longitud}'

    location_text = ''
    if comunidad.region:
        location_text = f'Regi√≥n {comunidad.region.codigo} ‚Ä¢ {comunidad.region.nombre}'

    return {
        'id': str(comunidad.id),
        'codigo': comunidad.codigo,
        'nombre': comunidad.nombre,
        'descripcion': comunidad.descripcion or '',
        'region': {
            'id': str(comunidad.region.id) if comunidad.region else None,
            'codigo': comunidad.region.codigo if comunidad.region else None,
            'nombre': comunidad.region.nombre if comunidad.region else None,
        } if comunidad.region else None,
        'tipo': comunidad.tipo.get_nombre_display() if comunidad.tipo else '',
        'poblacion': comunidad.poblacion,
        'latitud': float(comunidad.latitud) if comunidad.latitud is not None else None,
        'longitud': float(comunidad.longitud) if comunidad.longitud is not None else None,
        'coordinates': coordinates,
        'cocode': comunidad.cocode or '',
        'telefono_cocode': comunidad.telefono_cocode or '',
        'photos': fotos,
        'files': archivos,
        'projects': proyectos,
        'autoridades': autoridades,
        'data': data_cards,
        'custom_cards': tarjetas_custom,
        'cover_image': cover_image,
        'location': location_text,
        'actualizado_en': comunidad.actualizado_en.isoformat() if comunidad.actualizado_en else None,
        'creado_en': comunidad.creado_en.isoformat() if comunidad.creado_en else None,
    }
def api_comunidad_detalle(request, comunidad_id):
    """API: Obtener detalle completo de una comunidad"""
    try:
        comunidad = (
            Comunidad.objects.select_related('region', 'tipo')
            .prefetch_related(
                Prefetch(
                    'galeria',
                    queryset=ComunidadGaleria.objects.order_by('-creado_en'),
                    to_attr='galeria_api',
                ),
                Prefetch(
                    'archivos',
                    queryset=ComunidadArchivo.objects.order_by('-creado_en'),
                    to_attr='archivos_api',
                ),
                Prefetch(
                    'autoridades',
                    queryset=ComunidadAutoridad.objects.filter(activo=True).order_by('nombre'),
                    to_attr='autoridades_api',
                ),
            )
            .get(id=comunidad_id, activo=True)
        )
    except Comunidad.DoesNotExist:
        return JsonResponse({'error': 'Comunidad no encontrada'}, status=404)

    payload = _serialize_comunidad_detalle(comunidad, request)

    response = JsonResponse(payload)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response
@require_http_methods(["POST"])
@permiso_admin_o_personal_api
def api_actualizar_region_descripcion(request, region_id):
    """API: Actualizar la descripci√≥n de una regi√≥n"""
    try:
        region = Region.objects.get(id=region_id)
    except Region.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Regi√≥n no encontrada'
        }, status=404)

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inv√°lidos'
        }, status=400)

    descripcion = (payload.get('descripcion') or '').strip()

    if not descripcion:
        return JsonResponse({
            'success': False,
            'error': 'La descripci√≥n es requerida'
        }, status=400)

    region.descripcion = descripcion
    region.actualizado_en = timezone.now()
    region.save(update_fields=['descripcion', 'actualizado_en'])

    return JsonResponse({
        'success': True,
        'message': 'Descripci√≥n actualizada exitosamente',
        'descripcion': region.descripcion
    })


@require_http_methods(["POST"])
def api_agregar_imagen_region(request, region_id):
    """API: Agregar imagen a la galer√≠a de una regi√≥n"""
    try:
        region = Region.objects.get(id=region_id)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        # Validar que se haya enviado una imagen
        imagen = request.FILES.get('imagen')
        if not imagen:
            return JsonResponse({
                'success': False,
                'error': 'No se ha enviado ninguna imagen'
            }, status=400)
        
        # Validar que sea una imagen
        if not imagen.content_type or not imagen.content_type.startswith('image/'):
            return JsonResponse({
                'success': False,
                'error': 'El archivo debe ser una imagen (JPG, PNG, GIF, etc.)'
            }, status=400)
        
        # COMPRIMIR IMAGEN autom√°ticamente
        if getattr(settings, 'COMPRESS_IMAGES', True):
            from .image_compression import compress_if_needed, get_compression_settings
            compression_settings = get_compression_settings('gallery')
            imagen = compress_if_needed(imagen, **compression_settings)
        
        # Obtener descripci√≥n (opcional)
        descripcion = request.POST.get('descripcion', '').strip()
        
        # Crear carpeta si no existe
        portada_dir = os.path.join(str(settings.MEDIA_ROOT), 'regiones_portada_img')
        os.makedirs(portada_dir, exist_ok=True)
        
        print(f"üì§ Subiendo portada a regi√≥n {region_id}")
        print(f"üìÅ Directorio: {portada_dir}")
        print(f"‚úÖ Archivo recibido: {imagen.name}, tama√±o: {imagen.size}, tipo: {imagen.content_type}")
        
        # Verificar permisos de escritura
        if not os.access(portada_dir, os.W_OK):
            print(f"‚ùå No hay permisos de escritura en {portada_dir}")
            return JsonResponse({
                'success': False,
                'error': f'No se tienen permisos de escritura en {portada_dir}'
            }, status=500)
        
        # Guardar archivo
        fs = FileSystemStorage(location=portada_dir)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
        file_extension = os.path.splitext(imagen.name)[1]
        filename = f"{timestamp}_{region_id}{file_extension}"
        print(f"üíæ Intentando guardar archivo: {filename}")
        try:
            saved_name = fs.save(filename, imagen)
            print(f"‚úÖ Archivo guardado exitosamente: {saved_name}")
        except Exception as e:
            import traceback
            print(f"‚ùå Error al guardar archivo: {str(e)}")
            print(f"üìã Traceback:\n{traceback.format_exc()}")
            raise
        file_url = f"/media/regiones_portada_img/{saved_name}"
        
        # Crear registro en la BD usando RegionGaleria
        imagen_galeria = RegionGaleria.objects.create(
            region=region,
            archivo_nombre=imagen.name,
            url_almacenamiento=file_url,
            descripcion=descripcion,
            creado_por=usuario_maga
        )
        
        # Actualizar timestamp de la regi√≥n para reflejar el cambio
        region.actualizado_en = timezone.now()
        region.save(update_fields=['actualizado_en'])
        
        return JsonResponse({
            'success': True,
            'message': 'Imagen agregada exitosamente',
            'imagen': {
                'id': str(imagen_galeria.id),
                'url': file_url,
                'nombre': imagen.name,
                'descripcion': descripcion
            }
        })
        
    except Region.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Regi√≥n no encontrada'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al agregar imagen: {str(e)}'
        }, status=500)
@require_http_methods(["DELETE"])
@permiso_admin_o_personal_api
def api_eliminar_imagen_region(request, region_id, imagen_id):
    try:
        imagen = RegionGaleria.objects.get(id=imagen_id, region_id=region_id)
    except RegionGaleria.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Imagen no encontrada',
            },
            status=404,
        )

    ruta = imagen.url_almacenamiento or ''
    ruta_media = (settings.MEDIA_URL or '/media/').rstrip('/')
    if ruta.startswith(ruta_media):
        relative_path = ruta[len(ruta_media):].lstrip('/')
    elif ruta.startswith('/'):
        ruta_sin_slash = ruta.lstrip('/')
        media_prefix = ruta_media.lstrip('/')
        if ruta_sin_slash.startswith(media_prefix):
            relative_path = ruta_sin_slash[len(media_prefix):].lstrip('/')
        else:
            relative_path = ruta_sin_slash
    else:
        relative_path = ruta

    if relative_path:
        file_path = os.path.join(str(settings.MEDIA_ROOT), relative_path.replace('/', os.sep))
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                pass

    imagen.delete()

    region = imagen.region
    if region:
        region.actualizado_en = timezone.now()
        region.save(update_fields=['actualizado_en'])

    return JsonResponse(
        {
            'success': True,
            'message': 'Imagen eliminada correctamente',
        }
    )
@require_http_methods(["POST"])
@permiso_admin_o_personal_api
def api_agregar_archivo_region(request, region_id):
    """API: Agregar un archivo a la regi√≥n y almacenarlo en media/regiones_archivos"""
    try:
        region = Region.objects.get(id=region_id)
    except Region.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Regi√≥n no encontrada'
        }, status=404)

    usuario_maga = get_usuario_maga(request.user)
    if not usuario_maga:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no autenticado'
        }, status=401)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({
            'success': False,
            'error': 'No se ha enviado ning√∫n archivo'
        }, status=400)

    nombre = (request.POST.get('nombre') or '').strip()
    descripcion = (request.POST.get('descripcion') or '').strip()

    # Determinar extensi√≥n y nombre visible
    extension = os.path.splitext(archivo.name)[1].lower()
    if not nombre:
        nombre = os.path.splitext(archivo.name)[0]
    archivo_tipo = extension.replace('.', '') if extension else (archivo.content_type or '')

    # Preparar almacenamiento
    archivos_dir = os.path.join(str(settings.MEDIA_ROOT), 'regiones_archivos')
    os.makedirs(archivos_dir, exist_ok=True)

    fs = FileSystemStorage(location=archivos_dir)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
    safe_extension = extension or ''
    filename = f"{timestamp}_{region_id}{safe_extension}"
    saved_name = fs.save(filename, archivo)

    relative_path = f"regiones_archivos/{saved_name}"
    media_url = (settings.MEDIA_URL or '/media/').rstrip('/')
    file_url = f"{media_url}/{relative_path}"
    if not file_url.startswith('/'):
        file_url = f"/{file_url}"

    region_archivo = RegionArchivo.objects.create(
        region=region,
        nombre_archivo=nombre,
        archivo_tipo=archivo_tipo,
        url_almacenamiento=file_url,
        descripcion=descripcion,
        creado_por=usuario_maga
    )

    return JsonResponse({
        'success': True,
        'message': 'Archivo agregado exitosamente',
        'archivo': {
            'id': str(region_archivo.id),
            'name': nombre,
            'description': descripcion,
            'type': archivo_tipo or 'archivo',
            'url': file_url,
            'date': region_archivo.creado_en.isoformat() if region_archivo.creado_en else None
        }
    })
@require_http_methods(["DELETE"])
@permiso_admin_o_personal_api
def api_eliminar_archivo_region(request, region_id, archivo_id):
    """API: Eliminar un archivo de la regi√≥n y removerlo del sistema de archivos"""
    try:
        archivo = RegionArchivo.objects.get(id=archivo_id, region_id=region_id)
    except RegionArchivo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Archivo no encontrado'
        }, status=404)

    ruta_url = archivo.url_almacenamiento or ''
    media_url = (settings.MEDIA_URL or '/media/').rstrip('/')
    relative_path = ''

    if ruta_url.startswith(media_url):
        relative_path = ruta_url[len(media_url):].lstrip('/')
    elif ruta_url.startswith('/'):
        ruta_sin_slash = ruta_url.lstrip('/')
        media_prefix = media_url.lstrip('/')
        if ruta_sin_slash.startswith(media_prefix):
            relative_path = ruta_sin_slash[len(media_prefix):].lstrip('/')
        else:
            relative_path = ruta_sin_slash
    else:
        relative_path = ruta_url

    if relative_path:
        file_path = os.path.join(str(settings.MEDIA_ROOT), relative_path.replace('/', os.sep))
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                # Si no se puede eliminar, continuar con la eliminaci√≥n l√≥gica
                pass

    archivo.delete()

    return JsonResponse({
        'success': True,
        'message': 'Archivo eliminado correctamente'
    })


def api_tipos_actividad(request):
    """API: Listar todos los tipos de actividad activos"""
    tipos = TipoActividad.objects.filter(activo=True).values('id', 'nombre', 'descripcion').order_by('nombre')
    return JsonResponse(list(tipos), safe=False)


def api_comunidades(request):
    """API: Listar todas las comunidades"""
    region_id = request.GET.get('region_id')

    comunidades_query = (
        Comunidad.objects.filter(activo=True)
        .select_related('tipo', 'region')
        .annotate(
            num_actividades=Count(
                'actividades',
                filter=Q(actividades__eliminado_en__isnull=True),
                distinct=True,
            )
        )
        .prefetch_related(
            Prefetch(
                'galeria',
                queryset=ComunidadGaleria.objects.order_by('-creado_en'),
                to_attr='galeria_api',
            )
        )
        .order_by('region__codigo', 'nombre')
    )

    if region_id:
        comunidades_query = comunidades_query.filter(region_id=region_id)

    comunidades = []
    for com in comunidades_query:
        primera_imagen = None
        galeria = getattr(com, 'galeria_api', None)
        if galeria:
            primera_imagen = galeria[0].url_almacenamiento

        comunidades.append({
            'id': str(com.id),
            'codigo': com.codigo,
            'nombre': com.nombre,
            'tipo': com.tipo.get_nombre_display() if com.tipo else None,
            'region': {
                'id': str(com.region.id) if com.region else None,
                'codigo': com.region.codigo if com.region else None,
                'nombre': com.region.nombre if com.region else None
            } if com.region else None,
            'poblacion': com.poblacion,
            'cocode': com.cocode,
            'telefono_cocode': com.telefono_cocode,
            'has_projects': com.num_actividades > 0,
            'num_actividades': com.num_actividades,
            'actualizado_en': com.actualizado_en.isoformat() if com.actualizado_en else None,
            'creado_en': com.creado_en.isoformat() if com.creado_en else None,
            'imagen_url': primera_imagen or DEFAULT_COMUNIDAD_IMAGE_SMALL,
        })

    response = JsonResponse(comunidades, safe=False)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def api_actividades(request):
    """API: Listar actividades"""
    tipo_id = request.GET.get('tipo_id')
    estado = request.GET.get('estado')
    
    actividades_query = Actividad.objects.filter(
        eliminado_en__isnull=True
    ).select_related('tipo', 'comunidad', 'responsable')
    
    if tipo_id:
        actividades_query = actividades_query.filter(tipo_id=tipo_id)
    
    if estado:
        actividades_query = actividades_query.filter(estado=estado)
    
    actividades = []
    for act in actividades_query.order_by('-fecha')[:50]:  # Limitar a 50 m√°s recientes
        actividades.append({
            'id': str(act.id),
            'nombre': act.nombre,
            'fecha': act.fecha.isoformat(),
            'descripcion': act.descripcion,
            'estado': act.estado,
            'tipo': act.tipo.nombre if act.tipo else None,
            'comunidad': {
                'nombre': act.comunidad.nombre,
                'codigo': act.comunidad.codigo
            } if act.comunidad else None,
            'responsable': (act.responsable.nombre if act.responsable.nombre else act.responsable.username) if act.responsable else None
        })
    
    return JsonResponse(actividades, safe=False)
# =====================================================
# VISTAS DE GESTI√ìN DE EVENTOS
# =====================================================
@permiso_gestionar_eventos
@require_http_methods(["POST"])
@api_ratelimit_bulk(rate='15/m')  # 15 creaciones por minuto
def api_crear_evento(request):
    """
    API: Crear un nuevo evento/actividad
    Maneja el formulario con archivos, personal y beneficiarios
    Rate limit: 15 peticiones por minuto (puede incluir 50+ evidencias y 100+ comunidades por petici√≥n)
    """
    usuario_maga = get_usuario_maga(request.user)
    
    if not usuario_maga:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no autenticado'
        }, status=401)
    
    try:
        # Procesar los datos del formulario
        data = request.POST
        
        # Validar campos requeridos
        campos_requeridos = ['nombre', 'tipo_actividad_id', 'fecha', 'descripcion']
        for campo in campos_requeridos:
            if not data.get(campo):
                return JsonResponse({
                    'success': False,
                    'error': f'El campo {campo} es requerido'
                }, status=400)
        
        # Validar que no exista otro evento con el mismo nombre
        nombre_evento = data.get('nombre', '').strip()
        if nombre_evento:
            evento_existente = Actividad.objects.filter(
                nombre__iexact=nombre_evento,
                eliminado_en__isnull=True
            ).exclude(id=None).first()
            if evento_existente:
                return JsonResponse({
                    'success': False,
                    'error': f'Ya existe un evento con el nombre "{nombre_evento}". Por favor, elige un nombre diferente.'
                }, status=400)
        
        # Comunidades seleccionadas (principal + adicionales)
        comunidades_payload = []
        if data.get('comunidades_seleccionadas'):
            try:
                comunidades_payload = json.loads(data['comunidades_seleccionadas'])
            except (json.JSONDecodeError, ValueError):
                comunidades_payload = []

        comunidad_principal_id = data.get('comunidad_id')
        if not comunidad_principal_id and comunidades_payload:
            comunidad_principal_id = comunidades_payload[0].get('comunidad_id')

        portada_eliminar_flag = data.get('portada_eliminar') == 'true'

        if not comunidad_principal_id:
            return JsonResponse({
                'success': False,
                'error': 'Debe seleccionar al menos una comunidad asociada al evento'
            }, status=400)
        
        # Usar transacci√≥n para asegurar integridad de datos
        with transaction.atomic():
            # 1. Crear la actividad
            actividad = Actividad.objects.create(
                nombre=data['nombre'],
                tipo_id=data['tipo_actividad_id'],
                comunidad_id=comunidad_principal_id,
                responsable=usuario_maga,
                fecha=data['fecha'],
                descripcion=data['descripcion'],
                estado=data.get('estado', 'planificado')
            )
            
            # 2. Asignar personal
            if data.get('personal_ids'):
                try:
                    personal_raw = json.loads(data['personal_ids'])
                except (json.JSONDecodeError, ValueError):
                    personal_raw = []

                for item in personal_raw:
                    if isinstance(item, dict):
                        obj_id = item.get('id')
                        tipo = item.get('tipo', 'colaborador')
                        rol = item.get('rol', 'Colaborador')
                    else:
                        obj_id = item
                        tipo = 'colaborador'
                        rol = 'Colaborador'

                    if not obj_id:
                        continue

                    if tipo == 'usuario':
                        ActividadPersonal.objects.create(
                            actividad=actividad,
                            usuario_id=obj_id,
                            rol_en_actividad=rol
                        )
                    else:
                        colaborador = Colaborador.objects.filter(id=obj_id, activo=True).select_related('usuario').first()
                        if colaborador:
                            ActividadPersonal.objects.create(
                                actividad=actividad,
                                colaborador=colaborador,
                                usuario=colaborador.usuario if colaborador.usuario else None,
                                rol_en_actividad=rol
                            )
            
            # 3. Asociar comunidades seleccionadas (principal + adicionales)
            comunidades_a_registrar = comunidades_payload.copy()
            if comunidad_principal_id and not any(item.get('comunidad_id') == comunidad_principal_id for item in comunidades_a_registrar):
                comunidades_a_registrar.insert(0, {'comunidad_id': comunidad_principal_id})

            for item in comunidades_a_registrar:
                if not item.get('agregado_en'):
                    item['agregado_en'] = timezone.now().isoformat()

            if comunidades_a_registrar:
                comunidades_ids = [item.get('comunidad_id') for item in comunidades_a_registrar if item.get('comunidad_id')]
                comunidades_map = {
                    str(com.id): com for com in Comunidad.objects.filter(id__in=comunidades_ids).select_related('region')
                }

                for item in comunidades_a_registrar:
                    comunidad_id = item.get('comunidad_id')
                    if not comunidad_id:
                        continue

                    comunidad_obj = comunidades_map.get(str(comunidad_id))
                    if comunidad_obj is None:
                        raise Comunidad.DoesNotExist(f'Comunidad no v√°lida: {comunidad_id}')

                    region_id = item.get('region_id')
                    if not region_id and comunidad_obj.region_id:
                        region_id = comunidad_obj.region_id

                    agregado_en = parse_fecha_agregacion(item.get('agregado_en'))
                    if agregado_en is None:
                        agregado_en = timezone.now()

                    relacion, creada = ActividadComunidad.objects.get_or_create(
                        actividad=actividad,
                        comunidad=comunidad_obj,
                        defaults={'region_id': region_id, 'creado_en': agregado_en}
                    )
                    if not creada:
                        update_fields = []
                        if region_id and relacion.region_id != region_id:
                            relacion.region_id = region_id
                            update_fields.append('region_id')
                        if agregado_en and (not relacion.creado_en or relacion.creado_en != agregado_en):
                            relacion.creado_en = agregado_en
                            update_fields.append('creado_en')
                        if update_fields:
                            relacion.save(update_fields=update_fields)

            # Actualizar el campo actualizado_en de las comunidades y regiones asociadas
            # para que aparezcan en "√öltimas Regiones/Comunidades"
            comunidades_ids_a_actualizar = [item.get('comunidad_id') for item in comunidades_a_registrar if item.get('comunidad_id')]
            if comunidades_ids_a_actualizar:
                # Actualizar comunidades
                Comunidad.objects.filter(id__in=comunidades_ids_a_actualizar).update(actualizado_en=timezone.now())
                
                # Obtener las regiones de estas comunidades y actualizarlas tambi√©n
                regiones_ids = Comunidad.objects.filter(
                    id__in=comunidades_ids_a_actualizar,
                    region__isnull=False
                ).values_list('region_id', flat=True).distinct()
                
                if regiones_ids:
                    Region.objects.filter(id__in=regiones_ids).update(actualizado_en=timezone.now())

            tarjetas_creadas = []
            if data.get('tarjetas_datos_nuevas'):
                try:
                    tarjetas_payload = json.loads(data['tarjetas_datos_nuevas'])
                except json.JSONDecodeError:
                    tarjetas_payload = []

                for idx, tarjeta in enumerate(tarjetas_payload):
                    titulo = tarjeta.get('titulo')
                    valor = tarjeta.get('valor')
                    icono = tarjeta.get('icono')
                    if not titulo or not valor:
                        continue
                    
                    tarjeta_inst = TarjetaDato.objects.create(
                        entidad_tipo='actividad',
                        entidad_id=actividad.id,
                        titulo=titulo,
                        valor=valor,
                        icono=icono,
                        orden=idx
                    )
                    tarjetas_creadas.append({
                        'id': str(tarjeta_inst.id),
                        'titulo': tarjeta_inst.titulo,
                        'valor': tarjeta_inst.valor,
                        'icono': tarjeta_inst.icono,
                        'orden': tarjeta_inst.orden
                    })

            # 4. Crear y asignar beneficiarios nuevos
            if data.get('beneficiarios_nuevos'):
                try:
                    beneficiarios_nuevos = json.loads(data['beneficiarios_nuevos'])
                    print(f"üìã Beneficiarios nuevos a crear: {len(beneficiarios_nuevos)}")
                    
                    for benef_data in beneficiarios_nuevos:
                        tipo = benef_data.get('tipo')
                        print(f"üîç Procesando beneficiario tipo: {tipo}")
                        print(f"üîç Datos completos: {benef_data}")
                        comunidad_id = benef_data.get('comunidad_id') or data.get('comunidad_id')
                        
                        # Obtener el tipo de beneficiario (mapear "otro" a "institucion")
                        tipo_lookup = tipo if tipo != 'otro' else 'instituci√≥n'
                        tipo_benef = TipoBeneficiario.objects.get(nombre=tipo_lookup)

                        # Validaciones de DPI antes de crear registros
                        dpi_individual = dpi_jefe = dpi_rep = ''
                        if tipo == 'individual':
                            dpi_individual = normalizar_dpi(benef_data.get('dpi'))
                            if dpi_individual:
                                conflicto = buscar_conflicto_dpi(
                                    BeneficiarioIndividual.objects.all(),
                                    'dpi',
                                    dpi_individual,
                                )
                                if conflicto:
                                    raise ValueError(f'Ya existe un beneficiario individual con el DPI {dpi_individual}.')
                        elif tipo == 'familia':
                            dpi_jefe = normalizar_dpi(benef_data.get('dpi_jefe_familia'))
                            if dpi_jefe:
                                conflicto = buscar_conflicto_dpi(
                                    BeneficiarioFamilia.objects.all(),
                                    'dpi_jefe_familia',
                                    dpi_jefe,
                                )
                                if conflicto:
                                    raise ValueError(f'Ya existe una familia con el DPI del jefe {dpi_jefe}.')
                        elif tipo == 'institucion' or tipo == 'instituci√≥n':
                            dpi_rep = normalizar_dpi(benef_data.get('dpi_representante'))
                            if dpi_rep:
                                conflicto = buscar_conflicto_dpi(
                                    BeneficiarioInstitucion.objects.all(),
                                    'dpi_representante',
                                    dpi_rep,
                                )
                                if conflicto:
                                    raise ValueError(f'Ya existe una instituci√≥n con el DPI del representante {dpi_rep}.')

                        # Crear beneficiario principal
                        beneficiario = Beneficiario.objects.create(
                            tipo=tipo_benef,
                            comunidad_id=comunidad_id,
                            activo=True
                        )
                        
                        # Crear registro espec√≠fico seg√∫n tipo
                        if tipo == 'individual':
                            BeneficiarioIndividual.objects.create(
                                beneficiario=beneficiario,
                                nombre=benef_data.get('nombre', ''),
                                apellido=benef_data.get('apellido', ''),
                                dpi=dpi_individual or None,
                                fecha_nacimiento=benef_data.get('fecha_nacimiento'),
                                genero=benef_data.get('genero'),
                                telefono=benef_data.get('telefono')
                            )
                        elif tipo == 'familia':
                            BeneficiarioFamilia.objects.create(
                                beneficiario=beneficiario,
                                nombre_familia=benef_data.get('nombre_familia', ''),
                                jefe_familia=benef_data.get('jefe_familia', ''),
                                dpi_jefe_familia=dpi_jefe or None,
                                telefono=benef_data.get('telefono'),
                                numero_miembros=benef_data.get('numero_miembros')
                            )
                        elif tipo == 'institucion' or tipo == 'instituci√≥n':
                            inst = BeneficiarioInstitucion.objects.create(
                                beneficiario=beneficiario,
                                nombre_institucion=benef_data.get('nombre_institucion', ''),
                                tipo_institucion=benef_data.get('tipo_institucion', 'otro'),
                                representante_legal=benef_data.get('representante_legal'),
                                dpi_representante=dpi_rep or None,
                                telefono=benef_data.get('telefono'),
                                email=benef_data.get('email'),
                                numero_beneficiarios_directos=benef_data.get('numero_beneficiarios_directos')
                            )
                            print(f"‚úÖ Instituci√≥n creada: {inst.nombre_institucion} (ID: {inst.id})")
                        elif tipo == 'otro':
                            # Guardar como instituci√≥n con tipo 'otro' y usar campos flexibles
                            # No validamos DPI para tipo "otro" ya que puede no tener representante legal con DPI
                            BeneficiarioInstitucion.objects.create(
                                beneficiario=beneficiario,
                                nombre_institucion=benef_data.get('nombre', ''),
                                tipo_institucion='otro',
                                representante_legal=benef_data.get('contacto'),
                                telefono=benef_data.get('telefono'),
                                email=benef_data.get('tipo_descripcion')  # Usar email para guardar el tipo_descripcion
                            )
                        
                        # Asociar beneficiario con la actividad
                        ActividadBeneficiario.objects.create(
                            actividad=actividad,
                            beneficiario=beneficiario
                        )
                        
                except (json.JSONDecodeError, TipoBeneficiario.DoesNotExist) as e:
                    print(f"Error al crear beneficiarios: {e}")
                    pass  # Ignorar si hay errores en beneficiarios
                except (ValueError, IntegrityError) as e:
                    # Retornar error de validaci√≥n (duplicado)
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    }, status=400)

            # 3b. Asociar beneficiarios existentes seleccionados
            if data.get('beneficiarios_existentes'):
                try:
                    beneficiarios_existentes_ids = json.loads(data['beneficiarios_existentes'])
                    asociados = 0
                    for benef_id in beneficiarios_existentes_ids:
                        if not benef_id:
                            continue
                        beneficiario = Beneficiario.objects.filter(id=benef_id, activo=True).first()
                        if not beneficiario:
                            continue
                        _, creado_rel = ActividadBeneficiario.objects.get_or_create(
                            actividad=actividad,
                            beneficiario=beneficiario
                        )
                        if creado_rel:
                            asociados += 1
                    if asociados:
                        print(f"‚úÖ Asociados {asociados} beneficiarios existentes al evento")
                except (json.JSONDecodeError, ValueError) as e:
                    print(f"Error al asociar beneficiarios existentes: {e}")
                    pass

            # 3c. Aplicar modificaciones a beneficiarios existentes (si se editaron desde el formulario)
            if data.get('beneficiarios_modificados'):
                try:
                    beneficiarios_modificados = json.loads(data['beneficiarios_modificados'])
                    cambios_aplicados = aplicar_modificaciones_beneficiarios(beneficiarios_modificados)
                    if cambios_aplicados:
                        print(f"‚úèÔ∏è Beneficiarios existentes actualizados: {cambios_aplicados}")
                except json.JSONDecodeError as e:
                    print(f"Error al actualizar beneficiarios: {e}")
                    pass
                except (ValueError, IntegrityError) as e:
                    # Retornar error de validaci√≥n (duplicado)
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    }, status=400)
            
            # 5. Guardar evidencias (archivos/im√°genes)
            archivos_guardados = []
            evidencias_guardadas = request.FILES.getlist('evidences')
            
            if evidencias_guardadas:
                evidencias_dir = os.path.join(str(settings.MEDIA_ROOT), 'evidencias')
                os.makedirs(evidencias_dir, exist_ok=True)
                fs = FileSystemStorage(location=evidencias_dir)
                
                for index, file in enumerate(evidencias_guardadas):
                    # COMPRIMIR IMAGEN autom√°ticamente si es una imagen
                    es_imagen = file.content_type.startswith('image/') if hasattr(file, 'content_type') else False
                    if es_imagen and getattr(settings, 'COMPRESS_IMAGES', True):
                        from .image_compression import compress_if_needed, get_compression_settings
                        compression_settings = get_compression_settings('evidence')
                        file = compress_if_needed(file, **compression_settings)
                    
                    # Generar nombre √∫nico con microsegundos para evitar duplicados
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
                    # Obtener extensi√≥n del archivo
                    file_extension = os.path.splitext(file.name)[1]
                    filename = f"{timestamp}_{index}{file_extension}"
                    
                    # Guardar archivo f√≠sicamente
                    saved_name = fs.save(filename, file)
                    file_url = f"/media/evidencias/{saved_name}"
                    
                    # Crear registro de evidencia en la BD
                    evidencia = Evidencia.objects.create(
                        actividad=actividad,
                        archivo_nombre=file.name,
                        archivo_tipo=file.content_type if hasattr(file, 'content_type') else 'application/octet-stream',
                        archivo_tamanio=file.size if hasattr(file, 'size') else 0,
                        url_almacenamiento=file_url,
                        es_imagen=es_imagen,
                        creado_por=usuario_maga
                    )
                    
                    archivos_guardados.append({
                        'id': str(evidencia.id),
                        'nombre': file.name,
                        'url': file_url,
                        'tipo': 'imagen' if es_imagen else 'archivo',
                        'tamanio': file.size
                    })
            
            # 6. Guardar imagen de portada (si se envi√≥)
            portada_info = None
            portada_file = request.FILES.get('portada_evento')
            if portada_file:
                try:
                    portada_inst = guardar_portada_evento(actividad, portada_file)
                    portada_info = obtener_portada_evento(portada_inst.actividad)
                except ValueError as portada_error:
                    raise ValueError(str(portada_error))
            else:
                portada_info = obtener_portada_evento(actividad)
            
            # 6. Registrar cambio/creaci√≥n
            ActividadCambio.objects.create(
                actividad=actividad,
                responsable=usuario_maga,
                descripcion_cambio=f'Evento "{actividad.nombre}" creado por {usuario_maga.username}'
            )
            
            # Respuesta exitosa
            comunidades_evento = obtener_comunidades_evento(actividad)
            comunidad_principal_nombre = comunidades_evento[0]['comunidad_nombre'] if comunidades_evento else (actividad.comunidad.nombre if actividad.comunidad else None)
            
            return JsonResponse({
                'success': True,
                'message': 'Evento creado exitosamente',
                'actividad': {
                    'id': str(actividad.id),
                    'nombre': actividad.nombre,
                    'tipo': actividad.tipo.nombre,
                    'fecha': str(actividad.fecha),
                    'comunidad': comunidad_principal_nombre,
                    'comunidades': comunidades_evento,
                    'estado': actividad.estado,
                    'portada': portada_info,
                    'tarjetas_datos': tarjetas_creadas or obtener_tarjetas_datos(actividad)
                },
                'archivos': archivos_guardados,
                'total_archivos': len(archivos_guardados)
            })
    
    except TipoActividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Tipo de actividad no v√°lido'
        }, status=400)
    
    except Comunidad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Comunidad no v√°lida'
        }, status=400)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear evento: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
def api_listar_personal(request):
    """API: Listar colaboradores disponibles para asignar a eventos"""

    colaboradores = (
        Colaborador.objects.filter(activo=True)
        .select_related('puesto', 'usuario')
        .order_by('nombre')
    )

    personal_list = []
    for colaborador in colaboradores:
        personal_list.append({
            'id': str(colaborador.id),
            'nombre': colaborador.nombre,
            'username': colaborador.usuario.username if colaborador.usuario else '',
            'correo': colaborador.correo,
            'telefono': colaborador.telefono,
            'rol': 'Colaborador',
            'rol_display': 'Personal Fijo' if colaborador.es_personal_fijo else 'Colaborador Externo',
            'puesto': colaborador.puesto.nombre if colaborador.puesto else 'Sin puesto asignado',
            'tipo': 'colaborador',
            'es_personal_fijo': colaborador.es_personal_fijo,
            'usuario_id': str(colaborador.usuario_id) if colaborador.usuario_id else None,
        })

    return JsonResponse(personal_list, safe=False)
@permiso_gestionar_eventos
def api_listar_beneficiarios(request):
    """API: Listar beneficiarios disponibles
    - Sin b√∫squeda: Retorna los primeros 50 (para carga inicial r√°pida)
    - Con b√∫squeda: SIN L√çMITE - busca en TODA la base de datos
    """
    search = (request.GET.get('q') or '').strip()

    beneficiarios = Beneficiario.objects.filter(activo=True).select_related(
        'tipo', 'comunidad'
    ).prefetch_related('individual', 'familia', 'institucion')

    if search:
        # Con b√∫squeda: Buscar en TODA la base de datos sin l√≠mites
        beneficiarios = beneficiarios.filter(
            Q(individual__nombre__icontains=search) |
            Q(individual__apellido__icontains=search) |
            Q(individual__dpi__icontains=search) |
            Q(familia__nombre_familia__icontains=search) |
            Q(familia__jefe_familia__icontains=search) |
            Q(familia__dpi_jefe_familia__icontains=search) |
            Q(institucion__nombre_institucion__icontains=search) |
            Q(institucion__representante_legal__icontains=search) |
            Q(institucion__tipo_institucion__icontains=search) |
            Q(comunidad__nombre__icontains=search)
        ).order_by('tipo__nombre', 'comunidad__nombre', 'id')
        # NO aplicar l√≠mite cuando hay b√∫squeda - retornar TODOS los resultados
    else:
        # Sin b√∫squeda: Solo 50 para carga inicial r√°pida
        beneficiarios = beneficiarios.order_by('tipo__nombre', 'comunidad__nombre', 'id')[:50]

    beneficiarios_list = []
    for ben in beneficiarios:
        nombre_display, info_adicional, detalles, tipo_envio = obtener_detalle_beneficiario(ben)
        if ben.tipo and hasattr(ben.tipo, 'get_nombre_display'):
            tipo_display = ben.tipo.get_nombre_display()
        else:
            tipo_display = tipo_envio.title() if tipo_envio else ''
        beneficiarios_list.append({
            'id': str(ben.id),
            'nombre': nombre_display,
            'display_name': nombre_display,
            'tipo': tipo_envio,
            'tipo_display': tipo_display,
            'comunidad_id': str(ben.comunidad_id) if ben.comunidad_id else None,
            'comunidad_nombre': ben.comunidad.nombre if ben.comunidad else None,
            'region_id': str(ben.comunidad.region_id) if ben.comunidad and ben.comunidad.region_id else None,
            'region_nombre': ben.comunidad.region.nombre if ben.comunidad and ben.comunidad.region else None,
            'region_sede': ben.comunidad.region.comunidad_sede if ben.comunidad and ben.comunidad.region and ben.comunidad.region.comunidad_sede else None,
            'comunidad': ben.comunidad.nombre if ben.comunidad else None,
            'info_adicional': info_adicional,
            'detalles': detalles
        })

    return JsonResponse(beneficiarios_list, safe=False)
@permiso_gestionar_eventos
@require_http_methods(["GET"])
def api_obtener_beneficiario(request, beneficiario_id):
    """API: Obtener detalles completos de un beneficiario por ID"""
    try:
        beneficiario = Beneficiario.objects.filter(
            id=beneficiario_id, activo=True
        ).select_related(
            'tipo', 'comunidad', 'comunidad__region'
        ).prefetch_related('individual', 'familia', 'institucion').first()
        
        if not beneficiario:
            return JsonResponse({
                'success': False,
                'error': 'Beneficiario no encontrado'
            }, status=404)
        
        nombre_display, info_adicional, detalles, tipo_envio = obtener_detalle_beneficiario(beneficiario)
        
        if beneficiario.tipo and hasattr(beneficiario.tipo, 'get_nombre_display'):
            tipo_display = beneficiario.tipo.get_nombre_display()
        else:
            tipo_display = tipo_envio.title() if tipo_envio else ''
        
        return JsonResponse({
            'success': True,
            'beneficiario': {
                'id': str(beneficiario.id),
                'nombre': nombre_display,
                'display_name': nombre_display,
                'tipo': tipo_envio,
                'tipo_display': tipo_display,
                'comunidad_id': str(beneficiario.comunidad_id) if beneficiario.comunidad_id else None,
                'comunidad_nombre': beneficiario.comunidad.nombre if beneficiario.comunidad else None,
                'region_id': str(beneficiario.comunidad.region_id) if beneficiario.comunidad and beneficiario.comunidad.region_id else None,
                'region_nombre': beneficiario.comunidad.region.nombre if beneficiario.comunidad and beneficiario.comunidad.region else None,
                'region_sede': beneficiario.comunidad.region.comunidad_sede if beneficiario.comunidad and beneficiario.comunidad.region and beneficiario.comunidad.region.comunidad_sede else None,
                'info_adicional': info_adicional,
                'detalles': detalles,
                'creado_en': beneficiario.creado_en.isoformat() if beneficiario.creado_en else None,
                'actualizado_en': beneficiario.actualizado_en.isoformat() if beneficiario.actualizado_en else None
            }
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener beneficiario: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["POST"])
@api_ratelimit_upload(rate='10/m')  # 10 importaciones por minuto
def api_importar_beneficiarios_excel(request):
    """API: Importar beneficiarios individuales desde archivo Excel
    Rate limit: 10 peticiones por minuto (solo valida, no guarda en BD)"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No se proporcion√≥ ning√∫n archivo Excel'
            }, status=400)
        
        excel_file = request.FILES['excel_file']
        
        # Validar extensi√≥n
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'error': 'El archivo debe ser un Excel (.xlsx o .xls)'
            }, status=400)
        
        # Leer el archivo Excel
        import openpyxl
        from openpyxl import load_workbook
        
        workbook = load_workbook(excel_file, data_only=True)
        sheet = workbook.active
        
        # Validar encabezados esperados (primera fila)
        expected_headers = ['tipo', 'comunidad', 'nombre', 'apellido', 'dpi', 'fecha de nacimiento', 'edad', 'genero', 'telefono']
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else '')
        
        # Verificar que los encabezados coincidan (case-insensitive)
        headers_normalized = [h.lower().strip() for h in headers]
        expected_normalized = [h.lower().strip() for h in expected_headers]
        
        if headers_normalized != expected_normalized:
            return JsonResponse({
                'success': False,
                'error': f'Los encabezados del Excel no coinciden con la plantilla esperada. Encabezados encontrados: {headers}'
            }, status=400)
        
        # Obtener tipo de beneficiario "individual"
        try:
            tipo_individual = TipoBeneficiario.objects.get(nombre='individual')
        except TipoBeneficiario.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Tipo de beneficiario "individual" no encontrado en la base de datos'
            }, status=500)
        
        # Funci√≥n para normalizar nombres (sin tildes, sin espacios, case-insensitive)
        def normalizar_nombre_comunidad(nombre):
            """Normaliza un nombre removiendo tildes, espacios y convirtiendo a min√∫sculas"""
            if not nombre:
                return ''
            # Convertir a min√∫sculas y remover espacios
            nombre = nombre.lower().strip()
            # Remover tildes y caracteres especiales
            nombre = unicodedata.normalize('NFD', nombre)
            nombre = ''.join(c for c in nombre if unicodedata.category(c) != 'Mn')
            # Remover espacios m√∫ltiples y convertir a uno solo
            nombre = ' '.join(nombre.split())
            return nombre
        
        # Obtener todas las comunidades para b√∫squeda flexible
        comunidades_db = Comunidad.objects.filter(activo=True).select_related('region')
        comunidades_dict = {}
        for com in comunidades_db:
            # Crear un diccionario con nombre normalizado como clave
            nombre_normalizado = normalizar_nombre_comunidad(com.nombre)
            if nombre_normalizado not in comunidades_dict:
                comunidades_dict[nombre_normalizado] = com
        
        # Procesar filas (empezando desde la fila 2, ya que la 1 son encabezados)
        resultados = {
            'exitosos': [],
            'advertencias': [],
            'errores': [],
            'total_procesados': 0,
            'total_exitosos': 0,
            'total_advertencias': 0,
            'total_errores': 0
        }
        
        # Procesar filas SIN guardar en BD (solo validar y preparar datos)
        # Cerrar workbook despu√©s de procesar para liberar memoria
        try:
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                    # Saltar filas vac√≠as
                    if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                        continue
                    
                    resultados['total_procesados'] += 1
                    
                    try:
                        # Extraer valores de las celdas
                        valores = [cell.value for cell in row[:9]]  # Solo las primeras 9 columnas
                        
                        tipo_val = str(valores[0]).strip().lower() if valores[0] else ''
                        comunidad_nombre = str(valores[1]).strip() if valores[1] else ''
                        nombre = str(valores[2]).strip() if valores[2] else ''
                        apellido = str(valores[3]).strip() if valores[3] else ''
                        dpi_val = str(valores[4]).strip() if valores[4] else ''
                        fecha_nac_str = valores[5]
                        edad_val = valores[6]
                        genero_val = str(valores[7]).strip().lower() if valores[7] else ''
                        telefono_val = str(valores[8]).strip() if valores[8] else ''
                        
                        # Validar tipo (debe ser "individual")
                        if tipo_val != 'individual':
                            resultados['errores'].append({
                                'fila': row_num,
                                'error': f'Tipo debe ser "individual", se encontr√≥: "{tipo_val}"'
                            })
                            resultados['total_errores'] += 1
                            continue
                        
                        # Validar campos obligatorios
                        if not nombre:
                            resultados['errores'].append({
                                'fila': row_num,
                                'error': 'El campo "nombre" es obligatorio'
                            })
                            resultados['total_errores'] += 1
                            continue
                        
                        if not apellido:
                            resultados['errores'].append({
                                'fila': row_num,
                                'error': 'El campo "apellido" es obligatorio'
                            })
                            resultados['total_errores'] += 1
                            continue
                        
                        if not comunidad_nombre:
                            resultados['errores'].append({
                                'fila': row_num,
                                'error': 'El campo "comunidad" es obligatorio'
                            })
                            resultados['total_errores'] += 1
                            continue
                        
                        # Buscar comunidad (flexible: sin tildes, sin espacios, case-insensitive)
                        comunidad_nombre_normalizado = normalizar_nombre_comunidad(comunidad_nombre)
                        comunidad_obj = comunidades_dict.get(comunidad_nombre_normalizado)
                        
                        if not comunidad_obj:
                            # Intentar b√∫squeda m√°s flexible si no se encontr√≥ exacto
                            comunidad_obj = None
                            for com in comunidades_db:
                                if normalizar_nombre_comunidad(com.nombre) == comunidad_nombre_normalizado:
                                    comunidad_obj = com
                                    break
                            
                            if not comunidad_obj:
                                resultados['errores'].append({
                                    'fila': row_num,
                                    'error': f'Comunidad "{comunidad_nombre}" no encontrada en la base de datos'
                                })
                                resultados['total_errores'] += 1
                                continue
                        
                        # Normalizar y validar DPI
                        dpi_normalizado = normalizar_dpi(dpi_val) if dpi_val else ''
                        if dpi_normalizado and len(dpi_normalizado) != 13:
                            resultados['errores'].append({
                                'fila': row_num,
                                'error': f'DPI debe tener 13 d√≠gitos, se encontr√≥: {len(dpi_normalizado)}'
                            })
                            resultados['total_errores'] += 1
                            continue
                        
                        # Procesar fecha de nacimiento
                        fecha_nacimiento = None
                        if fecha_nac_str:
                            try:
                                if isinstance(fecha_nac_str, datetime):
                                    fecha_nacimiento = fecha_nac_str.date()
                                elif isinstance(fecha_nac_str, str):
                                    # Intentar parsear diferentes formatos de fecha
                                    fecha_str = fecha_nac_str.strip()
                                    formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']
                                    for formato in formatos:
                                        try:
                                            fecha_nacimiento = datetime.strptime(fecha_str, formato).date()
                                            break
                                        except ValueError:
                                            continue
                                elif hasattr(fecha_nac_str, 'date'):
                                    # Si es un objeto date de Python
                                    fecha_nacimiento = fecha_nac_str.date() if isinstance(fecha_nac_str, datetime) else fecha_nac_str
                            except (ValueError, AttributeError, TypeError):
                                # Si no se puede parsear, se deja como None
                                pass
                        
                        # Normalizar g√©nero
                        genero_normalizado = None
                        if genero_val:
                            genero_lower = genero_val.lower().strip()
                            if genero_lower in ['masculino', 'm', 'hombre', 'h']:
                                genero_normalizado = 'masculino'
                            elif genero_lower in ['femenino', 'f', 'mujer', 'm']:
                                genero_normalizado = 'femenino'
                            elif genero_lower in ['otro', 'o', 'otra']:
                                genero_normalizado = 'otro'
                        
                        # Normalizar tel√©fono (solo n√∫meros)
                        telefono_normalizado = normalizar_dpi(telefono_val) if telefono_val else ''
                        if telefono_normalizado and len(telefono_normalizado) != 8:
                            # Si tiene tel√©fono pero no es v√°lido, se guarda como est√° pero se registra advertencia
                            pass
                        
                        # Verificar si el DPI ya existe (para actualizar en vez de crear)
                        beneficiario_existente = None
                        beneficiario_individual_existente = None
                        es_actualizacion = False
                        
                        if dpi_normalizado:
                            conflicto = buscar_conflicto_dpi(
                                BeneficiarioIndividual.objects.all(),
                                'dpi',
                                dpi_normalizado
                            )
                            if conflicto:
                                # DPI encontrado - actualizar beneficiario existente
                                beneficiario_individual_existente = conflicto
                                beneficiario_existente = conflicto.beneficiario
                                es_actualizacion = True
                        
                        # Preparar datos del beneficiario (NO guardar a√∫n)
                        datos_beneficiario = {
                            'fila': row_num,
                            'tipo': 'individual',
                            'comunidad_id': str(comunidad_obj.id),
                            'comunidad_nombre': comunidad_obj.nombre,
                            'nombre': nombre,
                            'apellido': apellido,
                            'dpi': dpi_normalizado or None,
                            'fecha_nacimiento': fecha_nacimiento.isoformat() if fecha_nacimiento else None,
                            'genero': genero_normalizado,
                            'telefono': telefono_normalizado or None,
                            'es_actualizacion': es_actualizacion,
                            'beneficiario_existente_id': str(beneficiario_existente.id) if es_actualizacion and beneficiario_existente else None
                        }
                        
                        if es_actualizacion:
                            resultados['advertencias'].append({
                                'fila': row_num,
                                'mensaje': f'DPI detectado: {dpi_normalizado}. Se actualizar√° el beneficiario con la informaci√≥n nueva.',
                                'nombre': f'{nombre} {apellido}',
                                'comunidad': comunidad_obj.nombre,
                                'dpi': dpi_normalizado,
                                'datos': datos_beneficiario
                            })
                            resultados['total_advertencias'] += 1
                        else:
                            resultados['exitosos'].append({
                                'fila': row_num,
                                'nombre': f'{nombre} {apellido}',
                                'comunidad': comunidad_obj.nombre,
                                'dpi': dpi_normalizado or 'N/A',
                                'datos': datos_beneficiario
                            })
                            resultados['total_exitosos'] += 1
                        
                    except Exception as e:
                        resultados['errores'].append({
                            'fila': row_num,
                            'error': f'Error al procesar fila: {str(e)}'
                        })
                        resultados['total_errores'] += 1
                        continue
        finally:
            # Cerrar workbook para liberar memoria
            workbook.close()
        
        # Construir mensaje final
        mensaje_partes = []
        if resultados['total_exitosos'] > 0:
            mensaje_partes.append(f"{resultados['total_exitosos']} exitosos")
        if resultados['total_advertencias'] > 0:
            mensaje_partes.append(f"{resultados['total_advertencias']} actualizados")
        if resultados['total_errores'] > 0:
            mensaje_partes.append(f"{resultados['total_errores']} errores")
        
        mensaje_final = f'Importaci√≥n completada: {", ".join(mensaje_partes)}' if mensaje_partes else 'No se procesaron registros'
        
        return JsonResponse({
            'success': True,
            'message': mensaje_final,
            'resultados': resultados,
            'pendientes': {
                'exitosos': resultados['exitosos'],
                'advertencias': resultados['advertencias']
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al procesar el archivo Excel: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["POST"])
@api_ratelimit_bulk(rate='30/m')  # 30 guardados masivos por minuto
def api_guardar_beneficiarios_general(request):
    """API: Guardar beneficiarios pendientes de Excel sin vincularlos a ning√∫n evento
    Rate limit: 30 peticiones por minuto (operaci√≥n masiva - puede procesar 1000+ beneficiarios)"""
    try:
        usuario_maga = get_usuario_maga(request.user)
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        data = _parse_request_data(request)
        beneficiarios_pendientes = data.get('beneficiarios_pendientes', [])
        
        if not beneficiarios_pendientes:
            return JsonResponse({
                'success': False,
                'error': 'No hay beneficiarios pendientes para guardar'
            }, status=400)
        
        # Obtener tipo de beneficiario "individual"
        try:
            tipo_individual = TipoBeneficiario.objects.get(nombre='individual')
        except TipoBeneficiario.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Tipo de beneficiario "individual" no encontrado'
            }, status=500)
        
        resultados = {
            'creados': 0,
            'actualizados': 0,
            'errores': []
        }
        
        with transaction.atomic():
            for benef_data in beneficiarios_pendientes:
                try:
                    comunidad_id = benef_data.get('comunidad_id')
                    nombre = benef_data.get('nombre')
                    apellido = benef_data.get('apellido')
                    dpi = benef_data.get('dpi')
                    fecha_nacimiento = benef_data.get('fecha_nacimiento')
                    genero = benef_data.get('genero')
                    telefono = benef_data.get('telefono')
                    es_actualizacion = benef_data.get('es_actualizacion', False)
                    beneficiario_existente_id = benef_data.get('beneficiario_existente_id')
                    
                    # Validar campos obligatorios
                    if not nombre or not apellido or not comunidad_id:
                        resultados['errores'].append({
                            'error': 'Faltan campos obligatorios',
                            'datos': benef_data
                        })
                        continue
                    
                    # Obtener comunidad
                    try:
                        comunidad = Comunidad.objects.get(id=comunidad_id, activo=True)
                    except Comunidad.DoesNotExist:
                        resultados['errores'].append({
                            'error': f'Comunidad {comunidad_id} no encontrada',
                            'datos': benef_data
                        })
                        continue
                    
                    # Procesar fecha de nacimiento
                    fecha_nac = None
                    if fecha_nacimiento:
                        try:
                            if isinstance(fecha_nacimiento, str):
                                fecha_nac = datetime.fromisoformat(fecha_nacimiento).date()
                            elif hasattr(fecha_nacimiento, 'date'):
                                fecha_nac = fecha_nacimiento.date() if isinstance(fecha_nacimiento, datetime) else fecha_nacimiento
                        except (ValueError, AttributeError, TypeError):
                            pass
                    
                    if es_actualizacion and beneficiario_existente_id:
                        # Actualizar beneficiario existente
                        try:
                            beneficiario = Beneficiario.objects.get(id=beneficiario_existente_id)
                            beneficiario.comunidad = comunidad
                            beneficiario.activo = True
                            beneficiario.save()
                            
                            beneficiario_individual = beneficiario.individual
                            beneficiario_individual.nombre = nombre
                            beneficiario_individual.apellido = apellido
                            if fecha_nac:
                                beneficiario_individual.fecha_nacimiento = fecha_nac
                            if genero:
                                beneficiario_individual.genero = genero
                            if telefono:
                                beneficiario_individual.telefono = telefono
                            beneficiario_individual.save()
                            
                            resultados['actualizados'] += 1
                                
                        except Beneficiario.DoesNotExist:
                            resultados['errores'].append({
                                'error': f'Beneficiario {beneficiario_existente_id} no encontrado para actualizar',
                                'datos': benef_data
                            })
                            continue
                    else:
                        # Crear nuevo beneficiario
                        beneficiario = Beneficiario.objects.create(
                            tipo=tipo_individual,
                            comunidad=comunidad,
                            activo=True
                        )
                        
                        BeneficiarioIndividual.objects.create(
                            beneficiario=beneficiario,
                            nombre=nombre,
                            apellido=apellido,
                            dpi=dpi,
                            fecha_nacimiento=fecha_nac,
                            genero=genero,
                            telefono=telefono
                        )
                        
                        resultados['creados'] += 1
                        
                except Exception as e:
                    resultados['errores'].append({
                        'error': f'Error al procesar beneficiario: {str(e)}',
                        'datos': benef_data
                    })
                    continue
        
        mensaje = f'Proceso completado: {resultados["creados"]} creados, {resultados["actualizados"]} actualizados'
        if resultados['errores']:
            mensaje += f', {len(resultados["errores"])} errores'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'resultados': resultados
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al guardar beneficiarios: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["POST"])
@api_ratelimit_bulk(rate='30/m')  # 30 guardados masivos por minuto
def api_guardar_beneficiarios_pendientes(request):
    """API: Guardar beneficiarios pendientes de Excel y vincularlos al evento
    Rate limit: 30 peticiones por minuto (operaci√≥n masiva - puede procesar 1000+ beneficiarios)"""
    try:
        usuario_maga = get_usuario_maga(request.user)
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        data = _parse_request_data(request)
        actividad_id = data.get('actividad_id')
        beneficiarios_pendientes = data.get('beneficiarios_pendientes', [])
        
        if not actividad_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de actividad no proporcionado'
            }, status=400)
        
        if not beneficiarios_pendientes:
            return JsonResponse({
                'success': False,
                'error': 'No hay beneficiarios pendientes para guardar'
            }, status=400)
        
        # Validar que la actividad existe
        try:
            actividad = Actividad.objects.get(id=actividad_id, eliminado_en__isnull=True)
        except Actividad.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Actividad no encontrada'
            }, status=404)
        
        # Obtener tipo de beneficiario "individual"
        try:
            tipo_individual = TipoBeneficiario.objects.get(nombre='individual')
        except TipoBeneficiario.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Tipo de beneficiario "individual" no encontrado'
            }, status=500)
        
        resultados = {
            'creados': 0,
            'actualizados': 0,
            'vinculados': 0,
            'errores': []
        }
        
        with transaction.atomic():
            for benef_data in beneficiarios_pendientes:
                try:
                    comunidad_id = benef_data.get('comunidad_id')
                    nombre = benef_data.get('nombre')
                    apellido = benef_data.get('apellido')
                    dpi = benef_data.get('dpi')
                    fecha_nacimiento = benef_data.get('fecha_nacimiento')
                    genero = benef_data.get('genero')
                    telefono = benef_data.get('telefono')
                    es_actualizacion = benef_data.get('es_actualizacion', False)
                    beneficiario_existente_id = benef_data.get('beneficiario_existente_id')
                    
                    # Validar campos obligatorios
                    if not nombre or not apellido or not comunidad_id:
                        resultados['errores'].append({
                            'error': 'Faltan campos obligatorios',
                            'datos': benef_data
                        })
                        continue
                    
                    # Obtener comunidad
                    try:
                        comunidad = Comunidad.objects.get(id=comunidad_id, activo=True)
                    except Comunidad.DoesNotExist:
                        resultados['errores'].append({
                            'error': f'Comunidad {comunidad_id} no encontrada',
                            'datos': benef_data
                        })
                        continue
                    
                    # Procesar fecha de nacimiento
                    fecha_nac = None
                    if fecha_nacimiento:
                        try:
                            if isinstance(fecha_nacimiento, str):
                                fecha_nac = datetime.fromisoformat(fecha_nacimiento).date()
                            elif hasattr(fecha_nacimiento, 'date'):
                                fecha_nac = fecha_nacimiento.date() if isinstance(fecha_nacimiento, datetime) else fecha_nacimiento
                        except (ValueError, AttributeError, TypeError):
                            pass
                    
                    if es_actualizacion and beneficiario_existente_id:
                        # Actualizar beneficiario existente
                        try:
                            beneficiario = Beneficiario.objects.get(id=beneficiario_existente_id)
                            beneficiario.comunidad = comunidad
                            beneficiario.activo = True
                            beneficiario.save()
                            
                            beneficiario_individual = beneficiario.individual
                            beneficiario_individual.nombre = nombre
                            beneficiario_individual.apellido = apellido
                            if fecha_nac:
                                beneficiario_individual.fecha_nacimiento = fecha_nac
                            if genero:
                                beneficiario_individual.genero = genero
                            if telefono:
                                beneficiario_individual.telefono = telefono
                            beneficiario_individual.save()
                            
                            resultados['actualizados'] += 1
                            
                            # Vincular al evento
                            _, creado = ActividadBeneficiario.objects.get_or_create(
                                actividad=actividad,
                                beneficiario=beneficiario
                            )
                            if creado:
                                resultados['vinculados'] += 1
                                
                        except Beneficiario.DoesNotExist:
                            resultados['errores'].append({
                                'error': f'Beneficiario {beneficiario_existente_id} no encontrado para actualizar',
                                'datos': benef_data
                            })
                            continue
                    else:
                        # Crear nuevo beneficiario
                        beneficiario = Beneficiario.objects.create(
                            tipo=tipo_individual,
                            comunidad=comunidad,
                            activo=True
                        )
                        
                        BeneficiarioIndividual.objects.create(
                            beneficiario=beneficiario,
                            nombre=nombre,
                            apellido=apellido,
                            dpi=dpi,
                            fecha_nacimiento=fecha_nac,
                            genero=genero,
                            telefono=telefono
                        )
                        
                        resultados['creados'] += 1
                        
                        # Vincular al evento
                        ActividadBeneficiario.objects.create(
                            actividad=actividad,
                            beneficiario=beneficiario
                        )
                        resultados['vinculados'] += 1
                        
                except Exception as e:
                    resultados['errores'].append({
                        'error': f'Error al procesar beneficiario: {str(e)}',
                        'datos': benef_data
                    })
                    continue
        
        mensaje = f'Proceso completado: {resultados["creados"]} creados, {resultados["actualizados"]} actualizados, {resultados["vinculados"]} vinculados al evento'
        if resultados['errores']:
            mensaje += f', {len(resultados["errores"])} errores'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'resultados': resultados
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al guardar beneficiarios: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["GET"])
def api_descargar_plantilla_beneficiarios(request):
    """API: Descargar plantilla Excel para importar beneficiarios individuales"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Beneficiarios Individuales"
        
        # Definir encabezados
        headers = ['tipo', 'comunidad', 'nombre', 'apellido', 'dpi', 'fecha de nacimiento', 'edad', 'genero', 'telefono']
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajustar ancho de columnas
        column_widths = [12, 25, 20, 20, 15, 18, 8, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Agregar fila de ejemplo
        ejemplo = ['individual', 'Ejemplo Comunidad', 'Juan', 'P√©rez Garc√≠a', '1234567890101', '1990-01-15', '34', 'masculino', '55123456']
        for col_num, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = valor
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Crear respuesta HTTP
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_beneficiarios_individuales.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al generar la plantilla: {str(e)}'
        }, status=500)


# =====================================================
# APIs PARA GESTI√ìN DE EVENTOS (Listar, Editar, Eliminar)
# =====================================================
@permiso_gestionar_eventos
@require_http_methods(["GET"])
def api_listar_eventos(request):
    """Lista todos los eventos no eliminados"""
    try:
        eventos = Actividad.objects.filter(
            eliminado_en__isnull=True
        ).select_related(
            'tipo', 'comunidad', 'comunidad__region', 'responsable', 'colaborador', 'portada'
        ).prefetch_related(
            'personal__usuario__puesto',
            'personal__colaborador',
            'beneficiarios__beneficiario',
            'comunidades_relacionadas__comunidad__region',
            'comunidades_relacionadas__region'
         ).order_by('-creado_en')
        
        eventos_data = []
        for evento in eventos:
            # Contar personal y beneficiarios
            personal_count = evento.personal.count()
            beneficiarios_count = evento.beneficiarios.count()
            
            # Obtener nombres del personal (usar nombre completo o username como fallback)
            personal_nombres = []
            for ap in evento.personal.all()[:3]:
                if ap.usuario:
                    personal_nombres.append(ap.usuario.nombre if ap.usuario.nombre else ap.usuario.username)
                elif ap.colaborador:
                    personal_nombres.append(ap.colaborador.nombre)
            if personal_count > 3:
                personal_nombres.append(f'+{personal_count - 3} m√°s')
            
            # Convertir creado_en a zona horaria local
            from django.utils.timezone import localtime, is_aware, make_aware
            import pytz
            
            # Si la fecha no tiene timezone, agregarla (asumiendo UTC)
            if not is_aware(evento.creado_en):
                creado_en_aware = make_aware(evento.creado_en, pytz.UTC)
            else:
                creado_en_aware = evento.creado_en
            
            creado_en_local = localtime(creado_en_aware)
            
            comunidades_detalle = obtener_comunidades_evento(evento)
            if comunidades_detalle:
                comunidades_resumen = ', '.join([c['comunidad_nombre'] for c in comunidades_detalle if c['comunidad_nombre']])
                comunidad_principal = comunidades_detalle[0]['comunidad_nombre']
            else:
                comunidades_resumen = evento.comunidad.nombre if evento.comunidad else 'Sin comunidades'
                comunidad_principal = evento.comunidad.nombre if evento.comunidad else 'Sin comunidad'
            
            portada_evento = obtener_portada_evento(evento)
             
            eventos_data.append({
                'id': str(evento.id),
                'nombre': evento.nombre,
                'tipo': evento.tipo.nombre if evento.tipo else 'Sin tipo',
                'comunidad': comunidad_principal,
                'comunidades': comunidades_detalle,
                'comunidades_resumen': comunidades_resumen,
                'fecha': str(evento.fecha),
                'estado': evento.estado,
                'descripcion': evento.descripcion[:100] + '...' if len(evento.descripcion) > 100 else evento.descripcion,
                'personal_count': personal_count,
                'personal_nombres': ', '.join(personal_nombres) if personal_nombres else 'Sin personal',
                'beneficiarios_count': beneficiarios_count,
                'responsable': (evento.responsable.nombre if evento.responsable.nombre else evento.responsable.username) if evento.responsable else 'Sin responsable',
                'creado_en': creado_en_local.strftime('%d/%m/%Y %H:%M'),
                'portada': portada_evento
            })
        
        return JsonResponse({
            'success': True,
            'eventos': eventos_data,
            'total': len(eventos_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al listar eventos: {str(e)}'
        }, status=500)
@api_ratelimit_read(rate='30/m')  # 30 peticiones por minuto
def api_obtener_evento(request, evento_id):
    """Obtiene los detalles completos de un evento (OPTIMIZADO con paginaci√≥n)
    Rate limit: 30 peticiones por minuto"""
    try:
        # Par√°metros de paginaci√≥n
        evidencias_limite = int(request.GET.get('evidencias_limite', 50))
        evidencias_offset = int(request.GET.get('evidencias_offset', 0))
        # Aumentar l√≠mite de beneficiarios a 2000 para soportar importaciones masivas
        beneficiarios_limite = int(request.GET.get('beneficiarios_limite', 2000))
        beneficiarios_offset = int(request.GET.get('beneficiarios_offset', 0))
        
        # Optimizaci√≥n: Usar Prefetch para cargar relaciones de forma eficiente
        # IMPORTANTE: No usar slice dentro de Prefetch, causa error con filtros
        evento = Actividad.objects.select_related(
            'tipo', 'comunidad', 'responsable', 'colaborador', 'portada'
        ).prefetch_related(
            'personal__usuario__puesto',
            'personal__colaborador__puesto',
            'personal__colaborador__usuario',
            Prefetch(
                'beneficiarios',
                queryset=ActividadBeneficiario.objects.select_related(
                    'beneficiario__tipo',
                    'beneficiario__comunidad__region',
                    'beneficiario__individual',
                    'beneficiario__familia',
                    'beneficiario__institucion'
                ).order_by('id')
            ),
            Prefetch(
                'evidencias',
                queryset=Evidencia.objects.filter(
                    url_almacenamiento__icontains='/media/evidencias/'
                ).order_by('-creado_en')
            ),
            'comunidades_relacionadas__comunidad__region',
            'comunidades_relacionadas__region'
        ).get(id=evento_id, eliminado_en__isnull=True)
        
        # Los totales se calcular√°n despu√©s de cargar los datos con Prefetch
        
        # Personal asignado
        personal_data = []
        for ap in evento.personal.all():
            # IMPORTANTE: Si tiene colaborador, SIEMPRE usar el ID del colaborador (aunque tambi√©n tenga usuario)
            # Esto asegura que coincida con la lista de personal disponible desde api_listar_personal
            # que siempre retorna colaboradores con su ID de colaborador
            if ap.colaborador:
                personal_data.append({
                    'id': str(ap.colaborador.id),
                    'username': ap.colaborador.usuario.username if ap.colaborador.usuario else (ap.colaborador.correo or ''),
                    'nombre': ap.colaborador.nombre,
                    'rol': ap.rol_en_actividad,
                    'rol_display': 'Personal Fijo' if ap.colaborador.es_personal_fijo else 'Colaborador Externo',
                    'puesto': ap.colaborador.puesto.nombre if ap.colaborador.puesto else None,
                    'tipo': 'colaborador'
                })
            elif ap.usuario:
                # Solo usuarios directos (sin colaborador asociado)
                personal_data.append({
                    'id': str(ap.usuario.id),
                    'username': ap.usuario.username,
                    'nombre': ap.usuario.nombre or '',
                    'rol': ap.rol_en_actividad,
                    'rol_display': ap.usuario.get_rol_display(),
                    'puesto': ap.usuario.puesto.nombre if ap.usuario.puesto else None,
                    'tipo': 'usuario'
                })
        
        # Beneficiarios con detalles completos (aplicar paginaci√≥n manual)
        beneficiarios_data = []
        beneficiarios_list = list(evento.beneficiarios.all())
        beneficiarios_paginados = beneficiarios_list[beneficiarios_offset:beneficiarios_offset + beneficiarios_limite]
        
        for ab in beneficiarios_paginados:
            benef = ab.beneficiario
            nombre_display, _, detalles, tipo_envio = obtener_detalle_beneficiario(benef)
            if benef.tipo and hasattr(benef.tipo, 'get_nombre_display'):
                tipo_display = benef.tipo.get_nombre_display()
            else:
                tipo_display = tipo_envio.title() if tipo_envio else ''
            beneficiarios_data.append({
                'id': str(benef.id),
                'nombre': nombre_display,
                'tipo': tipo_envio,
                'tipo_display': tipo_display,
                'comunidad_id': detalles.get('comunidad_id'),
                'comunidad_nombre': detalles.get('comunidad_nombre'),
                'region_id': detalles.get('region_id'),
                'region_nombre': detalles.get('region_nombre'),
                'region_sede': detalles.get('region_sede'),
                'detalles': detalles
            })
        
        # Evidencias (aplicar paginaci√≥n manual)
        evidencias_data = []
        evidencias_list = list(evento.evidencias.all())
        evidencias_paginadas = evidencias_list[evidencias_offset:evidencias_offset + evidencias_limite]
        
        for evidencia in evidencias_paginadas:
            evidencias_data.append({
                'id': str(evidencia.id),
                'nombre': evidencia.archivo_nombre,
                'archivo_nombre': evidencia.archivo_nombre,
                'url': evidencia.url_almacenamiento or '',
                'tipo': evidencia.archivo_tipo,
                'es_imagen': evidencia.es_imagen,
                'descripcion': evidencia.descripcion or '',
                'es_galeria': False,
                'es_evidencia': True,
                'es_archivo': True
            })
        
        comunidades_data = obtener_comunidades_evento(evento)
        comunidad_principal_id = comunidades_data[0]['comunidad_id'] if comunidades_data else (str(evento.comunidad.id) if evento.comunidad else None)
        portada_data = obtener_portada_evento(evento)
        
        # Calcular totales desde las listas ya cargadas
        total_beneficiarios = len(beneficiarios_list)
        total_evidencias = len(evidencias_list)
        
        evento_data = {
            'id': str(evento.id),
            'nombre': evento.nombre,
            'tipo_id': str(evento.tipo.id) if evento.tipo else None,
            'comunidad_id': comunidad_principal_id,
            'fecha': str(evento.fecha),
            'estado': evento.estado,
            'descripcion': evento.descripcion,
            'personal': personal_data,
            'beneficiarios': beneficiarios_data,
            'evidencias': evidencias_data,
            'comunidades': comunidades_data,
            'portada': portada_data,
            'tarjetas_datos': obtener_tarjetas_datos(evento),
            'paginacion': {
                'beneficiarios': {
                    'total': total_beneficiarios,
                    'limite': beneficiarios_limite,
                    'offset': beneficiarios_offset,
                    'tiene_mas': (beneficiarios_offset + beneficiarios_limite) < total_beneficiarios
                },
                'evidencias': {
                    'total': total_evidencias,
                    'limite': evidencias_limite,
                    'offset': evidencias_offset,
                    'tiene_mas': (evidencias_offset + evidencias_limite) < total_evidencias
                }
            }
        }
        
        return JsonResponse({
            'success': True,
            'evento': evento_data
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener evento: {str(e)}'
        }, status=500)
@permiso_gestionar_eventos_api
@require_http_methods(["POST"])
@api_ratelimit_bulk(rate='20/m')  # 20 actualizaciones por minuto
def api_actualizar_evento(request, evento_id):
    """Actualiza un evento existente
    Rate limit: 20 peticiones por minuto (puede incluir 50+ evidencias y 100+ comunidades por petici√≥n)"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        usuario_maga = get_usuario_maga(request.user)
        
        data = request.POST
        cambios_realizados = []
        
        comunidades_payload = []
        if data.get('comunidades_seleccionadas'):
            try:
                comunidades_payload = json.loads(data['comunidades_seleccionadas'])
            except (json.JSONDecodeError, ValueError):
                comunidades_payload = []

        comunidad_principal_id = data.get('comunidad_id')
        if not comunidad_principal_id and comunidades_payload:
            comunidad_principal_id = comunidades_payload[0].get('comunidad_id')

        portada_eliminar_flag = data.get('portada_eliminar') == 'true'

        # Solo validar comunidad si se est√° intentando actualizar la comunidad expl√≠citamente
        # Si solo se est√° actualizando la descripci√≥n u otros campos, usar la comunidad existente del evento
        if not comunidad_principal_id:
            # Si el evento ya tiene una comunidad, usar esa. Solo validar si no tiene ninguna.
            if evento.comunidad:
                # Usar la comunidad existente del evento
                comunidad_principal_id = str(evento.comunidad.id)
            else:
                # El evento no tiene comunidad y no se est√° enviando una, entonces es un error
                return JsonResponse({
                    'success': False,
                    'error': 'Debe seleccionar al menos una comunidad asociada al evento'
                }, status=400)
        
        with transaction.atomic():
            # Actualizar campos b√°sicos
            if data.get('nombre') and data.get('nombre') != evento.nombre:
                nuevo_nombre = data.get('nombre').strip()
                # Validar que no exista otro evento con el mismo nombre (excluyendo el actual)
                evento_existente = Actividad.objects.filter(
                    nombre__iexact=nuevo_nombre,
                    eliminado_en__isnull=True
                ).exclude(id=evento.id).first()
                if evento_existente:
                    return JsonResponse({
                        'success': False,
                        'error': f'Ya existe otro evento con el nombre "{nuevo_nombre}". Por favor, elige un nombre diferente.'
                    }, status=400)
                cambios_realizados.append(f"Nombre: '{evento.nombre}' ‚Üí '{nuevo_nombre}'")
                evento.nombre = nuevo_nombre
            
            if data.get('tipo_actividad_id'):
                nuevo_tipo = TipoActividad.objects.get(id=data.get('tipo_actividad_id'))
                if nuevo_tipo != evento.tipo:
                    cambios_realizados.append(f"Tipo: '{evento.tipo.nombre}' ‚Üí '{nuevo_tipo.nombre}'")
                    evento.tipo = nuevo_tipo
            
            # Solo actualizar la comunidad si se est√° enviando expl√≠citamente un cambio
            # Si solo se est√° actualizando la descripci√≥n, mantener la comunidad existente
            if data.get('comunidad_id') or comunidades_payload:
                # Se est√° enviando un cambio expl√≠cito de comunidad
                if comunidad_principal_id:
                    nueva_comunidad = Comunidad.objects.get(id=comunidad_principal_id)
                    if nueva_comunidad != evento.comunidad:
                        cambios_realizados.append(f"Comunidad: '{evento.comunidad.nombre if evento.comunidad else 'Sin comunidad'}' ‚Üí '{nueva_comunidad.nombre}'")
                        evento.comunidad = nueva_comunidad
                elif data.get('comunidad_id') == '' or (data.get('comunidades_seleccionadas') and not comunidades_payload):
                    # Se est√° removiendo expl√≠citamente la comunidad
                    if evento.comunidad is not None:
                        cambios_realizados.append(f"Comunidad principal removida: '{evento.comunidad.nombre}'")
                        evento.comunidad = None
            # Si no se env√≠a comunidad_id ni comunidades_seleccionadas, mantener la comunidad existente
            
            if data.get('fecha') and data.get('fecha') != str(evento.fecha):
                cambios_realizados.append(f"Fecha: '{evento.fecha}' ‚Üí '{data.get('fecha')}'")
                evento.fecha = data.get('fecha')
            
            if data.get('estado') and data.get('estado') != evento.estado:
                cambios_realizados.append(f"Estado: '{evento.estado}' ‚Üí '{data.get('estado')}'")
                evento.estado = data.get('estado')
            
            if data.get('descripcion') and data.get('descripcion') != evento.descripcion:
                cambios_realizados.append("Descripci√≥n actualizada")
                evento.descripcion = data.get('descripcion')
            
            evento.actualizado_en = timezone.now()
            evento.save()
            
            # Actualizar personal
            if data.get('personal_ids') is not None:
                try:
                    personal_raw = json.loads(data.get('personal_ids'))
                except (json.JSONDecodeError, ValueError):
                    personal_raw = []

                parsed_incoming = []
                incoming_keys = set()
                for item in personal_raw:
                    if isinstance(item, dict):
                        obj_id = item.get('id')
                        tipo = item.get('tipo', 'colaborador')
                        rol = item.get('rol', 'Colaborador')
                    else:
                        obj_id = item
                        tipo = 'colaborador'
                        rol = 'Colaborador'

                    if not obj_id:
                        continue

                    key = f"{tipo}:{obj_id}"
                    incoming_keys.add(key)
                    parsed_incoming.append({'id': obj_id, 'tipo': tipo, 'rol': rol, 'key': key})

                # Mapear personal actual
                personal_actual = {}
                for ap in evento.personal.all():
                    if ap.colaborador_id:
                        key = f"colaborador:{ap.colaborador_id}"
                    elif ap.usuario_id:
                        key = f"usuario:{ap.usuario_id}"
                    else:
                        continue
                    personal_actual[key] = ap

                actuales_keys = set(personal_actual.keys())

                # Eliminar los que ya no est√°n
                keys_a_eliminar = actuales_keys - incoming_keys
                if keys_a_eliminar:
                    for key in keys_a_eliminar:
                        personal_actual[key].delete()
                    cambios_realizados.append(f"Removido {len(keys_a_eliminar)} personal")

                # Agregar o actualizar
                parsed_map = {item['key']: item for item in parsed_incoming}
                for key, item in parsed_map.items():
                    rol = item.get('rol', 'Colaborador')
                    if key in personal_actual:
                        ap = personal_actual[key]
                        if ap.rol_en_actividad != rol:
                            ap.rol_en_actividad = rol
                            ap.save(update_fields=['rol_en_actividad'])
                        continue

                    if item['tipo'] == 'usuario':
                        try:
                            usuario = Usuario.objects.get(id=item['id'])
                        except Usuario.DoesNotExist:
                            continue

                        ActividadPersonal.objects.create(
                            actividad=evento,
                            usuario=usuario,
                            rol_en_actividad=rol
                        )
                    else:
                        colaborador = Colaborador.objects.filter(id=item['id'], activo=True).select_related('usuario').first()
                        if not colaborador:
                            continue
                        ActividadPersonal.objects.create(
                            actividad=evento,
                            colaborador=colaborador,
                            usuario=colaborador.usuario if colaborador.usuario else None,
                            rol_en_actividad=rol
                        )

                nuevos_agregados = max(len(incoming_keys - actuales_keys), 0)
                if nuevos_agregados:
                    cambios_realizados.append(f"Agregado {nuevos_agregados} personal")
            
            # Actualizar comunidades asociadas
            comunidades_a_registrar = comunidades_payload.copy()
            if comunidad_principal_id and not any(item.get('comunidad_id') == comunidad_principal_id for item in comunidades_a_registrar):
                comunidades_a_registrar.insert(0, {'comunidad_id': comunidad_principal_id})

            relaciones_existentes_queryset = ActividadComunidad.objects.filter(actividad=evento)
            comunidades_existentes_map = {
                str(rel.comunidad_id): rel for rel in relaciones_existentes_queryset
            }
            comunidades_existentes_ids = set(comunidades_existentes_map.keys())

            if comunidades_a_registrar:
                for item in comunidades_a_registrar:
                    comunidad_id_tmp = item.get('comunidad_id')
                    if not item.get('agregado_en'):
                        if comunidad_id_tmp and str(comunidad_id_tmp) in comunidades_existentes_map and comunidades_existentes_map[str(comunidad_id_tmp)].creado_en:
                            item['agregado_en'] = comunidades_existentes_map[str(comunidad_id_tmp)].creado_en.isoformat()
                        else:
                            item['agregado_en'] = timezone.now().isoformat()

                comunidades_ids = [item.get('comunidad_id') for item in comunidades_a_registrar if item.get('comunidad_id')]
                comunidades_map = {
                    str(com.id): com for com in Comunidad.objects.filter(id__in=comunidades_ids).select_related('region')
                }

                nuevas_ids = set()

                for item in comunidades_a_registrar:
                    comunidad_id = item.get('comunidad_id')
                    if not comunidad_id:
                        continue

                    comunidad_obj = comunidades_map.get(str(comunidad_id))
                    if comunidad_obj is None:
                        raise Comunidad.DoesNotExist(f'Comunidad no v√°lida: {comunidad_id}')

                    region_id = item.get('region_id')
                    if not region_id and comunidad_obj.region_id:
                        region_id = comunidad_obj.region_id

                    nuevas_ids.add(str(comunidad_obj.id))

                    relacion, creada = ActividadComunidad.objects.get_or_create(
                        actividad=evento,
                        comunidad=comunidad_obj,
                        defaults={'region_id': region_id}
                    )

                    agregado_en = parse_fecha_agregacion(item.get('agregado_en'))
                    update_fields = []

                    if region_id and relacion.region_id != region_id:
                        relacion.region_id = region_id
                        update_fields.append('region_id')

                    if agregado_en and (not relacion.creado_en or relacion.creado_en != agregado_en):
                        relacion.creado_en = agregado_en
                        update_fields.append('creado_en')

                    if update_fields:
                        relacion.save(update_fields=update_fields)

                ids_a_eliminar = comunidades_existentes_ids - nuevas_ids
                if ids_a_eliminar:
                    ActividadComunidad.objects.filter(
                        actividad=evento,
                        comunidad_id__in=ids_a_eliminar
                    ).delete()

                # Actualizar el campo actualizado_en de las comunidades y regiones asociadas
                # para que aparezcan en "√öltimas Regiones/Comunidades"
                if comunidades_ids:
                    # Actualizar comunidades
                    Comunidad.objects.filter(id__in=comunidades_ids).update(actualizado_en=timezone.now())
                    
                    # Obtener las regiones de estas comunidades y actualizarlas tambi√©n
                    regiones_ids = Comunidad.objects.filter(
                        id__in=comunidades_ids,
                        region__isnull=False
                    ).values_list('region_id', flat=True).distinct()
                    
                    if regiones_ids:
                        Region.objects.filter(id__in=regiones_ids).update(actualizado_en=timezone.now())

            tarjetas_creadas = []
            if data.get('tarjetas_datos_nuevas'):
                try:
                    tarjetas_payload = json.loads(data['tarjetas_datos_nuevas'])
                except json.JSONDecodeError:
                    tarjetas_payload = []

                for idx, tarjeta in enumerate(tarjetas_payload):
                    titulo = tarjeta.get('titulo')
                    valor = tarjeta.get('valor')
                    icono = tarjeta.get('icono')
                    if not titulo or not valor:
                        continue
                    tarjeta_inst = TarjetaDato.objects.create(
                        entidad_tipo='actividad',
                        entidad_id=evento.id,
                        titulo=titulo,
                        valor=valor,
                        icono=icono,
                        orden=idx
                    )
                    tarjetas_creadas.append({
                        'id': str(tarjeta_inst.id),
                        'titulo': tarjeta_inst.titulo,
                        'valor': tarjeta_inst.valor,
                        'icono': tarjeta_inst.icono,
                        'orden': tarjeta_inst.orden
                    })

            # Eliminar beneficiarios (si se marcaron)
            if data.get('beneficiarios_eliminados'):
                try:
                    beneficiarios_ids_eliminar = json.loads(data['beneficiarios_eliminados'])
                    if beneficiarios_ids_eliminar:
                        # Eliminar las relaciones ActividadBeneficiario
                        ActividadBeneficiario.objects.filter(
                            actividad=evento,
                            beneficiario_id__in=beneficiarios_ids_eliminar
                        ).delete()
                        cambios_realizados.append(f"Removido {len(beneficiarios_ids_eliminar)} beneficiarios")
                        print(f"‚úÖ Removidos {len(beneficiarios_ids_eliminar)} beneficiarios del evento")
                except (json.JSONDecodeError, ValueError) as e:
                    print(f"Error al eliminar beneficiarios: {e}")
                    pass
            
            # Agregar beneficiarios existentes seleccionados durante la edici√≥n
            if data.get('beneficiarios_existentes_agregar'):
                try:
                    beneficiarios_agregar = json.loads(data['beneficiarios_existentes_agregar'])
                    agregados = 0
                    for benef_id in beneficiarios_agregar:
                        if not benef_id:
                            continue
                        beneficiario = Beneficiario.objects.filter(id=benef_id, activo=True).first()
                        if not beneficiario:
                            continue
                        _, created_rel = ActividadBeneficiario.objects.get_or_create(
                            actividad=evento,
                            beneficiario=beneficiario
                        )
                        if created_rel:
                            agregados += 1
                    if agregados:
                        cambios_realizados.append(f"Asociado {agregados} beneficiarios existentes")
                        print(f"‚úÖ Asociados {agregados} beneficiarios existentes al evento")
                except (json.JSONDecodeError, ValueError) as e:
                    print(f"Error al agregar beneficiarios existentes: {e}")
                    pass

            # Actualizar beneficiarios modificados (si hay)
            if data.get('beneficiarios_modificados'):
                try:
                    beneficiarios_modificados = json.loads(data['beneficiarios_modificados'])
                    cambios_aplicados = aplicar_modificaciones_beneficiarios(beneficiarios_modificados)
                    if cambios_aplicados:
                        cambios_realizados.append(f"Modificado {cambios_aplicados} beneficiarios")
                except json.JSONDecodeError as e:
                    print(f"Error al actualizar beneficiarios: {e}")
                    pass
                except (ValueError, IntegrityError) as e:
                    # Retornar error de validaci√≥n (duplicado)
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    }, status=400)
            
            # Eliminar evidencias (si se marcaron)
            if data.get('evidencias_eliminadas'):
                try:
                    evidencias_ids_eliminar = json.loads(data['evidencias_eliminadas'])
                    if evidencias_ids_eliminar:
                        evidencias_a_eliminar = Evidencia.objects.filter(
                            id__in=evidencias_ids_eliminar,
                            actividad=evento
                        )
                        eliminadas = _eliminar_queryset_con_archivos(evidencias_a_eliminar)
                        if eliminadas:
                            cambios_realizados.append(f"Removido {eliminadas} evidencias")
                            print(f"‚úÖ Removidas {eliminadas} evidencias del evento")
                except (json.JSONDecodeError, ValueError) as e:
                    print(f"Error al eliminar evidencias: {e}")
                    pass
            
            # Actualizar beneficiarios nuevos (si se agregaron)
            if data.get('beneficiarios_nuevos'):
                try:
                    beneficiarios_nuevos = json.loads(data['beneficiarios_nuevos'])
                    
                    for benef_data in beneficiarios_nuevos:
                        tipo = benef_data.get('tipo')
                        comunidad_id = data.get('comunidad_id')
                        
                        # Obtener el tipo de beneficiario
                        tipo_lookup = tipo if tipo != 'otro' else 'instituci√≥n'
                        tipo_benef = TipoBeneficiario.objects.get(nombre=tipo_lookup)
                        
                        # Validaciones de DPI
                        dpi_individual = dpi_jefe = dpi_rep = ''
                        if tipo == 'individual':
                            dpi_individual = normalizar_dpi(benef_data.get('dpi'))
                            if dpi_individual:
                                conflicto = buscar_conflicto_dpi(
                                    BeneficiarioIndividual.objects.all(),
                                    'dpi',
                                    dpi_individual,
                                )
                                if conflicto:
                                    raise ValueError(f'Ya existe un beneficiario individual con el DPI {dpi_individual}.')
                        elif tipo == 'familia':
                            dpi_jefe = normalizar_dpi(benef_data.get('dpi_jefe_familia'))
                            if dpi_jefe:
                                conflicto = buscar_conflicto_dpi(
                                    BeneficiarioFamilia.objects.all(),
                                    'dpi_jefe_familia',
                                    dpi_jefe,
                                )
                                if conflicto:
                                    raise ValueError(f'Ya existe una familia con el DPI del jefe {dpi_jefe}.')
                        elif tipo == 'institucion' or tipo == 'instituci√≥n':
                            dpi_rep = normalizar_dpi(benef_data.get('dpi_representante'))
                            if dpi_rep:
                                conflicto = buscar_conflicto_dpi(
                                    BeneficiarioInstitucion.objects.all(),
                                    'dpi_representante',
                                    dpi_rep,
                                )
                                if conflicto:
                                    raise ValueError(f'Ya existe una instituci√≥n con el DPI del representante {dpi_rep}.')

                        # Crear beneficiario principal
                        beneficiario = Beneficiario.objects.create(
                            tipo=tipo_benef,
                            comunidad_id=comunidad_id,
                            activo=True
                        )
                        
                        # Crear registro espec√≠fico seg√∫n tipo
                        if tipo == 'individual':
                            BeneficiarioIndividual.objects.create(
                                beneficiario=beneficiario,
                                nombre=benef_data.get('nombre', ''),
                                apellido=benef_data.get('apellido', ''),
                                dpi=dpi_individual or None,
                                fecha_nacimiento=benef_data.get('fecha_nacimiento'),
                                genero=benef_data.get('genero'),
                                telefono=benef_data.get('telefono')
                            )
                        elif tipo == 'familia':
                            BeneficiarioFamilia.objects.create(
                                beneficiario=beneficiario,
                                nombre_familia=benef_data.get('nombre_familia', ''),
                                jefe_familia=benef_data.get('jefe_familia', ''),
                                dpi_jefe_familia=dpi_jefe or None,
                                telefono=benef_data.get('telefono'),
                                numero_miembros=benef_data.get('numero_miembros')
                            )
                        elif tipo == 'institucion' or tipo == 'instituci√≥n':
                            BeneficiarioInstitucion.objects.create(
                                beneficiario=beneficiario,
                                nombre_institucion=benef_data.get('nombre_institucion', ''),
                                tipo_institucion=benef_data.get('tipo_institucion', 'otro'),
                                representante_legal=benef_data.get('representante_legal'),
                                dpi_representante=dpi_rep or None,
                                telefono=benef_data.get('telefono'),
                                email=benef_data.get('email'),
                                numero_beneficiarios_directos=benef_data.get('numero_beneficiarios_directos')
                            )
                        elif tipo == 'otro':
                            BeneficiarioInstitucion.objects.create(
                                beneficiario=beneficiario,
                                nombre_institucion=benef_data.get('nombre', ''),
                                tipo_institucion='otro',
                                representante_legal=benef_data.get('contacto'),
                                telefono=benef_data.get('telefono'),
                                email=benef_data.get('tipo_descripcion')
                            )
                        
                        # Asociar beneficiario con la actividad
                        ActividadBeneficiario.objects.create(
                            actividad=evento,
                            beneficiario=beneficiario
                        )
                    
                    cambios_realizados.append(f"Agregado {len(beneficiarios_nuevos)} beneficiarios")
                    
                except (json.JSONDecodeError, TipoBeneficiario.DoesNotExist) as e:
                    print(f"Error al agregar beneficiarios: {e}")
                    pass
                except (ValueError, IntegrityError) as e:
                    # Retornar error de validaci√≥n (duplicado)
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    }, status=400)
            
            # Agregar nuevas evidencias
            if request.FILES.getlist('evidencias_nuevas'):
                archivos = request.FILES.getlist('evidencias_nuevas')
                evidencias_dir = os.path.join(str(settings.MEDIA_ROOT), 'evidencias')
                os.makedirs(evidencias_dir, exist_ok=True)
                fs = FileSystemStorage(location=evidencias_dir)
                
                for idx, archivo in enumerate(archivos):
                    # COMPRIMIR IMAGEN autom√°ticamente si es una imagen
                    es_imagen = archivo.content_type.startswith('image/')
                    if es_imagen and getattr(settings, 'COMPRESS_IMAGES', True):
                        from .image_compression import compress_if_needed, get_compression_settings
                        compression_settings = get_compression_settings('evidence')
                        archivo = compress_if_needed(archivo, **compression_settings)
                    
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
                    nombre_unico = f"{timestamp}_{idx}.{archivo.name.split('.')[-1]}"
                    nombre_guardado = fs.save(nombre_unico, archivo)
                    url_archivo = f"/media/evidencias/{nombre_guardado}"
                    
                    Evidencia.objects.create(
                        actividad=evento,
                        archivo_nombre=archivo.name,
                        archivo_tipo=archivo.content_type,
                        url_almacenamiento=url_archivo,
                        es_imagen=es_imagen
                    )
                
                cambios_realizados.append(f"Agregado {len(archivos)} evidencias")
            
            # Registrar cambios
            if cambios_realizados:
                descripcion = f"Evento actualizado por {usuario_maga.username}: " + "; ".join(cambios_realizados)
                ActividadCambio.objects.create(
                    actividad=evento,
                    responsable=usuario_maga,
                    descripcion_cambio=descripcion[:500]
                )
            
            # Gestionar imagen de portada
            portada_actual = getattr(evento, 'portada', None)
            portada_file = request.FILES.get('portada_evento')
            if portada_eliminar_flag and portada_actual and not portada_file:
                eliminar_portada_evento(portada_actual)
                cambios_realizados.append('Imagen de portada eliminada')

            if portada_file:
                if portada_actual:
                    eliminar_portada_evento(portada_actual)
                try:
                    guardar_portada_evento(evento, portada_file)
                    cambios_realizados.append('Imagen de portada actualizada')
                except ValueError as portada_error:
                    raise ValueError(str(portada_error))
            if data.get('tarjetas_datos_nuevas'):
                try:
                    tarjetas_nuevas = json.loads(data['tarjetas_datos_nuevas'])
                except json.JSONDecodeError:
                    tarjetas_nuevas = []

                if tarjetas_nuevas:
                    orden_inicial = TarjetaDato.objects.filter(entidad_tipo='actividad', entidad_id=evento.id).count()
                    tarjetas_creadas_ids = set()  # Para evitar duplicados por t√≠tulo
                    
                    for idx, tarjeta in enumerate(tarjetas_nuevas):
                        titulo = tarjeta.get('titulo', '').strip()
                        valor = tarjeta.get('valor', '').strip()
                        icono = tarjeta.get('icono')
                        
                        if not titulo or not valor:
                            continue
                        
                        # Verificar si ya existe una tarjeta con el mismo t√≠tulo
                        titulo_normalizado = titulo.lower().strip()
                        existe = TarjetaDato.objects.filter(
                            entidad_tipo='actividad',
                            entidad_id=evento.id,
                            titulo__iexact=titulo_normalizado
                        ).exists()
                        
                        if existe:
                            print(f'‚ö†Ô∏è Tarjeta con t√≠tulo "{titulo}" ya existe, omitiendo creaci√≥n duplicada')
                            continue
                        
                        # Verificar si ya se est√° creando una con el mismo t√≠tulo en esta misma operaci√≥n
                        if titulo_normalizado in tarjetas_creadas_ids:
                            print(f'‚ö†Ô∏è Tarjeta con t√≠tulo "{titulo}" duplicada en la misma operaci√≥n, omitiendo')
                            continue
                        
                        tarjetas_creadas_ids.add(titulo_normalizado)
                        
                        TarjetaDato.objects.create(
                            entidad_tipo='actividad',
                            entidad_id=evento.id,
                            titulo=titulo,
                            valor=valor,
                            icono=icono,
                            orden=orden_inicial + idx
                        )
                    cambios_realizados.append(f"Agregados {len(tarjetas_creadas_ids)} datos del proyecto")
            if data.get('tarjetas_datos_actualizadas'):
                try:
                    tarjetas_actualizadas = json.loads(data['tarjetas_datos_actualizadas'])
                except json.JSONDecodeError:
                    tarjetas_actualizadas = []

                actualizados = 0
                for tarjeta in tarjetas_actualizadas:
                    tarjeta_id = tarjeta.get('id')
                    if not tarjeta_id:
                        continue
                    nuevo_titulo = tarjeta.get('titulo', '').strip()
                    
                    filas = TarjetaDato.objects.filter(id=tarjeta_id, entidad_tipo='actividad', entidad_id=evento.id)
                    if filas.exists():
                        filas.update(
                            titulo=nuevo_titulo,
                            valor=tarjeta.get('valor', ''),
                            icono=tarjeta.get('icono')
                        )
                        actualizados += 1
                if actualizados:
                    cambios_realizados.append(f"Actualizados {actualizados} datos del proyecto")

            if data.get('tarjetas_datos_eliminadas'):
                try:
                    tarjetas_eliminar = json.loads(data['tarjetas_datos_eliminadas'])
                except json.JSONDecodeError:
                    tarjetas_eliminar = []

                if tarjetas_eliminar:
                    eliminados, _ = TarjetaDato.objects.filter(
                        entidad_tipo='actividad',
                        entidad_id=evento.id,
                        id__in=tarjetas_eliminar
                    ).delete()
                    if eliminados:
                        cambios_realizados.append(f"Eliminados {eliminados} datos del proyecto")

            if cambios_realizados:
                descripcion = f"Evento actualizado por {usuario_maga.username}: " + "; ".join(cambios_realizados)
                ActividadCambio.objects.create(
                    actividad=evento,
                    responsable=usuario_maga,
                    descripcion_cambio=descripcion[:500]
                )
            
            return JsonResponse({
                'success': True,
                'message': 'Evento actualizado exitosamente',
                'evento': {
                    'id': str(evento.id),
                    'nombre': evento.nombre,
                    'fecha': str(evento.fecha),
                    'portada': obtener_portada_evento(evento),
                    'tarjetas_datos': obtener_tarjetas_datos(evento)
                },
                'cambios': cambios_realizados
            })
    
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except TipoActividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Tipo de actividad no v√°lido'
        }, status=400)
    except Comunidad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Comunidad no v√°lida'
        }, status=400)
    except (ValueError, IntegrityError) as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar evento: {str(e)}'
        }, status=500)
@permiso_gestionar_eventos
@require_http_methods(["GET"])
def api_cambios_recientes(request):
    """Obtiene los cambios recientes de actividades"""
    try:
        # Obtener los √∫ltimos 10 cambios
        limite = int(request.GET.get('limite', 10))
        
        # Obtener TODOS los cambios, incluyendo de eventos eliminados
        cambios = ActividadCambio.objects.select_related(
            'actividad', 'responsable'
        ).order_by('-fecha_cambio')[:limite]
        
        cambios_data = []
        for cambio in cambios:
            try:
                # Verificar que la actividad y responsable existen
                if not cambio.actividad:
                    continue
                
                # Determinar si el evento est√° eliminado
                evento_eliminado = cambio.actividad.eliminado_en is not None
                    
                # Convertir a la zona horaria local (Guatemala)
                from django.utils.timezone import localtime, is_aware, make_aware
                import pytz
                
                # Si la fecha no tiene timezone, agregarla (asumiendo UTC)
                if not is_aware(cambio.fecha_cambio):
                    fecha_aware = make_aware(cambio.fecha_cambio, pytz.UTC)
                else:
                    fecha_aware = cambio.fecha_cambio
                
                # Ahora convertir a hora local de Guatemala
                fecha_local = localtime(fecha_aware)
                
                cambios_data.append({
                    'id': str(cambio.id),
                    'actividad_id': str(cambio.actividad.id),
                    'actividad_nombre': cambio.actividad.nombre,
                    'responsable': (cambio.responsable.nombre if cambio.responsable.nombre else cambio.responsable.username) if cambio.responsable else 'Sistema',
                    'descripcion': cambio.descripcion_cambio,
                    'fecha': fecha_local.strftime('%d/%m/%Y %H:%M'),
                    'evento_eliminado': evento_eliminado
                })
            except Exception as item_error:
                # Si hay error con un cambio espec√≠fico, lo saltamos y continuamos
                print(f"‚ö†Ô∏è Error procesando cambio {cambio.id}: {item_error}")
                continue
        
        return JsonResponse({
            'success': True,
            'cambios': cambios_data,
            'total': len(cambios_data)
        })
        
    except Exception as e:
        # Log del error para debugging
        print(f"‚ùå Error en api_cambios_recientes: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener cambios: {str(e)}'
        }, status=500)
def api_verificar_admin(request):
    """Verifica las credenciales del administrador antes de permitir acciones cr√≠ticas"""
    try:
        data = json.loads(request.body or '{}')
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return JsonResponse({
                'success': False,
                'error': 'Usuario y contrase√±a son requeridos'
            }, status=400)
        
        user = authenticate(request, username=username, password=password)
        
        if user is None:
            return JsonResponse({
                'success': False,
                'error': 'Credenciales inv√°lidas'
            }, status=401)
        
        rol = getattr(user, 'rol_maga', None)
        
        if rol != 'admin':
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos de administrador'
            }, status=403)
        
        return JsonResponse({
            'success': True,
            'message': 'Credenciales verificadas correctamente',
            'user': {
                'username': user.username,
                'rol': rol
            }
        })
            
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inv√°lidos'
        }, status=400)
    except Exception as e:
        print(f"‚ùå Error en api_verificar_admin: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al verificar credenciales: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["DELETE"])
def api_eliminar_evento(request, evento_id):
    """Elimina un evento (soft delete)"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        usuario_maga = get_usuario_maga(request.user)
        
        # Eliminar portada (archivo f√≠sico + registro)
        portada = getattr(evento, 'portada', None)
        if portada:
            eliminar_portada_evento(portada)

        # Eliminar evidencias principales
        _eliminar_queryset_con_archivos(evento.evidencias.all())

        # Eliminar im√°genes de galer√≠a
        _eliminar_queryset_con_archivos(evento.galeria_imagenes.all())

        # Eliminar archivos adjuntos
        _eliminar_queryset_con_archivos(evento.archivos.all())

        # Eliminar evidencias asociadas a cambios internos
        for cambio in list(evento.cambios.all()):
            _eliminar_queryset_con_archivos(cambio.evidencias.all())
            cambio.delete()

        # Eliminar evidencias asociadas a cambios de colaboradores
        for cambio_colaborador in list(evento.cambios_colaboradores.all()):
            _eliminar_queryset_con_archivos(cambio_colaborador.evidencias.all())
            cambio_colaborador.delete()

        # Limpieza adicional de evidencias de cambios (por seguridad)
        _eliminar_queryset_con_archivos(evento.evidencias_cambios.all())

        # Soft delete - Marca el evento como eliminado con timestamp timezone-aware
        evento.eliminado_en = timezone.now()
        evento.save()
        
        # Registrar el cambio
        ActividadCambio.objects.create(
            actividad=evento,
            responsable=usuario_maga,
            descripcion_cambio=f'Evento "{evento.nombre}" eliminado por {usuario_maga.username if usuario_maga else "sistema"}'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Evento eliminado exitosamente'
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar evento: {str(e)}'
        }, status=500)
@require_http_methods(["GET"])
def api_listar_proyectos_por_tipo(request, tipo_actividad):
    """
    Lista proyectos/eventos filtrados por tipo de actividad (ej: Capacitaci√≥n, Entrega, etc.)
    Devuelve informaci√≥n completa incluyendo imagen principal (primera evidencia)
    """
    try:
        from django.utils.timezone import localtime, is_aware, make_aware
        import pytz
        
        # Normalizar el tipo de actividad para la b√∫squeda
        tipo_map = {
            'capacitaciones': 'Capacitaci√≥n',
            'entregas': 'Entrega',
            'proyectos-ayuda': 'Proyecto de Ayuda'
        }
        
        tipo_nombre = tipo_map.get(tipo_actividad.lower())
        if not tipo_nombre:
            return JsonResponse({
                'success': False,
                'error': f'Tipo de actividad no v√°lido: {tipo_actividad}'
            }, status=400)
        
        # Buscar eventos de ese tipo que no est√©n eliminados
        eventos = Actividad.objects.filter(
            tipo__nombre=tipo_nombre,
            eliminado_en__isnull=True
        ).select_related(
            'tipo', 'comunidad', 'comunidad__region', 'responsable', 'portada'
        ).prefetch_related(
             'personal__usuario',
             'beneficiarios__beneficiario',
             'evidencias'
        ).order_by('-actualizado_en')
        
        proyectos_data = []
        for evento in eventos:
            try:
                # Obtener la primera evidencia como imagen principal
                primera_evidencia = evento.evidencias.filter(es_imagen=True).first()
                imagen_url = None
                if primera_evidencia and primera_evidencia.url_almacenamiento:
                    try:
                        imagen_url = primera_evidencia.url_almacenamiento
                    except:
                        imagen_url = None
                
                # Convertir fechas a zona horaria local
                try:
                    if evento.creado_en:
                        if not is_aware(evento.creado_en):
                            creado_en_aware = make_aware(evento.creado_en, pytz.UTC)
                        else:
                            creado_en_aware = evento.creado_en
                        creado_en_local = localtime(creado_en_aware)
                    else:
                        creado_en_local = None
                except:
                    creado_en_local = None
                
                try:
                    if evento.actualizado_en:
                        if not is_aware(evento.actualizado_en):
                            actualizado_en_aware = make_aware(evento.actualizado_en, pytz.UTC)
                        else:
                            actualizado_en_aware = evento.actualizado_en
                        actualizado_en_local = localtime(actualizado_en_aware)
                    else:
                        actualizado_en_local = None
                except:
                    actualizado_en_local = None
                
                # Obtener nombres del personal
                try:
                    personal_nombres = []
                    for ap in evento.personal.all()[:3]:
                        if ap.usuario:
                            personal_nombres.append(ap.usuario.nombre if ap.usuario.nombre else ap.usuario.username)
                        elif ap.colaborador:
                            personal_nombres.append(ap.colaborador.nombre)
                    personal_count = evento.personal.count()
                    if personal_count > 3:
                        personal_nombres.append(f'+{personal_count - 3} m√°s')
                except:
                    personal_nombres = []
                    personal_count = 0
                
                # Construir ubicaci√≥n de forma segura
                ubicacion = 'Sin ubicaci√≥n'
                region_nombre = None
                if evento.comunidad:
                    if evento.comunidad.region:
                        ubicacion = f"{evento.comunidad.nombre}, {evento.comunidad.region.nombre}"
                        region_nombre = evento.comunidad.region.nombre
                    else:
                        ubicacion = evento.comunidad.nombre
                
                portada_info = obtener_portada_evento(evento)
                imagen_url = portada_info['url'] if portada_info else None
                
                if not imagen_url:
                    primera_evidencia = evento.evidencias.filter(es_imagen=True).first()
                    if primera_evidencia and primera_evidencia.url_almacenamiento:
                        try:
                            imagen_url = primera_evidencia.url_almacenamiento
                        except Exception:
                            imagen_url = None
                
                proyectos_data.append({
                    'id': str(evento.id),
                    'nombre': evento.nombre or 'Sin nombre',
                    'tipo': evento.tipo.nombre if evento.tipo else 'Sin tipo',
                    'comunidad': evento.comunidad.nombre if evento.comunidad else 'Sin comunidad',
                    'region': region_nombre,
                    'ubicacion': ubicacion,
                    'fecha': str(evento.fecha) if evento.fecha else '',
                    'estado': evento.estado or 'planificado',
                    'estado_display': evento.get_estado_display() if hasattr(evento, 'get_estado_display') else evento.estado,
                    'descripcion': evento.descripcion or '',
                    'imagen_principal': imagen_url,
                    'tarjetas_datos': obtener_tarjetas_datos(evento),
                    'personal_count': personal_count,
                    'personal_nombres': ', '.join(personal_nombres) if personal_nombres else 'Sin personal',
                    'beneficiarios_count': evento.beneficiarios.count() if hasattr(evento, 'beneficiarios') else 0,
                    'evidencias_count': evento.evidencias.count() if hasattr(evento, 'evidencias') else 0,
                    'creado_en': creado_en_local.strftime('%Y-%m-%d') if creado_en_local else '',
                    'actualizado_en': actualizado_en_local.strftime('%Y-%m-%d') if actualizado_en_local else '',
                    'creado_en_formatted': creado_en_local.strftime('%d de %B de %Y') if creado_en_local else '',
                    'actualizado_en_formatted': actualizado_en_local.strftime('%d de %B de %Y') if actualizado_en_local else ''
                })
            except Exception as item_error:
                # Si hay error con un evento espec√≠fico, continuar con el siguiente
                print(f"Error procesando evento {evento.id}: {item_error}")
                import traceback
                traceback.print_exc()
                continue
        
        return JsonResponse({
            'success': True,
            'tipo': tipo_nombre,
            'proyectos': proyectos_data,
            'total': len(proyectos_data)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al listar proyectos: {str(e)}'
        }, status=500)
@require_http_methods(["GET"])
def api_ultimos_proyectos(request):
    """
    Devuelve los √∫ltimos proyectos/eventos creados o actualizados (m√°ximo 3)
    """
    try:
        from django.utils.timezone import localtime, is_aware, make_aware
        import pytz
        
        # Obtener los √∫ltimos 3 eventos actualizados
        eventos = Actividad.objects.filter(
            eliminado_en__isnull=True
        ).select_related(
            'tipo', 'comunidad', 'comunidad__region', 'responsable', 'portada'
        ).prefetch_related(
            'personal__usuario',
            'beneficiarios__beneficiario',
            'evidencias'
        ).order_by('-actualizado_en')[:3]
        
        proyectos_data = []
        for evento in eventos:
            try:
                # Obtener la primera evidencia como imagen principal
                primera_evidencia = evento.evidencias.filter(es_imagen=True).first()
                imagen_url = None
                if primera_evidencia and primera_evidencia.url_almacenamiento:
                    try:
                        imagen_url = primera_evidencia.url_almacenamiento
                    except:
                        imagen_url = None
                
                # Convertir fecha a zona horaria local
                try:
                    if evento.fecha:
                        fecha_obj = evento.fecha
                    elif evento.actualizado_en:
                        fecha_obj = evento.actualizado_en.date() if hasattr(evento.actualizado_en, 'date') else evento.actualizado_en
                    else:
                        fecha_obj = evento.creado_en.date() if hasattr(evento.creado_en, 'date') else evento.creado_en
                except:
                    fecha_obj = None
                
                # Construir ubicaci√≥n de forma segura
                ubicacion = 'Sin ubicaci√≥n'
                if evento.comunidad:
                    if evento.comunidad.region:
                        ubicacion = f"{evento.comunidad.nombre}, {evento.comunidad.region.nombre}"
                    else:
                        ubicacion = evento.comunidad.nombre
                
                # Obtener conteo de personal
                try:
                    personal_count = evento.personal.count()
                    personal_nombres = []
                    for ap in evento.personal.all()[:3]:
                        if ap.usuario:
                            personal_nombres.append(ap.usuario.nombre if ap.usuario.nombre else ap.usuario.username)
                        elif ap.colaborador:
                            personal_nombres.append(ap.colaborador.nombre)
                    if personal_count > 3:
                        personal_nombres.append(f'+{personal_count - 3} m√°s')
                except:
                    personal_count = 0
                    personal_nombres = []
                
                proyectos_data.append({
                    'id': str(evento.id),
                    'nombre': evento.nombre or 'Sin nombre',
                    'tipo': evento.tipo.nombre if evento.tipo else 'Sin tipo',
                    'ubicacion': ubicacion,
                    'fecha': str(fecha_obj) if fecha_obj else '',
                    'estado': evento.estado or 'planificado',
                    'descripcion': evento.descripcion or '',
                    'imagen_principal': imagen_url,
                    'portada': obtener_portada_evento(evento),
                    'tarjetas_datos': obtener_tarjetas_datos(evento),
                    'personal_count': personal_count,
                    'personal_nombres': ', '.join(personal_nombres) if personal_nombres else 'Sin personal',
                })
            except Exception as item_error:
                print(f"Error procesando evento para √∫ltimos proyectos {evento.id}: {item_error}")
                import traceback
                traceback.print_exc()
                continue
        
        return JsonResponse({
            'success': True,
            'proyectos': proyectos_data,
            'total': len(proyectos_data)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener √∫ltimos proyectos: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def api_ultimos_eventos_inicio(request):
    """
    Devuelve los √∫ltimos 6 eventos para mostrar en el inicio
    """
    try:
        from django.utils.timezone import localtime, is_aware, make_aware
        import pytz
        
        # Obtener los √∫ltimos 6 eventos no eliminados
        eventos = Actividad.objects.filter(
            eliminado_en__isnull=True
        ).select_related(
            'tipo', 'comunidad', 'comunidad__region', 'responsable', 'portada'
        ).prefetch_related(
            'evidencias'
        ).order_by('-creado_en')[:6]
        
        eventos_data = []
        for evento in eventos:
            try:
                portada_info = obtener_portada_evento(evento)
                imagen_url = portada_info['url'] if portada_info else None

                if not imagen_url:
                    primera_evidencia = evento.evidencias.filter(es_imagen=True).first()
                    if primera_evidencia and primera_evidencia.url_almacenamiento:
                        imagen_url = primera_evidencia.url_almacenamiento

                try:
                    if evento.fecha:
                        fecha_obj = evento.fecha
                    elif evento.creado_en:
                        if not is_aware(evento.creado_en):
                            creado_aware = make_aware(evento.creado_en, pytz.UTC)
                        else:
                            creado_aware = evento.creado_en
                        fecha_local = localtime(creado_aware)
                        fecha_obj = fecha_local.date()
                    else:
                        fecha_obj = None
                except Exception:
                    fecha_obj = None

                if fecha_obj:
                    meses = {
                        1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
                        7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
                    }
                    mes = meses.get(fecha_obj.month, '')
                    dia = fecha_obj.day
                    anio = fecha_obj.year
                else:
                    mes = ''
                    dia = ''
                    anio = ''

                eventos_data.append({
                    'id': str(evento.id),
                    'nombre': evento.nombre or 'Sin nombre',
                    'descripcion': evento.descripcion[:100] + '...' if evento.descripcion and len(evento.descripcion) > 100 else (evento.descripcion or 'Sin descripci√≥n'),
                    'imagen_url': imagen_url,
                    'portada': portada_info,
                    'fecha_mes': mes,
                    'fecha_dia': dia,
                    'fecha_anio': anio,
                    'tipo': evento.tipo.nombre if evento.tipo else 'Sin tipo',
                    'comunidad': evento.comunidad.nombre if evento.comunidad else 'Sin comunidad',
                    'tarjetas_datos': obtener_tarjetas_datos(evento)
                })
            except Exception as item_error:
                print(f"Error procesando evento para inicio {evento.id}: {item_error}")
                import traceback
                traceback.print_exc()
                continue
        
        return JsonResponse({
            'success': True,
            'eventos': eventos_data,
            'total': len(eventos_data)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener √∫ltimos eventos: {str(e)}'
        }, status=500)
@require_http_methods(["GET"])
def api_calendar_events(request):
    """Eventos para el calendario entre start y end (YYYY-MM-DD). Muestra todos los eventos sin importar estado."""
    start = request.GET.get('start')
    end = request.GET.get('end')
    if not start or not end:
        return JsonResponse([], safe=False)
    try:
        # Mostrar todos los eventos sin importar estado (en_progreso, completado, planificado, cancelado)
        # Mostrar TODOS los eventos para TODOS los usuarios, sin filtros por usuario o colaborador
        # Incluir eventos con fecha en el rango O eventos sin fecha pero creados en el rango
        from django.db.models import Q
        actividades = Actividad.objects.filter(
            eliminado_en__isnull=True
        ).filter(
            Q(fecha__gte=start, fecha__lte=end) | 
            Q(fecha__isnull=True, creado_en__date__gte=start, creado_en__date__lte=end)
        ).select_related('tipo', 'comunidad', 'comunidad__region', 'responsable')

        # Obtener IDs de actividades para consultas optimizadas
        actividad_ids = list(actividades.values_list('id', flat=True))
        
        # Personal asignado (tabla ActividadPersonal)
        # Obtener usuarios asignados a trav√©s de ActividadPersonal
        usuarios_asignados_map = {}
        for ap in ActividadPersonal.objects.filter(actividad_id__in=actividad_ids).select_related('usuario'):
            key = str(ap.actividad_id)
            if ap.usuario_id:
                if key not in usuarios_asignados_map:
                    usuarios_asignados_map[key] = set()
                usuarios_asignados_map[key].add(str(ap.usuario_id))

        data = []
        for a in actividades:
            # IDs de usuarios asignados a trav√©s de ActividadPersonal
            usuarios_ids = usuarios_asignados_map.get(str(a.id), set()).copy()
            
            # Agregar responsable si existe
            if a.responsable_id:
                usuarios_ids.add(str(a.responsable_id))
            
            # Obtener nombres de responsables/encargados (usuario.nombre o username)
            responsables_nombres = []
            if a.responsable:
                responsables_nombres.append(a.responsable.nombre if a.responsable.nombre else a.responsable.username)
            
            # Obtener nombres de personal asignado
            personal_asignado = ActividadPersonal.objects.filter(actividad_id=a.id).select_related('usuario')
            for ap in personal_asignado:
                if ap.usuario:
                    nombre = ap.usuario.nombre if ap.usuario.nombre else ap.usuario.username
                    if nombre not in responsables_nombres:
                        responsables_nombres.append(nombre)
            
            responsables_texto = ', '.join(responsables_nombres) if responsables_nombres else None

            # Usar fecha del evento si existe, sino usar fecha de creaci√≥n como fallback
            event_date = None
            if a.fecha:
                event_date = a.fecha.isoformat()
            elif a.creado_en:
                # Si no tiene fecha, usar la fecha de creaci√≥n
                event_date = a.creado_en.date().isoformat()
            
            data.append({
                'date': event_date,
                'name': a.nombre,
                'description': a.descripcion,
                'status': a.estado,
                'responsable': a.responsable.username if a.responsable else None,
                'owners': responsables_texto,  # Texto con todos los responsables/encargados
                'usuarios_asignados_ids': list(usuarios_ids),
            })
        return JsonResponse(data, safe=False)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET", "POST"])
@login_required
def api_avances(request):
    """Obtiene avances/cambios de eventos para el calendario por fecha. Consulta la tabla eventos_cambios_colaboradores.
    Agrupa por grupo_id para evitar duplicados cuando hay m√∫ltiples colaboradores responsables."""
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse([], safe=False)
    
    try:
        from django.utils.timezone import localtime, is_aware, make_aware
        import pytz
        from django.db import connection
        
        # Obtener informaci√≥n del usuario actual
        is_admin = False
        colaborador_usuario_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                is_admin = usuario_maga.rol == 'admin'
                if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
                    colaborador_usuario_id = str(usuario_maga.colaborador.id)
        except Exception:
            pass
        
        # Construir la consulta base agrupando por grupo_id para evitar duplicados
        # Agrupamos por grupo_id y concatenamos todos los colaboradores del mismo grupo
        query_base = """
            SELECT ecc.grupo_id::text,
                   ecc.actividad_id::text,
                   ecc.descripcion_cambio,
                   MIN(ecc.fecha_cambio) as fecha_cambio,
                   a.nombre as evento_nombre,
                   STRING_AGG(DISTINCT COALESCE(c.nombre, 'Colaborador'), ', ' ORDER BY COALESCE(c.nombre, 'Colaborador')) as colaboradores_nombres,
                   STRING_AGG(DISTINCT COALESCE(com.nombre, ''), ', ' ORDER BY COALESCE(com.nombre, '')) FILTER (WHERE com.nombre IS NOT NULL) as comunidades_nombres,
                   STRING_AGG(DISTINCT COALESCE(r.nombre, ''), ', ' ORDER BY COALESCE(r.nombre, '')) FILTER (WHERE r.nombre IS NOT NULL) as regiones_nombres
            FROM eventos_cambios_colaboradores ecc
            INNER JOIN actividades a ON a.id = ecc.actividad_id
            LEFT JOIN colaboradores c ON c.id = ecc.colaborador_id
            LEFT JOIN comunidades com ON com.id = a.comunidad_id
            LEFT JOIN regiones r ON r.id = com.region_id
            WHERE (ecc.fecha_cambio AT TIME ZONE 'America/Guatemala')::date = %s::date
              AND a.eliminado_en IS NULL
        """
        
        # Si no es admin, filtrar por actividades donde el colaborador del usuario est√° asignado
        if not is_admin and colaborador_usuario_id:
            query_base += """
              AND ecc.actividad_id IN (
                  SELECT DISTINCT ap.actividad_id
                  FROM actividad_personal ap
                  WHERE ap.colaborador_id = %s
              )
            """
            params = [date_str, colaborador_usuario_id]
        else:
            params = [date_str]
        
        query_base += """
            GROUP BY ecc.grupo_id, ecc.actividad_id, ecc.descripcion_cambio, a.nombre
            ORDER BY MIN(ecc.fecha_cambio) ASC
        """
        
        with connection.cursor() as cur:
            cur.execute(query_base, params)
            rows = cur.fetchall()
        
        results = []
        for row in rows:
            grupo_id, actividad_id, descripcion, fecha_cambio, evento_nombre, colaboradores_nombres, comunidades_nombres, regiones_nombres = row
            
            # Si no hay evento_nombre desde la consulta, obtenerlo manualmente (fallback)
            if not evento_nombre:
                try:
                    actividad = Actividad.objects.get(id=actividad_id)
                    evento_nombre = actividad.nombre
                except Actividad.DoesNotExist:
                    continue
            
            # Usar grupo_id como id para evitar duplicados en el frontend
            # Los colaboradores ya est√°n concatenados en colaboradores_nombres
            responsable_display = colaboradores_nombres if colaboradores_nombres else 'Sistema'
            
            # Convertir fecha_cambio a zona horaria de Guatemala
            if fecha_cambio:
                guatemala_tz = pytz.timezone('America/Guatemala')
                if fecha_cambio.tzinfo is None:
                    fecha_guatemala = guatemala_tz.localize(fecha_cambio)
                else:
                    fecha_guatemala = fecha_cambio.astimezone(guatemala_tz)
                fecha_iso = fecha_guatemala.date().isoformat()
                hora_str = fecha_guatemala.strftime('%H:%M')
            else:
                fecha_iso = date_str
                hora_str = ''
            
            results.append({
                'id': grupo_id,  # Usar grupo_id como id para evitar duplicados
                'actividad_id': actividad_id,
                'evento_nombre': evento_nombre,
                'fecha': fecha_iso,
                'hora': hora_str,
                'responsable': responsable_display,  # Ya contiene todos los colaboradores concatenados
                'descripcion': descripcion or '',
                'comunidades': comunidades_nombres if comunidades_nombres else '',  # Comunidades donde se trabaj√≥
                'regiones': regiones_nombres if regiones_nombres else ''  # Regiones donde se trabaj√≥
            })
        
        return JsonResponse(results, safe=False)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
def api_reminders(request):
    """Lista o crea recordatorios usando tablas recordatorios y recordatorio_colaboradores."""
    if request.method == 'GET':
        date_str = request.GET.get('date')
        if not date_str:
            return JsonResponse([], safe=False)
        
        # Obtener informaci√≥n del usuario actual
        is_admin = False
        user_id = None
        colaborador_usuario_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                user_id = str(usuario_maga.id)
                is_admin = usuario_maga.rol == 'admin'
                if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
                    colaborador_usuario_id = str(usuario_maga.colaborador.id)
        except Exception:
            pass
        
        try:
            with connection.cursor() as cur:
                # Si es admin, mostrar todos los recordatorios del d√≠a EXCEPTO los personales
                # Un recordatorio es personal si:
                # - El creador es personal (no admin)
                # - Solo tiene un colaborador asignado
                # - Ese colaborador pertenece al mismo usuario que cre√≥ el recordatorio
                if is_admin:
                    cur.execute(
                        """
                        SELECT r.id::text,
                               r.actividad_id::text,
                               r.created_by::text,
                               COALESCE(u.nombre, u.username, 'Usuario desconocido') as created_by_name,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado
                        FROM recordatorios r
                        LEFT JOIN usuarios u ON u.id = r.created_by
                        WHERE (r.due_at AT TIME ZONE 'America/Guatemala')::date = %s::date
                          AND NOT (
                            -- Excluir recordatorios personales
                            u.rol = 'personal'
                            AND (
                              -- Caso 1: Solo tiene un colaborador y ese colaborador pertenece al creador
                              (
                                SELECT COUNT(*)
                                FROM recordatorio_colaboradores rc
                                WHERE rc.recordatorio_id = r.id
                              ) = 1
                              AND EXISTS (
                                SELECT 1
                                FROM recordatorio_colaboradores rc
                                JOIN colaboradores c ON c.id = rc.colaborador_id
                                WHERE rc.recordatorio_id = r.id
                                  AND c.usuario_id = r.created_by
                              )
                            )
                          )
                        ORDER BY r.due_at ASC
                        """,
                        [date_str]
                    )
                else:
                    # Para usuarios personales: mostrar recordatorios creados por ellos O donde su colaborador est√° incluido
                    if user_id:
                        if colaborador_usuario_id:
                            # Mostrar recordatorios creados por el usuario O donde su colaborador est√° en recordatorio_colaboradores
                            cur.execute(
                                """
                                SELECT DISTINCT r.id::text,
                                       r.actividad_id::text,
                                       r.created_by::text,
                                       COALESCE(u.nombre, u.username, 'Usuario desconocido') as created_by_name,
                                       r.titulo,
                                       r.descripcion,
                                       r.due_at,
                                       r.enviar_notificacion,
                                       r.enviado
                                FROM recordatorios r
                                LEFT JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                                LEFT JOIN usuarios u ON u.id = r.created_by
                                WHERE (r.due_at AT TIME ZONE 'America/Guatemala')::date = %s::date
                                  AND (r.created_by = %s OR rc.colaborador_id = %s)
                                ORDER BY r.due_at ASC
                                """,
                                [date_str, user_id, colaborador_usuario_id]
                            )
                        else:
                            # Si no tiene colaborador vinculado, solo mostrar los que cre√≥
                            cur.execute(
                                """
                                SELECT r.id::text,
                                       r.actividad_id::text,
                                       r.created_by::text,
                                       COALESCE(u.nombre, u.username, 'Usuario desconocido') as created_by_name,
                                       r.titulo,
                                       r.descripcion,
                                       r.due_at,
                                       r.enviar_notificacion,
                                       r.enviado
                                FROM recordatorios r
                                LEFT JOIN usuarios u ON u.id = r.created_by
                                WHERE (r.due_at AT TIME ZONE 'America/Guatemala')::date = %s::date
                                  AND r.created_by = %s
                                ORDER BY r.due_at ASC
                                """,
                                [date_str, user_id]
                            )
                    else:
                        # Si no hay user_id, no mostrar nada
                        rows = []
                        results = []
                        return JsonResponse(results, safe=False)
                rows = cur.fetchall()

            results = []
            for row in rows:
                rid, act_id, created_by, created_by_name, titulo, desc, due_at, enviar, enviado = row
                # owners text
                with connection.cursor() as cur2:
                    cur2.execute(
                        """
                        SELECT c.nombre
                        FROM recordatorio_colaboradores rc
                        JOIN colaboradores c ON c.id = rc.colaborador_id
                        WHERE rc.recordatorio_id = %s
                        ORDER BY c.nombre
                        """,
                        [rid]
                    )
                    owners_names = [r[0] for r in cur2.fetchall()]
                owners_text = ', '.join(owners_names)
                
                # Extraer fecha y hora correctamente en zona horaria de Guatemala
                if due_at:
                    import pytz
                    guatemala_tz = pytz.timezone('America/Guatemala')
                    # Si due_at ya es timezone-aware, convertir a Guatemala
                    if due_at.tzinfo is None:
                        due_at_guatemala = guatemala_tz.localize(due_at)
                    else:
                        due_at_guatemala = due_at.astimezone(guatemala_tz)
                    
                    # Extraer hora y fecha por separado para evitar problemas de cambio de d√≠a
                    time_str = due_at_guatemala.strftime('%H:%M')
                    # Usar directamente la fecha del datetime en zona horaria de Guatemala
                    # Importante: usar .date() despu√©s de la conversi√≥n de timezone
                    date_iso = due_at_guatemala.date().isoformat()
                else:
                    time_str = ''
                    date_iso = date_str
                    
                results.append({
                    'id': rid,
                    'date': date_iso,
                    'time': time_str,
                    'description': desc or '',
                    'event_name': titulo or None,
                    'created_by': created_by,
                    'created_by_name': created_by_name or 'Usuario desconocido',
                    'owners_text': owners_text
                })
            return JsonResponse(results, safe=False)
        except Exception as e:
            return JsonResponse([], safe=False)

    # POST crear recordatorio
    try:
        payload = json.loads(request.body or '{}')
        date = payload.get('date')
        time = payload.get('time') or '08:00'
        event_name = payload.get('event_name')
        description = payload.get('description') or ''
        owners = payload.get('owners') or []
        recordar = payload.get('recordar', False)  # Opci√≥n de reenv√≠o 10 minutos despu√©s

        # due_at timestamptz (zona horaria de Guatemala)
        try:
            from django.utils.timezone import make_aware
            import pytz
            # Crear datetime naive desde la fecha y hora (parsear por separado para evitar problemas)
            date_parts = date.split('-')
            time_parts = time.split(':')
            if len(date_parts) == 3 and len(time_parts) >= 2:
                year, month, day = int(date_parts[0]), int(date_parts[1]), int(date_parts[2])
                hour, minute = int(time_parts[0]), int(time_parts[1])
                dt_naive = datetime(year, month, day, hour, minute)
            else:
                dt_naive = datetime.strptime(f"{date} {time}", "%Y-%m-%d %H:%M")
            # Convertir a timezone-aware usando zona horaria de Guatemala
            guatemala_tz = pytz.timezone('America/Guatemala')
            due_at = guatemala_tz.localize(dt_naive)
        except Exception as e:
            # Fallback: usar hora actual con timezone de Guatemala
            import pytz
            guatemala_tz = pytz.timezone('America/Guatemala')
            due_at = timezone.now().astimezone(guatemala_tz)

        # Resolver created_by (UUID en tabla usuarios). Si el auth user no es el mismo modelo, buscar por username
        created_by = None
        try:
            with connection.cursor() as curu:
                curu.execute("SELECT id::text FROM usuarios WHERE username = %s LIMIT 1", [getattr(request.user, 'username', None)])
                rowu = curu.fetchone()
                if rowu:
                    created_by = rowu[0]
        except Exception:
            created_by = None
        actividad_id = None
        if event_name:
            act = Actividad.objects.filter(nombre__iexact=event_name).order_by('-creado_en').first()
            if act:
                actividad_id = str(act.id)

        # Insertar recordatorio (enviar_notificacion se usa para activar notificaciones, 
        # y guardamos recordar en un campo adicional si existe, o lo guardamos en enviar_notificacion)
        # Por ahora usaremos enviar_notificacion para activar notificaciones y agregaremos un campo para recordar
        # Si no existe el campo recordar en la BD, lo guardaremos como metadata o usaremos enviar_notificacion
        with connection.cursor() as cur:
            # Verificar si existe columna recordar, si no, usar enviar_notificacion para ambos
            try:
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name='recordatorios' AND column_name='recordar'
                """)
                tiene_recordar = cur.fetchone() is not None
            except:
                tiene_recordar = False
            
            if tiene_recordar:
                cur.execute(
                    """
                    INSERT INTO recordatorios (actividad_id, created_by, titulo, descripcion, due_at, enviar_notificacion, recordar)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    [actividad_id, created_by, event_name, description, due_at, True, recordar]
                )
            else:
                # Si no existe el campo recordar, solo guardamos enviar_notificacion
                cur.execute(
                    """
                    INSERT INTO recordatorios (actividad_id, created_by, titulo, descripcion, due_at, enviar_notificacion)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    [actividad_id, created_by, event_name, description, due_at, True]
                )
            rid = cur.fetchone()[0]
        # Insertar involucrados
        if owners:
            with connection.cursor() as cur2:
                cur2.executemany(
                    """
                    INSERT INTO recordatorio_colaboradores (recordatorio_id, colaborador_id)
                    VALUES (%s, %s)
                    ON CONFLICT (recordatorio_id, colaborador_id) DO NOTHING
                    """,
                    [(rid, o) for o in owners]
                )
        return JsonResponse({'id': str(rid)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
@require_http_methods(["GET"])
@login_required
def api_reminders_pending(request):
    """Obtiene recordatorios pendientes para el usuario actual (para notificaciones).
    Solo devuelve recordatorios donde el usuario est√° involucrado."""
    try:
        from django.utils.timezone import now
        import pytz
        from django.db import connection
        
        # Obtener informaci√≥n del usuario actual
        user_id = None
        colaborador_usuario_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                user_id = str(usuario_maga.id)
                if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
                    colaborador_usuario_id = str(usuario_maga.colaborador.id)
        except Exception:
            pass
        
        if not user_id:
            return JsonResponse([], safe=False)
        
        # Obtener recordatorios pendientes donde el usuario est√° involucrado
        guatemala_tz = pytz.timezone('America/Guatemala')
        ahora_guatemala = now().astimezone(guatemala_tz)
        
        with connection.cursor() as cur:
            # Verificar si existe columna recordar
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='recordatorios' AND column_name='recordar'
            """)
            tiene_recordar = cur.fetchone() is not None
            
            if colaborador_usuario_id:
                # Buscar recordatorios donde el colaborador del usuario est√° involucrado
                # Incluir recordatorios que ya pasaron pero que tienen "recordar" activado (para reenv√≠o)
                # Solo incluir recordatorios que est√°n dentro de los 15 minutos despu√©s de la hora
                if tiene_recordar:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               COALESCE(r.recordar, FALSE) as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        INNER JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND rc.colaborador_id = %s
                          AND (
                            -- Recordatorios futuros que no han sido enviados
                            (r.due_at >= %s AND (r.enviado = FALSE OR r.enviado IS NULL))
                            OR
                            -- Recordatorios pasados con "recordar" activado, dentro de los 15 minutos
                            (r.due_at < %s 
                             AND r.due_at >= %s - INTERVAL '15 minutes'
                             AND COALESCE(r.recordar, FALSE) = TRUE)
                          )
                        ORDER BY r.due_at ASC
                        """,
                        [colaborador_usuario_id, ahora_guatemala, ahora_guatemala, ahora_guatemala]
                    )
                else:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               FALSE as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        INNER JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.due_at >= %s
                          AND rc.colaborador_id = %s
                          AND (r.enviado = FALSE OR r.enviado IS NULL)
                        ORDER BY r.due_at ASC
                        """,
                        [ahora_guatemala, colaborador_usuario_id]
                    )
            else:
                # Si no tiene colaborador, solo buscar los que cre√≥
                if tiene_recordar:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               COALESCE(r.recordar, FALSE) as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.created_by = %s
                          AND (
                            -- Recordatorios futuros que no han sido enviados
                            (r.due_at >= %s AND (r.enviado = FALSE OR r.enviado IS NULL))
                            OR
                            -- Recordatorios pasados con "recordar" activado, dentro de los 15 minutos
                            (r.due_at < %s 
                             AND r.due_at >= %s - INTERVAL '15 minutes'
                             AND COALESCE(r.recordar, FALSE) = TRUE)
                          )
                        ORDER BY r.due_at ASC
                        """,
                        [user_id, ahora_guatemala, ahora_guatemala, ahora_guatemala]
                    )
                else:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               FALSE as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.due_at >= %s
                          AND r.created_by = %s
                          AND (r.enviado = FALSE OR r.enviado IS NULL)
                        ORDER BY r.due_at ASC
                        """,
                        [ahora_guatemala, user_id]
                    )
            
            rows = cur.fetchall()
        
        results = []
        for row in rows:
            # Extraer datos de la fila seg√∫n si tiene el campo recordar o no
            if tiene_recordar:
                if len(row) >= 8:
                    rid, titulo, desc, due_at, enviar, enviado, recordar, evento_nombre = row
                else:
                    rid, titulo, desc, due_at, enviar, enviado, recordar = row
                    evento_nombre = None
            else:
                if len(row) >= 8:
                    rid, titulo, desc, due_at, enviar, enviado, recordar, evento_nombre = row[0], row[1], row[2], row[3], row[4], row[5], False, row[7]
                else:
                    rid, titulo, desc, due_at, enviar, enviado = row[0], row[1], row[2], row[3], row[4], row[5]
                    recordar = False
                    evento_nombre = row[7] if len(row) > 7 else None
            
            # Obtener personal involucrado
            owners_names = []
            with connection.cursor() as cur2:
                cur2.execute(
                    """
                    SELECT c.nombre
                    FROM recordatorio_colaboradores rc
                    JOIN colaboradores c ON c.id = rc.colaborador_id
                    WHERE rc.recordatorio_id = %s
                    ORDER BY c.nombre
                    """,
                    [rid]
                )
                owners_names = [r[0] for r in cur2.fetchall()]
            owners_text = ', '.join(owners_names) if owners_names else 'Sin personal asignado'
            
            # Convertir due_at a zona horaria de Guatemala
            if due_at:
                if due_at.tzinfo is None:
                    due_at_guatemala = guatemala_tz.localize(due_at)
                else:
                    due_at_guatemala = due_at.astimezone(guatemala_tz)
                
                # Calcular tiempo hasta el recordatorio
                tiempo_restante = (due_at_guatemala - ahora_guatemala).total_seconds()
                
                # L√≠mite de 15 minutos (900 segundos) despu√©s de la hora del recordatorio
                LIMITE_MINUTOS = 15
                LIMITE_SEGUNDOS = LIMITE_MINUTOS * 60  # 900 segundos
                
                # Incluir recordatorios:
                # 1. Que a√∫n no han llegado (tiempo_restante > 0)
                # 2. Que ya pasaron pero est√°n dentro de los 15 minutos despu√©s de la hora (tiempo_restante < 0 pero abs(tiempo_restante) <= 900)
                tiempo_absoluto = abs(tiempo_restante)
                incluir = False
                
                if tiempo_restante > 0:
                    # A√∫n no ha llegado, incluir
                    incluir = True
                elif tiempo_restante < 0:
                    # Ya pas√≥, solo incluir si est√° dentro de los 15 minutos
                    if tiempo_absoluto <= LIMITE_SEGUNDOS:
                        incluir = True
                    else:
                        # Pasaron m√°s de 15 minutos, no incluir
                        incluir = False
                
                if incluir:
                    # Formatear fecha y hora
                    fecha_str = due_at_guatemala.strftime('%d/%m/%Y')
                    hora_str = due_at_guatemala.strftime('%H:%M')
                    
                    results.append({
                        'id': rid,
                        'titulo': titulo or 'Recordatorio',
                        'descripcion': desc or '',
                        'due_at': due_at_guatemala.isoformat(),
                        'due_at_timestamp': int(due_at_guatemala.timestamp() * 1000),  # En milisegundos
                        'tiempo_restante_segundos': int(tiempo_restante),
                        'recordar': bool(recordar) if recordar is not None else False,
                        'evento_nombre': evento_nombre or 'Sin evento',
                        'fecha': fecha_str,
                        'hora': hora_str,
                        'owners_text': owners_text,
                        'enviado': bool(enviado) if enviado is not None else False
                    })
        
        return JsonResponse(results, safe=False)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
@login_required
def api_reminders_check_background(request):
    """Endpoint para que el Service Worker verifique recordatorios en segundo plano.
    Solo devuelve recordatorios si el usuario tiene sesi√≥n activa.
    Este endpoint puede ser llamado incluso cuando la p√°gina est√° cerrada."""
    try:
        from django.utils.timezone import now
        import pytz
        from django.db import connection
        
        # Verificar que el usuario tiene sesi√≥n activa
        if not request.user.is_authenticated:
            return JsonResponse({'reminders': [], 'session_active': False}, safe=False)
        
        # Obtener informaci√≥n del usuario actual
        user_id = None
        colaborador_usuario_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                user_id = str(usuario_maga.id)
                if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
                    colaborador_usuario_id = str(usuario_maga.colaborador.id)
        except Exception:
            pass
        
        if not user_id:
            return JsonResponse({'reminders': [], 'session_active': False}, safe=False)
        
        # Obtener recordatorios pendientes (misma l√≥gica que api_reminders_pending)
        guatemala_tz = pytz.timezone('America/Guatemala')
        ahora_guatemala = now().astimezone(guatemala_tz)
        
        with connection.cursor() as cur:
            # Verificar si existe columna recordar
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='recordatorios' AND column_name='recordar'
            """)
            tiene_recordar = cur.fetchone() is not None
            
            if colaborador_usuario_id:
                if tiene_recordar:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               COALESCE(r.recordar, FALSE) as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        INNER JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND rc.colaborador_id = %s
                          AND (
                            (r.due_at >= %s AND (r.enviado = FALSE OR r.enviado IS NULL))
                            OR
                            (r.due_at < %s 
                             AND r.due_at >= %s - INTERVAL '15 minutes'
                             AND COALESCE(r.recordar, FALSE) = TRUE)
                          )
                        ORDER BY r.due_at ASC
                        """,
                        [colaborador_usuario_id, ahora_guatemala, ahora_guatemala, ahora_guatemala]
                    )
                else:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               FALSE as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        INNER JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.due_at >= %s
                          AND rc.colaborador_id = %s
                          AND (r.enviado = FALSE OR r.enviado IS NULL)
                        ORDER BY r.due_at ASC
                        """,
                        [ahora_guatemala, colaborador_usuario_id]
                    )
            else:
                if tiene_recordar:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               COALESCE(r.recordar, FALSE) as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.created_by = %s
                          AND (
                            (r.due_at >= %s AND (r.enviado = FALSE OR r.enviado IS NULL))
                            OR
                            (r.due_at < %s 
                             AND r.due_at >= %s - INTERVAL '15 minutes'
                             AND COALESCE(r.recordar, FALSE) = TRUE)
                          )
                        ORDER BY r.due_at ASC
                        """,
                        [user_id, ahora_guatemala, ahora_guatemala, ahora_guatemala]
                    )
                else:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               FALSE as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.due_at >= %s
                          AND r.created_by = %s
                          AND (r.enviado = FALSE OR r.enviado IS NULL)
                        ORDER BY r.due_at ASC
                        """,
                        [ahora_guatemala, user_id]
                    )
            
            rows = cur.fetchall()
        
        results = []
        for row in rows:
            if tiene_recordar:
                if len(row) >= 8:
                    rid, titulo, desc, due_at, enviar, enviado, recordar, evento_nombre = row
                else:
                    rid, titulo, desc, due_at, enviar, enviado, recordar = row
                    evento_nombre = None
            else:
                if len(row) >= 8:
                    rid, titulo, desc, due_at, enviar, enviado, recordar, evento_nombre = row[0], row[1], row[2], row[3], row[4], row[5], False, row[7]
                else:
                    rid, titulo, desc, due_at, enviar, enviado = row[0], row[1], row[2], row[3], row[4], row[5]
                    recordar = False
                    evento_nombre = row[7] if len(row) > 7 else None
            
            # Obtener personal involucrado
            owners_names = []
            with connection.cursor() as cur2:
                cur2.execute(
                    """
                    SELECT c.nombre
                    FROM recordatorio_colaboradores rc
                    JOIN colaboradores c ON c.id = rc.colaborador_id
                    WHERE rc.recordatorio_id = %s
                    ORDER BY c.nombre
                    """,
                    [rid]
                )
                owners_names = [r[0] for r in cur2.fetchall()]
            owners_text = ', '.join(owners_names) if owners_names else 'Sin personal asignado'
            
            # Convertir due_at a zona horaria de Guatemala
            if due_at:
                if due_at.tzinfo is None:
                    due_at_guatemala = guatemala_tz.localize(due_at)
                else:
                    due_at_guatemala = due_at.astimezone(guatemala_tz)
                
                # Calcular tiempo hasta el recordatorio
                tiempo_restante = (due_at_guatemala - ahora_guatemala).total_seconds()
                
                # L√≠mite de 15 minutos (900 segundos) despu√©s de la hora del recordatorio
                LIMITE_MINUTOS = 15
                LIMITE_SEGUNDOS = LIMITE_MINUTOS * 60
                
                tiempo_absoluto = abs(tiempo_restante)
                incluir = False
                
                # SOLO incluir recordatorios que YA PASARON (no futuros)
                # Esto evita que se env√≠en notificaciones cuando se crea el recordatorio
                if tiempo_restante <= 0:
                    # Ya pas√≥, solo incluir si est√° dentro de los 15 minutos
                    if tiempo_absoluto <= LIMITE_SEGUNDOS:
                        incluir = True
                # NO incluir recordatorios futuros - esperar a que llegue la hora
                
                if incluir:
                    # Formatear fecha y hora
                    fecha_str = due_at_guatemala.strftime('%d/%m/%Y')
                    hora_str = due_at_guatemala.strftime('%H:%M')
                    
                    results.append({
                        'id': rid,
                        'titulo': titulo or 'Recordatorio',
                        'descripcion': desc or '',
                        'due_at': due_at_guatemala.isoformat(),
                        'due_at_timestamp': int(due_at_guatemala.timestamp() * 1000),
                        'tiempo_restante_segundos': int(tiempo_restante),
                        'recordar': bool(recordar) if recordar is not None else False,
                        'evento_nombre': evento_nombre or 'Sin evento',
                        'fecha': fecha_str,
                        'hora': hora_str,
                        'owners_text': owners_text,
                        'enviado': bool(enviado) if enviado is not None else False
                    })
        
        response = JsonResponse({
            'reminders': results,
            'session_active': True
        }, safe=False)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'reminders': [], 'session_active': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
@login_required
def api_reminders_notifications(request):
    """API intermediaria para verificar recordatorios y enviar notificaciones en todas las p√°ginas.
    Esta API consolida la l√≥gica de verificaci√≥n de recordatorios para ser usada desde base.html.
    Devuelve recordatorios que deben mostrar notificaciones (futuros y pasados dentro de la ventana de 15 minutos)."""
    try:
        from django.utils.timezone import now
        import pytz
        from django.db import connection
        
        # Verificar que el usuario tiene sesi√≥n activa
        if not request.user.is_authenticated:
            return JsonResponse({'reminders': [], 'session_active': False}, safe=False)
        
        # Obtener informaci√≥n del usuario actual
        user_id = None
        colaborador_usuario_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                user_id = str(usuario_maga.id)
                if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
                    colaborador_usuario_id = str(usuario_maga.colaborador.id)
        except Exception:
            pass
        
        if not user_id:
            return JsonResponse({'reminders': [], 'session_active': False}, safe=False)
        
        # Obtener recordatorios pendientes (similar a api_reminders_pending pero optimizado para notificaciones)
        guatemala_tz = pytz.timezone('America/Guatemala')
        ahora_guatemala = now().astimezone(guatemala_tz)
        
        with connection.cursor() as cur:
            # Verificar si existe columna recordar
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='recordatorios' AND column_name='recordar'
            """)
            tiene_recordar = cur.fetchone() is not None
            
            if colaborador_usuario_id:
                if tiene_recordar:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               COALESCE(r.recordar, FALSE) as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        INNER JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND rc.colaborador_id = %s
                          AND (
                            -- Recordatorios futuros que no han sido enviados
                            (r.due_at >= %s AND (r.enviado = FALSE OR r.enviado IS NULL))
                            OR
                            -- Recordatorios pasados con "recordar" activado, dentro de los 15 minutos
                            (r.due_at < %s 
                             AND r.due_at >= %s - INTERVAL '15 minutes'
                             AND COALESCE(r.recordar, FALSE) = TRUE)
                          )
                        ORDER BY r.due_at ASC
                        """,
                        [colaborador_usuario_id, ahora_guatemala, ahora_guatemala, ahora_guatemala]
                    )
                else:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               FALSE as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        INNER JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.due_at >= %s
                          AND rc.colaborador_id = %s
                          AND (r.enviado = FALSE OR r.enviado IS NULL)
                        ORDER BY r.due_at ASC
                        """,
                        [ahora_guatemala, colaborador_usuario_id]
                    )
            else:
                if tiene_recordar:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               COALESCE(r.recordar, FALSE) as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.created_by = %s
                          AND (
                            -- Recordatorios futuros que no han sido enviados
                            (r.due_at >= %s AND (r.enviado = FALSE OR r.enviado IS NULL))
                            OR
                            -- Recordatorios pasados con "recordar" activado, dentro de los 15 minutos
                            (r.due_at < %s 
                             AND r.due_at >= %s - INTERVAL '15 minutes'
                             AND COALESCE(r.recordar, FALSE) = TRUE)
                          )
                        ORDER BY r.due_at ASC
                        """,
                        [user_id, ahora_guatemala, ahora_guatemala, ahora_guatemala]
                    )
                else:
                    cur.execute(
                        """
                        SELECT DISTINCT r.id::text,
                               r.titulo,
                               r.descripcion,
                               r.due_at,
                               r.enviar_notificacion,
                               r.enviado,
                               FALSE as recordar,
                               a.nombre as evento_nombre
                        FROM recordatorios r
                        LEFT JOIN actividades a ON a.id = r.actividad_id
                        WHERE r.enviar_notificacion = TRUE
                          AND r.due_at >= %s
                          AND r.created_by = %s
                          AND (r.enviado = FALSE OR r.enviado IS NULL)
                        ORDER BY r.due_at ASC
                        """,
                        [ahora_guatemala, user_id]
                    )
            
            rows = cur.fetchall()
        
        results = []
        for row in rows:
            if tiene_recordar:
                if len(row) >= 8:
                    rid, titulo, desc, due_at, enviar, enviado, recordar, evento_nombre = row
                else:
                    rid, titulo, desc, due_at, enviar, enviado, recordar = row
                    evento_nombre = None
            else:
                if len(row) >= 8:
                    rid, titulo, desc, due_at, enviar, enviado, recordar, evento_nombre = row[0], row[1], row[2], row[3], row[4], row[5], False, row[7]
                else:
                    rid, titulo, desc, due_at, enviar, enviado = row[0], row[1], row[2], row[3], row[4], row[5]
                    recordar = False
                    evento_nombre = row[7] if len(row) > 7 else None
            
            # Obtener personal involucrado
            owners_names = []
            with connection.cursor() as cur2:
                cur2.execute(
                    """
                    SELECT c.nombre
                    FROM recordatorio_colaboradores rc
                    JOIN colaboradores c ON c.id = rc.colaborador_id
                    WHERE rc.recordatorio_id = %s
                    ORDER BY c.nombre
                    """,
                    [rid]
                )
                owners_names = [r[0] for r in cur2.fetchall()]
            owners_text = ', '.join(owners_names) if owners_names else 'Sin personal asignado'
            
            # Convertir due_at a zona horaria de Guatemala
            if due_at:
                if due_at.tzinfo is None:
                    due_at_guatemala = guatemala_tz.localize(due_at)
                else:
                    due_at_guatemala = due_at.astimezone(guatemala_tz)
                
                # Calcular tiempo hasta el recordatorio
                tiempo_restante = (due_at_guatemala - ahora_guatemala).total_seconds()
                
                # L√≠mite de 15 minutos (900 segundos) despu√©s de la hora del recordatorio
                LIMITE_MINUTOS = 15
                LIMITE_SEGUNDOS = LIMITE_MINUTOS * 60
                
                tiempo_absoluto = abs(tiempo_restante)
                incluir = False
                
                # Incluir recordatorios futuros (para programar notificaciones)
                # Y recordatorios pasados dentro de los 15 minutos (para reenv√≠o con "recordar")
                if tiempo_restante > 0:
                    # Futuro: incluir todos los que a√∫n no han pasado
                    incluir = True
                elif tiempo_restante <= 0:
                    # Ya pas√≥, solo incluir si est√° dentro de los 15 minutos
                    if tiempo_absoluto <= LIMITE_SEGUNDOS:
                        incluir = True
                
                if incluir:
                    # Formatear fecha y hora
                    fecha_str = due_at_guatemala.strftime('%d/%m/%Y')
                    hora_str = due_at_guatemala.strftime('%H:%M')
                    
                    results.append({
                        'id': rid,
                        'titulo': titulo or 'Recordatorio',
                        'descripcion': desc or '',
                        'due_at': due_at_guatemala.isoformat(),
                        'due_at_timestamp': int(due_at_guatemala.timestamp() * 1000),
                        'tiempo_restante_segundos': int(tiempo_restante),
                        'recordar': bool(recordar) if recordar is not None else False,
                        'evento_nombre': evento_nombre or 'Sin evento',
                        'fecha': fecha_str,
                        'hora': hora_str,
                        'owners_text': owners_text,
                        'enviado': bool(enviado) if enviado is not None else False
                    })
        
        response = JsonResponse({
            'reminders': results,
            'session_active': True
        }, safe=False)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'reminders': [], 'session_active': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
@login_required
def api_marcar_notificacion_enviada(request, reminder_id):
    """Marca una notificaci√≥n como enviada para evitar duplicados"""
    try:
        from django.db import connection
        
        # Verificar que el recordatorio pertenece al usuario o est√° involucrado
        user_id = None
        colaborador_usuario_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                user_id = str(usuario_maga.id)
                if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
                    colaborador_usuario_id = str(usuario_maga.colaborador.id)
        except Exception:
            pass
        
        if not user_id:
            return JsonResponse({'error': 'Usuario no autenticado'}, status=401)
        
        # Verificar que el usuario est√° involucrado en el recordatorio
        with connection.cursor() as cur:
            if colaborador_usuario_id:
                cur.execute("""
                    SELECT COUNT(*) 
                    FROM recordatorios r
                    INNER JOIN recordatorio_colaboradores rc ON rc.recordatorio_id = r.id
                    WHERE r.id = %s AND rc.colaborador_id = %s
                """, [reminder_id, colaborador_usuario_id])
            else:
                cur.execute("""
                    SELECT COUNT(*) 
                    FROM recordatorios r
                    WHERE r.id = %s AND r.created_by = %s
                """, [reminder_id, user_id])
            
            count = cur.fetchone()[0]
            if count == 0:
                return JsonResponse({'error': 'No autorizado'}, status=403)
            
            # Marcar como enviado
            cur.execute("""
                UPDATE recordatorios 
                SET enviado = TRUE 
                WHERE id = %s
            """, [reminder_id])
        
        return JsonResponse({'success': True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["DELETE"])
@login_required
def api_reminder_detail(request, reminder_id):
    """Eliminar recordatorio. Solo admin puede eliminar cualquier recordatorio. Usuarios personales solo pueden eliminar los que crearon."""
    try:
        # Obtener informaci√≥n del recordatorio
        with connection.cursor() as cur:
            cur.execute("SELECT created_by::text FROM recordatorios WHERE id = %s", [reminder_id])
            row = cur.fetchone()
            if not row:
                return JsonResponse({'error': 'Not found'}, status=404)
            created_by = row[0]
        
        # Obtener informaci√≥n del usuario actual
        is_admin = False
        user_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                user_id = str(usuario_maga.id)
                is_admin = usuario_maga.rol == 'admin'
        except Exception:
            pass
        
        # Verificar permisos: admin puede eliminar cualquier recordatorio, usuarios personales solo los que crearon
        is_creator = (created_by == user_id) if user_id else False
        
        if not (is_admin or is_creator):
            return JsonResponse({'error': 'Forbidden'}, status=403)
        
        # Eliminar el recordatorio (CASCADE eliminar√° tambi√©n los colaboradores)
        with connection.cursor() as cur2:
            cur2.execute("DELETE FROM recordatorios WHERE id = %s", [reminder_id])
        
        return JsonResponse({'deleted': True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["GET"])
@login_required
def api_events_list(request):
    """Lista simple de eventos (id, name) para el formulario de recordatorios."""
    try:
        # Obtener informaci√≥n del usuario actual
        is_admin = False
        colaborador_usuario_id = None
        try:
            usuario_maga = get_usuario_maga(request.user)
            if usuario_maga:
                is_admin = usuario_maga.rol == 'admin'
                if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
                    colaborador_usuario_id = str(usuario_maga.colaborador.id)
        except Exception:
            pass
        
        # Si es admin, mostrar todos los eventos
        if is_admin:
            qs = Actividad.objects.filter(eliminado_en__isnull=True).order_by('-creado_en')[:300]
        else:
            # Si no es admin, solo mostrar eventos donde el colaborador del usuario est√° asignado
            if colaborador_usuario_id:
                qs = Actividad.objects.filter(
                    eliminado_en__isnull=True,
                    personal__colaborador_id=colaborador_usuario_id
                ).distinct().order_by('-creado_en')[:300]
            else:
                # Si no tiene colaborador vinculado, no mostrar eventos
                qs = Actividad.objects.none()
        
        data = [{'id': str(a.id), 'name': a.nombre} for a in qs]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse([], safe=False)


@require_http_methods(["GET"])
@login_required
def api_collaborators(request):
    """Lista de colaboradores activos para el formulario (id, nombre, puesto). Filtra por evento si se proporciona evento_id."""
    try:
        evento_id = request.GET.get('evento_id') or request.GET.get('event_id')
        
        with connection.cursor() as cur:
            if evento_id:
                # Filtrar colaboradores del evento espec√≠fico
                # Incluir:
                # 1. Colaboradores asignados directamente (ap.colaborador_id)
                # 2. Colaboradores vinculados a usuarios asignados (ap.usuario_id -> usuario.colaborador_id)
                cur.execute(
                    """
                    SELECT DISTINCT c.id::text, 
                           c.nombre, 
                           COALESCE(p.nombre,'') as puesto,
                           COALESCE(u.username, '') as username
                    FROM (
                        -- Colaboradores asignados directamente
                        SELECT ap.actividad_id, ap.colaborador_id as colab_id
                        FROM actividad_personal ap
                        WHERE ap.actividad_id = %s AND ap.colaborador_id IS NOT NULL
                        
                        UNION
                        
                        -- Colaboradores vinculados a usuarios asignados
                        SELECT ap.actividad_id, c_linked.id as colab_id
                        FROM actividad_personal ap
                        INNER JOIN usuarios u ON u.id = ap.usuario_id
                        INNER JOIN colaboradores c_linked ON c_linked.usuario_id = u.id
                        WHERE ap.actividad_id = %s AND ap.usuario_id IS NOT NULL
                    ) colaboradores_evento
                    INNER JOIN colaboradores c ON c.id = colaboradores_evento.colab_id
                    LEFT JOIN puestos p ON p.id = c.puesto_id
                    LEFT JOIN usuarios u ON u.id = c.usuario_id
                    WHERE c.activo = TRUE
                    ORDER BY c.nombre
                    """,
                    [evento_id, evento_id]
                )
            else:
                # Si no se proporciona evento_id, devolver lista vac√≠a (el formulario requiere seleccionar evento primero)
                return JsonResponse([], safe=False)
            
            rows = cur.fetchall()
        
        data = []
        for r in rows:
            colaborador_id, nombre, puesto, username = r
            display_name = nombre
            if username:
                display_name = f"{nombre} ({username})"
            data.append({
                'id': colaborador_id,
                'name': nombre,
                'displayName': display_name,
                'puesto': puesto,
                'username': username
            })
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse([], safe=False)
@require_http_methods(["GET"])
def api_obtener_detalle_proyecto(request, evento_id):
    """
    Obtiene los detalles completos de un proyecto/evento espec√≠fico
    """
    try:
        from django.utils.timezone import localtime, is_aware, make_aware
        import pytz
        
        # Verificar si las columnas comunidad_id y region_id existen en eventos_cambios_colaboradores
        from django.db import connection
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'eventos_cambios_colaboradores' 
                    AND column_name IN ('comunidad_id', 'region_id');
                """)
                existing_columns = {row[0] for row in cursor.fetchall()}
                has_cambios_colaboradores_comunidad = 'comunidad_id' in existing_columns
                has_cambios_colaboradores_region = 'region_id' in existing_columns
        except Exception as e:
            # Si hay un error al verificar, asumir que las columnas no existen
            print(f'‚ö†Ô∏è Error al verificar columnas en eventos_cambios_colaboradores: {e}')
            has_cambios_colaboradores_comunidad = False
            has_cambios_colaboradores_region = False

        # Construir prefetch_related din√°micamente seg√∫n las columnas disponibles
        # OPTIMIZACI√ìN: NO cargar beneficiarios aqu√≠ - solo se necesita el conteo para la tarjeta de datos
        # Los beneficiarios solo se cargan cuando se edita un evento, no para ver detalles
        # NOTA: No usar slice en Prefetch - causa errores. Limitaremos en el loop.
        prefetch_fields = [
            'personal__usuario__puesto',
            'personal__colaborador__puesto',
            # Beneficiarios removidos - solo calcular conteo sin cargar objetos
            # Evidencias, archivos y galer√≠a se limitar√°n en el loop (ver MAX_EVIDENCIAS_VISTA)
            'evidencias',
            'archivos',
            'galeria_imagenes',
            'cambios__responsable',
            'comunidades_relacionadas__comunidad__region',
            'comunidades_relacionadas__region'
        ]
        
        # Solo agregar prefetch de cambios_colaboradores si las columnas existen
        # Si no existen, omitir el prefetch y dejar que obtener_cambios_evento maneje la carga
        if has_cambios_colaboradores_comunidad or has_cambios_colaboradores_region:
            cambios_colaboradores_queryset = EventoCambioColaborador.objects.select_related('colaborador').prefetch_related('evidencias')
            if has_cambios_colaboradores_comunidad:
                cambios_colaboradores_queryset = cambios_colaboradores_queryset.select_related('comunidad')
            if has_cambios_colaboradores_region:
                cambios_colaboradores_queryset = cambios_colaboradores_queryset.select_related('region')
            prefetch_fields.append(
                Prefetch('cambios_colaboradores', queryset=cambios_colaboradores_queryset)
            )
        # Si las columnas no existen, no hacer prefetch para evitar errores
        # La funci√≥n obtener_cambios_evento manejar√° la carga de manera segura
        
        evento = Actividad.objects.select_related(
            'tipo', 'comunidad', 'comunidad__region', 'responsable', 'colaborador'
        ).prefetch_related(*prefetch_fields).get(id=evento_id, eliminado_en__isnull=True)
        
        # Personal asignado
        personal_data = []
        for ap in evento.personal.all():
            if ap.usuario:
                # Si el usuario tiene un colaborador vinculado, incluir informaci√≥n del colaborador
                colaborador_vinculado = None
                if hasattr(ap.usuario, 'colaborador') and ap.usuario.colaborador:
                    colaborador_vinculado = ap.usuario.colaborador
                    # Incluir como colaborador (con el ID del colaborador, no del usuario)
                    personal_data.append({
                        'id': str(colaborador_vinculado.id),  # ID del colaborador
                        'username': ap.usuario.username,
                        'nombre': colaborador_vinculado.nombre,
                        'rol': ap.rol_en_actividad,
                        'rol_display': 'Personal Fijo' if colaborador_vinculado.es_personal_fijo else 'Colaborador Externo',
                        'puesto': colaborador_vinculado.puesto.nombre if colaborador_vinculado.puesto else 'Sin puesto',
                        'tipo': 'colaborador',  # Tipo colaborador aunque sea usuario
                        'usuario_id': str(ap.usuario.id),  # Guardar tambi√©n el ID del usuario para referencia
                        'tiene_colaborador': True
                    })
                else:
                    # Usuario sin colaborador vinculado - no incluirlo en la lista de cambios
                    pass
            elif ap.colaborador:
                personal_data.append({
                    'id': str(ap.colaborador.id),
                    'username': ap.colaborador.correo or '',
                    'nombre': ap.colaborador.nombre,
                    'rol': ap.rol_en_actividad,
                    'rol_display': 'Personal Fijo' if ap.colaborador.es_personal_fijo else 'Colaborador Externo',
                    'puesto': ap.colaborador.puesto.nombre if ap.colaborador.puesto else 'Sin puesto',
                    'tipo': 'colaborador',
                    'tiene_colaborador': True
                })
        
        # OPTIMIZACI√ìN: NO cargar todos los beneficiarios - solo calcular el conteo
        # Los beneficiarios solo se muestran en la tarjeta de datos (que usa el conteo)
        # Cargar 900+ beneficiarios consume mucho tiempo y CPU innecesariamente
        beneficiarios_count = evento.beneficiarios.count()
        beneficiarios_data = []  # Array vac√≠o - no se usa en la vista de detalles
        
        # OPTIMIZACI√ìN: Limitar evidencias de galer√≠a a las primeras 50 para vista de detalles
        # Las evidencias adicionales se cargan bajo demanda si es necesario
        MAX_EVIDENCIAS_VISTA = 50
        evidencias_data = []
        galeria_urls = set()
        galeria_nombres = set()
        for imagen in evento.galeria_imagenes.all()[:MAX_EVIDENCIAS_VISTA]:
            if imagen.url_almacenamiento:
                galeria_urls.add(imagen.url_almacenamiento)
            if imagen.archivo_nombre:
                galeria_nombres.add(imagen.archivo_nombre)
            evidencias_data.append({
                'id': str(imagen.id),
                'nombre': imagen.archivo_nombre,
                'url': imagen.url_almacenamiento,
                'tipo': imagen.archivo_tipo or '',
                'es_imagen': True,
                'descripcion': imagen.descripcion or '',
                'es_galeria': True,
                'es_evidencia': False
            })
        
        # OPTIMIZACI√ìN: Limitar evidencias propias y archivos a los primeros 50
        # Archivos del proyecto (evidencias + archivos de actividad_archivos)
        # Evidencias: ordenar por fecha antigua primero (orden original)
        archivos_data = []
        evidencias_propias = []
        for evidencia in evento.evidencias.all().order_by('creado_en')[:MAX_EVIDENCIAS_VISTA]:
            url = evidencia.url_almacenamiento or ''
            url_lower = url.lower()
            esta_en_galeria = url in galeria_urls or evidencia.archivo_nombre in galeria_nombres
            es_archivo_evidencia = '/media/evidencias/' in url_lower

            evidencias_propias.append({
                'id': str(evidencia.id),
                'nombre': evidencia.archivo_nombre,
                'url': url,
                'tipo': evidencia.archivo_tipo or '',
                'es_imagen': evidencia.es_imagen,
                'descripcion': evidencia.descripcion or '',
                'es_galeria': esta_en_galeria,
                'es_evidencia': True,
                'es_archivo': es_archivo_evidencia
            })

            if es_archivo_evidencia:
                archivos_data.append({
                    'id': str(evidencia.id),
                    'nombre': evidencia.archivo_nombre,
                    'url': url,
                    'tipo': evidencia.archivo_tipo or 'application/octet-stream',
                    'tamanio': evidencia.archivo_tamanio,
                    'descripcion': evidencia.descripcion or '',
                    'es_evidencia': True,  # Marca que es de evidencias (no se puede eliminar)
                    'es_imagen': evidencia.es_imagen,
                    'es_galeria': esta_en_galeria,
                    'creado_en': evidencia.creado_en.isoformat() if evidencia.creado_en else None
                })
        # OPTIMIZACI√ìN: Limitar archivos de actividad_archivos a los primeros 50
        # Ordenar por fecha reciente primero (los nuevos aparecen arriba)
        for archivo in evento.archivos.all().order_by('-creado_en')[:MAX_EVIDENCIAS_VISTA]:
            archivos_data.append({
                'id': str(archivo.id),
                'nombre': archivo.nombre_archivo,
                'url': archivo.url_almacenamiento,
                'tipo': archivo.archivo_tipo or 'application/octet-stream',
                'tamanio': archivo.archivo_tamanio,
                'descripcion': archivo.descripcion or '',
                'es_evidencia': False,  # Marca que es de actividad_archivos (se puede eliminar)
                'es_imagen': False,
                'es_galeria': False,
                'creado_en': archivo.creado_en.isoformat() if archivo.creado_en else None
            })
        evidencias_data.extend(evidencias_propias)
        
        # Ubicaci√≥n
        ubicacion = 'Sin ubicaci√≥n'
        if evento.comunidad:
            if evento.comunidad.region:
                ubicacion = f"{evento.comunidad.nombre}, {evento.comunidad.region.nombre}"
            else:
                ubicacion = evento.comunidad.nombre
        
        # Convertir fechas
        try:
            if evento.fecha:
                fecha_str = evento.fecha.strftime('%Y-%m-%d')
                fecha_display = evento.fecha.strftime('%d de %B de %Y')
            else:
                fecha_str = ''
                fecha_display = ''
        except:
            fecha_str = ''
            fecha_display = ''
        
        # Verificar permisos del usuario actual para este evento
        usuario_maga = get_usuario_maga(request.user) if request.user.is_authenticated else None
        puede_gestionar = False
        
        if usuario_maga:
            puede_gestionar = usuario_puede_gestionar_evento(usuario_maga, evento_id)
        
        proyecto_data = {
            'id': str(evento.id),
            'nombre': evento.nombre,
            'tipo': evento.tipo.nombre if evento.tipo else 'Sin tipo',
            'descripcion': evento.descripcion or 'Sin descripci√≥n',
            'ubicacion': ubicacion,
            'comunidad': evento.comunidad.nombre if evento.comunidad else 'Sin comunidad',
            'region': evento.comunidad.region.nombre if evento.comunidad and evento.comunidad.region else None,
            'fecha': fecha_str,
            'fecha_display': fecha_display,
            'estado': evento.estado,
            'estado_display': evento.get_estado_display() if hasattr(evento, 'get_estado_display') else evento.estado,
            'responsable': (evento.responsable.nombre if evento.responsable.nombre else evento.responsable.username) if evento.responsable else 'Sin responsable',
            'personal': personal_data,
            'beneficiarios': beneficiarios_data,  # Array vac√≠o - optimizado (no se usa en vista de detalles)
            'beneficiarios_count': beneficiarios_count,  # Solo el conteo para la tarjeta de datos
            'evidencias': evidencias_data,  # Limitadas a 50 para optimizaci√≥n
            'archivos': archivos_data,  # Limitados a 50 para optimizaci√≥n
            'evidencias_count': evento.evidencias.count(),  # Conteo total de evidencias
            'archivos_count': evento.archivos.count(),  # Conteo total de archivos
            'galeria_count': evento.galeria_imagenes.count(),  # Conteo total de galer√≠a
            'portada': obtener_portada_evento(evento),
            'tarjetas_datos': obtener_tarjetas_datos(evento),
            'comunidades': obtener_comunidades_evento(evento),
            'cambios': obtener_cambios_evento(evento),
            'puede_gestionar': puede_gestionar,
            'permisos': {
                'puede_gestionar': puede_gestionar,
                'es_admin': usuario_maga.rol == 'admin' if usuario_maga else False,
                'es_personal': usuario_maga.rol == 'personal' if usuario_maga else False,
            }
        }
        
        print(f'üì§ Retornando proyecto con {len(proyecto_data.get("cambios", []))} cambios')
        
        return JsonResponse({
            'success': True,
            'proyecto': proyecto_data
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Proyecto no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener proyecto: {str(e)}'
        }, status=500)
@permiso_gestionar_eventos_api
@require_http_methods(["POST"])
def api_agregar_imagen_galeria(request, evento_id):
    """API: Agregar imagen a la galer√≠a de un evento"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        # Validar que se haya enviado una imagen
        imagen = request.FILES.get('imagen')
        if not imagen:
            return JsonResponse({
                'success': False,
                'error': 'No se ha enviado ninguna imagen'
            }, status=400)
        
        # Validar que sea una imagen
        if not imagen.content_type or not imagen.content_type.startswith('image/'):
            return JsonResponse({
                'success': False,
                'error': 'El archivo debe ser una imagen (JPG, PNG, GIF, etc.)'
            }, status=400)
        
        # COMPRIMIR IMAGEN autom√°ticamente
        if getattr(settings, 'COMPRESS_IMAGES', True):
            from .image_compression import compress_if_needed, get_compression_settings
            compression_settings = get_compression_settings('gallery')
            imagen = compress_if_needed(imagen, **compression_settings)
        
        # Obtener descripci√≥n (opcional)
        descripcion = request.POST.get('descripcion', '').strip()
        
        # Crear carpeta si no existe
        galeria_dir = os.path.join(str(settings.MEDIA_ROOT), 'galeria_img')
        os.makedirs(galeria_dir, exist_ok=True)
        
        print(f"üì§ Subiendo imagen a evento {evento_id}")
        print(f"üìÅ Directorio: {galeria_dir}")
        print(f"‚úÖ Archivo recibido: {imagen.name}, tama√±o: {imagen.size}, tipo: {imagen.content_type}")
        
        # Verificar permisos de escritura
        if not os.access(galeria_dir, os.W_OK):
            print(f"‚ùå No hay permisos de escritura en {galeria_dir}")
            return JsonResponse({
                'success': False,
                'error': f'No se tienen permisos de escritura en {galeria_dir}'
            }, status=500)
        
        # Guardar archivo
        fs = FileSystemStorage(location=galeria_dir)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
        file_extension = os.path.splitext(imagen.name)[1]
        filename = f"{timestamp}_{evento_id}{file_extension}"
        print(f"üíæ Intentando guardar archivo: {filename}")
        try:
            saved_name = fs.save(filename, imagen)
            print(f"‚úÖ Archivo guardado exitosamente: {saved_name}")
        except Exception as e:
            import traceback
            print(f"‚ùå Error al guardar archivo: {str(e)}")
            print(f"üìã Traceback:\n{traceback.format_exc()}")
            raise
        file_url = f"/media/galeria_img/{saved_name}"
        
        # Crear registro en la BD usando EventosGaleria
        imagen_galeria = EventosGaleria.objects.create(
            actividad=evento,
            archivo_nombre=imagen.name,
            archivo_tipo=imagen.content_type,
            archivo_tamanio=imagen.size,
            url_almacenamiento=file_url,
            descripcion=descripcion,
            creado_por=usuario_maga
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Imagen agregada exitosamente',
            'imagen': {
                'id': str(imagen_galeria.id),
                'url': file_url,
                'nombre': imagen.name,
                'descripcion': descripcion
            }
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al agregar imagen: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos_api
@require_http_methods(["DELETE", "POST"])
def api_eliminar_imagen_galeria(request, evento_id, imagen_id):
    """API: Eliminar imagen de la galer√≠a de un evento"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        # Obtener la imagen de la galer√≠a
        imagen_galeria = EventosGaleria.objects.filter(
            id=imagen_id,
            actividad=evento
        ).first()
        
        if not imagen_galeria:
            return JsonResponse({
                'success': False,
                'error': 'Imagen no encontrada'
            }, status=404)
        
        # Eliminar archivo f√≠sico
        _eliminar_archivo_media(getattr(imagen_galeria, 'url_almacenamiento', None))
        
        # Eliminar registro de la BD
        imagen_galeria.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Imagen eliminada exitosamente'
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar imagen: {str(e)}'
        }, status=500)
@permiso_gestionar_eventos
@require_http_methods(["POST"])
def api_agregar_archivo(request, evento_id):
    """API: Agregar archivo a un evento"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        # Validar que se haya enviado un archivo
        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({
                'success': False,
                'error': 'No se ha enviado ning√∫n archivo'
            }, status=400)
        
        # Obtener descripci√≥n (opcional)
        descripcion = request.POST.get('descripcion', '').strip()
        
        # Crear carpeta si no existe
        archivos_dir = os.path.join(str(settings.MEDIA_ROOT), 'archivos_eventos')
        os.makedirs(archivos_dir, exist_ok=True)
        
        # Guardar archivo
        fs = FileSystemStorage(location=archivos_dir)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
        file_extension = os.path.splitext(archivo.name)[1]
        filename = f"{timestamp}_{evento_id}{file_extension}"
        saved_name = fs.save(filename, archivo)
        file_url = f"/media/archivos_eventos/{saved_name}"
        
        # Crear registro en la tabla actividad_archivos
        archivo_registro = ActividadArchivo.objects.create(
            actividad=evento,
            nombre_archivo=archivo.name,
            archivo_tipo=archivo.content_type or 'application/octet-stream',
            archivo_tamanio=archivo.size,
            url_almacenamiento=file_url,
            descripcion=descripcion,
            creado_por=usuario_maga
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Archivo agregado exitosamente',
            'archivo': {
                'id': str(archivo_registro.id),
                'nombre': archivo.name,
                'url': file_url,
                'tipo': archivo.content_type or 'application/octet-stream',
                'tamanio': archivo.size,
                'descripcion': descripcion,
                'es_evidencia': False
            }
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al agregar archivo: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["DELETE", "POST"])
def api_eliminar_archivo(request, evento_id, archivo_id):
    """API: Eliminar archivo de un evento (solo de actividad_archivos, no evidencias)"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        # Obtener el archivo de actividad_archivos (NO evidencias)
        archivo = ActividadArchivo.objects.filter(
            id=archivo_id,
            actividad=evento
        ).first()
        
        if not archivo:
            return JsonResponse({
                'success': False,
                'error': 'Archivo no encontrado o no se puede eliminar (es una evidencia)'
            }, status=404)
        
        # Eliminar archivo f√≠sico
        _eliminar_archivo_media(getattr(archivo, 'url_almacenamiento', None))
        
        # Eliminar registro de la BD
        archivo.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Archivo eliminado exitosamente'
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar archivo: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["POST"])
def api_actualizar_archivo_evento(request, evento_id, archivo_id):
    """API: Actualizar descripci√≥n de un archivo del evento"""
    try:
        archivo = ActividadArchivo.objects.get(id=archivo_id, actividad_id=evento_id)
    except ActividadArchivo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Archivo no encontrado'
        }, status=404)

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        payload = {}

    descripcion = (payload.get('descripcion') or '').strip()

    archivo.descripcion = descripcion or ''
    archivo.save(update_fields=['descripcion'])

    actividad = archivo.actividad
    if actividad:
        actividad.actualizado_en = timezone.now()
        actividad.save(update_fields=['actualizado_en'])

    return JsonResponse({
        'success': True,
        'message': 'Descripci√≥n actualizada correctamente',
        'archivo': {
            'id': str(archivo.id),
            'nombre': archivo.nombre_archivo,
            'url': archivo.url_almacenamiento,
            'tipo': archivo.archivo_tipo or 'application/octet-stream',
            'tamanio': archivo.archivo_tamanio,
            'descripcion': archivo.descripcion or '',
            'es_evidencia': False,
        }
    })
# =====================================================
# APIs PARA GESTI√ìN DE CAMBIOS
# =====================================================
def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ('1', 'true', 'yes', 'on')
    return bool(value)


@solo_administrador
@require_http_methods(["GET"])
def api_listar_usuarios(request):
    """API: Listar todos los usuarios del sistema"""
    try:
        usuarios = Usuario.objects.select_related('puesto', 'colaborador__puesto').order_by('username')
        
        usuarios_list = []
        for usuario in usuarios:
            # Obtener colaborador vinculado si existe (usando la relaci√≥n inversa OneToOne)
            colaborador = usuario.colaborador if hasattr(usuario, 'colaborador') else None
            
            # Determinar el puesto a mostrar: primero del usuario, luego del colaborador
            puesto_nombre = None
            puesto_id = None
            
            if usuario.puesto:
                puesto_nombre = usuario.puesto.nombre
                puesto_id = str(usuario.puesto.id)
            elif colaborador and colaborador.puesto:
                puesto_nombre = colaborador.puesto.nombre
                puesto_id = str(colaborador.puesto.id)
            
            usuarios_list.append({
                'id': str(usuario.id),
                'username': usuario.username,
                'nombre': usuario.nombre or '',
                'email': usuario.email,
                'telefono': usuario.telefono or '',
                'rol': usuario.rol,
                'rol_display': usuario.get_rol_display(),
                'puesto_id': puesto_id,
                'puesto_nombre': puesto_nombre,
                'activo': usuario.activo,
                'tiene_colaborador': colaborador is not None,
                'colaborador_id': str(colaborador.id) if colaborador else None,
                'colaborador_nombre': colaborador.nombre if colaborador else None,
                'colaborador_puesto_id': str(colaborador.puesto.id) if colaborador and colaborador.puesto else None,
                'creado_en': usuario.creado_en.isoformat() if usuario.creado_en else None,
            })
        
        return JsonResponse({
            'success': True,
            'usuarios': usuarios_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al listar usuarios: {str(e)}'
        }, status=500)
@solo_administrador
@require_http_methods(["POST"])
def api_crear_usuario(request):
    """API: Crear un nuevo usuario del sistema"""
    try:
        data = json.loads(request.body or '{}')
        
        username = (data.get('username') or '').strip()
        nombre = (data.get('nombre') or '').strip()
        email = (data.get('email') or '').strip()
        telefono = (data.get('telefono') or '').strip()
        password = data.get('password') or ''
        password_confirm = data.get('password_confirm') or ''
        rol = (data.get('rol') or '').strip()
        puesto_id = (data.get('puesto_id') or '').strip()
        colaborador_id = (data.get('colaborador_id') or '').strip()
        
        # Validaciones
        if not username or not email or not password:
            return JsonResponse({
                'success': False,
                'error': 'Username, email y contrase√±a son requeridos'
            }, status=400)
        
        if password != password_confirm:
            return JsonResponse({
                'success': False,
                'error': 'Las contrase√±as no coinciden'
            }, status=400)
        
        if len(password) < 8:
            return JsonResponse({
                'success': False,
                'error': 'La contrase√±a debe tener al menos 8 caracteres'
            }, status=400)
        
        if rol not in ['admin', 'personal']:
            return JsonResponse({
                'success': False,
                'error': 'Rol inv√°lido'
            }, status=400)
        
        # Validar puesto si es personal
        puesto = None
        if rol == 'personal':
            if not puesto_id:
                return JsonResponse({
                    'success': False,
                    'error': 'El puesto es requerido para usuarios con rol "Personal"'
                }, status=400)
            try:
                puesto = Puesto.objects.get(id=puesto_id, activo=True)
            except Puesto.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Puesto no encontrado'
                }, status=404)
        
        # Validar colaborador si se proporciona
        colaborador = None
        if colaborador_id:
            try:
                colaborador = Colaborador.objects.get(id=colaborador_id)
                if colaborador.usuario_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'Este colaborador ya tiene un usuario asignado'
                    }, status=400)
                if not colaborador.activo:
                    return JsonResponse({
                        'success': False,
                        'error': 'No se puede vincular un colaborador inactivo.'
                    }, status=400)
            except Colaborador.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Colaborador no encontrado'
                }, status=404)
        
        # Verificar username √∫nico
        if Usuario.objects.filter(username=username).exists():
            return JsonResponse({
                'success': False,
                'error': 'El username ya existe'
            }, status=400)
        
        # Verificar email √∫nico
        if Usuario.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'error': 'El email ya existe'
            }, status=400)
        
        # Hashear contrase√±a usando Django
        from django.contrib.auth.hashers import make_password
        password_hash = make_password(password)
        
        # Crear usuario
        usuario = Usuario.objects.create(
            username=username,
            nombre=nombre if nombre else None,
            email=email,
            telefono=telefono if telefono else None,
            password_hash=password_hash,
            rol=rol,
            puesto=puesto,
            activo=True
        )
        
        # Vincular con colaborador si se proporciona
        if colaborador:
            colaborador.usuario = usuario
            if not colaborador.es_personal_fijo:
                colaborador.es_personal_fijo = True
            colaborador.save(update_fields=['usuario', 'es_personal_fijo'])
        
        return JsonResponse({
            'success': True,
            'message': 'Usuario creado exitosamente',
            'usuario': {
                'id': str(usuario.id),
                'username': usuario.username,
                'nombre': usuario.nombre or '',
                'email': usuario.email,
                'rol': usuario.rol,
                'rol_display': usuario.get_rol_display(),
                'puesto_id': str(usuario.puesto.id) if usuario.puesto else None,
                'puesto_nombre': usuario.puesto.nombre if usuario.puesto else None,
                'colaborador_id': str(colaborador.id) if colaborador else None,
                'colaborador_nombre': colaborador.nombre if colaborador else None
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al crear usuario: {str(e)}'
        }, status=500)
@permiso_gestionar_eventos
@require_http_methods(["POST"])
def api_crear_cambio(request, evento_id):
    """API: Crear un cambio en un evento"""
    try:
        print(f'üîµ ========== INICIO api_crear_cambio para evento {evento_id} ==========')
        print(f'üîµ M√©todo: {request.method}')
        print(f'üîµ POST keys: {list(request.POST.keys())}')
        print(f'üîµ FILES keys: {list(request.FILES.keys())}')
        
        # Log de todos los datos POST
        for key in request.POST.keys():
            value = request.POST.get(key)
            if len(str(value)) < 200:  # Solo mostrar valores cortos
                print(f'üîµ POST[{key}]: {value}')
            else:
                print(f'üîµ POST[{key}]: (valor largo, {len(str(value))} caracteres)')
        
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        descripcion = request.POST.get('descripcion', '').strip()
        if not descripcion:
            return JsonResponse({
                'success': False,
                'error': 'La descripci√≥n del cambio es obligatoria'
        }, status=400)
        
        # Obtener lista de colaboradores seleccionados
        colaboradores_ids_str = request.POST.get('colaboradores_ids')
        print(f'üì• colaboradores_ids recibido: {colaboradores_ids_str}')
        colaboradores_ids = []
        if colaboradores_ids_str:
            try:
                colaboradores_ids = json.loads(colaboradores_ids_str)
                print(f'‚úÖ Colaboradores parseados: {colaboradores_ids}')
            except json.JSONDecodeError as e:
                print(f'‚ùå Error al parsear JSON: {e}')
                # Fallback: intentar como colaborador_id individual (compatibilidad)
                colaborador_id_legacy = request.POST.get('colaborador_id')
                if colaborador_id_legacy:
                    colaboradores_ids = [colaborador_id_legacy]
                    print(f'‚úÖ Usando colaborador_id legacy: {colaboradores_ids}')
        
        # Validar que se haya proporcionado al menos un colaborador
        if not colaboradores_ids or len(colaboradores_ids) == 0:
            print('‚ùå No se proporcionaron colaboradores')
            return JsonResponse({
                'success': False,
                'error': 'Debe seleccionar al menos un colaborador responsable'
            }, status=400)
        
        # Validar y obtener colaboradores
        colaboradores = []
        for colaborador_id in colaboradores_ids:
            try:
                colaborador = Colaborador.objects.get(id=colaborador_id, activo=True)
                # Verificar que el colaborador est√© asignado al evento
                if not ActividadPersonal.objects.filter(actividad=evento, colaborador=colaborador).exists():
                    return JsonResponse({
                        'success': False,
                        'error': f'El colaborador {colaborador.nombre} no est√° asignado a este evento'
                    }, status=400)
                colaboradores.append(colaborador)
            except Colaborador.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': f'Colaborador con ID {colaborador_id} no encontrado'
                }, status=404)
        
        # Obtener comunidades seleccionadas
        comunidades_ids_str = request.POST.get('comunidades_ids')
        comunidades_ids = []
        if comunidades_ids_str:
            try:
                comunidades_ids = json.loads(comunidades_ids_str)
                print(f'‚úÖ Comunidades parseadas: {comunidades_ids}')
            except json.JSONDecodeError as e:
                print(f'‚ùå Error al parsear JSON de comunidades: {e}')
        
        # Validar que las comunidades pertenezcan al evento
        comunidades_validas = []
        if comunidades_ids:
            from webmaga.models import ActividadComunidad, Comunidad
            # Obtener comunidades relacionadas con el evento
            comunidades_evento = ActividadComunidad.objects.filter(
                actividad=evento
            ).values_list('comunidad_id', flat=True)
            
            # Convertir a lista para comparaci√≥n (los IDs pueden ser UUIDs, no enteros)
            comunidades_evento_list = list(comunidades_evento)
            print(f'üîç Comunidades relacionadas con el evento: {comunidades_evento_list} (cantidad: {len(comunidades_evento_list)})')
            print(f'üîç Tipos de IDs en comunidades_evento: {[type(cid) for cid in comunidades_evento_list[:3]] if comunidades_evento_list else "lista vac√≠a"}')
            
            for comunidad_id_str in comunidades_ids:
                try:
                    # Los IDs pueden ser UUIDs (strings) o enteros, intentar ambos
                    comunidad_id_parsed = None
                    try:
                        # Intentar convertir a UUID si es string UUID
                        import uuid as uuid_module
                        if isinstance(comunidad_id_str, str):
                            comunidad_id_parsed = uuid_module.UUID(comunidad_id_str)
                        else:
                            comunidad_id_parsed = comunidad_id_str
                    except (ValueError, AttributeError):
                        # Si no es UUID, intentar como entero
                        try:
                            comunidad_id_parsed = int(comunidad_id_str) if isinstance(comunidad_id_str, str) else comunidad_id_str
                        except (ValueError, TypeError):
                            comunidad_id_parsed = comunidad_id_str
                    
                    print(f'üîç Verificando comunidad ID: {comunidad_id_str} -> {comunidad_id_parsed} (tipo: {type(comunidad_id_parsed)})')
                    
                    # Verificar que la comunidad est√© relacionada con el evento
                    # Comparar tanto el ID original como el parseado
                    comunidad_encontrada = False
                    for cid_evento in comunidades_evento_list:
                        # Comparar como strings para UUIDs
                        if str(cid_evento) == str(comunidad_id_parsed) or cid_evento == comunidad_id_parsed:
                            comunidad_encontrada = True
                            break
                    
                    if comunidad_encontrada:
                        # Intentar obtener la comunidad con el ID parseado
                        comunidad = Comunidad.objects.select_related('region').get(id=comunidad_id_parsed, activo=True)
                        comunidades_validas.append(comunidad)
                        print(f'‚úÖ Comunidad v√°lida: {comunidad.nombre} (ID: {comunidad.id}, Regi√≥n: {comunidad.region.nombre if comunidad.region else "Sin regi√≥n"})')
                    else:
                        print(f'‚ö†Ô∏è Comunidad {comunidad_id_str} ({comunidad_id_parsed}) no est√° relacionada con el evento.')
                        print(f'   Comunidades del evento: {[str(cid) for cid in comunidades_evento_list[:5]]}')
                except Comunidad.DoesNotExist:
                    print(f'‚ö†Ô∏è Comunidad con ID {comunidad_id_str} no encontrada en la base de datos')
                except Exception as e:
                    print(f'‚ö†Ô∏è Error con comunidad ID {comunidad_id_str}: {e}')
                    import traceback
                    traceback.print_exc()
        
        # Obtener fecha_cambio si se proporciona, de lo contrario usar la fecha actual
        fecha_cambio_str = request.POST.get('fecha_cambio')
        fecha_cambio = None
        if fecha_cambio_str:
            try:
                from django.utils.dateparse import parse_datetime
                import pytz
                fecha_cambio_naive = parse_datetime(fecha_cambio_str)
                if fecha_cambio_naive:
                    # Convertir a zona horaria de Guatemala
                    guatemala_tz = pytz.timezone('America/Guatemala')
                    if timezone.is_aware(fecha_cambio_naive):
                        fecha_cambio = fecha_cambio_naive.astimezone(guatemala_tz)
                    else:
                        fecha_cambio = guatemala_tz.localize(fecha_cambio_naive)
            except Exception as e:
                print(f'‚ö†Ô∏è Error al parsear fecha_cambio: {e}, usando fecha actual')
                fecha_cambio = None
        
        usar_fecha_actual = request.POST.get('usar_fecha_actual') == 'true'

        fecha_cambio_final = timezone.now() if usar_fecha_actual else (fecha_cambio if fecha_cambio else timezone.now())
        
        # Crear un cambio por cada colaborador seleccionado, todos compartiendo el mismo grupo
        grupo_uuid = uuid.uuid4()
        # Crear un cambio por cada colaborador seleccionado
        cambios_creados = []
        evidencias_data = []
        
        # Verificar archivos ANTES de crear cambios
        archivos_recibidos = bool(request.FILES)
        archivos_keys_list = []
        if archivos_recibidos:
            archivos_keys_list = [key for key in request.FILES.keys() if key.startswith('archivo_')]
            if not archivos_keys_list:
                archivos_keys_list = list(request.FILES.keys())
            print(f'üìé Archivos detectados ANTES de crear cambios: {len(archivos_keys_list)} archivos')
            print(f'üìé Claves de archivos: {archivos_keys_list}')
        
        with transaction.atomic():
            # Crear cambios para cada colaborador
            for colaborador in colaboradores:
                # Si hay comunidades, crear un cambio por cada combinaci√≥n colaborador-comunidad
                # Si no hay comunidades, crear un cambio sin comunidad
                if comunidades_validas:
                    for comunidad in comunidades_validas:
                        # Obtener regi√≥n de la comunidad si existe (ya viene con select_related)
                        region_id = None
                        if comunidad.region:
                            region_id = comunidad.region.id
                            print(f'üîç Regi√≥n obtenida de comunidad: {comunidad.region.nombre} (ID: {region_id})')
                        else:
                            print(f'‚ö†Ô∏è Comunidad {comunidad.nombre} no tiene regi√≥n asociada')
                        
                        cambio_colaborador = EventoCambioColaborador.objects.create(
                            actividad=evento,
                            colaborador=colaborador,
                            descripcion_cambio=descripcion,
                            fecha_cambio=fecha_cambio_final,
                            grupo_id=grupo_uuid,
                            comunidad_id=comunidad.id,
                            region_id=region_id
                        )
                        cambios_creados.append(cambio_colaborador)
                        print(f'‚úÖ Cambio creado para colaborador {colaborador.nombre} y comunidad {comunidad.nombre} (ID: {comunidad.id}, Regi√≥n ID: {region_id}): Cambio ID {cambio_colaborador.id}')
                else:
                    cambio_colaborador = EventoCambioColaborador.objects.create(
                        actividad=evento,
                        colaborador=colaborador,
                        descripcion_cambio=descripcion,
                        fecha_cambio=fecha_cambio_final,
                        grupo_id=grupo_uuid
                    )
                    cambios_creados.append(cambio_colaborador)
                    print(f'‚úÖ Cambio creado para colaborador {colaborador.nombre}: {cambio_colaborador.id}')
            
            # Procesar evidencias y asociarlas a TODOS los cambios creados
            # IMPORTANTE: Verificar que se crearon cambios antes de procesar evidencias
            if len(cambios_creados) == 0:
                print('‚ùå ERROR CR√çTICO: No se crearon cambios, no se pueden procesar evidencias')
                raise ValueError('No se pudieron crear los cambios de colaboradores')
            
            if archivos_recibidos and archivos_keys_list:
                evidencias_dir = os.path.join(str(settings.MEDIA_ROOT), 'evidencias_cambios_eventos')
                os.makedirs(evidencias_dir, exist_ok=True)
                
                fs = FileSystemStorage(location=evidencias_dir)
                print(f'üìé Archivos recibidos en request.FILES: {list(request.FILES.keys())}')
                print(f'üìé Total de cambios creados: {len(cambios_creados)}')
                
                # Usar la lista de archivos que ya detectamos
                archivos_keys = archivos_keys_list
                print(f'üìé Procesando {len(archivos_keys)} archivo(s) de evidencias')
                
                for index, key in enumerate(archivos_keys):
                    try:
                        archivo = request.FILES[key]
                        print(f'üìé Procesando archivo {key}: {archivo.name} ({archivo.size} bytes, tipo: {archivo.content_type})')
                        
                        # COMPRIMIR IMAGEN autom√°ticamente si es una imagen
                        es_imagen = archivo.content_type.startswith('image/') if hasattr(archivo, 'content_type') else False
                        if es_imagen and getattr(settings, 'COMPRESS_IMAGES', True):
                            from .image_compression import compress_if_needed, get_compression_settings
                            compression_settings = get_compression_settings('evidence')
                            archivo = compress_if_needed(archivo, **compression_settings)
                            print(f'‚úÖ Imagen comprimida: {archivo.size} bytes')
                        
                        # Obtener descripci√≥n de la evidencia
                        # Primero intentar con el √≠ndice
                        descripcion_evidencia = request.POST.get(f'descripcion_evidencia_{index}', '').strip()
                        if not descripcion_evidencia:
                            # Intentar obtener por el nombre del campo sin el prefijo "archivo_"
                            index_num = key.replace('archivo_', '')
                            descripcion_evidencia = request.POST.get(f'descripcion_evidencia_{index_num}', '').strip()
                        if not descripcion_evidencia:
                            # Intentar obtener descripci√≥n gen√©rica
                            descripcion_evidencia = request.POST.get('descripcion_evidencia', '').strip()
                        
                        print(f'üìé Descripci√≥n de evidencia obtenida: "{descripcion_evidencia}"')
                        
                        # Guardar el archivo f√≠sico UNA VEZ (usando el ID del primer cambio para el nombre)
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
                        file_extension = os.path.splitext(archivo.name)[1]
                        # Usar el ID del primer cambio para el nombre base del archivo
                        filename = f"{timestamp}_{cambios_creados[0].id}_{index}{file_extension}"
                        saved_name = fs.save(filename, archivo)
                        file_url = f"/media/evidencias_cambios_eventos/{saved_name}"
                        print(f'üìé Archivo guardado: {saved_name} -> {file_url}')
                        
                        # Crear un registro de evidencia en la BD para cada cambio creado
                        # Todos apuntan al mismo archivo f√≠sico
                        primera_evidencia_id = None
                        evidencias_creadas_count = 0
                        for cambio_colaborador in cambios_creados:
                            try:
                                # Validar que el cambio existe y tiene ID
                                if not cambio_colaborador.id:
                                    raise ValueError(f'El cambio {cambio_colaborador} no tiene ID v√°lido')
                                
                                # Validar que el evento existe
                                if not evento.id:
                                    raise ValueError(f'El evento {evento} no tiene ID v√°lido')
                                
                                # Validar que el usuario existe
                                if not usuario_maga or not usuario_maga.id:
                                    raise ValueError('El usuario no tiene ID v√°lido')
                                
                                print(f'üìù Creando evidencia para cambio {cambio_colaborador.id}, actividad {evento.id}, usuario {usuario_maga.id}')
                                
                                evidencia = EventosEvidenciasCambios.objects.create(
                                    actividad=evento,
                                    cambio=cambio_colaborador,
                                    archivo_nombre=archivo.name,
                                    archivo_tipo=archivo.content_type or 'application/octet-stream',
                                    archivo_tamanio=archivo.size,
                                    url_almacenamiento=file_url,  # Misma URL para todos
                                    descripcion=descripcion_evidencia if descripcion_evidencia else None,
                                    creado_por=usuario_maga
                                )
                                
                                # Verificar que se cre√≥ correctamente
                                if not evidencia.id:
                                    raise ValueError('La evidencia se cre√≥ pero no tiene ID')
                                
                                if primera_evidencia_id is None:
                                    primera_evidencia_id = evidencia.id
                                
                                evidencias_creadas_count += 1
                                print(f'‚úÖ Evidencia creada exitosamente para cambio {cambio_colaborador.id}: {evidencia.id} - {evidencia.archivo_nombre}')
                                
                                # Verificar que se puede recuperar de la BD
                                evidencia_verificada = EventosEvidenciasCambios.objects.filter(id=evidencia.id).first()
                                if not evidencia_verificada:
                                    raise ValueError(f'La evidencia {evidencia.id} no se puede recuperar de la BD despu√©s de crearla')
                                print(f'‚úÖ Evidencia verificada en BD: {evidencia_verificada.id}')
                                
                            except Exception as e:
                                print(f'‚ùå Error al crear evidencia para cambio {cambio_colaborador.id}: {e}')
                                print(f'‚ùå Tipo de error: {type(e).__name__}')
                                import traceback
                                traceback.print_exc()
                                raise Exception(f'Error al crear evidencia para cambio {cambio_colaborador.id}: {str(e)}')
                        
                        if evidencias_creadas_count == 0:
                            raise ValueError(f'No se pudo crear ninguna evidencia para los {len(cambios_creados)} cambios')
                        
                        print(f'‚úÖ Total de evidencias creadas para este archivo: {evidencias_creadas_count} (una por cada cambio)')
                        
                        # Agregar a evidencias_data solo una vez (para evitar duplicados en la respuesta)
                        if primera_evidencia_id:
                            evidencias_data.append({
                                'id': str(primera_evidencia_id),
                                'nombre': archivo.name,
                                'url': file_url,
                                'tipo': archivo.content_type or 'application/octet-stream',
                                'descripcion': descripcion_evidencia
                            })
                    except Exception as e:
                        print(f'‚ùå Error al procesar archivo {key}: {e}')
                        import traceback
                        traceback.print_exc()
                        # NO continuar - lanzar el error para que se revierta la transacci√≥n
                        # Si hay un error al guardar evidencias, es mejor que falle todo
                        raise Exception(f'Error al procesar evidencia {key}: {str(e)}')
                
                print(f'‚úÖ Total de evidencias procesadas exitosamente: {len(evidencias_data)}')
            elif archivos_recibidos:
                print('‚ö†Ô∏è Se recibieron archivos pero no se encontraron con el formato esperado.')
                print(f'‚ö†Ô∏è Cambios creados: {len(cambios_creados)}')
                print(f'‚ö†Ô∏è Claves de archivos recibidos: {list(request.FILES.keys())}')
            else:
                print('‚ÑπÔ∏è No se recibieron archivos en request.FILES (opcional)')
        
        # Obtener nombres de los responsables (todos los colaboradores)
        responsables_nombres = [colab.nombre for colab in colaboradores]
        responsable_nombre = ', '.join(responsables_nombres)
        
        # Formatear fecha en zona horaria de Guatemala
        fecha_display = ''
        if fecha_cambio_final:
            import pytz
            guatemala_tz = pytz.timezone('America/Guatemala')
            if timezone.is_aware(fecha_cambio_final):
                fecha_local = fecha_cambio_final.astimezone(guatemala_tz)
            else:
                fecha_local = timezone.make_aware(fecha_cambio_final, guatemala_tz)
            fecha_display = fecha_local.strftime('%d/%m/%Y %H:%M')
        
        # Retornar informaci√≥n del primer cambio creado (para compatibilidad) y lista de todos
        cambios_ids = [str(c.id) for c in cambios_creados]
        
        return JsonResponse({
            'success': True,
            'message': f'Cambio creado exitosamente para {len(cambios_creados)} colaborador(es)',
            'cambio': {
                'id': str(cambios_creados[0].id),  # Primer cambio para compatibilidad
                'ids': cambios_ids,  # Todos los IDs de cambios creados
                'grupo_id': str(grupo_uuid),
                'descripcion': descripcion,
                'fecha_cambio': fecha_cambio_final.isoformat() if fecha_cambio_final else None,
                'fecha_display': fecha_display,
                'responsable': responsable_nombre,
                'responsables': responsables_nombres,  # Lista de nombres
                'responsables_display': responsable_nombre,
                'colaboradores_ids': [str(c.id) for c in colaboradores],
                'colaboradores': [
                    {
                        'id': str(colaborador.id),
                        'nombre': colaborador.nombre,
                        'puesto': colaborador.puesto.nombre if colaborador.puesto else '',
                        'rol_display': 'Personal Fijo' if colaborador.es_personal_fijo else 'Colaborador Externo',
                    }
                    for colaborador in colaboradores
                ],
                'cantidad_colaboradores': len(colaboradores),
                'evidencias': evidencias_data
            }
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al crear cambio: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["POST"])
def api_actualizar_cambio(request, evento_id, cambio_id):
    """API: Actualizar un cambio existente"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        cambio_colaborador = None
        cambio = None
        
        # Intentar buscar por ID primero
        try:
            cambio_colaborador = EventoCambioColaborador.objects.get(id=cambio_id, actividad=evento)
            print(f'‚úÖ Cambio encontrado por ID: {cambio_id}')
        except EventoCambioColaborador.DoesNotExist:
            print(f'‚ö†Ô∏è Cambio no encontrado por ID: {cambio_id}, intentando buscar por grupo_id o cambio_ids')
            
            # Si no se encuentra por ID, intentar buscar por grupo_id si viene en el POST
            grupo_id_str = request.POST.get('grupo_id')
            cambio_ids_str = request.POST.get('cambio_ids')
            
            # Intentar buscar por grupo_id primero
            if grupo_id_str:
                try:
                    import uuid as uuid_module
                    grupo_id = uuid_module.UUID(grupo_id_str)
                    # Buscar cualquier cambio del grupo
                    cambios_del_grupo = EventoCambioColaborador.objects.filter(
                        grupo_id=grupo_id,
                        actividad=evento
                    )
                    cambio_colaborador = cambios_del_grupo.first()
                    if cambio_colaborador:
                        print(f'‚úÖ Cambio encontrado por grupo_id: {grupo_id} (encontrados {cambios_del_grupo.count()} cambios en el grupo)')
                    else:
                        print(f'‚ö†Ô∏è No se encontr√≥ cambio por grupo_id: {grupo_id} para evento {evento_id}')
                except (ValueError, TypeError) as e:
                    print(f'‚ö†Ô∏è Error al parsear grupo_id: {e}, grupo_id_str: {grupo_id_str}')
                    cambio_colaborador = None
            
            # Si a√∫n no se encuentra, intentar buscar por cambio_ids del POST
            if not cambio_colaborador and cambio_ids_str:
                try:
                    cambio_ids = json.loads(cambio_ids_str)
                    print(f'üîç Intentando buscar por cambio_ids: {cambio_ids}')
                    # Buscar cualquier cambio que coincida con alguno de los IDs
                    for cambio_id_attempt in cambio_ids:
                        try:
                            cambio_colaborador = EventoCambioColaborador.objects.get(
                                id=cambio_id_attempt,
                                actividad=evento
                            )
                            print(f'‚úÖ Cambio encontrado por cambio_id en lista: {cambio_id_attempt}')
                            break
                        except EventoCambioColaborador.DoesNotExist:
                            continue
                except (json.JSONDecodeError, ValueError, TypeError) as e:
                    print(f'‚ö†Ô∏è Error al parsear cambio_ids: {e}, cambio_ids_str: {cambio_ids_str}')
            
            # Si a√∫n no se encuentra, intentar buscar como ActividadCambio
            if not cambio_colaborador:
                try:
                    cambio = ActividadCambio.objects.get(id=cambio_id, actividad=evento)
                    print(f'‚úÖ Cambio encontrado como ActividadCambio: {cambio_id}')
                except ActividadCambio.DoesNotExist:
                    print(f'‚ùå No se encontr√≥ cambio con ID: {cambio_id}, grupo_id: {grupo_id_str}, cambio_ids: {cambio_ids_str}, ni como ActividadCambio')
                    cambio = None
        
        # Validar que se haya encontrado un cambio
        if not cambio_colaborador and not cambio:
            print(f'‚ùå Error: No se encontr√≥ cambio con ID {cambio_id} para evento {evento_id}')
            return JsonResponse({
                'success': False,
                'error': 'Cambio no encontrado'
            }, status=404)
        
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        descripcion = request.POST.get('descripcion', '').strip()
        if not descripcion:
            return JsonResponse({
                'success': False,
                'error': 'La descripci√≥n del cambio es obligatoria'
            }, status=400)
        
        # Obtener fecha_cambio si se proporciona
        fecha_cambio_str = request.POST.get('fecha_cambio')
        fecha_cambio = None
        if fecha_cambio_str:
            try:
                from django.utils.dateparse import parse_datetime
                import pytz
                fecha_cambio_naive = parse_datetime(fecha_cambio_str)
                if fecha_cambio_naive:
                    # Convertir a zona horaria de Guatemala
                    guatemala_tz = pytz.timezone('America/Guatemala')
                    if timezone.is_aware(fecha_cambio_naive):
                        fecha_cambio = fecha_cambio_naive.astimezone(guatemala_tz)
                    else:
                        fecha_cambio = guatemala_tz.localize(fecha_cambio_naive)
            except Exception as e:
                print(f'‚ö†Ô∏è Error al parsear fecha_cambio: {e}, manteniendo fecha existente')
                fecha_cambio = None
        
        usar_fecha_actual = request.POST.get('usar_fecha_actual') == 'true'
        if usar_fecha_actual:
            fecha_cambio = timezone.now()

        colaboradores_ids_str = request.POST.get('colaboradores_ids')
        colaboradores_ids = []
        if colaboradores_ids_str:
            try:
                colaboradores_ids = json.loads(colaboradores_ids_str)
            except json.JSONDecodeError:
                colaboradores_ids = []
        elif request.POST.get('colaborador_id'):
            colaboradores_ids = [request.POST.get('colaborador_id')]

        colaboradores_objs = []
        if colaboradores_ids:
            for colaborador_id in colaboradores_ids:
                try:
                    colaborador = Colaborador.objects.get(id=colaborador_id, activo=True)
                except Colaborador.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': f'Colaborador con ID {colaborador_id} no encontrado'
                    }, status=404)
                if not ActividadPersonal.objects.filter(actividad=evento, colaborador=colaborador).exists():
                    return JsonResponse({
                        'success': False,
                        'error': f'El colaborador {colaborador.nombre} no est√° asignado a este evento'
                    }, status=400)
                colaboradores_objs.append(colaborador)

        # Obtener comunidades seleccionadas
        comunidades_ids_str = request.POST.get('comunidades_ids')
        print(f'üîµ comunidades_ids_str recibido (actualizaci√≥n): {comunidades_ids_str} (tipo: {type(comunidades_ids_str)})')
        comunidades_ids = []
        if comunidades_ids_str:
            try:
                comunidades_ids = json.loads(comunidades_ids_str)
                print(f'‚úÖ Comunidades parseadas (actualizaci√≥n): {comunidades_ids} (cantidad: {len(comunidades_ids)})')
            except json.JSONDecodeError as e:
                print(f'‚ùå Error al parsear JSON de comunidades: {e}')
                print(f'‚ùå String recibido: {comunidades_ids_str}')
        else:
            print(f'‚ö†Ô∏è No se recibi√≥ comunidades_ids_str en POST (actualizaci√≥n)')
            print(f'üîµ Todas las keys POST: {list(request.POST.keys())}')
        
        # Validar que las comunidades pertenezcan al evento
        comunidades_validas = []
        print(f'üîµ comunidades_ids tiene {len(comunidades_ids)} elementos (actualizaci√≥n)')
        if comunidades_ids:
            from webmaga.models import ActividadComunidad, Comunidad
            # Obtener comunidades relacionadas con el evento
            comunidades_evento = ActividadComunidad.objects.filter(
                actividad=evento
            ).values_list('comunidad_id', flat=True)
            
            # Convertir a lista para comparaci√≥n (los IDs pueden ser UUIDs, no enteros)
            comunidades_evento_list = list(comunidades_evento)
            print(f'üîç Comunidades relacionadas con el evento (actualizaci√≥n): {comunidades_evento_list} (cantidad: {len(comunidades_evento_list)})')
            print(f'üîç Tipos de IDs en comunidades_evento: {[type(cid) for cid in comunidades_evento_list[:3]] if comunidades_evento_list else "lista vac√≠a"}')
            
            for comunidad_id_str in comunidades_ids:
                try:
                    # Los IDs pueden ser UUIDs (strings) o enteros, intentar ambos
                    comunidad_id_parsed = None
                    try:
                        # Intentar convertir a UUID si es string UUID
                        import uuid as uuid_module
                        if isinstance(comunidad_id_str, str):
                            comunidad_id_parsed = uuid_module.UUID(comunidad_id_str)
                        else:
                            comunidad_id_parsed = comunidad_id_str
                    except (ValueError, AttributeError):
                        # Si no es UUID, intentar como entero
                        try:
                            comunidad_id_parsed = int(comunidad_id_str) if isinstance(comunidad_id_str, str) else comunidad_id_str
                        except (ValueError, TypeError):
                            comunidad_id_parsed = comunidad_id_str
                    
                    print(f'üîç Verificando comunidad ID (actualizaci√≥n): {comunidad_id_str} -> {comunidad_id_parsed} (tipo: {type(comunidad_id_parsed)})')
                    
                    # Verificar que la comunidad est√© relacionada con el evento
                    # Comparar tanto el ID original como el parseado
                    comunidad_encontrada = False
                    for cid_evento in comunidades_evento_list:
                        # Comparar como strings para UUIDs
                        if str(cid_evento) == str(comunidad_id_parsed) or cid_evento == comunidad_id_parsed:
                            comunidad_encontrada = True
                            break
                    
                    if comunidad_encontrada:
                        # Intentar obtener la comunidad con el ID parseado
                        comunidad = Comunidad.objects.select_related('region').get(id=comunidad_id_parsed, activo=True)
                        comunidades_validas.append(comunidad)
                        print(f'‚úÖ Comunidad v√°lida (actualizaci√≥n): {comunidad.nombre} (ID: {comunidad.id}, Regi√≥n: {comunidad.region.nombre if comunidad.region else "Sin regi√≥n"})')
                    else:
                        print(f'‚ö†Ô∏è Comunidad {comunidad_id_str} ({comunidad_id_parsed}) no est√° relacionada con el evento.')
                        print(f'   Comunidades del evento: {[str(cid) for cid in comunidades_evento_list[:5]]}')
                except Comunidad.DoesNotExist:
                    print(f'‚ö†Ô∏è Comunidad con ID {comunidad_id_str} no encontrada en la base de datos')
                except Exception as e:
                    print(f'‚ö†Ô∏è Error con comunidad ID {comunidad_id_str} (actualizaci√≥n): {e}')
                    import traceback
                    traceback.print_exc()

        if cambio_colaborador:
            grupo_uuid = cambio_colaborador.grupo_id
            grupo_cambios_qs = EventoCambioColaborador.objects.filter(
                actividad=evento,
                grupo_id=grupo_uuid
            ).select_related('colaborador').prefetch_related('evidencias')

            # Guardar evidencias antes de eliminar cambios (para reasignarlas despu√©s)
            evidencias_por_cambio = {}
            for cambio_rel in grupo_cambios_qs:
                evidencias_por_cambio[str(cambio_rel.id)] = list(cambio_rel.evidencias.all())

            cambios_por_colaborador = {}
            for cambio_rel in grupo_cambios_qs:
                key = str(cambio_rel.colaborador.id) if cambio_rel.colaborador else None
                cambios_por_colaborador[key] = cambio_rel

            if colaboradores_objs:
                # Guardar los IDs de los cambios antiguos antes de crear nuevos
                cambios_antiguos_ids = list(grupo_cambios_qs.values_list('id', flat=True))
                print(f'üìé Guardando {len(cambios_antiguos_ids)} cambios antiguos para eliminar despu√©s')
                
                nuevos_o_actualizados = []
                fecha_cambio_final = fecha_cambio or cambio_colaborador.fecha_cambio
                
                print(f'üîµ Total colaboradores: {len(colaboradores_objs)}, Total comunidades v√°lidas: {len(comunidades_validas)}')
                for colaborador in colaboradores_objs:
                    print(f'üîµ Procesando colaborador: {colaborador.nombre}')
                    if comunidades_validas:
                        print(f'üîµ Creando cambios con {len(comunidades_validas)} comunidades')
                        # Crear un cambio por cada combinaci√≥n colaborador-comunidad
                        for comunidad in comunidades_validas:
                            # Obtener regi√≥n de la comunidad si existe (ya viene con select_related)
                            region_id = None
                            if comunidad.region:
                                region_id = comunidad.region.id
                                print(f'üîç Regi√≥n obtenida de comunidad (actualizaci√≥n): {comunidad.region.nombre} (ID: {region_id})')
                            else:
                                print(f'‚ö†Ô∏è Comunidad {comunidad.nombre} no tiene regi√≥n asociada (actualizaci√≥n)')
                            
                            cambio_nuevo = EventoCambioColaborador.objects.create(
                                actividad=evento,
                                colaborador=colaborador,
                                descripcion_cambio=descripcion,
                                fecha_cambio=fecha_cambio_final,
                                grupo_id=grupo_uuid,
                                comunidad_id=comunidad.id,
                                region_id=region_id
                            )
                            nuevos_o_actualizados.append(cambio_nuevo)
                            print(f'‚úÖ Cambio actualizado para colaborador {colaborador.nombre} y comunidad {comunidad.nombre} (ID: {comunidad.id}, Regi√≥n ID: {region_id}): Cambio ID {cambio_nuevo.id}')
                    else:
                        # Sin comunidades, crear cambio sin comunidad
                        print(f'‚ö†Ô∏è NO hay comunidades v√°lidas para colaborador {colaborador.nombre}, creando cambio sin comunidad')
                        cambio_nuevo = EventoCambioColaborador.objects.create(
                            actividad=evento,
                            colaborador=colaborador,
                            descripcion_cambio=descripcion,
                            fecha_cambio=fecha_cambio_final,
                            grupo_id=grupo_uuid
                        )
                        nuevos_o_actualizados.append(cambio_nuevo)
                        print(f'‚úÖ Cambio creado sin comunidad para colaborador {colaborador.nombre}: {cambio_nuevo.id}')
                
                # Reasignar evidencias ANTES de eliminar los cambios antiguos (para evitar CASCADE delete)
                if nuevos_o_actualizados and evidencias_por_cambio:
                    primer_cambio = nuevos_o_actualizados[0]
                    print(f'üìé Reasignando {sum(len(evs) for evs in evidencias_por_cambio.values())} evidencia(s) al nuevo cambio {primer_cambio.id}')
                    for cambio_id_antiguo, evidencias_list in evidencias_por_cambio.items():
                        for evidencia in evidencias_list:
                            # Verificar que la evidencia a√∫n existe antes de reasignarla
                            try:
                                evidencia.refresh_from_db()
                                evidencia.cambio = primer_cambio
                                evidencia.save(update_fields=['cambio'])
                                print(f'‚úÖ Evidencia {evidencia.id} reasignada al cambio {primer_cambio.id}')
                            except Exception as e:
                                print(f'‚ö†Ô∏è Error al reasignar evidencia {evidencia.id}: {e}')
                    
                    # Ahora s√≠ eliminar los cambios antiguos (las evidencias ya est√°n reasignadas)
                    if cambios_antiguos_ids:
                        EventoCambioColaborador.objects.filter(id__in=cambios_antiguos_ids).delete()
                        print(f'‚úÖ {len(cambios_antiguos_ids)} cambios antiguos eliminados despu√©s de reasignar evidencias')

                if nuevos_o_actualizados:
                    cambio_colaborador = nuevos_o_actualizados[0]
                    
                    # Procesar nuevas evidencias si se enviaron
                    archivos_recibidos = bool(request.FILES)
                    if archivos_recibidos:
                        archivos_keys_list = [key for key in request.FILES.keys() if key.startswith('archivo_')]
                        if not archivos_keys_list:
                            archivos_keys_list = list(request.FILES.keys())
                        
                        if archivos_keys_list:
                            print(f'üìé Procesando {len(archivos_keys_list)} nueva(s) evidencia(s) al actualizar cambio')
                            evidencias_dir = os.path.join(str(settings.MEDIA_ROOT), 'evidencias_cambios_eventos')
                            os.makedirs(evidencias_dir, exist_ok=True)
                            fs = FileSystemStorage(location=evidencias_dir)
                            
                            # Obtener todos los cambios del grupo actualizados para asociar evidencias
                            cambios_para_evidencias = EventoCambioColaborador.objects.filter(
                                actividad=evento,
                                grupo_id=grupo_uuid
                            )
                            
                            for index, key in enumerate(archivos_keys_list):
                                try:
                                    archivo = request.FILES[key]
                                    print(f'üìé Procesando nueva evidencia {key}: {archivo.name} ({archivo.size} bytes)')
                                    
                                    # Obtener descripci√≥n de la evidencia
                                    descripcion_evidencia = request.POST.get(f'descripcion_evidencia_{index}', '').strip()
                                    if not descripcion_evidencia:
                                        index_num = key.replace('archivo_', '')
                                        descripcion_evidencia = request.POST.get(f'descripcion_evidencia_{index_num}', '').strip()
                                    if not descripcion_evidencia:
                                        descripcion_evidencia = request.POST.get('descripcion_evidencia', '').strip()
                                    
                                    # Guardar el archivo f√≠sico
                                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
                                    file_extension = os.path.splitext(archivo.name)[1]
                                    filename = f"{timestamp}_{cambio_colaborador.id}_{index}{file_extension}"
                                    saved_name = fs.save(filename, archivo)
                                    file_url = f"/media/evidencias_cambios_eventos/{saved_name}"
                                    
                                    # Crear evidencia para cada cambio del grupo
                                    for cambio_evidencia in cambios_para_evidencias:
                                        EventosEvidenciasCambios.objects.create(
                                            actividad=evento,
                                            cambio=cambio_evidencia,
                                            archivo_nombre=archivo.name,
                                            archivo_tipo=archivo.content_type or 'application/octet-stream',
                                            archivo_tamanio=archivo.size,
                                            url_almacenamiento=file_url,
                                            descripcion=descripcion_evidencia if descripcion_evidencia else None,
                                            creado_por=usuario_maga
                                        )
                                    print(f'‚úÖ Nueva evidencia agregada: {archivo.name}')
                                except Exception as e:
                                    print(f'‚ùå Error al procesar nueva evidencia {key}: {e}')
                                    import traceback
                                    traceback.print_exc()

            else:
                # No se enviaron colaboradores, mantener colaborador existente
                cambio_colaborador.descripcion_cambio = descripcion
                if fecha_cambio:
                    cambio_colaborador.fecha_cambio = fecha_cambio
                cambio_colaborador.save(update_fields=['descripcion_cambio', 'fecha_cambio'])
                
                # Procesar nuevas evidencias si se enviaron (cuando NO se recrean cambios)
                archivos_recibidos = bool(request.FILES)
                if archivos_recibidos:
                    archivos_keys_list = [key for key in request.FILES.keys() if key.startswith('archivo_')]
                    if not archivos_keys_list:
                        archivos_keys_list = list(request.FILES.keys())
                    
                    if archivos_keys_list:
                        print(f'üìé Procesando {len(archivos_keys_list)} nueva(s) evidencia(s) al actualizar cambio (sin recrear)')
                        evidencias_dir = os.path.join(str(settings.MEDIA_ROOT), 'evidencias_cambios_eventos')
                        os.makedirs(evidencias_dir, exist_ok=True)
                        fs = FileSystemStorage(location=evidencias_dir)
                        
                        # Obtener todos los cambios del grupo para asociar evidencias
                        cambios_para_evidencias = EventoCambioColaborador.objects.filter(
                            actividad=evento,
                            grupo_id=grupo_uuid
                        )
                        
                        for index, key in enumerate(archivos_keys_list):
                            try:
                                archivo = request.FILES[key]
                                print(f'üìé Procesando nueva evidencia {key}: {archivo.name} ({archivo.size} bytes)')
                                
                                # Obtener descripci√≥n de la evidencia
                                descripcion_evidencia = request.POST.get(f'descripcion_evidencia_{index}', '').strip()
                                if not descripcion_evidencia:
                                    index_num = key.replace('archivo_', '')
                                    descripcion_evidencia = request.POST.get(f'descripcion_evidencia_{index_num}', '').strip()
                                if not descripcion_evidencia:
                                    descripcion_evidencia = request.POST.get('descripcion_evidencia', '').strip()
                                
                                # Guardar el archivo f√≠sico
                                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
                                file_extension = os.path.splitext(archivo.name)[1]
                                filename = f"{timestamp}_{cambio_colaborador.id}_{index}{file_extension}"
                                saved_name = fs.save(filename, archivo)
                                file_url = f"/media/evidencias_cambios_eventos/{saved_name}"
                                
                                # Crear evidencia para cada cambio del grupo
                                for cambio_evidencia in cambios_para_evidencias:
                                    EventosEvidenciasCambios.objects.create(
                                        actividad=evento,
                                        cambio=cambio_evidencia,
                                        archivo_nombre=archivo.name,
                                        archivo_tipo=archivo.content_type or 'application/octet-stream',
                                        archivo_tamanio=archivo.size,
                                        url_almacenamiento=file_url,
                                        descripcion=descripcion_evidencia if descripcion_evidencia else None,
                                        creado_por=usuario_maga
                                    )
                                print(f'‚úÖ Nueva evidencia agregada: {archivo.name}')
                            except Exception as e:
                                print(f'‚ùå Error al procesar nueva evidencia {key}: {e}')
                                import traceback
                                traceback.print_exc()

            responsables_nombres = []
            grupo_cambios_actualizados = EventoCambioColaborador.objects.filter(
                actividad=evento,
                grupo_id=grupo_uuid
            ).select_related('colaborador')
            for cambio_rel in grupo_cambios_actualizados:
                if cambio_rel.colaborador:
                    responsables_nombres.append(cambio_rel.colaborador.nombre)

            responsable_nombre = ', '.join(responsables_nombres) if responsables_nombres else 'Colaborador desconocido'
            fecha_base = cambio_colaborador.fecha_cambio
            colaborador_response_id = str(cambio_colaborador.colaborador.id) if cambio_colaborador.colaborador else None
            colaboradores_response_ids = [str(c.colaborador.id) for c in grupo_cambios_actualizados if c.colaborador]
            grupo_uuid_response = str(grupo_uuid)
        else:
            # Cambio de usuario/responsable
            cambio.descripcion_cambio = descripcion
            cambio.responsable = cambio.responsable or usuario_maga
            if fecha_cambio:
                cambio.fecha_cambio = fecha_cambio
            cambio.save()
            responsable_nombre = cambio.responsable.nombre or cambio.responsable.username if cambio.responsable else 'Sistema'
            fecha_base = cambio.fecha_cambio
            colaborador_response_id = None
            colaboradores_response_ids = []
            grupo_uuid_response = None
        
        # Formatear fecha en zona horaria de Guatemala
        fecha_display = ''
        if fecha_base:
            import pytz
            guatemala_tz = pytz.timezone('America/Guatemala')
            if timezone.is_aware(fecha_base):
                fecha_local = fecha_base.astimezone(guatemala_tz)
            else:
                fecha_local = timezone.make_aware(fecha_base, guatemala_tz)
            fecha_display = fecha_local.strftime('%d/%m/%Y %H:%M')
        
        return JsonResponse({
            'success': True,
            'message': 'Cambio actualizado exitosamente',
            'cambio': {
                'id': str((cambio_colaborador or cambio).id),
                'descripcion': descripcion,
                'fecha_cambio': fecha_base.isoformat() if fecha_base else None,
                'fecha_display': fecha_display,
                'responsable': responsable_nombre,
                'colaborador_id': colaborador_response_id,
                'responsable_id': colaborador_response_id if cambio_colaborador else (str(cambio.responsable.id) if cambio and cambio.responsable else None),
                'responsables_display': responsable_nombre,
                'colaboradores_ids': colaboradores_response_ids if cambio_colaborador else [],
                'grupo_id': grupo_uuid_response,
            }
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except (ActividadCambio.DoesNotExist, EventoCambioColaborador.DoesNotExist):
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar cambio: {str(e)}'
        }, status=500)
@permiso_gestionar_eventos
@require_http_methods(["DELETE", "POST"])
def api_eliminar_cambio(request, evento_id, cambio_id):
    """API: Eliminar un cambio y sus evidencias"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        cambio_colaborador = None
        cambio = None
        try:
            cambio_colaborador = EventoCambioColaborador.objects.get(id=cambio_id, actividad=evento)
        except EventoCambioColaborador.DoesNotExist:
            cambio = ActividadCambio.objects.get(id=cambio_id, actividad=evento)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        if cambio_colaborador:
            grupo_uuid = cambio_colaborador.grupo_id
            cambios_grupo = EventoCambioColaborador.objects.filter(
                actividad=evento,
                grupo_id=grupo_uuid
            ).select_related('colaborador').prefetch_related('evidencias')
            eliminados = 0
            for cambio_rel in cambios_grupo:
                evidencias_iter = cambio_rel.evidencias.all()
                for evidencia in evidencias_iter:
                    _eliminar_archivo_media(getattr(evidencia, 'url_almacenamiento', None))
                cambio_rel.delete()
                eliminados += 1
            print(f'üóëÔ∏è Eliminados {eliminados} cambios pertenecientes al grupo {grupo_uuid}')
            cambio_colaborador = None  # Ya eliminados
        else:
            evidencias_iter = cambio.evidencias.all()
            for evidencia in evidencias_iter:
                _eliminar_archivo_media(getattr(evidencia, 'url_almacenamiento', None))
            cambio.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Cambio eliminado exitosamente'
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except (ActividadCambio.DoesNotExist, EventoCambioColaborador.DoesNotExist):
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar cambio: {str(e)}'
        }, status=500)
@permiso_gestionar_eventos
@require_http_methods(["POST"])
def api_agregar_evidencia_cambio(request, evento_id, cambio_id):
    """API: Agregar evidencia a un cambio existente"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        cambio = EventoCambioColaborador.objects.get(id=cambio_id, actividad=evento)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({
                'success': False,
                'error': 'No se ha enviado ning√∫n archivo'
            }, status=400)
        
        descripcion = request.POST.get('descripcion', '').strip()
        
        # Crear carpeta si no existe
        evidencias_dir = os.path.join(settings.MEDIA_ROOT, 'evidencias_cambios_eventos')
        os.makedirs(evidencias_dir, exist_ok=True)
        
        # Guardar archivo
        fs = FileSystemStorage(location=evidencias_dir)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
        file_extension = os.path.splitext(archivo.name)[1]
        filename = f"{timestamp}_{cambio_id}{file_extension}"
        saved_name = fs.save(filename, archivo)
        file_url = f"/media/evidencias_cambios_eventos/{saved_name}"
        
        # Crear registro en la BD usando la nueva tabla eventos_evidencias_cambios
        evidencia = EventosEvidenciasCambios.objects.create(
            actividad=evento,
            cambio=cambio,
            archivo_nombre=archivo.name,
            archivo_tipo=archivo.content_type or 'application/octet-stream',
            archivo_tamanio=archivo.size,
            url_almacenamiento=file_url,
            descripcion=descripcion,
            creado_por=usuario_maga
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Evidencia agregada exitosamente',
            'evidencia': {
                'id': str(evidencia.id),
                'nombre': evidencia.archivo_nombre,
                'url': evidencia.url_almacenamiento,
                'tipo': evidencia.archivo_tipo or '',
                'descripcion': evidencia.descripcion or ''
            }
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except EventoCambioColaborador.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al agregar evidencia: {str(e)}'
        }, status=500)


@solo_administrador
@require_http_methods(["GET"])
def api_listar_colaboradores(request):
    """API: Listar todos los colaboradores para vincular con usuarios"""
    try:
        colaboradores = Colaborador.objects.select_related('puesto', 'usuario').order_by('nombre')
        
        colaboradores_list = []
        for colaborador in colaboradores:
            colaboradores_list.append({
                'id': str(colaborador.id),
                'nombre': colaborador.nombre,
                'puesto_id': str(colaborador.puesto.id) if colaborador.puesto else None,
                'puesto_nombre': colaborador.puesto.nombre if colaborador.puesto else None,
                'puesto_codigo': colaborador.puesto.codigo if colaborador.puesto else None,
                'descripcion': colaborador.descripcion or '',
                'telefono': colaborador.telefono or '',
                'correo': colaborador.correo or '',
                'dpi': colaborador.dpi or '',
                'es_personal_fijo': colaborador.es_personal_fijo,
                'activo': colaborador.activo,
                'tiene_usuario': colaborador.usuario_id is not None,
                'usuario_id': str(colaborador.usuario.id) if colaborador.usuario else None,
                'usuario_username': colaborador.usuario.username if colaborador.usuario else None,
            })
        
        return JsonResponse({
            'success': True,
            'colaboradores': colaboradores_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al listar colaboradores: {str(e)}'
        }, status=500)


@solo_administrador
@require_http_methods(["GET"])
def api_obtener_colaborador(request, colaborador_id):
    """API: Obtener informaci√≥n de un colaborador espec√≠fico (solo admin)"""
    try:
        colaborador = Colaborador.objects.select_related('puesto', 'usuario').get(id=colaborador_id)
        
        return JsonResponse({
            'success': True,
            'colaborador': {
                'id': str(colaborador.id),
                'nombre': colaborador.nombre,
                'puesto_id': str(colaborador.puesto.id) if colaborador.puesto else None,
                'puesto_nombre': colaborador.puesto.nombre if colaborador.puesto else None,
                'puesto_codigo': colaborador.puesto.codigo if colaborador.puesto else None,
                'descripcion': colaborador.descripcion or '',
                'telefono': colaborador.telefono or '',
                'correo': colaborador.correo or '',
                'dpi': colaborador.dpi or '',
                'es_personal_fijo': colaborador.es_personal_fijo,
                'activo': colaborador.activo,
                'tiene_usuario': colaborador.usuario_id is not None,
                'usuario_id': str(colaborador.usuario.id) if colaborador.usuario else None,
                'usuario_username': colaborador.usuario.username if colaborador.usuario else None,
            }
        })
    except Colaborador.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Colaborador no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener colaborador: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
@permiso_admin_o_personal_api
def api_usuario_colaborador(request):
    """API: Obtener informaci√≥n del colaborador vinculado al usuario actual"""
    usuario_maga = get_usuario_maga(request.user)
    
    if not usuario_maga:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    
    try:
        # Obtener colaborador vinculado al usuario
        colaborador = None
        if hasattr(usuario_maga, 'colaborador') and usuario_maga.colaborador:
            colaborador = usuario_maga.colaborador
            
            return JsonResponse({
                'success': True,
                'colaborador': {
                    'id': str(colaborador.id),
                    'nombre': colaborador.nombre,
                    'puesto_id': str(colaborador.puesto.id) if colaborador.puesto else None,
                    'puesto_nombre': colaborador.puesto.nombre if colaborador.puesto else None,
                    'puesto_codigo': colaborador.puesto.codigo if colaborador.puesto else None,
                    'descripcion': colaborador.descripcion or '',
                    'telefono': colaborador.telefono or '',
                    'correo': colaborador.correo or '',
                    'dpi': colaborador.dpi or '',
                    'es_personal_fijo': colaborador.es_personal_fijo,
                    'activo': colaborador.activo,
                }
            })
        else:
            return JsonResponse({
                'success': True,
                'colaborador': None
            })
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener colaborador: {str(e)}'
        }, status=500)


@solo_administrador
@require_http_methods(["POST"])
def api_crear_colaborador(request):
    """API: Crear un nuevo colaborador"""
    try:
        usuario_maga = get_usuario_maga(request.user)
        data = json.loads(request.body or '{}')
        
        nombre = (data.get('nombre') or '').strip()
        puesto_id = (data.get('puesto_id') or '').strip()
        descripcion = (data.get('descripcion') or '').strip()
        telefono = (data.get('telefono') or '').strip()
        correo = (data.get('correo') or '').strip()
        dpi = (data.get('dpi') or '').strip()
        es_personal_fijo = _to_bool(data.get('es_personal_fijo'), False)
        activo = _to_bool(data.get('activo'), True)
        
        # Validaciones
        if not nombre:
            return JsonResponse({
                'success': False,
                'error': 'El nombre es requerido'
            }, status=400)
        
        # Validar puesto si se proporciona
        puesto = None
        if puesto_id:
            try:
                puesto = Puesto.objects.get(id=puesto_id, activo=True)
            except Puesto.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Puesto no encontrado'
                }, status=404)
        
        # Normalizar correo y DPI
        correo_normalizado = correo.lower() if correo else ''
        dpi_normalizado = dpi.replace(' ', '') if dpi else ''
        correo = correo_normalizado
        dpi = dpi_normalizado
        correo = correo_normalizado
        dpi = dpi_normalizado

        # Validar unicidad de correo
        if correo_normalizado:
            if Colaborador.objects.filter(correo__iexact=correo_normalizado).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe un colaborador con ese correo electr√≥nico'
                }, status=400)

        # Validar unicidad de DPI
        if dpi_normalizado:
            if Colaborador.objects.filter(dpi=dpi_normalizado).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe un colaborador con ese n√∫mero de DPI'
                }, status=400)

        # Validar restricci√≥n de personal fijo: solo se permite si se vincula un usuario
        if es_personal_fijo:
            return JsonResponse({
                'success': False,
                'error': 'No se puede marcar como personal fijo sin un usuario vinculado. Primero cree el colaborador y luego genere el usuario correspondiente desde la gesti√≥n de usuarios.'
            }, status=400)

        # Crear colaborador
        colaborador = Colaborador.objects.create(
            nombre=nombre,
            puesto=puesto,
            descripcion=descripcion if descripcion else None,
            telefono=telefono if telefono else None,
            correo=correo if correo else None,
            dpi=dpi if dpi else None,
            es_personal_fijo=False,
            activo=activo,
            creado_por=usuario_maga if usuario_maga else None
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Colaborador creado exitosamente',
            'colaborador': {
                'id': str(colaborador.id),
                'nombre': colaborador.nombre,
                'puesto_nombre': colaborador.puesto.nombre if colaborador.puesto else None,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inv√°lidos'
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al crear colaborador: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def api_listar_puestos(request):
    """API: Listar todos los puestos activos"""
    try:
        puestos = Puesto.objects.filter(activo=True).order_by('nombre')
        
        puestos_list = []
        for puesto in puestos:
            puestos_list.append({
                'id': str(puesto.id),
                'codigo': puesto.codigo,
                'nombre': puesto.nombre,
                'descripcion': puesto.descripcion or '',
            })
        
        return JsonResponse({
            'success': True,
            'puestos': puestos_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al listar puestos: {str(e)}'
        }, status=500)


@solo_administrador
@require_http_methods(["POST"])
def api_crear_puesto(request):
    """API: Crear un nuevo puesto"""
    try:
        data = json.loads(request.body or '{}')
        
        codigo = data.get('codigo', '').strip().upper()
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        
        # Validaciones
        if not codigo or not nombre:
            return JsonResponse({
                'success': False,
                'error': 'C√≥digo y nombre son requeridos'
            }, status=400)
        
        # Verificar c√≥digo √∫nico
        if Puesto.objects.filter(codigo=codigo).exists():
            return JsonResponse({
                'success': False,
                'error': 'El c√≥digo del puesto ya existe'
            }, status=400)
        
        # Crear puesto
        puesto = Puesto.objects.create(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion if descripcion else None,
            activo=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Puesto creado exitosamente',
            'puesto': {
                'id': str(puesto.id),
                'codigo': puesto.codigo,
                'nombre': puesto.nombre,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inv√°lidos'
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al crear puesto: {str(e)}'
        }, status=500)


@solo_administrador
@require_http_methods(["GET"])
def api_obtener_usuario(request, usuario_id):
    """API: Obtener un usuario espec√≠fico"""
    try:
        usuario = Usuario.objects.select_related('puesto', 'colaborador__puesto').get(id=usuario_id)
        
        # Obtener colaborador vinculado si existe (usando la relaci√≥n inversa OneToOne)
        colaborador = usuario.colaborador if hasattr(usuario, 'colaborador') else None
        colaborador_id = str(colaborador.id) if colaborador else None
        colaborador_nombre = colaborador.nombre if colaborador else None
        tiene_colaborador = colaborador is not None
        colaborador_puesto_id = str(colaborador.puesto.id) if colaborador and colaborador.puesto else None
        
        # Determinar el puesto a mostrar: primero del usuario, luego del colaborador
        puesto_id = None
        puesto_nombre = None
        
        if usuario.puesto:
            puesto_id = str(usuario.puesto.id)
            puesto_nombre = usuario.puesto.nombre
        elif colaborador and colaborador.puesto:
            puesto_id = str(colaborador.puesto.id)
            puesto_nombre = colaborador.puesto.nombre
        
        return JsonResponse({
            'success': True,
            'usuario': {
                'id': str(usuario.id),
                'username': usuario.username,
                'nombre': usuario.nombre,
                'email': usuario.email,
                'telefono': usuario.telefono,
                'rol': usuario.rol,
                'rol_display': 'Administrador' if usuario.rol == 'admin' else 'Personal',
                'puesto_id': puesto_id,
                'puesto_nombre': puesto_nombre,
                'activo': usuario.activo,
                'tiene_colaborador': tiene_colaborador,
                'colaborador_id': colaborador_id,
                'colaborador_nombre': colaborador_nombre,
                'colaborador_puesto_id': colaborador_puesto_id,
            }
        })
    except Usuario.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener usuario: {str(e)}'
        }, status=500)


def api_actualizar_usuario(request, usuario_id):
    """API: Actualizar un usuario existente"""
    try:
        usuario = Usuario.objects.get(id=usuario_id)
        data = json.loads(request.body or '{}')
        
        nombre = (data.get('nombre') or '').strip()
        email = (data.get('email') or '').strip()
        telefono = (data.get('telefono') or '').strip()
        password = data.get('password') or ''
        password_confirm = data.get('password_confirm') or ''
        rol = (data.get('rol') or '').strip()
        puesto_id = (data.get('puesto_id') or '').strip()
        
        # Validaciones
        if not email:
            return JsonResponse({
                'success': False,
                'error': 'El email es requerido'
            }, status=400)
        
        # Validar email √∫nico (excepto el mismo usuario)
        if Usuario.objects.filter(email=email).exclude(id=usuario_id).exists():
            return JsonResponse({
                'success': False,
                'error': 'El email ya est√° en uso'
            }, status=400)
        
        # Validar contrase√±a si se proporciona
        if password:
            if password != password_confirm:
                return JsonResponse({
                    'success': False,
                    'error': 'Las contrase√±as no coinciden'
                }, status=400)
            
            if len(password) < 8:
                return JsonResponse({
                    'success': False,
                    'error': 'La contrase√±a debe tener al menos 8 caracteres'
                }, status=400)
        
        # Validar rol
        if rol not in ['admin', 'personal']:
            return JsonResponse({
                'success': False,
                'error': 'Rol inv√°lido'
            }, status=400)
        
        # Obtener colaborador vinculado si existe (usando la relaci√≥n inversa OneToOne)
        colaborador = usuario.colaborador if hasattr(usuario, 'colaborador') else None
        tiene_colaborador = colaborador is not None
        
        # Validar y asignar puesto seg√∫n la restricci√≥n de base de datos:
        # - Si rol es 'admin': puesto del usuario DEBE ser NULL
        # - Si rol es 'personal': puesto del usuario DEBE tener un valor
        puesto = None
        puesto_colaborador = None
        
        if rol == 'personal':
            # Para rol personal, el puesto es requerido
            if puesto_id:
                try:
                    puesto = Puesto.objects.get(id=puesto_id, activo=True)
                except Puesto.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Puesto no encontrado'
                    }, status=404)
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'El puesto es requerido para el rol personal'
                }, status=400)
        elif rol == 'admin':
            # Para rol admin, el puesto del usuario DEBE ser NULL
            # Si tiene colaborador vinculado y se proporciona puesto, actualizar solo el colaborador
            if puesto_id:
                try:
                    puesto_colaborador = Puesto.objects.get(id=puesto_id, activo=True)
                except Puesto.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Puesto no encontrado'
                    }, status=404)
            # puesto sigue siendo None para el usuario (cumple la restricci√≥n)
        
        # Actualizar usuario
        usuario.nombre = nombre if nombre else None
        usuario.email = email
        usuario.telefono = telefono if telefono else None
        usuario.rol = rol
        usuario.puesto = puesto  # None para admin, valor para personal
        
        # Actualizar contrase√±a si se proporciona
        if password:
            from django.contrib.auth.hashers import make_password
            usuario.password_hash = make_password(password)
        
        usuario.save()
        
        # Si el usuario tiene colaborador vinculado, tambi√©n actualizar campos comunes
        if colaborador:
            # Actualizar puesto del colaborador si se proporcion√≥
            if puesto_id:
                puesto_a_asignar = puesto if puesto else puesto_colaborador
                if puesto_a_asignar:
                    colaborador.puesto = puesto_a_asignar
            
            # Sincronizar campos comunes: nombre, email, tel√©fono
            # Solo actualizar si el valor no est√° vac√≠o
            if nombre:
                colaborador.nombre = nombre
            if email:
                colaborador.correo = email
            if telefono:
                colaborador.telefono = telefono
            
            colaborador.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Usuario actualizado exitosamente',
            'usuario': {
                'id': str(usuario.id),
                'username': usuario.username,
                'email': usuario.email,
                'rol': usuario.rol
            }
        })
        
    except Usuario.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inv√°lidos'
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar usuario: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["POST"])
def api_actualizar_evidencia_cambio(request, evento_id, cambio_id, evidencia_id):
    """API: Actualizar descripci√≥n de una evidencia de cambio"""
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        cambio = EventoCambioColaborador.objects.get(id=cambio_id, actividad=evento)
        evidencia = EventosEvidenciasCambios.objects.get(id=evidencia_id, actividad=evento, cambio=cambio)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        descripcion = request.POST.get('descripcion', '').strip()
        evidencia.descripcion = descripcion
        evidencia.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Descripci√≥n actualizada exitosamente',
            'evidencia': {
                'id': str(evidencia.id),
                'nombre': evidencia.archivo_nombre,
                'url': evidencia.url_almacenamiento,
                'tipo': evidencia.archivo_tipo or '',
                'descripcion': evidencia.descripcion or ''
            }
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except EventoCambioColaborador.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        }, status=404)
    except EventosEvidenciasCambios.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evidencia no encontrada'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar evidencia: {str(e)}'
        }, status=500)


@solo_administrador
@require_http_methods(["POST"])
def api_eliminar_usuario(request, usuario_id):
    """API: Eliminar (desactivar) un usuario"""
    try:
        usuario = Usuario.objects.get(id=usuario_id)
        
        # Verificar si tiene colaborador vinculado
        try:
            colaborador = Colaborador.objects.get(usuario=usuario)
            # Desvincular colaborador
            colaborador.usuario = None
            colaborador.es_personal_fijo = False
            colaborador.save()
        except Colaborador.DoesNotExist:
            pass
        
        # Desactivar usuario en lugar de eliminarlo
        usuario.activo = False
        usuario.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Usuario eliminado exitosamente'
        })
        
    except Usuario.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Usuario no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar usuario: {str(e)}'
        }, status=500)


@permiso_gestionar_eventos
@require_http_methods(["DELETE", "POST"])
def api_eliminar_evidencia_cambio(request, evento_id, cambio_id, evidencia_id):
    try:
        evento = Actividad.objects.get(id=evento_id, eliminado_en__isnull=True)
        cambio = EventoCambioColaborador.objects.get(id=cambio_id, actividad=evento)
        evidencia = EventosEvidenciasCambios.objects.get(id=evidencia_id, actividad=evento, cambio=cambio)
        usuario_maga = get_usuario_maga(request.user)
        
        if not usuario_maga:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no autenticado'
            }, status=401)
        
        _eliminar_queryset_con_archivos(
            EventosEvidenciasCambios.objects.filter(pk=evidencia.pk)
        )

        return JsonResponse({
            'success': True,
            'message': 'Evidencia eliminada exitosamente'
        })
        
    except Actividad.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evento no encontrado'
        }, status=404)
    except EventoCambioColaborador.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cambio no encontrado'
        }, status=404)
    except EventosEvidenciasCambios.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evidencia no encontrada'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar evidencia: {str(e)}'
        }, status=500)
@solo_administrador
@require_http_methods(["POST"])
def api_actualizar_colaborador(request, colaborador_id):
    """API: Actualizar un colaborador existente"""
    try:
        colaborador = Colaborador.objects.get(id=colaborador_id)
        data = json.loads(request.body or '{}')
        
        nombre = (data.get('nombre') or '').strip()
        puesto_id = (data.get('puesto_id') or '').strip()
        descripcion = (data.get('descripcion') or '').strip()
        telefono = (data.get('telefono') or '').strip()
        correo = (data.get('correo') or '').strip()
        dpi = (data.get('dpi') or '').strip()
        es_personal_fijo = _to_bool(data.get('es_personal_fijo'), False)
        activo = _to_bool(data.get('activo'), True)
        
        # Validaciones
        if not nombre:
            return JsonResponse({
                'success': False,
                'error': 'El nombre es requerido'
            }, status=400)
        
        # Validar puesto si se proporciona
        puesto = None
        if puesto_id:
            try:
                puesto = Puesto.objects.get(id=puesto_id, activo=True)
            except Puesto.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Puesto no encontrado'
                }, status=404)
        
        # Validar coherencia entre personal fijo y usuario
        if colaborador.usuario:
            if not es_personal_fijo:
                return JsonResponse({
                    'success': False,
                    'error': 'No se puede desmarcar como personal fijo si tiene usuario asignado'
                }, status=400)
        else:
            if es_personal_fijo:
                return JsonResponse({
                    'success': False,
                    'error': 'Para marcar a un colaborador como personal fijo primero debes vincularlo a un usuario del sistema desde la gesti√≥n de usuarios.'
                }, status=400)
        
        correo_normalizado = correo.lower() if correo else ''
        dpi_normalizado = dpi.replace(' ', '') if dpi else ''

        # Validar unicidad de correo y DPI
        if correo:
            if Colaborador.objects.filter(correo__iexact=correo).exclude(id=colaborador.id).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe otro colaborador con ese correo electr√≥nico'
                }, status=400)

        if dpi:
            if Colaborador.objects.filter(dpi=dpi).exclude(id=colaborador.id).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe otro colaborador con ese n√∫mero de DPI'
                }, status=400)

        # Actualizar colaborador
        colaborador.nombre = nombre
        colaborador.puesto = puesto
        colaborador.descripcion = descripcion if descripcion else None
        colaborador.telefono = telefono if telefono else None
        colaborador.correo = correo if correo else None
        colaborador.dpi = dpi if dpi else None
        colaborador.es_personal_fijo = es_personal_fijo
        colaborador.activo = activo
        colaborador.save()
        
        # Si el colaborador tiene usuario vinculado, tambi√©n actualizar campos comunes
        if colaborador.usuario:
            usuario = colaborador.usuario
            
            # Sincronizar campos comunes: nombre, email, tel√©fono
            # Solo actualizar si el valor no est√° vac√≠o
            if nombre:
                usuario.nombre = nombre
            if correo:
                usuario.email = correo
            if telefono:
                usuario.telefono = telefono
            
            # Actualizar puesto del usuario solo si el rol es 'personal'
            # (Los administradores no pueden tener puesto seg√∫n la restricci√≥n)
            if puesto and usuario.rol == 'personal':
                usuario.puesto = puesto
            
            usuario.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Colaborador actualizado exitosamente',
            'colaborador': {
                'id': str(colaborador.id),
                'nombre': colaborador.nombre,
                'puesto_nombre': colaborador.puesto.nombre if colaborador.puesto else None,
            }
        })
        
    except Colaborador.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Colaborador no encontrado'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inv√°lidos'
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar colaborador: {str(e)}'
        }, status=500)


@solo_administrador
@require_http_methods(["POST"])
def api_eliminar_colaborador(request, colaborador_id):
    """API: Eliminar (desactivar) un colaborador"""
    try:
        colaborador = Colaborador.objects.get(id=colaborador_id)
        
        # Verificar si tiene usuario vinculado
        if colaborador.usuario:
            return JsonResponse({
                'success': False,
                'error': 'No se puede eliminar un colaborador que tiene usuario asignado. Primero elimine el usuario.'
            }, status=400)
        
        # Desactivar colaborador en lugar de eliminarlo
        colaborador.activo = False
        colaborador.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Colaborador eliminado exitosamente'
        })
        
    except Colaborador.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Colaborador no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar colaborador: {str(e)}'
        }, status=500)
def api_dashboard_stats(request):
    """API para obtener estad√≠sticas del dashboard ejecutivo"""
    try:
        from datetime import datetime, timedelta
        from django.utils import timezone as tz
        
        # Obtener fecha actual y mes actual
        ahora = tz.now()
        inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Filtro base para actividades no eliminadas
        actividades_base = Actividad.objects.filter(eliminado_en__isnull=True)
        
        # Total de actividades
        total_actividades = actividades_base.count()
        
        # Comunidades alcanzadas (distintas)
        comunidades_alcanzadas = actividades_base.values('comunidad_id').distinct().count()
        
        # Total de comunidades del municipio (activas)
        total_comunidades_municipio = Comunidad.objects.filter(activo=True).count()
        
        # Trabajadores activos (solo colaboradores activos)
        trabajadores_activos = Colaborador.objects.filter(activo=True).count()
        
        # Beneficiarios alcanzados (distintos)
        beneficiarios_alcanzados = ActividadBeneficiario.objects.values('beneficiario_id').distinct().count()
        
        # Actividades completadas este mes
        actividades_completadas_mes = actividades_base.filter(
            estado='completado',
            fecha__gte=inicio_mes
        ).count()
        
        # Actividades pendientes (planificadas o en progreso)
        actividades_pendientes = actividades_base.filter(
            estado__in=['planificado', 'en_progreso']
        ).count()
        
        # Actividades por mes (√∫ltimos 12 meses)
        actividades_por_mes = actividades_base.filter(
            fecha__gte=ahora - timedelta(days=365)
        ).annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            total=Count('id')
        ).order_by('mes')
        
        actividades_por_mes_list = []
        for item in actividades_por_mes:
            if item['mes']:
                fecha = item['mes']
                actividades_por_mes_list.append({
                    'mes': fecha.strftime('%Y-%m'),
                    'total': item['total']
                })
        
        # Distribuci√≥n por tipo
        distribucion_por_tipo = {}
        tipos_actividad = TipoActividad.objects.filter(activo=True)
        for tipo in tipos_actividad:
            count = actividades_base.filter(tipo_id=tipo.id).count()
            if count > 0:
                distribucion_por_tipo[tipo.nombre] = count
        
        # Actividades por regi√≥n (Top 5)
        actividades_por_region = actividades_base.select_related('comunidad__region').values(
            'comunidad__region__nombre'
        ).annotate(
            total=Count('id')
        ).order_by('-total')[:5]
        
        actividades_por_region_list = [
            {
                'region': item['comunidad__region__nombre'] or 'Sin regi√≥n',
                'total': item['total']
            }
            for item in actividades_por_region
        ]
        
        # Estado de actividades
        estado_actividades = {}
        estados = ['planificado', 'en_progreso', 'completado', 'cancelado']
        for estado in estados:
            estado_actividades[estado] = actividades_base.filter(estado=estado).count()
        
        # Top 5 comunidades m√°s activas
        top_comunidades = actividades_base.select_related('comunidad__region').values(
            'comunidad__nombre',
            'comunidad__region__nombre'
        ).annotate(
            total_actividades=Count('id'),
            ultima_actividad=Max('fecha')
        ).order_by('-total_actividades')[:5]
        
        top_comunidades_list = [
            {
                'comunidad': item['comunidad__nombre'] or 'Sin comunidad',
                'region': item['comunidad__region__nombre'] or 'Sin regi√≥n',
                'total_actividades': item['total_actividades'],
                'ultima_actividad': item['ultima_actividad'].strftime('%Y-%m-%d') if item['ultima_actividad'] else '-'
            }
            for item in top_comunidades
        ]
        
        # Top 5 responsables m√°s productivos
        top_responsables = actividades_base.select_related('responsable__puesto').filter(
            responsable__isnull=False
        ).values(
            'responsable__username',
            'responsable__puesto__nombre'
        ).annotate(
            completadas=Count('id', filter=Q(estado='completado')),
            en_progreso=Count('id', filter=Q(estado='en_progreso'))
        ).order_by('-completadas')[:5]
        
        top_responsables_list = [
            {
                'responsable': item['responsable__username'] or 'Sin responsable',
                'puesto': item['responsable__puesto__nombre'] or '-',
                'completadas': item['completadas'],
                'en_progreso': item['en_progreso']
            }
            for item in top_responsables
        ]
        
        # Actividades trabajadas recientemente (√∫ltimos 7 d√≠as con cambios/avances)
        # Obtener actividades que han tenido cambios recientes (avances)
        fecha_limite = ahora - timedelta(days=7)
        
        # Actividades con cambios de usuarios recientes
        actividades_con_cambios_usuarios = ActividadCambio.objects.filter(
            fecha_cambio__gte=fecha_limite
        ).values_list('actividad_id', flat=True).distinct()
        
        # Actividades con cambios de colaboradores recientes
        actividades_con_cambios_colaboradores = EventoCambioColaborador.objects.filter(
            fecha_cambio__gte=fecha_limite
        ).values_list('actividad_id', flat=True).distinct()
        
        # Combinar ambas listas
        actividad_ids_recientes = set(list(actividades_con_cambios_usuarios) + list(actividades_con_cambios_colaboradores))
        
        # Obtener las actividades con sus datos m√°s recientes
        actividades_recientes = actividades_base.filter(
            id__in=actividad_ids_recientes
        ).select_related('comunidad').distinct()
        
        # Ordenar por la fecha del cambio m√°s reciente
        actividades_recientes_list = []
        for act in actividades_recientes[:10]:
            # Obtener la fecha del cambio m√°s reciente
            ultimo_cambio_usuario = ActividadCambio.objects.filter(
                actividad_id=act.id
            ).order_by('-fecha_cambio').first()
            
            ultimo_cambio_colaborador = EventoCambioColaborador.objects.filter(
                actividad_id=act.id
            ).order_by('-fecha_cambio').first()
            
            fecha_cambio_reciente = None
            if ultimo_cambio_usuario and ultimo_cambio_colaborador:
                fecha_cambio_reciente = max(ultimo_cambio_usuario.fecha_cambio, ultimo_cambio_colaborador.fecha_cambio)
            elif ultimo_cambio_usuario:
                fecha_cambio_reciente = ultimo_cambio_usuario.fecha_cambio
            elif ultimo_cambio_colaborador:
                fecha_cambio_reciente = ultimo_cambio_colaborador.fecha_cambio
            
            actividades_recientes_list.append({
                'nombre': act.nombre,
                'fecha': act.fecha.strftime('%Y-%m-%d') if act.fecha else '-',
                'comunidad': act.comunidad.nombre if act.comunidad else 'Sin comunidad',
                'estado': act.estado,
                'fecha_ultimo_cambio': fecha_cambio_reciente.strftime('%Y-%m-%d %H:%M') if fecha_cambio_reciente else '-'
            })
        
        # Ordenar por fecha de √∫ltimo cambio (m√°s reciente primero)
        actividades_recientes_list.sort(key=lambda x: x['fecha_ultimo_cambio'], reverse=True)
        
        # Limitar a 10
        actividades_recientes_list = actividades_recientes_list[:10]
        
        return JsonResponse({
            'total_actividades': total_actividades,
            'comunidades_alcanzadas': comunidades_alcanzadas,
            'total_comunidades_municipio': total_comunidades_municipio,
            'trabajadores_activos': trabajadores_activos,
            'beneficiarios_alcanzados': beneficiarios_alcanzados,
            'actividades_completadas_mes': actividades_completadas_mes,
            'actividades_pendientes': actividades_pendientes,
            'actividades_por_mes': actividades_por_mes_list,
            'distribucion_por_tipo': distribucion_por_tipo,
            'actividades_por_region': actividades_por_region_list,
            'estado_actividades': estado_actividades,
            'top_comunidades': top_comunidades_list,
            'top_responsables': top_responsables_list,
            'actividades_trabajadas_recientemente': actividades_recientes_list
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': f'Error al obtener estad√≠sticas: {str(e)}'
        }, status=500)
@solo_administrador
def api_generar_reporte(request, report_type):
    """API para generar reporte seg√∫n tipo"""
    try:
        from datetime import datetime
        
        # Obtener filtros de la request
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        region_ids = request.GET.get('region', '').split(',') if request.GET.get('region') else []
        comunidad_ids = request.GET.get('comunidad', '').split(',') if request.GET.get('comunidad') else []
        estados = request.GET.get('estado', '').split(',') if request.GET.get('estado') else []
        # Obtener par√°metro de tipo de actividad - Si no se env√≠a, debe ser None o cadena vac√≠a
        tipo_actividad_param = request.GET.get('tipo_actividad', None)
        # Si el par√°metro no existe o est√° vac√≠o, tipo_actividad debe ser lista vac√≠a
        if tipo_actividad_param is None or tipo_actividad_param == '' or tipo_actividad_param.strip() == '':
            tipo_actividad = []
        else:
            # Procesar el par√°metro y filtrar valores vac√≠os
            tipo_actividad = [t.strip() for t in tipo_actividad_param.split(',') if t and t.strip() and t.strip() != '']
        responsable_id = request.GET.get('responsable')
        tipo_beneficiario_param = request.GET.get('tipo_beneficiario', '')
        tipo_beneficiario = [t.strip() for t in tipo_beneficiario_param.split(',') if t.strip()] if tipo_beneficiario_param else []
        
        # Nuevos filtros para reportes unificados
        agrupar_por = request.GET.get('agrupar_por', 'region')  # 'region' o 'comunidad'
        comunidades_filtro_param = request.GET.get('comunidades', '')
        comunidades_filtro = [c.strip() for c in comunidades_filtro_param.split(',') if c.strip()] if comunidades_filtro_param else []
        evento_filtro_param = request.GET.get('evento', '')
        evento_filtro = [e.strip() for e in evento_filtro_param.split(',') if e.strip()] if evento_filtro_param else []
        
        # Limpiar listas vac√≠as
        region_ids = [r.strip() for r in region_ids if r.strip()]
        comunidad_ids = [c.strip() for c in comunidad_ids if c.strip()]
        estados = [e.strip() for e in estados if e.strip()]
        
        # Construir query base - TODAS las actividades no eliminadas
        actividades_query = Actividad.objects.filter(eliminado_en__isnull=True)
        
        # Aplicar filtros b√°sicos de fecha
        # Si no se proporcionan fechas (per√≠odo "todo_el_tiempo"), NO aplicar filtros de fecha
        # Esto permitir√° mostrar todas las actividades, incluyendo futuras
        if fecha_inicio and fecha_inicio.strip():
            actividades_query = actividades_query.filter(fecha__gte=fecha_inicio)
        if fecha_fin and fecha_fin.strip():
            actividades_query = actividades_query.filter(fecha__lte=fecha_fin)
        
        # NOTA: Los filtros de region_ids y comunidad_ids se aplican SOLO para reportes gen√©ricos
        # Para el reporte 'actividades-por-region-comunidad', estos filtros NO se aplican aqu√≠
        # porque se manejan m√°s adelante considerando relaciones M2M
        if report_type != 'actividades-por-region-comunidad':
            if region_ids:
                actividades_query = actividades_query.filter(comunidad__region_id__in=region_ids)
            if comunidad_ids:
                actividades_query = actividades_query.filter(comunidad_id__in=comunidad_ids)
        
        if comunidades_filtro:
            # Este filtro se aplica m√°s adelante en la l√≥gica espec√≠fica del reporte
            pass
        
        if estados:
            actividades_query = actividades_query.filter(estado__in=estados)
        
        # IMPORTANTE: Solo filtrar por tipo de actividad si hay tipos seleccionados v√°lidos
        # Si no hay filtro o est√° vac√≠o, mostrar TODOS los tipos (Capacitaci√≥n, Entrega, Proyecto de Ayuda)
        # VERIFICACI√ìN EXPL√çCITA: Asegurarse de que tipo_actividad tenga valores v√°lidos antes de filtrar
        if tipo_actividad and isinstance(tipo_actividad, list) and len(tipo_actividad) > 0:
            # Verificar que todos los valores sean strings no vac√≠os
            tipos_validos = [t for t in tipo_actividad if isinstance(t, str) and t.strip() != '']
            if tipos_validos and len(tipos_validos) > 0:
                # Filtrar por los tipos seleccionados
                tipos = TipoActividad.objects.filter(nombre__in=tipos_validos, activo=True)
                if tipos.exists():
                    tipo_ids = [t.id for t in tipos]
                    actividades_query = actividades_query.filter(tipo_id__in=tipo_ids)
        # Si tipo_actividad est√° vac√≠o o no tiene valores v√°lidos, NO aplicar filtro (mostrar TODOS los tipos)
        
        if responsable_id:
            actividades_query = actividades_query.filter(responsable_id=responsable_id)
        if evento_filtro and len(evento_filtro) > 0:
            actividades_query = actividades_query.filter(id__in=evento_filtro)
        
        # Generar reporte seg√∫n tipo
        data = []
        
        # NUEVO: Actividades por Regi√≥n o Comunidad (unificado)
        if report_type == 'actividades-por-region-comunidad':
            # IMPORTANTE: Para este reporte, NO aplicar filtros de region_ids/comunidad_ids aqu√≠
            # porque se manejan m√°s adelante considerando relaciones M2M
            # Asegurarse de que actividades_query tenga TODAS las actividades (excepto eliminadas)
            # que cumplan los filtros de fecha, estado, tipo (si hay selecci√≥n), etc.
            
            if agrupar_por == 'comunidad':
                # Agrupar por comunidad
                # Obtener TODAS las actividades que cumplan los filtros (fecha, estado, tipo, etc.)
                # Sin filtro de eliminadas (ya est√° en actividades_query)
                
                # Obtener todas las actividades que cumplen los filtros
                actividades_filtradas = actividades_query
                
                # Si hay filtro de comunidades, filtrar actividades que tengan esa comunidad (directa o M2M)
                if comunidades_filtro and comunidades_filtro[0]:
                    # Actividades con comunidad directa en el filtro
                    actividad_ids_directas = set(actividades_filtradas.filter(
                        comunidad_id__in=comunidades_filtro
                    ).values_list('id', flat=True))
                    
                    # Actividades con comunidades M2M en el filtro
                    actividad_ids_m2m = set(ActividadComunidad.objects.filter(
                        comunidad_id__in=comunidades_filtro,
                        actividad_id__in=actividades_filtradas.values_list('id', flat=True)
                    ).values_list('actividad_id', flat=True).distinct())
                    
                    # Combinar IDs (una actividad puede estar en ambas)
                    actividad_ids_combinadas = actividad_ids_directas | actividad_ids_m2m
                    actividades_filtradas = actividades_filtradas.filter(id__in=actividad_ids_combinadas)
                
                # Obtener IDs de todas las actividades filtradas
                actividad_ids_filtradas = list(actividades_filtradas.values_list('id', flat=True))
                
                # Obtener todas las comunidades √∫nicas que tienen actividades
                # 1. Comunidades de actividades con comunidad directa
                comunidades_directas = actividades_filtradas.filter(
                    comunidad_id__isnull=False
                ).values(
                    'comunidad__id',
                    'comunidad__nombre',
                    'comunidad__region__nombre'
                ).distinct()
                
                # 2. Comunidades de actividades M2M
                comunidades_m2m = ActividadComunidad.objects.filter(
                    actividad_id__in=actividad_ids_filtradas
                ).select_related('comunidad', 'comunidad__region').values(
                    'comunidad__id',
                    'comunidad__nombre',
                    'comunidad__region__nombre'
                ).distinct()
                
                # Combinar y obtener comunidades √∫nicas
                comunidades_dict = {}
                for com in comunidades_directas:
                    com_id = com['comunidad__id']
                    if com_id:
                        comunidades_dict[com_id] = {
                            'id': com_id,
                            'nombre': com['comunidad__nombre'],
                            'region': com['comunidad__region__nombre'] or 'Sin regi√≥n'
                        }
                
                for com in comunidades_m2m:
                    com_id = com['comunidad__id']
                    if com_id:
                        comunidades_dict[com_id] = {
                            'id': com_id,
                            'nombre': com['comunidad__nombre'],
                            'region': com['comunidad__region__nombre'] or 'Sin regi√≥n'
                        }
                
                # PRIMERO: Recolectar TODAS las actividades √∫nicas con todas sus comunidades
                # Esto evita que una actividad aparezca m√∫ltiples veces en los detalles
                todas_actividades_unicas = {}
                todas_actividades_ids = set(actividad_ids_filtradas)
                
                for act in actividades_filtradas.filter(id__in=todas_actividades_ids).select_related(
                    'comunidad', 'comunidad__region', 'responsable', 'colaborador', 'tipo'
                ).distinct():
                    if act.id in todas_actividades_unicas:
                        continue
                    
                    # Obtener todas las comunidades de esta actividad (directa + M2M)
                    comunidades_actividad = []
                    if act.comunidad:
                        comunidades_actividad.append(act.comunidad.nombre)
                    comunidades_m2m_act = ActividadComunidad.objects.filter(
                        actividad_id=act.id
                    ).select_related('comunidad').values_list('comunidad__nombre', flat=True)
                    comunidades_actividad.extend(comunidades_m2m_act)
                    comunidades_actividad = sorted(list(set(comunidades_actividad)))  # Eliminar duplicados y ordenar
                    
                    ben_count = ActividadBeneficiario.objects.filter(actividad_id=act.id).count()
                    
                    todas_actividades_unicas[act.id] = {
                        'nombre': act.nombre,
                        'fecha': act.fecha.strftime('%Y-%m-%d') if act.fecha else '-',
                        'estado': act.estado,
                        'tipo_actividad': act.tipo.nombre if act.tipo else '-',
                        'comunidad': ', '.join(comunidades_actividad) if comunidades_actividad else '-',
                        'responsable': act.responsable.username if act.responsable else '-',
                        'colaborador': act.colaborador.nombre if act.colaborador else '-',
                        'total_beneficiarios': ben_count
                    }
                
                # Convertir a lista ordenada por fecha (m√°s reciente primero)
                todas_actividades_lista = list(todas_actividades_unicas.values())
                todas_actividades_lista.sort(key=lambda x: x['fecha'], reverse=True)
                
                # SEGUNDO: Procesar cada comunidad para los totales/resumen
                for comunidad_info in comunidades_dict.values():
                    comunidad_id = comunidad_info['id']
                    
                    # Obtener actividades para esta comunidad (directa O M2M) - solo para c√°lculos
                    actividades_directas_com = actividades_filtradas.filter(comunidad_id=comunidad_id)
                    actividades_m2m_ids_com = ActividadComunidad.objects.filter(
                        actividad_id__in=actividad_ids_filtradas,
                        comunidad_id=comunidad_id
                    ).values_list('actividad_id', flat=True).distinct()
                    
                    # Combinar actividades (una actividad puede estar en ambas)
                    actividades_comunidad_ids = set(
                        list(actividades_directas_com.values_list('id', flat=True)) +
                        list(actividades_m2m_ids_com)
                    )
                    
                    # Obtener beneficiarios SOLO de las actividades de esta comunidad
                    beneficiarios_query = ActividadBeneficiario.objects.filter(
                        actividad_id__in=actividades_comunidad_ids
                    ).select_related('beneficiario__tipo', 'beneficiario__comunidad')
                    
                    total_benef = beneficiarios_query.values('beneficiario_id').distinct().count()
                    benef_ind = beneficiarios_query.filter(beneficiario__tipo__nombre='individual').values('beneficiario_id').distinct().count()
                    benef_fam = beneficiarios_query.filter(beneficiario__tipo__nombre='familia').values('beneficiario_id').distinct().count()
                    benef_inst = beneficiarios_query.filter(beneficiario__tipo__nombre='instituci√≥n').values('beneficiario_id').distinct().count()
                    
                    # CALCULAR BENEFICIARIOS EXCLUSIVOS: Solo beneficiarios individuales cuya comunidad_id coincide con esta comunidad
                    beneficiarios_exclusivos = beneficiarios_query.filter(
                        beneficiario__tipo__nombre='individual',
                        beneficiario__comunidad_id=comunidad_id
                    ).values('beneficiario_id').distinct().count()
                    
                    # Obtener responsables y colaboradores √∫nicos de esta comunidad
                    actividades_comunidad = actividades_filtradas.filter(id__in=actividades_comunidad_ids)
                    responsables = actividades_comunidad.filter(responsable__isnull=False).values_list(
                        'responsable__username', flat=True
                    ).distinct()
                    
                    colaboradores = actividades_comunidad.filter(colaborador__isnull=False).values_list(
                        'colaborador__nombre', flat=True
                    ).distinct()
                    
                    # Agregar datos de la comunidad (sin actividades en detalle, se agregar√°n al final)
                    data.append({
                        'nombre': comunidad_info['nombre'],
                        'region': comunidad_info['region'],
                        'total_actividades': len(actividades_comunidad_ids),
                        'total_beneficiarios': total_benef,
                        'beneficiarios_individuales': benef_ind,
                        'beneficiarios_familias': benef_fam,
                        'beneficiarios_instituciones': benef_inst,
                        'beneficiarios_exclusivos': beneficiarios_exclusivos,
                        'responsables': ', '.join(responsables) if responsables else '-',
                        'colaboradores': ', '.join(colaboradores) if colaboradores else '-',
                        'actividades': []  # Las actividades se mostrar√°n una sola vez al final
                    })
                
                # Agregar TODAS las actividades √∫nicas UNA SOLA VEZ
                # Crear una entrada consolidada con todas las actividades √∫nicas
                # Esta entrada se mostrar√° despu√©s de todas las comunidades
                if todas_actividades_lista:
                    # Calcular totales consolidados de todas las actividades
                    todas_actividades_ids_unicos = set(actividad_ids_filtradas)
                    todos_beneficiarios_query = ActividadBeneficiario.objects.filter(
                        actividad_id__in=todas_actividades_ids_unicos
                    ).select_related('beneficiario__tipo', 'beneficiario__comunidad')
                    
                    total_benef_todos = todos_beneficiarios_query.values('beneficiario_id').distinct().count()
                    benef_ind_todos = todos_beneficiarios_query.filter(beneficiario__tipo__nombre='individual').values('beneficiario_id').distinct().count()
                    benef_fam_todos = todos_beneficiarios_query.filter(beneficiario__tipo__nombre='familia').values('beneficiario_id').distinct().count()
                    benef_inst_todos = todos_beneficiarios_query.filter(beneficiario__tipo__nombre='instituci√≥n').values('beneficiario_id').distinct().count()
                    
                    # Beneficiarios exclusivos para "Todas las Actividades" = suma de todos los exclusivos √∫nicos de cada comunidad
                    # Obtener todas las comunidades √∫nicas que tienen actividades
                    todas_comunidades_ids = set()
                    if comunidades_dict:
                        for com_info in comunidades_dict.values():
                            todas_comunidades_ids.add(com_info['id'])
                    
                    # Calcular beneficiarios exclusivos: solo beneficiarios individuales que tienen comunidad_id
                    # en alguna de las comunidades que tienen actividades
                    if todas_comunidades_ids:
                        beneficiarios_exclusivos_todos = todos_beneficiarios_query.filter(
                            beneficiario__tipo__nombre='individual',
                            beneficiario__comunidad_id__in=todas_comunidades_ids
                        ).values('beneficiario_id').distinct().count()
                    else:
                        beneficiarios_exclusivos_todos = 0
                    
                    todas_actividades_objs = actividades_filtradas.filter(id__in=todas_actividades_ids_unicos)
                    # ARREGLAR: Eliminar duplicados de responsables usando set y luego convertir a lista ordenada
                    todos_responsables_set = set(todas_actividades_objs.filter(
                        responsable__isnull=False
                    ).values_list('responsable__username', flat=True).distinct())
                    todos_responsables_list = sorted(list(todos_responsables_set))
                    
                    todos_colaboradores_set = set(todas_actividades_objs.filter(
                        colaborador__isnull=False
                    ).values_list('colaborador__nombre', flat=True).distinct())
                    todos_colaboradores_list = sorted(list(todos_colaboradores_set))
                    
                    # Agregar entrada consolidada con TODAS las actividades √∫nicas
                    data.append({
                        'nombre': 'Todas las Actividades',
                        'region': 'Todas las Regiones',
                        'total_actividades': len(todas_actividades_lista),
                        'total_beneficiarios': total_benef_todos,
                        'beneficiarios_individuales': benef_ind_todos,
                        'beneficiarios_familias': benef_fam_todos,
                        'beneficiarios_instituciones': benef_inst_todos,
                        'beneficiarios_exclusivos': beneficiarios_exclusivos_todos,
                        'responsables': ', '.join(todos_responsables_list) if todos_responsables_list else '-',
                        'colaboradores': ', '.join(todos_colaboradores_list) if todos_colaboradores_list else '-',
                        'actividades': todas_actividades_lista  # Todas las actividades √∫nicas, sin duplicados
                    })
                
                # Procesar actividades sin comunidad (solo si NO hay filtro de comunidades)
                if not (comunidades_filtro and comunidades_filtro[0]):
                    # Actividades sin comunidad directa Y sin comunidades M2M
                    actividades_con_comunidad_ids = set(
                        list(actividades_filtradas.filter(comunidad_id__isnull=False).values_list('id', flat=True)) +
                        list(ActividadComunidad.objects.filter(
                            actividad_id__in=actividad_ids_filtradas
                        ).values_list('actividad_id', flat=True).distinct())
                    )
                    actividades_sin_comunidad_ids = set(actividad_ids_filtradas) - actividades_con_comunidad_ids
                    
                    if actividades_sin_comunidad_ids:
                        actividades_sin_comunidad = actividades_filtradas.filter(id__in=actividades_sin_comunidad_ids)
                        
                        beneficiarios_sin_comunidad = ActividadBeneficiario.objects.filter(
                            actividad_id__in=actividades_sin_comunidad.values_list('id', flat=True)
                        ).select_related('beneficiario__tipo')
                        
                        total_benef_sc = beneficiarios_sin_comunidad.values('beneficiario_id').distinct().count()
                        benef_ind_sc = beneficiarios_sin_comunidad.filter(beneficiario__tipo__nombre='individual').values('beneficiario_id').distinct().count()
                        benef_fam_sc = beneficiarios_sin_comunidad.filter(beneficiario__tipo__nombre='familia').values('beneficiario_id').distinct().count()
                        benef_inst_sc = beneficiarios_sin_comunidad.filter(beneficiario__tipo__nombre='instituci√≥n').values('beneficiario_id').distinct().count()
                        
                        responsables_sc = actividades_sin_comunidad.filter(responsable__isnull=False).values_list(
                            'responsable__username', flat=True
                        ).distinct()
                        
                        colaboradores_sc = actividades_sin_comunidad.filter(colaborador__isnull=False).values_list(
                            'colaborador__nombre', flat=True
                        ).distinct()
                        
                        # Beneficiarios exclusivos para actividades sin comunidad = 0 (no tienen comunidad asignada)
                        beneficiarios_exclusivos_sc = 0
                        
                        # Las actividades sin comunidad ya est√°n incluidas en todas_actividades_lista
                        # Solo agregar entrada de resumen para "Sin comunidad asignada" (sin actividades en detalle)
                        # Las actividades detalladas ya est√°n en la secci√≥n "Todas las Actividades"
                        data.append({
                            'nombre': 'Sin comunidad asignada',
                            'region': 'Sin regi√≥n',
                            'total_actividades': len(actividades_sin_comunidad_ids),
                            'total_beneficiarios': total_benef_sc,
                            'beneficiarios_individuales': benef_ind_sc,
                            'beneficiarios_familias': benef_fam_sc,
                            'beneficiarios_instituciones': benef_inst_sc,
                            'beneficiarios_exclusivos': beneficiarios_exclusivos_sc,
                            'responsables': ', '.join(responsables_sc) if responsables_sc else '-',
                            'colaboradores': ', '.join(colaboradores_sc) if colaboradores_sc else '-',
                            'actividades': []  # Vac√≠o porque ya est√°n en "Todas las Actividades"
                        })
            else:
                # Agrupar por regi√≥n
                # Obtener TODAS las actividades que cumplen los filtros
                actividades_filtradas = actividades_query
                
                # Si hay filtro de comunidades, filtrar actividades que tengan esa comunidad (directa o M2M)
                if comunidades_filtro and comunidades_filtro[0]:
                    # Actividades con comunidad directa en el filtro
                    actividad_ids_directas = set(actividades_filtradas.filter(
                        comunidad_id__in=comunidades_filtro
                    ).values_list('id', flat=True))
                    
                    # Actividades con comunidades M2M en el filtro
                    actividad_ids_m2m = set(ActividadComunidad.objects.filter(
                        comunidad_id__in=comunidades_filtro,
                        actividad_id__in=actividades_filtradas.values_list('id', flat=True)
                    ).values_list('actividad_id', flat=True).distinct())
                    
                    # Combinar IDs
                    actividad_ids_combinadas = actividad_ids_directas | actividad_ids_m2m
                    actividades_filtradas = actividades_filtradas.filter(id__in=actividad_ids_combinadas)
                
                # Obtener IDs de todas las actividades filtradas
                actividad_ids_filtradas = list(actividades_filtradas.values_list('id', flat=True))
                
                # Obtener todas las regiones √∫nicas que tienen actividades
                # 1. Regiones de actividades con comunidad directa
                regiones_directas = actividades_filtradas.filter(
                    comunidad__region_id__isnull=False
                ).values(
                    'comunidad__region__id',
                    'comunidad__region__nombre'
                ).distinct()
                
                # 2. Regiones de actividades M2M (a trav√©s de actividad_comunidades)
                regiones_m2m = ActividadComunidad.objects.filter(
                    actividad_id__in=actividad_ids_filtradas,
                    region_id__isnull=False
                ).select_related('region').values(
                    'region__id',
                    'region__nombre'
                ).distinct()
                
                # 3. Regiones de comunidades relacionadas en M2M (si no tienen region_id en actividad_comunidades)
                comunidades_m2m = ActividadComunidad.objects.filter(
                    actividad_id__in=actividad_ids_filtradas
                ).select_related('comunidad', 'comunidad__region').filter(
                    comunidad__region_id__isnull=False
                ).values(
                    'comunidad__region__id',
                    'comunidad__region__nombre'
                ).distinct()
                
                # Combinar y obtener regiones √∫nicas
                regiones_dict = {}
                for reg in regiones_directas:
                    reg_id = reg['comunidad__region__id']
                    if reg_id:
                        regiones_dict[reg_id] = {
                            'id': reg_id,
                            'nombre': reg['comunidad__region__nombre']
                        }
                
                for reg in regiones_m2m:
                    reg_id = reg['region__id']
                    if reg_id:
                        regiones_dict[reg_id] = {
                            'id': reg_id,
                            'nombre': reg['region__nombre']
                        }
                
                for reg in comunidades_m2m:
                    reg_id = reg['comunidad__region__id']
                    if reg_id:
                        regiones_dict[reg_id] = {
                            'id': reg_id,
                            'nombre': reg['comunidad__region__nombre']
                        }
                
                # PRIMERO: Recolectar TODAS las actividades √∫nicas con todas sus comunidades
                # Esto evita que una actividad aparezca m√∫ltiples veces en los detalles
                todas_actividades_unicas_reg = {}
                todas_actividades_ids_reg = set(actividad_ids_filtradas)
                
                for act in actividades_filtradas.filter(id__in=todas_actividades_ids_reg).select_related(
                    'comunidad', 'comunidad__region', 'responsable', 'colaborador', 'tipo'
                ).distinct():
                    if act.id in todas_actividades_unicas_reg:
                        continue
                    
                    # Obtener todas las comunidades de esta actividad (directa + M2M)
                    comunidades_actividad = []
                    if act.comunidad:
                        comunidades_actividad.append(act.comunidad.nombre)
                    comunidades_m2m_act = ActividadComunidad.objects.filter(
                        actividad_id=act.id
                    ).select_related('comunidad').values_list('comunidad__nombre', flat=True)
                    comunidades_actividad.extend(comunidades_m2m_act)
                    comunidades_actividad = sorted(list(set(comunidades_actividad)))  # Eliminar duplicados y ordenar
                    
                    ben_count = ActividadBeneficiario.objects.filter(actividad_id=act.id).count()
                    
                    todas_actividades_unicas_reg[act.id] = {
                        'nombre': act.nombre,
                        'fecha': act.fecha.strftime('%Y-%m-%d') if act.fecha else '-',
                        'estado': act.estado,
                        'tipo_actividad': act.tipo.nombre if act.tipo else '-',
                        'comunidad': ', '.join(comunidades_actividad) if comunidades_actividad else '-',
                        'responsable': act.responsable.username if act.responsable else '-',
                        'colaborador': act.colaborador.nombre if act.colaborador else '-',
                        'total_beneficiarios': ben_count
                    }
                
                # Convertir a lista ordenada por fecha (m√°s reciente primero)
                todas_actividades_lista_reg = list(todas_actividades_unicas_reg.values())
                todas_actividades_lista_reg.sort(key=lambda x: x['fecha'], reverse=True)
                
                # SEGUNDO: Procesar cada regi√≥n para los totales/resumen
                for region_info in regiones_dict.values():
                    region_id = region_info['id']
                    
                    # Obtener actividades para esta regi√≥n (directa O M2M) - solo para c√°lculos
                    actividades_directas_reg = actividades_filtradas.filter(comunidad__region_id=region_id)
                    actividad_ids_directas_reg = set(actividades_directas_reg.values_list('id', flat=True))
                    
                    # Actividades M2M en esta regi√≥n
                    actividad_ids_m2m_reg = set(ActividadComunidad.objects.filter(
                        actividad_id__in=actividad_ids_filtradas
                    ).filter(
                        Q(region_id=region_id) | Q(comunidad__region_id=region_id)
                    ).values_list('actividad_id', flat=True).distinct())
                    
                    # Combinar actividades
                    actividad_ids_region = actividad_ids_directas_reg | actividad_ids_m2m_reg
                    actividades_region = actividades_filtradas.filter(id__in=actividad_ids_region)
                    
                    # Obtener beneficiarios SOLO de las actividades de esta regi√≥n
                    beneficiarios_query = ActividadBeneficiario.objects.filter(
                        actividad_id__in=actividad_ids_region
                    ).select_related('beneficiario__tipo', 'beneficiario__comunidad', 'beneficiario__comunidad__region')
                    
                    total_benef = beneficiarios_query.values('beneficiario_id').distinct().count()
                    benef_ind = beneficiarios_query.filter(beneficiario__tipo__nombre='individual').values('beneficiario_id').distinct().count()
                    benef_fam = beneficiarios_query.filter(beneficiario__tipo__nombre='familia').values('beneficiario_id').distinct().count()
                    benef_inst = beneficiarios_query.filter(beneficiario__tipo__nombre='instituci√≥n').values('beneficiario_id').distinct().count()
                    
                    # CALCULAR BENEFICIARIOS EXCLUSIVOS: Solo beneficiarios individuales cuya comunidad.region_id coincide con esta regi√≥n
                    beneficiarios_exclusivos = beneficiarios_query.filter(
                        beneficiario__tipo__nombre='individual',
                        beneficiario__comunidad__region_id=region_id
                    ).values('beneficiario_id').distinct().count()
                    
                    # Obtener responsables y colaboradores √∫nicos de esta regi√≥n
                    responsables = actividades_region.filter(responsable__isnull=False).values_list(
                        'responsable__username', flat=True
                    ).distinct()
                    
                    colaboradores = actividades_region.filter(colaborador__isnull=False).values_list(
                        'colaborador__nombre', flat=True
                    ).distinct()
                    
                    # Agregar datos de la regi√≥n (sin actividades en detalle, se agregar√°n al final)
                    data.append({
                        'nombre': region_info['nombre'],
                        'region': region_info['nombre'],
                        'total_actividades': len(actividad_ids_region),
                        'total_beneficiarios': total_benef,
                        'beneficiarios_individuales': benef_ind,
                        'beneficiarios_familias': benef_fam,
                        'beneficiarios_instituciones': benef_inst,
                        'beneficiarios_exclusivos': beneficiarios_exclusivos,
                        'responsables': ', '.join(responsables) if responsables else '-',
                        'colaboradores': ', '.join(colaboradores) if colaboradores else '-',
                        'actividades': []  # Las actividades se mostrar√°n una sola vez al final
                    })
                
                # Agregar TODAS las actividades √∫nicas UNA SOLA VEZ
                if todas_actividades_lista_reg:
                    # Calcular totales consolidados de todas las actividades
                    todas_actividades_ids_unicos_reg = set(actividad_ids_filtradas)
                    todos_beneficiarios_query_reg = ActividadBeneficiario.objects.filter(
                        actividad_id__in=todas_actividades_ids_unicos_reg
                    ).select_related('beneficiario__tipo', 'beneficiario__comunidad', 'beneficiario__comunidad__region')
                    
                    total_benef_todos_reg = todos_beneficiarios_query_reg.values('beneficiario_id').distinct().count()
                    benef_ind_todos_reg = todos_beneficiarios_query_reg.filter(beneficiario__tipo__nombre='individual').values('beneficiario_id').distinct().count()
                    benef_fam_todos_reg = todos_beneficiarios_query_reg.filter(beneficiario__tipo__nombre='familia').values('beneficiario_id').distinct().count()
                    benef_inst_todos_reg = todos_beneficiarios_query_reg.filter(beneficiario__tipo__nombre='instituci√≥n').values('beneficiario_id').distinct().count()
                    
                    # Beneficiarios exclusivos para "Todas las Actividades" = beneficiarios individuales que tienen
                    # comunidad en alguna de las regiones que tienen actividades
                    todas_regiones_ids = set()
                    if regiones_dict:
                        for reg_info in regiones_dict.values():
                            todas_regiones_ids.add(reg_info['id'])
                    
                    if todas_regiones_ids:
                        beneficiarios_exclusivos_todos_reg = todos_beneficiarios_query_reg.filter(
                            beneficiario__tipo__nombre='individual',
                            beneficiario__comunidad__region_id__in=todas_regiones_ids
                        ).values('beneficiario_id').distinct().count()
                    else:
                        beneficiarios_exclusivos_todos_reg = 0
                    
                    todas_actividades_objs_reg = actividades_filtradas.filter(id__in=todas_actividades_ids_unicos_reg)
                    # ARREGLAR: Eliminar duplicados de responsables usando set y ordenar
                    todos_responsables_reg_set = set(todas_actividades_objs_reg.filter(
                        responsable__isnull=False
                    ).values_list('responsable__username', flat=True).distinct())
                    todos_responsables_reg_list = sorted(list(todos_responsables_reg_set))
                    
                    todos_colaboradores_reg_set = set(todas_actividades_objs_reg.filter(
                        colaborador__isnull=False
                    ).values_list('colaborador__nombre', flat=True).distinct())
                    todos_colaboradores_reg_list = sorted(list(todos_colaboradores_reg_set))
                    
                    # Agregar entrada consolidada con TODAS las actividades √∫nicas
                    data.append({
                        'nombre': 'Todas las Actividades',
                        'region': 'Todas las Regiones',
                        'total_actividades': len(todas_actividades_lista_reg),
                        'total_beneficiarios': total_benef_todos_reg,
                        'beneficiarios_individuales': benef_ind_todos_reg,
                        'beneficiarios_familias': benef_fam_todos_reg,
                        'beneficiarios_instituciones': benef_inst_todos_reg,
                        'beneficiarios_exclusivos': beneficiarios_exclusivos_todos_reg,
                        'responsables': ', '.join(todos_responsables_reg_list) if todos_responsables_reg_list else '-',
                        'colaboradores': ', '.join(todos_colaboradores_reg_list) if todos_colaboradores_reg_list else '-',
                        'actividades': todas_actividades_lista_reg  # Todas las actividades √∫nicas, sin duplicados
                    })
                
                # Procesar actividades sin regi√≥n (solo si NO hay filtro de comunidades)
                if not (comunidades_filtro and comunidades_filtro[0]):
                    # Actividades sin regi√≥n directa Y sin regiones M2M
                    actividades_con_region_ids = set(
                        list(actividades_filtradas.filter(comunidad__region_id__isnull=False).values_list('id', flat=True)) +
                        list(ActividadComunidad.objects.filter(
                            actividad_id__in=actividad_ids_filtradas
                        ).filter(
                            Q(region_id__isnull=False) | Q(comunidad__region_id__isnull=False)
                        ).values_list('actividad_id', flat=True).distinct())
                    )
                    actividades_sin_region_ids = set(actividad_ids_filtradas) - actividades_con_region_ids
                    
                    if actividades_sin_region_ids:
                        actividades_sin_region = actividades_filtradas.filter(id__in=actividades_sin_region_ids)
                        
                        beneficiarios_sin_region = ActividadBeneficiario.objects.filter(
                            actividad_id__in=actividades_sin_region.values_list('id', flat=True)
                        ).select_related('beneficiario__tipo')
                        
                        total_benef_sr = beneficiarios_sin_region.values('beneficiario_id').distinct().count()
                        benef_ind_sr = beneficiarios_sin_region.filter(beneficiario__tipo__nombre='individual').values('beneficiario_id').distinct().count()
                        benef_fam_sr = beneficiarios_sin_region.filter(beneficiario__tipo__nombre='familia').values('beneficiario_id').distinct().count()
                        benef_inst_sr = beneficiarios_sin_region.filter(beneficiario__tipo__nombre='instituci√≥n').values('beneficiario_id').distinct().count()
                        
                        responsables_sr = actividades_sin_region.filter(responsable__isnull=False).values_list(
                            'responsable__username', flat=True
                        ).distinct()
                        
                        colaboradores_sr = actividades_sin_region.filter(colaborador__isnull=False).values_list(
                            'colaborador__nombre', flat=True
                        ).distinct()
                        
                        # Beneficiarios exclusivos para actividades sin regi√≥n = 0 (no tienen regi√≥n asignada)
                        beneficiarios_exclusivos_sr = 0
                        
                        # Las actividades sin regi√≥n ya est√°n incluidas en todas_actividades_lista_reg
                        # Solo agregar entrada de resumen para "Sin regi√≥n asignada" (sin actividades en detalle)
                        # Las actividades detalladas ya est√°n en la secci√≥n "Todas las Actividades"
                        data.append({
                            'nombre': 'Sin regi√≥n asignada',
                            'region': 'Sin regi√≥n',
                            'total_actividades': len(actividades_sin_region_ids),
                            'total_beneficiarios': total_benef_sr,
                            'beneficiarios_individuales': benef_ind_sr,
                            'beneficiarios_familias': benef_fam_sr,
                            'beneficiarios_instituciones': benef_inst_sr,
                            'beneficiarios_exclusivos': beneficiarios_exclusivos_sr,
                            'responsables': ', '.join(responsables_sr) if responsables_sr else '-',
                            'colaboradores': ', '.join(colaboradores_sr) if colaboradores_sr else '-',
                            'actividades': []  # Vac√≠o porque ya est√°n en "Todas las Actividades"
                        })
        
        # NUEVO: Beneficiarios por Regi√≥n o Comunidad (unificado)
        elif report_type == 'beneficiarios-por-region-comunidad':
            # IMPORTANTE: Para este reporte, NO filtrar por fechas (buscar globalmente)
            # Construir query base de actividades SIN filtros de fecha
            actividades_base = Actividad.objects.filter(eliminado_en__isnull=True)
            
            # NO aplicar filtro de estados para este reporte (mostrar todos los estados)
            
            # IMPORTANTE: Solo filtrar por tipo de actividad si hay tipos seleccionados v√°lidos
            if tipo_actividad and isinstance(tipo_actividad, list) and len(tipo_actividad) > 0:
                tipos_validos = [t for t in tipo_actividad if isinstance(t, str) and t.strip() != '']
                if tipos_validos and len(tipos_validos) > 0:
                    tipos = TipoActividad.objects.filter(nombre__in=tipos_validos, activo=True)
                    if tipos.exists():
                        tipo_ids = [t.id for t in tipos]
                        actividades_base = actividades_base.filter(tipo_id__in=tipo_ids)
            
            # Aplicar filtro de eventos si est√° presente
            if evento_filtro and len(evento_filtro) > 0:
                actividades_base = actividades_base.filter(id__in=evento_filtro)
            
            # Obtener todas las actividades que cumplen los filtros (sin fechas)
            actividad_ids = list(actividades_base.values_list('id', flat=True))
            
            # Si hay filtro de comunidades, filtrar actividades que tengan esa comunidad (directa o M2M)
            if comunidades_filtro and comunidades_filtro[0]:
                # Actividades con comunidad directa en el filtro
                actividad_ids_directas = set(actividades_base.filter(
                    comunidad_id__in=comunidades_filtro
                ).values_list('id', flat=True))
                
                # Actividades con comunidades M2M en el filtro
                actividad_ids_m2m = set(ActividadComunidad.objects.filter(
                    comunidad_id__in=comunidades_filtro,
                    actividad_id__in=actividad_ids
                ).values_list('actividad_id', flat=True).distinct())
                
                # Combinar IDs (una actividad puede estar en ambas)
                actividad_ids = list(actividad_ids_directas | actividad_ids_m2m)
            
            # Obtener beneficiarios de las actividades filtradas
            beneficiarios_query = ActividadBeneficiario.objects.filter(
                actividad_id__in=actividad_ids
            ).select_related('beneficiario', 'beneficiario__tipo', 'beneficiario__comunidad', 
                            'beneficiario__comunidad__region', 'actividad')
            
            # Aplicar filtro de comunidad del beneficiario si est√° presente
            if comunidades_filtro and comunidades_filtro[0]:
                beneficiarios_query = beneficiarios_query.filter(
                    beneficiario__comunidad_id__in=comunidades_filtro
                )
            
            # Aplicar filtro de tipo de beneficiario si est√° presente
            if tipo_beneficiario and len(tipo_beneficiario) > 0:
                # Obtener todos los tipos de beneficiario disponibles en la BD con sus nombres e IDs
                tipos_bd = TipoBeneficiario.objects.all()
                tipos_bd_dict = {tipo.nombre.lower(): tipo.id for tipo in tipos_bd}
                
                # Mapear valores del frontend a IDs de tipos en la base de datos
                # La base de datos tiene: 'individual', 'familia', 'instituci√≥n' (con tilde)
                # El frontend puede enviar: 'individual', 'familia', 'instituci√≥n', 'otro'
                tipo_ids_validos = []
                for t in tipo_beneficiario:
                    t_clean = t.strip()
                    t_lower = t_clean.lower()
                    
                    # Buscar el tipo en la BD usando el nombre (case-insensitive)
                    tipo_id = tipos_bd_dict.get(t_lower)
                    
                    # Si no se encontr√≥, intentar mapeo manual
                    if not tipo_id:
                        # Mapear variantes comunes
                        if t_lower == 'individual':
                            tipo_id = tipos_bd_dict.get('individual')
                        elif t_lower == 'familia':
                            tipo_id = tipos_bd_dict.get('familia')
                        elif t_lower in ['institucion', 'instituci√≥n']:
                            # Buscar 'instituci√≥n' (con tilde) en la BD
                            tipo_id = tipos_bd_dict.get('instituci√≥n')
                        # Ignorar "otro" ya que no existe en la BD
                    
                    # Agregar el ID si se encontr√≥
                    if tipo_id:
                        tipo_ids_validos.append(tipo_id)
                
                # Eliminar duplicados
                tipo_ids_validos = list(set(tipo_ids_validos))
                
                if tipo_ids_validos:
                    # Filtrar beneficiarios por los IDs de tipos v√°lidos
                    beneficiarios_query = beneficiarios_query.filter(
                        beneficiario__tipo_id__in=tipo_ids_validos
                    )
            
            # Agrupar beneficiarios por comunidad
            # Primero, obtener todos los beneficiarios √∫nicos con sus datos
            beneficiarios_dict = {}
            
            for ab in beneficiarios_query:
                benef = ab.beneficiario
                benef_id = str(benef.id)
                comunidad_id = str(benef.comunidad.id) if benef.comunidad else 'sin_comunidad'
                
                if benef_id not in beneficiarios_dict:
                    # Obtener datos del beneficiario seg√∫n tipo
                    # Usar hasattr directamente para detectar el tipo espec√≠fico
                    nombre = ''
                    dpi = ''
                    telefono = ''
                    email = ''
                    tipo_nombre = benef.tipo.nombre if benef.tipo else ''
                    
                    # Obtener datos seg√∫n el tipo espec√≠fico del beneficiario
                    if hasattr(benef, 'individual'):
                        ind = benef.individual
                        nombre = f"{ind.nombre} {ind.apellido}".strip()
                        dpi = ind.dpi or ''
                        telefono = ind.telefono or ''
                    elif hasattr(benef, 'familia'):
                        fam = benef.familia
                        nombre = fam.nombre_familia
                        dpi = fam.dpi_jefe_familia or ''
                        telefono = fam.telefono or ''
                    elif hasattr(benef, 'institucion'):
                        inst = benef.institucion
                        nombre = inst.nombre_institucion
                        dpi = inst.dpi_representante or ''
                        telefono = inst.telefono or ''
                        email = inst.email or ''
                    else:
                        # Si no tiene ning√∫n tipo espec√≠fico, usar valores por defecto
                        nombre = 'Sin nombre'
                        dpi = ''
                        telefono = ''
                        email = ''
                    
                    # Formatear tipo para mostrar (usar get_nombre_display si est√° disponible)
                    tipo_display = '-'
                    if benef.tipo:
                        try:
                            tipo_display = benef.tipo.get_nombre_display()
                        except:
                            # Si no tiene get_nombre_display, capitalizar manualmente
                            if tipo_nombre == 'individual':
                                tipo_display = 'Individual'
                            elif tipo_nombre == 'familia':
                                tipo_display = 'Familia'
                            elif tipo_nombre == 'instituci√≥n':
                                tipo_display = 'Instituci√≥n'
                            else:
                                tipo_display = tipo_nombre.capitalize() if tipo_nombre else '-'
                    
                    beneficiarios_dict[benef_id] = {
                        'nombre': nombre or 'Sin nombre',
                        'tipo': tipo_display,
                        'comunidad_id': comunidad_id,
                        'comunidad': benef.comunidad.nombre if benef.comunidad else 'Sin comunidad',
                        'region': benef.comunidad.region.nombre if benef.comunidad and benef.comunidad.region else 'Sin regi√≥n',
                        'dpi': dpi,
                        'documento': dpi,
                        'telefono': telefono,
                        'email': email,
                        'eventos': []
                    }
                
                # Agregar evento si existe
                if ab.actividad:
                    evento_nombre = ab.actividad.nombre
                    if evento_nombre and evento_nombre not in beneficiarios_dict[benef_id]['eventos']:
                        beneficiarios_dict[benef_id]['eventos'].append(evento_nombre)
            
            # Agrupar beneficiarios por comunidad
            comunidades_dict = {}
            for benef_data in beneficiarios_dict.values():
                comunidad_key = benef_data['comunidad_id']
                if comunidad_key not in comunidades_dict:
                    comunidades_dict[comunidad_key] = {
                        'comunidad': benef_data['comunidad'],
                        'region': benef_data['region'],
                        'beneficiarios': []
                    }
                
                # Formatear eventos
                benef_data['evento'] = ', '.join(benef_data['eventos']) if benef_data['eventos'] else '-'
                del benef_data['eventos']  # Remover la lista temporal
                del benef_data['comunidad_id']  # Remover el ID temporal
                comunidades_dict[comunidad_key]['beneficiarios'].append(benef_data)
            
            # Convertir a formato de lista agrupada por comunidad
            data = []
            for comunidad_data in comunidades_dict.values():
                # Agregar informaci√≥n de la comunidad y sus beneficiarios
                for beneficiario in comunidad_data['beneficiarios']:
                    data.append({
                        'comunidad': comunidad_data['comunidad'],
                        'region': comunidad_data['region'],
                        'nombre': beneficiario['nombre'],
                        'tipo': beneficiario['tipo'],
                        'dpi': beneficiario['dpi'],
                        'documento': beneficiario['documento'],
                        'telefono': beneficiario['telefono'],
                        'email': beneficiario['email'],
                        'evento': beneficiario['evento']
                    })
        
        elif report_type == 'actividades-por-region':
            resultado = actividades_query.select_related('comunidad__region').values(
                'comunidad__region__nombre'
            ).annotate(
                total=Count('id')
            ).order_by('-total')
            
            data = [
                {
                    'region': item['comunidad__region__nombre'] or 'Sin regi√≥n',
                    'total_actividades': item['total']
                }
                for item in resultado
            ]
        
        elif report_type == 'actividades-por-comunidad':
            resultado = actividades_query.select_related('comunidad', 'comunidad__region').values(
                'comunidad__nombre',
                'comunidad__region__nombre'
            ).annotate(
                total=Count('id')
            ).order_by('-total')
            
            data = [
                {
                    'comunidad': item['comunidad__nombre'] or 'Sin comunidad',
                    'region': item['comunidad__region__nombre'] or 'Sin regi√≥n',
                    'total_actividades': item['total']
                }
                for item in resultado
            ]
        elif report_type == 'actividad-de-personal':
            # Obtener filtros espec√≠ficos
            colaboradores_filtro = request.GET.get('colaboradores', '').split(',') if request.GET.get('colaboradores') else []
            colaboradores_filtro = [c for c in colaboradores_filtro if c]  # Limpiar valores vac√≠os
            eventos_filtro = request.GET.get('eventos', '').split(',') if request.GET.get('eventos') else []
            eventos_filtro = [e for e in eventos_filtro if e]  # Limpiar valores vac√≠os
            
            # Obtener per√≠odo para calcular fechas correctamente
            periodo = request.GET.get('periodo', 'todo')
            from django.utils import timezone
            from datetime import timedelta, datetime, time
            
            # Calcular fechas seg√∫n per√≠odo (para DateTimeField, usar datetime completo)
            if periodo == 'ultimo_mes':
                fecha_inicio_dt = timezone.now() - timedelta(days=30)
                fecha_fin_dt = timezone.now()  # Incluye todo el d√≠a actual
            elif periodo == 'ultima_semana':
                fecha_inicio_dt = timezone.now() - timedelta(days=7)
                fecha_fin_dt = timezone.now()  # Incluye todo el d√≠a actual
            elif periodo == 'rango' and fecha_inicio and fecha_fin:
                try:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                    # Para fecha_fin, usar el final del d√≠a (23:59:59.999999)
                    fecha_fin_dt = datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d').date(), time.max)
                    # Convertir a timezone-aware si es necesario
                    if timezone.is_naive(fecha_inicio_dt):
                        fecha_inicio_dt = timezone.make_aware(fecha_inicio_dt)
                    if timezone.is_naive(fecha_fin_dt):
                        fecha_fin_dt = timezone.make_aware(fecha_fin_dt)
                except:
                    fecha_inicio_dt = None
                    fecha_fin_dt = None
            else:
                fecha_inicio_dt = None
                fecha_fin_dt = None
            
            # Si hay filtro de colaboradores, obtener esos colaboradores primero
            colaboradores_seleccionados = {}
            if colaboradores_filtro:
                colaboradores_objs = Colaborador.objects.filter(
                    id__in=colaboradores_filtro,
                    activo=True
                ).select_related('puesto')
                
                for col in colaboradores_objs:
                    col_id = str(col.id)
                    colaboradores_seleccionados[col_id] = {
                        'colaborador_id': col_id,
                        'nombre': col.nombre,
                        'puesto': col.puesto.nombre if col.puesto else '-',
                        'telefono': col.telefono or '-',
                        'total_eventos': set(),
                        'total_avances': 0,
                        'eventos': {},
                        'tiene_avances': False
                    }
            
            # Construir query base para cambios de colaboradores
            cambios_query = EventoCambioColaborador.objects.filter(
                colaborador__isnull=False,
                colaborador__activo=True
            ).select_related(
                'colaborador', 'colaborador__puesto', 'actividad', 
                'actividad__comunidad', 'actividad__comunidad__region', 'actividad__tipo'
            )
            
            # Aplicar filtros de fecha (usar datetime para DateTimeField)
            if fecha_inicio_dt:
                cambios_query = cambios_query.filter(fecha_cambio__gte=fecha_inicio_dt)
            if fecha_fin_dt:
                cambios_query = cambios_query.filter(fecha_cambio__lte=fecha_fin_dt)
            # Si no hay filtros de fecha, se muestran todos los cambios (per√≠odo "todo el tiempo")
            
            # Aplicar filtros de actividades
            if comunidades_filtro and comunidades_filtro[0]:
                cambios_query = cambios_query.filter(actividad__comunidad_id__in=comunidades_filtro)
            if eventos_filtro and eventos_filtro[0]:
                cambios_query = cambios_query.filter(actividad_id__in=eventos_filtro)
            if estados and estados[0]:
                cambios_query = cambios_query.filter(actividad__estado__in=estados)
            if tipo_actividad and tipo_actividad[0]:
                tipos = TipoActividad.objects.filter(nombre__in=tipo_actividad, activo=True)
                cambios_query = cambios_query.filter(actividad__tipo_id__in=[t.id for t in tipos])
            if colaboradores_filtro:
                cambios_query = cambios_query.filter(colaborador_id__in=colaboradores_filtro)
            
            # Agrupar por colaborador
            # IMPORTANTE: Si hay filtro de colaboradores, SOLO trabajar con esos colaboradores
            if colaboradores_filtro:
                # Inicializar el diccionario SOLO con los colaboradores filtrados
                colaboradores_dict = colaboradores_seleccionados.copy()
            else:
                # Si NO hay filtro de colaboradores, inicializar vac√≠o
                colaboradores_dict = {}
            
            # Procesar TODOS los cambios una sola vez (la query ya est√° filtrada por colaboradores_filtro si hay filtro)
            for cambio in cambios_query:
                colaborador = cambio.colaborador
                actividad = cambio.actividad
                
                if not colaborador or not actividad:
                    continue
                
                col_id = str(colaborador.id)
                
                # Si hay filtro de colaboradores, verificar que est√© en el filtro
                if colaboradores_filtro and col_id not in colaboradores_filtro:
                    continue
                
                # Si el colaborador no est√° en el diccionario, crearlo (solo si no hay filtro)
                if col_id not in colaboradores_dict:
                    if not colaboradores_filtro:
                        # Solo crear si NO hay filtro (si hay filtro, todos deber√≠an estar ya inicializados)
                        colaboradores_dict[col_id] = {
                            'colaborador_id': col_id,
                            'nombre': colaborador.nombre,
                            'puesto': colaborador.puesto.nombre if colaborador.puesto else '-',
                            'telefono': colaborador.telefono or '-',
                            'total_eventos': set(),
                            'total_avances': 0,
                            'eventos': {},
                            'tiene_avances': False
                        }
                    else:
                        # Si hay filtro y el colaborador no est√° en el diccionario, saltarlo
                        continue
                
                # Marcar que tiene avances
                colaboradores_dict[col_id]['tiene_avances'] = True
                
                # Agregar evento
                actividad_id = str(actividad.id)
                if actividad_id not in colaboradores_dict[col_id]['eventos']:
                    colaboradores_dict[col_id]['eventos'][actividad_id] = {
                        'evento_id': actividad_id,
                        'nombre': actividad.nombre,
                        'estado': actividad.estado,
                        'tipo': actividad.tipo.nombre if actividad.tipo else '-',
                        'comunidad': actividad.comunidad.nombre if actividad.comunidad else '-',
                        'tipo_comunidad': actividad.comunidad.tipo.nombre if actividad.comunidad and actividad.comunidad.tipo else '-',
                        'avances': [],
                        'total_avances': 0,
                        'fecha_primer_avance': None,
                        'fecha_ultimo_avance': None
                    }
                    colaboradores_dict[col_id]['total_eventos'].add(actividad_id)
                
                # Agregar avance
                fecha_cambio = cambio.fecha_cambio
                colaboradores_dict[col_id]['eventos'][actividad_id]['avances'].append({
                    'fecha': fecha_cambio.isoformat() if fecha_cambio else None,
                    'descripcion': cambio.descripcion_cambio
                })
                colaboradores_dict[col_id]['eventos'][actividad_id]['total_avances'] += 1
                colaboradores_dict[col_id]['total_avances'] += 1
                
                # Actualizar fechas
                if fecha_cambio:
                    if not colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_primer_avance']:
                        colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_primer_avance'] = fecha_cambio
                    elif fecha_cambio < colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_primer_avance']:
                        colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_primer_avance'] = fecha_cambio
                    
                    if not colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_ultimo_avance']:
                        colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_ultimo_avance'] = fecha_cambio
                    elif fecha_cambio > colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_ultimo_avance']:
                        colaboradores_dict[col_id]['eventos'][actividad_id]['fecha_ultimo_avance'] = fecha_cambio
            
            # Formatear datos para respuesta
            data = []
            
            # Si hay filtro de colaboradores, asegurarse de que SOLO los colaboradores filtrados est√©n en el resultado
            if colaboradores_filtro:
                # Asegurarse de que todos los colaboradores filtrados est√©n en el diccionario
                # (ya deber√≠an estar porque los inicializamos al principio, pero por seguridad verificamos)
                for col_id in colaboradores_filtro:
                    if col_id not in colaboradores_dict:
                        # Este colaborador fue seleccionado pero no tiene cambios que coincidan con los filtros
                        # Buscar el colaborador en la base de datos
                        try:
                            col = Colaborador.objects.get(id=col_id, activo=True)
                            colaboradores_dict[col_id] = {
                                'colaborador_id': col_id,
                                'nombre': col.nombre,
                                'puesto': col.puesto.nombre if col.puesto else '-',
                                'telefono': col.telefono or '-',
                                'total_eventos': set(),
                                'total_avances': 0,
                                'eventos': {},
                                'tiene_avances': False
                            }
                        except Colaborador.DoesNotExist:
                            # Colaborador no existe, saltarlo
                            continue
                
                # FILTRAR ESTRICTAMENTE: Solo procesar colaboradores que est√°n en el filtro
                colaboradores_dict_filtrado = {}
                for col_id in colaboradores_filtro:
                    if col_id in colaboradores_dict:
                        colaboradores_dict_filtrado[col_id] = colaboradores_dict[col_id]
                
                colaboradores_dict = colaboradores_dict_filtrado
            
            # Procesar solo los colaboradores que est√°n en el diccionario (ya filtrados si hay filtro)
            for col_data in colaboradores_dict.values():
                eventos_list = []
                
                # Si tiene avances, procesar eventos normalmente
                if col_data.get('tiene_avances', False):
                    for evento_data in col_data['eventos'].values():
                        eventos_list.append({
                            'nombre': evento_data['nombre'],
                            'estado': evento_data['estado'],
                            'tipo': evento_data['tipo'],
                            'comunidad': evento_data['comunidad'],
                            'tipo_comunidad': evento_data['tipo_comunidad'],
                            'total_avances': evento_data['total_avances'],
                            'fecha_primer_avance': evento_data['fecha_primer_avance'].strftime('%d/%m/%Y') if evento_data['fecha_primer_avance'] else '-',
                            'fecha_ultimo_avance': evento_data['fecha_ultimo_avance'].strftime('%d/%m/%Y') if evento_data['fecha_ultimo_avance'] else '-'
                        })
                # Si no tiene avances, no agregar eventos a la lista
                
                data.append({
                    'colaborador_id': col_data['colaborador_id'],
                    'nombre': col_data['nombre'],
                    'puesto': col_data['puesto'],
                    'telefono': col_data['telefono'],
                    'total_eventos': len(col_data['total_eventos']),
                    'total_avances': col_data['total_avances'],
                    'eventos': eventos_list,
                    'sin_avances': not col_data.get('tiene_avances', False)
                })
            
            # Ordenar por total de avances descendente (los sin avances van al final)
            data.sort(key=lambda x: (x['total_avances'], x['nombre']), reverse=True)
        elif report_type == 'avances-eventos-generales':
            # Obtener filtros espec√≠ficos
            comunidades_filtro = request.GET.get('comunidades', '').split(',') if request.GET.get('comunidades') else []
            comunidades_filtro = [c for c in comunidades_filtro if c]
            eventos_filtro = request.GET.get('eventos', '').split(',') if request.GET.get('eventos') else []
            eventos_filtro = [e for e in eventos_filtro if e]
            colaboradores_filtro = request.GET.get('colaboradores', '').split(',') if request.GET.get('colaboradores') else []
            colaboradores_filtro = [c for c in colaboradores_filtro if c]
            mostrar_evidencias = request.GET.get('mostrar_evidencias', 'true').lower() == 'true'
            
            # Obtener per√≠odo para calcular fechas correctamente
            periodo = request.GET.get('periodo', 'todo')
            from django.utils import timezone
            from datetime import timedelta, datetime, time
            
            # Calcular fechas seg√∫n per√≠odo (para DateTimeField, usar datetime completo)
            if periodo == 'ultimo_mes':
                fecha_inicio_dt = timezone.now() - timedelta(days=30)
                fecha_fin_dt = timezone.now()  # Incluye todo el d√≠a actual
            elif periodo == 'ultima_semana':
                fecha_inicio_dt = timezone.now() - timedelta(days=7)
                fecha_fin_dt = timezone.now()  # Incluye todo el d√≠a actual
            elif periodo == 'rango' and fecha_inicio and fecha_fin:
                try:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                    # Para fecha_fin, usar el final del d√≠a (23:59:59.999999)
                    fecha_fin_dt = datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d').date(), time.max)
                    # Convertir a timezone-aware si es necesario
                    if timezone.is_naive(fecha_inicio_dt):
                        fecha_inicio_dt = timezone.make_aware(fecha_inicio_dt)
                    if timezone.is_naive(fecha_fin_dt):
                        fecha_fin_dt = timezone.make_aware(fecha_fin_dt)
                except:
                    fecha_inicio_dt = None
                    fecha_fin_dt = None
            else:
                fecha_inicio_dt = None
                fecha_fin_dt = None
            
            # Construir query base para cambios de colaboradores
            cambios_query = EventoCambioColaborador.objects.filter(
                colaborador__isnull=False,
                colaborador__activo=True,
                actividad__eliminado_en__isnull=True
            ).select_related(
                'colaborador', 'colaborador__puesto', 'actividad', 
                'actividad__tipo', 'actividad__comunidad', 'actividad__comunidad__region'
            ).prefetch_related(
                'actividad__comunidades_relacionadas__comunidad',
                'actividad__comunidades_relacionadas__comunidad__region',
                'actividad__comunidades_relacionadas__region'
            )
            
            # Aplicar filtros de fecha (usar datetime para DateTimeField)
            if fecha_inicio_dt:
                cambios_query = cambios_query.filter(fecha_cambio__gte=fecha_inicio_dt)
            if fecha_fin_dt:
                cambios_query = cambios_query.filter(fecha_cambio__lte=fecha_fin_dt)
            
            # Aplicar filtros de actividades
            if comunidades_filtro:
                cambios_query = cambios_query.filter(
                    Q(actividad__comunidad_id__in=comunidades_filtro) |
                    Q(actividad__comunidades_relacionadas__comunidad_id__in=comunidades_filtro)
                ).distinct()
            if eventos_filtro:
                cambios_query = cambios_query.filter(actividad_id__in=eventos_filtro)
            if tipo_actividad and tipo_actividad[0]:
                tipos = TipoActividad.objects.filter(nombre__in=tipo_actividad, activo=True)
                cambios_query = cambios_query.filter(actividad__tipo_id__in=[t.id for t in tipos])
            if colaboradores_filtro:
                cambios_query = cambios_query.filter(colaborador_id__in=colaboradores_filtro)
            
            # Nota: Las evidencias se obtendr√°n manualmente agrupadas por grupo_id
            # No usar prefetch_related aqu√≠ ya que agrupamos por grupo_id
            
            # Procesar cambios agrupados por grupo_id para evitar duplicados
            # Agrupar cambios por grupo_id (un avance puede tener m√∫ltiples colaboradores)
            cambios_por_grupo = {}
            for cambio in cambios_query:
                actividad = cambio.actividad
                colaborador = cambio.colaborador
                
                if not actividad or not colaborador:
                    continue
                
                grupo_id = str(cambio.grupo_id)
                
                # Si es el primer cambio de este grupo, inicializar
                if grupo_id not in cambios_por_grupo:
                    # Obtener comunidades del evento
                    comunidades_nombres = []
                    regiones_nombres = []
                    
                    if actividad.comunidad:
                        comunidades_nombres.append(actividad.comunidad.nombre)
                        if actividad.comunidad.region:
                            regiones_nombres.append(actividad.comunidad.region.nombre)
                    
                    # Tambi√©n obtener de relaciones de comunidades
                    if hasattr(actividad, 'comunidades_relacionadas'):
                        for relacion in actividad.comunidades_relacionadas.all():
                            if relacion.comunidad and relacion.comunidad.nombre not in comunidades_nombres:
                                comunidades_nombres.append(relacion.comunidad.nombre)
                            # Obtener regi√≥n de la relaci√≥n o de la comunidad
                            if relacion.region and relacion.region.nombre not in regiones_nombres:
                                regiones_nombres.append(relacion.region.nombre)
                            elif relacion.comunidad and relacion.comunidad.region and relacion.comunidad.region.nombre not in regiones_nombres:
                                regiones_nombres.append(relacion.comunidad.region.nombre)
                    
                    # Obtener todas las evidencias del grupo (pueden estar en cualquier cambio del grupo)
                    evidencias_data = []
                    if mostrar_evidencias:
                        # Obtener todos los cambios del mismo grupo
                        cambios_grupo_ids = EventoCambioColaborador.objects.filter(
                            grupo_id=cambio.grupo_id
                        ).values_list('id', flat=True)
                        
                        # Obtener todas las evidencias de todos los cambios del mismo grupo
                        evidencias_qs = EventosEvidenciasCambios.objects.filter(
                            cambio_id__in=cambios_grupo_ids
                        ).select_related('cambio')
                        
                        # Usar un diccionario para evitar evidencias duplicadas basado en URL (archivo √∫nico)
                        # La misma evidencia puede estar asociada a m√∫ltiples cambios del mismo grupo
                        evidencias_unicas = {}  # key: url_almacenamiento, value: evidencia_data
                        
                        for evidencia in evidencias_qs:
                            url_almacenamiento = evidencia.url_almacenamiento
                            # Solo agregar si no existe ya una evidencia con la misma URL
                            if url_almacenamiento not in evidencias_unicas:
                                evidencias_unicas[url_almacenamiento] = {
                                    'id': str(evidencia.id),
                                    'nombre': evidencia.archivo_nombre,
                                    'url': url_almacenamiento,
                                    'tipo': evidencia.archivo_tipo or '',
                                    'es_imagen': evidencia.archivo_tipo and evidencia.archivo_tipo.startswith('image/'),
                                    'descripcion': evidencia.descripcion or ''  # Descripci√≥n de la evidencia (no del avance)
                                }
                        
                        # Convertir el diccionario a lista
                        evidencias_data = list(evidencias_unicas.values())
                    
                    # Formatear fecha
                    fecha_display = ''
                    if cambio.fecha_cambio:
                        from django.utils.timezone import localtime
                        import pytz
                        guatemala_tz = pytz.timezone('America/Guatemala')
                        if timezone.is_aware(cambio.fecha_cambio):
                            fecha_local = cambio.fecha_cambio.astimezone(guatemala_tz)
                        else:
                            fecha_local = timezone.make_aware(cambio.fecha_cambio, guatemala_tz)
                        fecha_display = fecha_local.strftime('%d/%m/%Y %H:%M')
                    
                    # Obtener todos los colaboradores del grupo
                    colaboradores_grupo = EventoCambioColaborador.objects.filter(
                        grupo_id=cambio.grupo_id
                    ).select_related('colaborador').values('colaborador__id', 'colaborador__nombre')
                    
                    colaboradores_nombres = [c['colaborador__nombre'] for c in colaboradores_grupo if c['colaborador__nombre']]
                    
                    # Obtener la comunidad espec√≠fica del avance (desde cambio.comunidad)
                    comunidad_avance_nombre = None
                    comunidad_avance_id = None
                    if cambio.comunidad:
                        comunidad_avance_nombre = cambio.comunidad.nombre
                        comunidad_avance_id = str(cambio.comunidad.id)
                    
                    cambios_por_grupo[grupo_id] = {
                        'evento_id': str(actividad.id),
                        'evento': {
                            'id': str(actividad.id),
                            'nombre': actividad.nombre,
                            'estado': actividad.estado,
                            'tipo': actividad.tipo.nombre if actividad.tipo else '-'
                        },
                        'comunidad': ', '.join(comunidades_nombres) if comunidades_nombres else '-',
                        'region': ', '.join(regiones_nombres) if regiones_nombres else None,  # Regi√≥n/ubicaci√≥n del avance
                        'comunidad_avance': comunidad_avance_nombre,  # Comunidad espec√≠fica donde se realiz√≥ el avance
                        'comunidad_avance_id': comunidad_avance_id,  # ID de la comunidad espec√≠fica del avance
                        'colaboradores_ids': [str(c['colaborador__id']) for c in colaboradores_grupo if c['colaborador__id']],
                        'colaboradores_nombres': colaboradores_nombres,
                        'colaborador_nombre': ', '.join(colaboradores_nombres) if colaboradores_nombres else 'Sin nombre',  # Todos los colaboradores
                        'descripcion_cambio': cambio.descripcion_cambio,  # Descripci√≥n del avance (no de la evidencia)
                        'fecha_cambio': cambio.fecha_cambio.isoformat() if cambio.fecha_cambio else None,
                        'fecha_display': fecha_display,
                        'evidencias': evidencias_data
                    }
            
            # Convertir el diccionario agrupado a lista
            data = list(cambios_por_grupo.values())
            
            # El frontend espera data directamente (no en un campo 'data')
            return JsonResponse({
                'data': data,
                'total': len(data)
            })
        
        elif report_type == 'comunidades':
            # Obtener filtros espec√≠ficos para reporte de comunidades
            comunidades_filtro = request.GET.get('comunidades', '').split(',') if request.GET.get('comunidades') else []
            comunidades_filtro = [c.strip() for c in comunidades_filtro if c.strip()]
            
            evento_filtro_param = request.GET.get('evento', '')
            if evento_filtro_param:
                evento_filtro = [e.strip() for e in evento_filtro_param.split(',') if e.strip()]
            else:
                evento_filtro = []
            
            # Obtener per√≠odo para calcular fechas correctamente
            periodo = request.GET.get('periodo', 'todo')
            from django.utils import timezone
            from datetime import timedelta, datetime, time
            
            # Calcular fechas seg√∫n per√≠odo
            if periodo == 'ultimo_mes':
                fecha_inicio_dt = timezone.now() - timedelta(days=30)
                fecha_fin_dt = timezone.now()
            elif periodo == 'ultima_semana':
                fecha_inicio_dt = timezone.now() - timedelta(days=7)
                fecha_fin_dt = timezone.now()
            elif periodo == 'rango' and fecha_inicio and fecha_fin:
                try:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                    fecha_fin_dt = datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d').date(), time.max)
                    if timezone.is_naive(fecha_inicio_dt):
                        fecha_inicio_dt = timezone.make_aware(fecha_inicio_dt)
                    if timezone.is_naive(fecha_fin_dt):
                        fecha_fin_dt = timezone.make_aware(fecha_fin_dt)
                except:
                    fecha_inicio_dt = None
                    fecha_fin_dt = None
            else:
                fecha_inicio_dt = None
                fecha_fin_dt = None
            
            # Construir query base de actividades
            actividades_query = Actividad.objects.filter(eliminado_en__isnull=True)
            
            # Aplicar filtros de fecha
            if fecha_inicio_dt:
                actividades_query = actividades_query.filter(fecha__gte=fecha_inicio_dt.date())
            if fecha_fin_dt:
                actividades_query = actividades_query.filter(fecha__lte=fecha_fin_dt.date())
            
            # Aplicar filtro de tipo de actividad
            if tipo_actividad and isinstance(tipo_actividad, list) and len(tipo_actividad) > 0:
                tipos_validos = [t for t in tipo_actividad if isinstance(t, str) and t.strip() != '']
                if tipos_validos:
                    tipos = TipoActividad.objects.filter(nombre__in=tipos_validos, activo=True)
                    if tipos.exists():
                        tipo_ids = [t.id for t in tipos]
                        actividades_query = actividades_query.filter(tipo_id__in=tipo_ids)
            
            # Aplicar filtro de evento si est√° presente
            if evento_filtro and len(evento_filtro) > 0:
                actividades_query = actividades_query.filter(id__in=evento_filtro)
            
            # Obtener IDs de actividades filtradas
            actividad_ids = list(actividades_query.values_list('id', flat=True))
            
            # Si hay filtro de comunidades, filtrar actividades que tengan esa comunidad (directa o M2M)
            if comunidades_filtro:
                # Actividades con comunidad directa en el filtro
                actividad_ids_directas = set(actividades_query.filter(
                    comunidad_id__in=comunidades_filtro
                ).values_list('id', flat=True))
                
                # Actividades con comunidades M2M en el filtro
                actividad_ids_m2m = set(ActividadComunidad.objects.filter(
                    comunidad_id__in=comunidades_filtro,
                    actividad_id__in=actividad_ids
                ).values_list('actividad_id', flat=True).distinct())
                
                # Combinar IDs
                actividad_ids = list(actividad_ids_directas | actividad_ids_m2m)
            
            # Obtener comunidades a procesar
            if comunidades_filtro:
                comunidades_a_procesar = Comunidad.objects.filter(
                    id__in=comunidades_filtro,
                    activo=True
                ).select_related('tipo', 'region')
            else:
                # Si no hay filtro de comunidades, obtener todas las comunidades que tienen actividades
                comunidades_ids_directas = set(actividades_query.filter(
                    comunidad_id__isnull=False
                ).values_list('comunidad_id', flat=True).distinct())
                
                comunidades_ids_m2m = set(ActividadComunidad.objects.filter(
                    actividad_id__in=actividad_ids
                ).values_list('comunidad_id', flat=True).distinct())
                
                todas_comunidades_ids = list(comunidades_ids_directas | comunidades_ids_m2m)
                comunidades_a_procesar = Comunidad.objects.filter(
                    id__in=todas_comunidades_ids,
                    activo=True
                ).select_related('tipo', 'region')
            
            # Procesar cada comunidad
            comunidades_data = []
            for comunidad in comunidades_a_procesar:
                # Obtener actividades de esta comunidad (directa o M2M)
                actividad_ids_comunidad_directa = set(actividades_query.filter(
                    comunidad_id=comunidad.id
                ).values_list('id', flat=True))
                
                actividad_ids_comunidad_m2m = set(ActividadComunidad.objects.filter(
                    comunidad_id=comunidad.id,
                    actividad_id__in=actividad_ids
                ).values_list('actividad_id', flat=True).distinct())
                
                actividad_ids_comunidad = list(actividad_ids_comunidad_directa | actividad_ids_comunidad_m2m)
                
                # Obtener actividades de esta comunidad
                actividades_comunidad = Actividad.objects.filter(
                    id__in=actividad_ids_comunidad
                ).select_related('tipo', 'comunidad', 'comunidad__region')
                
                # Obtener beneficiarios de esta comunidad que est√°n en estas actividades
                beneficiarios_comunidad = Beneficiario.objects.filter(
                    comunidad_id=comunidad.id,
                    activo=True
                ).select_related('tipo')
                
                # Filtrar beneficiarios que est√°n en las actividades
                beneficiarios_en_actividades = ActividadBeneficiario.objects.filter(
                    actividad_id__in=actividad_ids_comunidad,
                    beneficiario_id__in=[b.id for b in beneficiarios_comunidad]
                ).values_list('beneficiario_id', flat=True).distinct()
                
                beneficiarios_filtrados = beneficiarios_comunidad.filter(
                    id__in=beneficiarios_en_actividades
                )
                
                # Procesar proyectos/eventos
                proyectos_data = []
                for actividad in actividades_comunidad:
                    proyectos_data.append({
                        'nombre': actividad.nombre,
                        'tipo': actividad.tipo.nombre if actividad.tipo else '-',
                        'estado': actividad.estado,
                        'fecha': actividad.fecha.strftime('%Y-%m-%d') if actividad.fecha else '-'
                    })
                
                # Procesar beneficiarios
                beneficiarios_data = []
                for beneficiario in beneficiarios_filtrados:
                    # Obtener eventos en los que participa este beneficiario
                    eventos_beneficiario = ActividadBeneficiario.objects.filter(
                        beneficiario_id=beneficiario.id,
                        actividad_id__in=actividad_ids_comunidad
                    ).select_related('actividad').values_list('actividad__nombre', flat=True).distinct()
                    
                    # Obtener nombre del beneficiario seg√∫n tipo
                    nombre_beneficiario = ''
                    if hasattr(beneficiario, 'individual'):
                        nombre_beneficiario = f"{beneficiario.individual.nombre} {beneficiario.individual.apellido}".strip()
                    elif hasattr(beneficiario, 'familia'):
                        nombre_beneficiario = beneficiario.familia.nombre_familia
                    elif hasattr(beneficiario, 'institucion'):
                        nombre_beneficiario = beneficiario.institucion.nombre_institucion
                    
                    tipo_beneficiario = beneficiario.tipo.nombre if beneficiario.tipo else '-'
                    
                    beneficiarios_data.append({
                        'nombre': nombre_beneficiario or 'Sin nombre',
                        'tipo': tipo_beneficiario.capitalize() if tipo_beneficiario else '-',
                        'eventos': list(eventos_beneficiario)
                    })
                
                comunidades_data.append({
                    'nombre': comunidad.nombre,
                    'cocode': comunidad.cocode or '-',
                    'region': comunidad.region.nombre if comunidad.region else '-',
                    'tipo': comunidad.tipo.nombre if comunidad.tipo else '-',
                    'numero_beneficiarios': len(beneficiarios_data),
                    'numero_proyectos': len(proyectos_data),
                    'proyectos': proyectos_data,
                    'beneficiarios': beneficiarios_data
                })
            
            # El frontend espera un objeto con 'comunidades'
            return JsonResponse({
                'data': {
                    'comunidades': comunidades_data
                },
                'total': len(comunidades_data)
            })
        
        elif report_type == 'reporte-evento-individual':
            # Obtener el evento espec√≠fico
            evento_id = request.GET.get('evento')
            if not evento_id:
                return JsonResponse({
                    'error': 'Debe seleccionar un evento'
                }, status=400)
            
            try:
                # Verificar si las columnas comunidad_id y region_id existen en eventos_cambios_colaboradores
                from django.db import connection
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT column_name 
                            FROM information_schema.columns 
                            WHERE table_name = 'eventos_cambios_colaboradores' 
                            AND column_name IN ('comunidad_id', 'region_id');
                        """)
                        existing_columns = {row[0] for row in cursor.fetchall()}
                        has_cambios_colaboradores_comunidad = 'comunidad_id' in existing_columns
                        has_cambios_colaboradores_region = 'region_id' in existing_columns
                except Exception as e:
                    # Si hay un error al verificar, asumir que las columnas no existen
                    print(f'‚ö†Ô∏è Error al verificar columnas en eventos_cambios_colaboradores: {e}')
                    has_cambios_colaboradores_comunidad = False
                    has_cambios_colaboradores_region = False

                # Construir prefetch_related din√°micamente seg√∫n las columnas disponibles
                prefetch_fields = [
                    'personal__usuario__puesto',
                    'personal__colaborador__puesto',
                    'beneficiarios__beneficiario__individual',
                    'beneficiarios__beneficiario__familia',
                    'beneficiarios__beneficiario__institucion',
                    'beneficiarios__beneficiario__tipo',
                    'evidencias',
                    'archivos',
                    'galeria_imagenes',
                    'cambios__responsable',
                    'comunidades_relacionadas__comunidad__region',
                    'comunidades_relacionadas__region'
                ]
                
                # Solo agregar prefetch de cambios_colaboradores si las columnas existen
                # Si no existen, omitir el prefetch y dejar que obtener_cambios_evento maneje la carga
                if has_cambios_colaboradores_comunidad or has_cambios_colaboradores_region:
                    cambios_colaboradores_queryset = EventoCambioColaborador.objects.select_related('colaborador').prefetch_related('evidencias')
                    if has_cambios_colaboradores_comunidad:
                        cambios_colaboradores_queryset = cambios_colaboradores_queryset.select_related('comunidad')
                    if has_cambios_colaboradores_region:
                        cambios_colaboradores_queryset = cambios_colaboradores_queryset.select_related('region')
                    prefetch_fields.append(
                        Prefetch('cambios_colaboradores', queryset=cambios_colaboradores_queryset)
                    )
                # Si las columnas no existen, no hacer prefetch para evitar errores
                # La funci√≥n obtener_cambios_evento manejar√° la carga de manera segura
                
                evento = Actividad.objects.select_related(
                    'tipo', 'comunidad', 'comunidad__region', 'responsable', 'colaborador', 'portada'
                ).prefetch_related(*prefetch_fields).get(id=evento_id, eliminado_en__isnull=True)
                
                # Personal asignado
                personal_data = []
                for ap in evento.personal.all():
                    if ap.usuario:
                        personal_data.append({
                            'id': str(ap.usuario.id),
                            'username': ap.usuario.username,
                            'nombre': ap.usuario.nombre or ap.usuario.username,
                            'rol': ap.rol_en_actividad,
                            'rol_display': ap.usuario.get_rol_display() if hasattr(ap.usuario, 'get_rol_display') else ap.usuario.rol,
                            'puesto': ap.usuario.puesto.nombre if ap.usuario.puesto else 'Sin puesto',
                            'tipo': 'usuario'
                        })
                    elif ap.colaborador:
                        personal_data.append({
                            'id': str(ap.colaborador.id),
                            'username': ap.colaborador.correo or '',
                            'nombre': ap.colaborador.nombre,
                            'rol': ap.rol_en_actividad,
                            'rol_display': 'Personal Fijo' if ap.colaborador.es_personal_fijo else 'Colaborador Externo',
                            'puesto': ap.colaborador.puesto.nombre if ap.colaborador.puesto else 'Sin puesto',
                            'tipo': 'colaborador'
                        })
                
                # Beneficiarios
                beneficiarios_data = []
                for ab in evento.beneficiarios.all():
                    benef = ab.beneficiario
                    nombre_display, info_adicional, detalles, tipo_envio = obtener_detalle_beneficiario(benef)
                    if benef.tipo and hasattr(benef.tipo, 'get_nombre_display'):
                        tipo_display = benef.tipo.get_nombre_display()
                    else:
                        tipo_display = tipo_envio.title() if tipo_envio else ''
                    beneficiarios_data.append({
                        'id': str(benef.id),
                        'tipo': tipo_envio,
                        'nombre': nombre_display,
                        'info_adicional': info_adicional,
                        'tipo_display': tipo_display,
                        'detalles': detalles
                    })
                
                # Galer√≠a de im√°genes (desde eventos_galeria)
                evidencias_data = []
                for imagen in evento.galeria_imagenes.all():
                    evidencias_data.append({
                        'id': str(imagen.id),
                        'nombre': imagen.archivo_nombre,
                        'url': imagen.url_almacenamiento,
                        'tipo': imagen.archivo_tipo or '',
                        'es_imagen': True,
                        'descripcion': imagen.descripcion or ''
                    })
                
                # Archivos del proyecto (evidencias en /media/evidencias + archivos de actividad_archivos)
                archivos_data = []
                for evidencia in evento.evidencias.all():
                    url = evidencia.url_almacenamiento or ''
                    url_lower = url.lower()
                    if '/media/evidencias/' not in url_lower:
                        continue
                    archivos_data.append({
                        'id': str(evidencia.id),
                        'nombre': evidencia.archivo_nombre,
                        'url': url,
                        'tipo': evidencia.archivo_tipo or 'application/octet-stream',
                        'tamanio': evidencia.archivo_tamanio,
                        'descripcion': evidencia.descripcion or '',
                        'es_evidencia': True,
                        'es_imagen': evidencia.es_imagen,
                        'es_galeria': False,
                        'creado_en': evidencia.creado_en.isoformat() if evidencia.creado_en else None
                    })
                # Archivos de actividad_archivos
                for archivo in evento.archivos.all():
                    archivos_data.append({
                        'id': str(archivo.id),
                        'nombre': archivo.nombre_archivo,
                        'url': archivo.url_almacenamiento,
                        'tipo': archivo.archivo_tipo or 'application/octet-stream',
                        'tamanio': archivo.archivo_tamanio,
                        'descripcion': archivo.descripcion or '',
                        'es_evidencia': False,
                        'es_imagen': False,
                        'es_galeria': False,
                        'creado_en': archivo.creado_en.isoformat() if archivo.creado_en else None
                    })
                
                # Ubicaci√≥n
                ubicacion = 'Sin ubicaci√≥n'
                if evento.comunidad:
                    if evento.comunidad.region:
                        ubicacion = f"{evento.comunidad.nombre}, {evento.comunidad.region.nombre}"
                    else:
                        ubicacion = evento.comunidad.nombre
                
                # Convertir fechas
                try:
                    if evento.fecha:
                        fecha_str = evento.fecha.strftime('%Y-%m-%d')
                        fecha_display = evento.fecha.strftime('%d/%m/%Y')
                    else:
                        fecha_str = ''
                        fecha_display = ''
                except:
                    fecha_str = ''
                    fecha_display = ''
                
                # Obtener cambios (siempre, incluso si est√° vac√≠o)
                cambios_data = obtener_cambios_evento(evento)
                
                # Construir objeto de evento
                evento_data = {
                    'id': str(evento.id),
                    'nombre': evento.nombre or 'Sin nombre',
                    'tipo': evento.tipo.nombre if evento.tipo else 'Sin tipo',
                    'descripcion': evento.descripcion or '',
                    'ubicacion': ubicacion,
                    'comunidad': evento.comunidad.nombre if evento.comunidad else 'Sin comunidad',
                    'region': evento.comunidad.region.nombre if evento.comunidad and evento.comunidad.region else None,
                    'fecha': fecha_str,
                    'fecha_display': fecha_display,
                    'estado': evento.estado,
                    'responsable': (evento.responsable.nombre if evento.responsable and evento.responsable.nombre else evento.responsable.username) if evento.responsable else 'Sin responsable',
                    'personal': personal_data,
                    'beneficiarios': beneficiarios_data,
                    'evidencias': evidencias_data,
                    'archivos': archivos_data,
                    'portada': obtener_portada_evento(evento),
                    'cambios': cambios_data  # Siempre incluir, incluso si est√° vac√≠o
                }
                
                # El frontend espera data.evento (no data.data)
                return JsonResponse({
                    'data': {
                        'evento': evento_data
                    },
                    'total': 1
                })
                
            except Actividad.DoesNotExist:
                return JsonResponse({
                    'error': 'Evento no encontrado'
                }, status=404)
        elif report_type == 'actividad-usuarios':
            # Obtener filtros espec√≠ficos para reporte de actividad de usuarios
            periodo = request.GET.get('periodo', 'todo')
            fecha_inicio_param = request.GET.get('fecha_inicio')
            fecha_fin_param = request.GET.get('fecha_fin')
            
            comunidades_filtro = request.GET.get('comunidades', '').split(',') if request.GET.get('comunidades') else []
            comunidades_filtro = [c.strip() for c in comunidades_filtro if c.strip()]
            
            evento_filtro_param = request.GET.get('evento', '')
            evento_filtro = evento_filtro_param.strip() if evento_filtro_param else None
            
            tipo_actividad_param = request.GET.get('tipo_actividad', '')
            tipo_actividad = [t.strip() for t in tipo_actividad_param.split(',') if t.strip()] if tipo_actividad_param else []
            
            usuarios_filtro = request.GET.get('usuarios', '').split(',') if request.GET.get('usuarios') else []
            usuarios_filtro = [u.strip() for u in usuarios_filtro if u.strip()]
            
            # Calcular fechas seg√∫n per√≠odo
            from django.utils import timezone
            from datetime import timedelta, datetime, time
            
            fecha_inicio_dt = None
            fecha_fin_dt = None
            
            if periodo == 'ultimo_mes':
                fecha_inicio_dt = timezone.now() - timedelta(days=30)
                fecha_fin_dt = timezone.now()
            elif periodo == 'ultima_semana':
                fecha_inicio_dt = timezone.now() - timedelta(days=7)
                fecha_fin_dt = timezone.now()
            elif periodo == 'rango' and fecha_inicio_param and fecha_fin_param:
                try:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio_param, '%Y-%m-%d')
                    fecha_fin_dt = datetime.combine(datetime.strptime(fecha_fin_param, '%Y-%m-%d').date(), time.max)
                    if timezone.is_naive(fecha_inicio_dt):
                        fecha_inicio_dt = timezone.make_aware(fecha_inicio_dt)
                    if timezone.is_naive(fecha_fin_dt):
                        fecha_fin_dt = timezone.make_aware(fecha_fin_dt)
                except:
                    fecha_inicio_dt = None
                    fecha_fin_dt = None
            
            # Obtener actividades base seg√∫n filtros
            actividades_base = Actividad.objects.filter(eliminado_en__isnull=True)
            
            # Aplicar filtro de tipo de actividad
            if tipo_actividad and len(tipo_actividad) > 0:
                tipos = TipoActividad.objects.filter(nombre__in=tipo_actividad, activo=True)
                if tipos.exists():
                    tipo_ids = [t.id for t in tipos]
                    actividades_base = actividades_base.filter(tipo_id__in=tipo_ids)
            
            # Aplicar filtro de evento
            if evento_filtro:
                actividades_base = actividades_base.filter(id=evento_filtro)
            
            # Aplicar filtro de comunidades (directa o M2M)
            if comunidades_filtro:
                actividad_ids_directas = set(actividades_base.filter(
                    comunidad_id__in=comunidades_filtro
                ).values_list('id', flat=True))
                
                actividad_ids_m2m = set(ActividadComunidad.objects.filter(
                    comunidad_id__in=comunidades_filtro,
                    actividad_id__in=actividades_base.values_list('id', flat=True)
                ).values_list('actividad_id', flat=True).distinct())
                
                actividad_ids = list(actividad_ids_directas | actividad_ids_m2m)
                actividades_base = actividades_base.filter(id__in=actividad_ids)
            
            actividad_ids = list(actividades_base.values_list('id', flat=True))
            
            # Obtener todos los usuarios del sistema y colaboradores con usuario
            usuarios_sistema = Usuario.objects.filter(activo=True).select_related('puesto')
            colaboradores_con_usuario = Colaborador.objects.filter(
                activo=True,
                usuario__isnull=False
            ).select_related('usuario', 'puesto')
            
            # Si hay filtro de usuarios, aplicar
            if usuarios_filtro:
                usuarios_sistema = usuarios_sistema.filter(id__in=usuarios_filtro)
                colaboradores_con_usuario = colaboradores_con_usuario.filter(usuario_id__in=usuarios_filtro)
            
            # Obtener cambios de usuarios (ActividadCambio)
            cambios_usuarios = ActividadCambio.objects.filter(
                actividad_id__in=actividad_ids
            ).select_related('responsable', 'actividad', 'actividad__tipo', 'actividad__comunidad')
            
            if fecha_inicio_dt:
                cambios_usuarios = cambios_usuarios.filter(fecha_cambio__gte=fecha_inicio_dt)
            if fecha_fin_dt:
                cambios_usuarios = cambios_usuarios.filter(fecha_cambio__lte=fecha_fin_dt)
            
            if usuarios_filtro:
                cambios_usuarios = cambios_usuarios.filter(responsable_id__in=usuarios_filtro)
            
            # Obtener cambios de colaboradores con usuario (EventoCambioColaborador)
            colaborador_ids_con_usuario = [c.id for c in colaboradores_con_usuario]
            cambios_colaboradores = EventoCambioColaborador.objects.filter(
                actividad_id__in=actividad_ids,
                colaborador_id__in=colaborador_ids_con_usuario
            ).select_related('colaborador', 'colaborador__usuario', 'actividad', 'actividad__tipo', 'actividad__comunidad')
            
            if fecha_inicio_dt:
                cambios_colaboradores = cambios_colaboradores.filter(fecha_cambio__gte=fecha_inicio_dt)
            if fecha_fin_dt:
                cambios_colaboradores = cambios_colaboradores.filter(fecha_cambio__lte=fecha_fin_dt)
            
            # Agrupar cambios por usuario
            usuarios_dict = {}
            
            # Procesar cambios de usuarios
            for cambio in cambios_usuarios:
                usuario = cambio.responsable
                if not usuario:
                    continue
                
                usuario_id = str(usuario.id)
                if usuario_id not in usuarios_dict:
                    usuarios_dict[usuario_id] = {
                        'id': usuario_id,
                        'username': usuario.username,
                        'nombre': usuario.nombre or usuario.username,
                        'rol': usuario.rol,
                        'rol_display': usuario.get_rol_display() if hasattr(usuario, 'get_rol_display') else usuario.rol,
                        'puesto': usuario.puesto.nombre if usuario.puesto else None,
                        'puesto_nombre': usuario.puesto.nombre if usuario.puesto else None,
                        'cambios': [],
                        'total_cambios': 0
                    }
                
                # Formatear fecha
                fecha_display = ''
                if cambio.fecha_cambio:
                    import pytz
                    guatemala_tz = pytz.timezone('America/Guatemala')
                    if timezone.is_aware(cambio.fecha_cambio):
                        fecha_local = cambio.fecha_cambio.astimezone(guatemala_tz)
                    else:
                        fecha_local = timezone.make_aware(cambio.fecha_cambio, guatemala_tz)
                    fecha_display = fecha_local.strftime('%d/%m/%Y %H:%M')
                
                usuarios_dict[usuario_id]['cambios'].append({
                    'tipo_cambio': 'Cambio de Actividad',
                    'descripcion': cambio.descripcion_cambio,
                    'fecha': cambio.fecha_cambio.isoformat() if cambio.fecha_cambio else None,
                    'fecha_display': fecha_display,
                    'evento_id': str(cambio.actividad.id),
                    'evento_nombre': cambio.actividad.nombre
                })
                usuarios_dict[usuario_id]['total_cambios'] += 1
            
            # Procesar cambios de colaboradores con usuario
            # Agrupar por grupo_id para evitar duplicados cuando un cambio tiene m√∫ltiples comunidades
            cambios_por_grupo = {}
            for cambio in cambios_colaboradores:
                colaborador = cambio.colaborador
                if not colaborador or not colaborador.usuario:
                    continue
                
                usuario = colaborador.usuario
                usuario_id = str(usuario.id)
                grupo_id = str(cambio.grupo_id) if hasattr(cambio, 'grupo_id') and cambio.grupo_id else str(cambio.id)
                
                # Crear clave √∫nica por usuario y grupo
                clave_grupo = f"{usuario_id}_{grupo_id}"
                
                # Si ya procesamos este grupo para este usuario, saltarlo
                if clave_grupo not in cambios_por_grupo:
                    cambios_por_grupo[clave_grupo] = {
                        'usuario_id': usuario_id,
                        'usuario': usuario,
                        'colaborador': colaborador,
                        'cambio': cambio
                    }
            
            # Procesar los cambios √∫nicos agrupados
            for clave_grupo, grupo_data in cambios_por_grupo.items():
                usuario = grupo_data['usuario']
                colaborador = grupo_data['colaborador']
                cambio = grupo_data['cambio']
                usuario_id = grupo_data['usuario_id']
                
                if usuario_id not in usuarios_dict:
                    usuarios_dict[usuario_id] = {
                        'id': usuario_id,
                        'username': usuario.username,
                        'nombre': usuario.nombre or usuario.username,
                        'rol': usuario.rol,
                        'rol_display': usuario.get_rol_display() if hasattr(usuario, 'get_rol_display') else usuario.rol,
                        'puesto': colaborador.puesto.nombre if colaborador.puesto else (usuario.puesto.nombre if usuario.puesto else None),
                        'puesto_nombre': colaborador.puesto.nombre if colaborador.puesto else (usuario.puesto.nombre if usuario.puesto else None),
                        'cambios': [],
                        'total_cambios': 0
                    }
                
                # Formatear fecha
                fecha_display = ''
                if cambio.fecha_cambio:
                    import pytz
                    guatemala_tz = pytz.timezone('America/Guatemala')
                    if timezone.is_aware(cambio.fecha_cambio):
                        fecha_local = cambio.fecha_cambio.astimezone(guatemala_tz)
                    else:
                        fecha_local = timezone.make_aware(cambio.fecha_cambio, guatemala_tz)
                    fecha_display = fecha_local.strftime('%d/%m/%Y %H:%M')
                
                usuarios_dict[usuario_id]['cambios'].append({
                    'tipo_cambio': 'Cambio por Colaborador',
                    'descripcion': cambio.descripcion_cambio,
                    'fecha': cambio.fecha_cambio.isoformat() if cambio.fecha_cambio else None,
                    'fecha_display': fecha_display,
                    'evento_id': str(cambio.actividad.id),
                    'evento_nombre': cambio.actividad.nombre
                })
                usuarios_dict[usuario_id]['total_cambios'] += 1
            
            # Si hay filtro de usuarios pero no tienen cambios, agregarlos sin actividad
            if usuarios_filtro:
                for usuario in usuarios_sistema:
                    usuario_id = str(usuario.id)
                    if usuario_id not in usuarios_dict:
                        usuarios_dict[usuario_id] = {
                            'id': usuario_id,
                            'username': usuario.username,
                            'nombre': usuario.nombre or usuario.username,
                            'rol': usuario.rol,
                            'rol_display': usuario.get_rol_display() if hasattr(usuario, 'get_rol_display') else usuario.rol,
                            'puesto': usuario.puesto.nombre if usuario.puesto else None,
                            'puesto_nombre': usuario.puesto.nombre if usuario.puesto else None,
                            'cambios': [],
                            'total_cambios': 0,
                            'sin_actividad': True
                        }
            
            # Convertir a lista y ordenar por total de cambios
            usuarios_lista = list(usuarios_dict.values())
            usuarios_lista.sort(key=lambda x: x['total_cambios'], reverse=True)
            
            return JsonResponse({
                'data': {
                    'usuarios': usuarios_lista
                },
                'total': len(usuarios_lista)
            })
        elif report_type == 'reporte-general':
            # Obtener filtros
            periodo = request.GET.get('periodo', 'todo')
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            apartados = request.GET.get('apartados', '').split(',') if request.GET.get('apartados') else []
            apartados = [a.strip() for a in apartados if a.strip()]
            
            # Si no hay apartados seleccionados, retornar vac√≠o
            if not apartados:
                return JsonResponse({
                    'data': {},
                    'total': 0
                })
            
            # Construir query base para actividades
            actividades_query = Actividad.objects.filter(eliminado_en__isnull=True)
            
            # Filtrar por fecha seg√∫n per√≠odo
            from django.utils import timezone
            from datetime import timedelta
            
            if periodo == 'ultimo_mes':
                fecha_inicio = (timezone.now() - timedelta(days=30)).date()
                fecha_fin = timezone.now().date()
            elif periodo == 'ultima_semana':
                fecha_inicio = (timezone.now() - timedelta(days=7)).date()
                fecha_fin = timezone.now().date()
            elif periodo == 'rango' and fecha_inicio and fecha_fin:
                try:
                    from datetime import datetime
                    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                except:
                    fecha_inicio = None
                    fecha_fin = None
            
            if fecha_inicio:
                actividades_query = actividades_query.filter(fecha__gte=fecha_inicio)
            if fecha_fin:
                actividades_query = actividades_query.filter(fecha__lte=fecha_fin)
            
            # Construir datos del reporte seg√∫n apartados seleccionados
            reporte_data = {}
            
            # Total de Eventos
            if 'total_eventos' in apartados:
                total_eventos = actividades_query.count()
                lista_eventos = []
                # Si se selecciona mostrar lista, obtener eventos
                if 'total_eventos' in apartados:  # Siempre incluir lista si est√° seleccionado
                    eventos_lista = actividades_query.select_related('tipo', 'comunidad', 'comunidad__region')[:100]
                    for evento in eventos_lista:
                        ben_count = ActividadBeneficiario.objects.filter(actividad_id=evento.id).count()
                        lista_eventos.append({
                            'nombre': evento.nombre,
                            'tipo': evento.tipo.nombre if evento.tipo else '-',
                            'estado': evento.get_estado_display() if hasattr(evento, 'get_estado_display') else evento.estado,
                            'fecha': evento.fecha.strftime('%d/%m/%Y') if evento.fecha else '-',
                            'comunidad': evento.comunidad.nombre if evento.comunidad else '-',
                            'beneficiarios': ben_count
                        })
                reporte_data['total_eventos'] = {
                    'total': total_eventos,
                    'lista_eventos': lista_eventos
                }
            
            # Total de Tipos de Eventos
            if 'tipos_eventos' in apartados:
                tipos_query = actividades_query.values('tipo__nombre').annotate(
                    total=Count('id')
                )
                tipos_eventos = {
                    'capacitacion': 0,
                    'entrega': 0,
                    'proyecto_ayuda': 0
                }
                for tipo in tipos_query:
                    nombre = tipo['tipo__nombre'] or ''
                    if 'capacitaci√≥n' in nombre.lower() or 'capacitacion' in nombre.lower():
                        tipos_eventos['capacitacion'] = tipo['total']
                    elif 'entrega' in nombre.lower():
                        tipos_eventos['entrega'] = tipo['total']
                    elif 'proyecto' in nombre.lower() or 'ayuda' in nombre.lower():
                        tipos_eventos['proyecto_ayuda'] = tipo['total']
                reporte_data['tipos_eventos'] = tipos_eventos
            
            # Total de Beneficiarios Alcanzados
            if 'total_beneficiarios' in apartados:
                actividad_ids = actividades_query.values_list('id', flat=True)
                total_benef = ActividadBeneficiario.objects.filter(
                    actividad_id__in=actividad_ids
                ).values('beneficiario_id').distinct().count()
                reporte_data['total_beneficiarios'] = total_benef
            
            # Total de Tipos de Beneficiarios
            if 'tipos_beneficiarios' in apartados:
                actividad_ids = actividades_query.values_list('id', flat=True)
                beneficiarios_query = ActividadBeneficiario.objects.filter(
                    actividad_id__in=actividad_ids
                ).select_related('beneficiario__tipo', 'beneficiario__individual', 'beneficiario__familia')
                
                tipos_benef = {
                    'individual': 0,
                    'familia': 0,
                    'institucion': 0,
                    'mujeres': 0,
                    'hombres': 0
                }
                
                beneficiarios_ids = beneficiarios_query.values_list('beneficiario_id', flat=True).distinct()
                for benef_id in beneficiarios_ids:
                    try:
                        benef = Beneficiario.objects.select_related('tipo', 'individual', 'familia').get(id=benef_id)
                        tipo_nombre = benef.tipo.nombre.lower() if benef.tipo else ''
                        if 'individual' in tipo_nombre:
                            tipos_benef['individual'] += 1
                            # Contar g√©nero si es individual
                            if hasattr(benef, 'individual') and benef.individual:
                                genero = benef.individual.genero.lower() if benef.individual.genero else ''
                                if 'femenino' in genero or 'mujer' in genero:
                                    tipos_benef['mujeres'] += 1
                                elif 'masculino' in genero or 'hombre' in genero:
                                    tipos_benef['hombres'] += 1
                        elif 'familia' in tipo_nombre:
                            tipos_benef['familia'] += 1
                        elif 'instituci√≥n' in tipo_nombre or 'institucion' in tipo_nombre:
                            tipos_benef['institucion'] += 1
                    except Beneficiario.DoesNotExist:
                        continue
                
                reporte_data['tipos_beneficiarios'] = tipos_benef
            
            # Total de Comunidades Alcanzadas
            if 'total_comunidades' in apartados:
                actividad_ids = actividades_query.values_list('id', flat=True)
                comunidades_ids_list = list(actividades_query.values_list('comunidad_id', flat=True).distinct())
                # Tambi√©n incluir comunidades relacionadas
                comunidades_relacionadas_ids_list = list(ActividadComunidad.objects.filter(
                    actividad_id__in=actividad_ids
                ).values_list('comunidad_id', flat=True).distinct())
                
                # Combinar y obtener IDs √∫nicos
                todas_comunidades_ids = list(set([c for c in comunidades_ids_list if c] + [c for c in comunidades_relacionadas_ids_list if c]))
                total_comunidades = len(todas_comunidades_ids)
                
                # Obtener informaci√≥n de las comunidades
                comunidades_lista = []
                for com_id in todas_comunidades_ids:
                    try:
                        comunidad = Comunidad.objects.select_related('region', 'tipo').get(id=com_id)
                        # Contar eventos de esta comunidad
                        eventos_count = actividades_query.filter(
                            Q(comunidad_id=com_id) | 
                            Q(comunidades_relacionadas__comunidad_id=com_id)
                        ).distinct().count()
                        
                        # Contar beneficiarios de esta comunidad
                        beneficiarios_count = Beneficiario.objects.filter(
                            comunidad_id=com_id
                        ).count()
                        
                        comunidades_lista.append({
                            'nombre': comunidad.nombre,
                            'cocode': comunidad.cocode or '-',
                            'region': comunidad.region.nombre if comunidad.region else '-',
                            'tipo': comunidad.tipo.nombre if comunidad.tipo else '-',
                            'total_eventos': eventos_count,
                            'total_beneficiarios': beneficiarios_count
                        })
                    except Comunidad.DoesNotExist:
                        continue
                
                # Ordenar por nombre
                comunidades_lista.sort(key=lambda x: x['nombre'])
                
                reporte_data['total_comunidades'] = {
                    'total': total_comunidades,
                    'lista_comunidades': comunidades_lista
                }
            
            # Total de Avances en Proyectos
            if 'total_avances' in apartados:
                actividad_ids = actividades_query.values_list('id', flat=True)
                total_avances = EventoCambioColaborador.objects.filter(
                    actividad_id__in=actividad_ids
                ).count()
                reporte_data['total_avances'] = total_avances
            
            # Comunidades con M√°s Beneficiarios y Eventos
            if 'comunidades_mas_beneficiarios' in apartados:
                actividad_ids = actividades_query.values_list('id', flat=True)
                comunidades_stats = {}
                
                # Obtener comunidades principales
                for actividad in actividades_query.select_related('comunidad', 'comunidad__region'):
                    if actividad.comunidad:
                        com_id = str(actividad.comunidad.id)
                        if com_id not in comunidades_stats:
                            comunidades_stats[com_id] = {
                                'nombre': actividad.comunidad.nombre,
                                'region': actividad.comunidad.region.nombre if actividad.comunidad.region else '-',
                                'total_eventos': 0,
                                'total_beneficiarios': set()
                            }
                        comunidades_stats[com_id]['total_eventos'] += 1
                
                # Obtener comunidades relacionadas
                actividades_comunidades = ActividadComunidad.objects.filter(
                    actividad_id__in=actividad_ids
                ).select_related('comunidad', 'comunidad__region')
                
                for ac in actividades_comunidades:
                    if ac.comunidad:
                        com_id = str(ac.comunidad.id)
                        if com_id not in comunidades_stats:
                            comunidades_stats[com_id] = {
                                'nombre': ac.comunidad.nombre,
                                'region': ac.comunidad.region.nombre if ac.comunidad.region else '-',
                                'total_eventos': 0,
                                'total_beneficiarios': set()
                            }
                        comunidades_stats[com_id]['total_eventos'] += 1
                
                # Contar beneficiarios por comunidad
                for actividad_id in actividad_ids:
                    ben_ids = ActividadBeneficiario.objects.filter(
                        actividad_id=actividad_id
                    ).values_list('beneficiario_id', flat=True)
                    
                    # Obtener comunidad de cada beneficiario
                    for ben_id in ben_ids:
                        try:
                            ben = Beneficiario.objects.select_related('comunidad').get(id=ben_id)
                            if ben.comunidad:
                                com_id = str(ben.comunidad.id)
                                if com_id in comunidades_stats:
                                    comunidades_stats[com_id]['total_beneficiarios'].add(ben_id)
                        except Beneficiario.DoesNotExist:
                            continue
                
                # Convertir a lista y ordenar
                comunidades_list = []
                for com_id, stats in comunidades_stats.items():
                    comunidades_list.append({
                        'nombre': stats['nombre'],
                        'region': stats['region'],
                        'total_beneficiarios': len(stats['total_beneficiarios']),
                        'total_eventos': stats['total_eventos']
                    })
                
                comunidades_list.sort(key=lambda x: (x['total_beneficiarios'], x['total_eventos']), reverse=True)
                reporte_data['comunidades_mas_beneficiarios'] = comunidades_list[:10]  # Top 10
            
            # Eventos con M√°s Beneficiarios
            if 'eventos_mas_beneficiarios' in apartados:
                eventos_stats = []
                for actividad in actividades_query.select_related('tipo', 'comunidad')[:100]:
                    ben_count = ActividadBeneficiario.objects.filter(actividad_id=actividad.id).count()
                    eventos_stats.append({
                        'nombre': actividad.nombre,
                        'tipo': actividad.tipo.nombre if actividad.tipo else '-',
                        'estado': actividad.get_estado_display() if hasattr(actividad, 'get_estado_display') else actividad.estado,
                        'total_beneficiarios': ben_count,
                        'comunidad': actividad.comunidad.nombre if actividad.comunidad else '-'
                    })
                
                eventos_stats.sort(key=lambda x: x['total_beneficiarios'], reverse=True)
                reporte_data['eventos_mas_beneficiarios'] = eventos_stats[:10]  # Top 10
            
            # Evento con M√°s Avances y Cambios
            if 'evento_mas_avances' in apartados:
                actividad_ids = actividades_query.values_list('id', flat=True)
                eventos_avances = {}
                
                for cambio in EventoCambioColaborador.objects.filter(actividad_id__in=actividad_ids).select_related('actividad', 'actividad__tipo', 'actividad__comunidad'):
                    act_id = str(cambio.actividad.id)
                    if act_id not in eventos_avances:
                        eventos_avances[act_id] = {
                            'nombre': cambio.actividad.nombre,
                            'tipo': cambio.actividad.tipo.nombre if cambio.actividad.tipo else '-',
                            'estado': cambio.actividad.get_estado_display() if hasattr(cambio.actividad, 'get_estado_display') else cambio.actividad.estado,
                            'comunidad': cambio.actividad.comunidad.nombre if cambio.actividad.comunidad else '-',
                            'total_avances': 0
                        }
                    eventos_avances[act_id]['total_avances'] += 1
                
                if eventos_avances:
                    evento_mas_avances = max(eventos_avances.values(), key=lambda x: x['total_avances'])
                    reporte_data['evento_mas_avances'] = evento_mas_avances
                else:
                    reporte_data['evento_mas_avances'] = {}
            
            return JsonResponse({
                'data': reporte_data,
                'total': len(apartados)
            })
        
        else:
            # Reporte gen√©rico: lista de actividades
            actividades = actividades_query.select_related(
                'comunidad', 'comunidad__region', 'tipo', 'responsable'
            )[:100]  # Limitar a 100 resultados
            
            data = [
                {
                    'nombre': act.nombre,
                    'fecha': act.fecha.strftime('%Y-%m-%d') if act.fecha else '-',
                    'estado': act.estado,
                    'tipo': act.tipo.nombre if act.tipo else '-',
                    'comunidad': act.comunidad.nombre if act.comunidad else '-',
                    'region': act.comunidad.region.nombre if act.comunidad and act.comunidad.region else '-',
                    'responsable': act.responsable.username if act.responsable else '-'
                }
                for act in actividades
            ]
        
        return JsonResponse({
            'data': data,
            'total': len(data)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': f'Error al generar reporte: {str(e)}'
        }, status=500)


@permiso_generar_reportes
def api_exportar_reporte(request, report_type):
    """API para exportar reporte en PDF o Word"""
    try:
        from .report_generator import generate_pdf_report, generate_word_report, get_report_title
        
        # Obtener formato solicitado
        formato = request.GET.get('formato', 'pdf').lower()

        if hasattr(request, 'user') and getattr(request.user, 'is_authenticated', False):
            user_label = f"{request.user.get_username()} (id={request.user.id})"
        else:
            user_label = 'usuario_no_autenticado'

        query_params = {}
        for key in request.GET.keys():
            values = request.GET.getlist(key)
            if not values:
                query_params[key] = None
            elif len(values) == 1:
                query_params[key] = values[0]
            else:
                query_params[key] = values

        logger.info(
            "Solicitud de exportaci√≥n de reporte recibida: usuario=%s, tipo=%s, formato=%s, params=%s",
            user_label,
            report_type,
            formato,
            query_params,
        )
        
        if formato not in ['pdf', 'word']:
            logger.warning(
                "Formato de exportaci√≥n no v√°lido: usuario=%s, tipo=%s, formato=%s",
                user_label,
                report_type,
                formato,
            )
            return JsonResponse({
                'error': 'Formato no v√°lido. Use "pdf" o "word"'
            }, status=400)
        
        # Obtener los mismos filtros que usa api_generar_reporte
        # Reutilizar la l√≥gica de generaci√≥n de datos
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        region_ids = request.GET.get('region', '').split(',') if request.GET.get('region') else []
        comunidad_ids = request.GET.get('comunidad', '').split(',') if request.GET.get('comunidad') else []
        estados = request.GET.get('estado', '').split(',') if request.GET.get('estado') else []
        tipo_actividad_param = request.GET.get('tipo_actividad', None)
        if tipo_actividad_param is None or tipo_actividad_param == '' or tipo_actividad_param.strip() == '':
            tipo_actividad = []
        else:
            tipo_actividad = [t.strip() for t in tipo_actividad_param.split(',') if t and t.strip() and t.strip() != '']
        responsable_id = request.GET.get('responsable')
        tipo_beneficiario_param = request.GET.get('tipo_beneficiario', '')
        tipo_beneficiario = [t.strip() for t in tipo_beneficiario_param.split(',') if t.strip()] if tipo_beneficiario_param else []
        
        # Nuevos filtros para reportes unificados
        agrupar_por = request.GET.get('agrupar_por', 'region')
        comunidades_filtro_param = request.GET.get('comunidades', '')
        comunidades_filtro = [c.strip() for c in comunidades_filtro_param.split(',') if c.strip()] if comunidades_filtro_param else []
        evento_filtro_param = request.GET.get('evento', '')
        evento_filtro = [e.strip() for e in evento_filtro_param.split(',') if e.strip()] if evento_filtro_param else []
        
        # Construir informaci√≥n de filtros para el reporte
        filters_info_parts = []
        if fecha_inicio:
            filters_info_parts.append(f'Desde: {fecha_inicio}')
        if fecha_fin:
            filters_info_parts.append(f'Hasta: {fecha_fin}')
        if estados:
            filters_info_parts.append(f'Estados: {", ".join(estados)}')
        if tipo_actividad:
            filters_info_parts.append(f'Tipos: {", ".join(tipo_actividad)}')
        if comunidades_filtro:
            filters_info_parts.append(f'Comunidades: {len(comunidades_filtro)} seleccionadas')
        if evento_filtro:
            filters_info_parts.append(f'Eventos: {len(evento_filtro)} seleccionados')
        
        filters_info = ' | '.join(filters_info_parts) if filters_info_parts else 'Sin filtros aplicados'
        
        # Generar datos del reporte llamando directamente a api_generar_reporte
        try:
            # Llamar a api_generar_reporte con la misma request
            report_response = api_generar_reporte(request, report_type)
            
            # Verificar si la respuesta es exitosa
            if not hasattr(report_response, 'status_code'):
                logger.error(
                    "Respuesta inv√°lida al generar datos previos del reporte: usuario=%s, tipo=%s",
                    user_label,
                    report_type,
                )
                return JsonResponse({
                    'error': 'Error: respuesta inv√°lida del servidor'
                }, status=500)
                
            if report_response.status_code != 200:
                # Intentar obtener el mensaje de error del response
                try:
                    import json
                    error_data = json.loads(report_response.content.decode('utf-8'))
                    error_msg = error_data.get('error', 'Error al generar datos del reporte')
                except:
                    error_msg = 'Error al generar datos del reporte'
                logger.error(
                    "Error HTTP al generar datos del reporte: usuario=%s, tipo=%s, detalle=%s",
                    user_label,
                    report_type,
                    error_msg,
                )
                return JsonResponse({
                    'error': error_msg
                }, status=500)
            
            # Obtener los datos del JSON response
            import json
            try:
                if hasattr(report_response, 'content'):
                    content = report_response.content
                    if isinstance(content, bytes):
                        content = content.decode('utf-8')
                    report_data_dict = json.loads(content)
                else:
                    logger.error(
                        "No se pudo obtener el contenido de la respuesta del reporte: usuario=%s, tipo=%s",
                        user_label,
                        report_type,
                    )
                    return JsonResponse({
                        'error': 'Error: no se pudo obtener el contenido de la respuesta'
                    }, status=500)
            except json.JSONDecodeError as e:
                logger.exception(
                    "Error al decodificar JSON del reporte: usuario=%s, tipo=%s",
                    user_label,
                    report_type,
                )
                return JsonResponse({
                    'error': f'Error al decodificar respuesta JSON: {str(e)}'
                }, status=500)
            
            if 'error' in report_data_dict:
                logger.error(
                    "La generaci√≥n de datos del reporte devolvi√≥ un error: usuario=%s, tipo=%s, detalle=%s",
                    user_label,
                    report_type,
                    report_data_dict['error'],
                )
                return JsonResponse({
                    'error': report_data_dict['error']
                }, status=500)
            
            # Extraer los datos seg√∫n el tipo de reporte
            if 'data' in report_data_dict:
                report_data = report_data_dict['data']
                # Validar que report_data no sea None
                if report_data is None:
                    report_data = []
            else:
                report_data = report_data_dict
            
            # Para evento individual, asegurar que la estructura sea correcta
            if report_type == 'reporte-evento-individual':
                print(f'üîµ reporte-evento-individual: report_data keys: {list(report_data.keys()) if isinstance(report_data, dict) else "No es dict"}')
                # Si report_data es {'evento': {...}}, est√° bien
                # Si report_data es directamente el evento, envolverlo
                if isinstance(report_data, dict) and 'evento' not in report_data and 'nombre' in report_data:
                    # Es el evento directamente, envolverlo
                    report_data = {'evento': report_data}
                    print(f'üîµ Envuelto evento en estructura correcta')
            
            # Para comunidades, extraer la lista de comunidades del objeto
            if report_type == 'comunidades':
                if isinstance(report_data, dict) and 'comunidades' in report_data:
                    report_data = report_data['comunidades']
                elif not isinstance(report_data, list):
                    report_data = []
            
            # Para actividad-usuarios, extraer la lista de usuarios del objeto
            if report_type == 'actividad-usuarios':
                if isinstance(report_data, dict) and 'usuarios' in report_data:
                    report_data = report_data['usuarios']
                elif not isinstance(report_data, list):
                    report_data = []
                
            # Validar que tenemos datos v√°lidos
            if report_data is None:
                report_data = []
                
        except Exception as e:
            import traceback
            logger.exception(
                "Error al obtener datos previos del reporte: usuario=%s, tipo=%s",
                user_label,
                report_type,
            )
            traceback.print_exc()
            return JsonResponse({
                'error': f'Error al obtener datos del reporte: {str(e)}'
            }, status=500)
        
        # Generar el archivo seg√∫n el formato
        try:
            if formato == 'pdf':
                file_buffer = generate_pdf_report(report_type, report_data, filters_info)
                content_type = 'application/pdf'
                file_extension = 'pdf'
            else:  # word
                file_buffer = generate_word_report(report_type, report_data, filters_info)
                content_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                file_extension = 'docx'
            
            # Verificar que file_buffer no sea None
            if file_buffer is None:
                logger.error(
                    "El generador de archivos devolvi√≥ un buffer vac√≠o: usuario=%s, tipo=%s, formato=%s",
                    user_label,
                    report_type,
                    formato,
                )
                return JsonResponse({
                    'error': 'Error: no se pudo generar el archivo del reporte'
                }, status=500)
            
            # Crear respuesta HTTP con el archivo
            file_size = None
            try:
                file_buffer.seek(0, os.SEEK_END)
                file_size = file_buffer.tell()
                file_buffer.seek(0)
            except Exception:
                file_size = None
            response_content = file_buffer.read()
            response = HttpResponse(response_content, content_type=content_type)
            report_title = get_report_title(report_type)
            filename = f"{report_title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}"
            
            # Configurar headers para evitar extensi√≥n incorrecta en navegadores
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Type'] = content_type
            response['Content-Length'] = len(response_content)
            
            # Headers adicionales para asegurar descarga correcta
            response['X-Content-Type-Options'] = 'nosniff'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'

            records_count = None
            try:
                records_count = len(report_data)
            except Exception:
                records_count = None

            logger.info(
                "Reporte exportado exitosamente: usuario=%s, tipo=%s, formato=%s, registros=%s, tama√±o=%s bytes, archivo=%s",
                user_label,
                report_type,
                formato,
                records_count,
                file_size if file_size is not None else 'desconocido',
                filename,
            )
            
            return response
            
        except Exception as e:
            import traceback
            logger.exception(
                "Error al generar el archivo del reporte: usuario=%s, tipo=%s, formato=%s",
                user_label,
                report_type,
                formato,
            )
            traceback.print_exc()
            return JsonResponse({
                'error': f'Error al generar el archivo del reporte: {str(e)}'
            }, status=500)
        
    except Exception as e:
        import traceback
        logger.exception(
            "Error inesperado en la exportaci√≥n del reporte: usuario=%s, tipo=%s",
            user_label if 'user_label' in locals() else 'desconocido',
            report_type,
        )
        traceback.print_exc()
        return JsonResponse({
            'error': f'Error al exportar reporte: {str(e)}'
        }, status=500)

def api_actualizar_comunidad_datos(request, comunidad_id):
    """API: Actualizar datos generales y tarjetas personalizadas de una comunidad."""
    try:
        comunidad = Comunidad.objects.get(id=comunidad_id, activo=True)
    except Comunidad.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Comunidad no encontrada',
            },
            status=404,
        )

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse(
            {
                'success': False,
                'error': 'Datos inv√°lidos',
            },
            status=400,
        )

    coordenadas = (payload.get('coordenadas') or '').strip()
    cocode = (payload.get('cocode') or '').strip()
    telefono_cocode = (payload.get('telefono_cocode') or '').strip()

    cocode = re.sub(r'\s+', ' ', cocode)
    telefono_cocode = re.sub(r'\s+', '', telefono_cocode)
    coordenadas = re.sub(r'\s+', ' ', coordenadas)
    coordenadas = re.sub(r'\s*,\s*', ', ', coordenadas)

    nombre_regex = re.compile(r"^[A-Za-z√Å√â√ç√ì√ö√ú√ë√°√©√≠√≥√∫√º√±'\s-]+$")
    coords_regex = re.compile(r"^-?\d{1,3}(?:\.\d+)?\s*,\s*-?\d{1,3}(?:\.\d+)?$")

    if not cocode:
        return JsonResponse(
            {
                'success': False,
                'error': 'El nombre del COCODE es obligatorio.',
            },
            status=400,
        )

    if len(cocode) > 100:
        return JsonResponse(
            {
                'success': False,
                'error': 'El nombre del COCODE no puede superar 100 caracteres.',
            },
            status=400,
        )

    if not nombre_regex.fullmatch(cocode):
        return JsonResponse(
            {
                'success': False,
                'error': 'El nombre del COCODE solo puede contener letras, espacios, ap√≥strofes o guiones.',
            },
            status=400,
        )

    if not telefono_cocode:
        return JsonResponse(
            {
                'success': False,
                'error': 'El tel√©fono del COCODE es obligatorio.',
            },
            status=400,
        )

    if not re.fullmatch(r"\d{8}", telefono_cocode):
        return JsonResponse(
            {
                'success': False,
                'error': 'El tel√©fono del COCODE debe tener exactamente 8 d√≠gitos num√©ricos.',
            },
            status=400,
        )

    if coordenadas and not coords_regex.fullmatch(coordenadas):
        return JsonResponse(
            {
                'success': False,
                'error': 'Las coordenadas deben tener el formato "latitud, longitud" con n√∫meros v√°lidos.',
            },
            status=400,
        )

    poblacion = payload.get('poblacion')
    if poblacion in (None, '', 'null'):
        poblacion_valor = None
    else:
        try:
            poblacion_valor = int(poblacion)
        except (TypeError, ValueError):
            return JsonResponse(
                {
                    'success': False,
                    'error': 'La poblaci√≥n debe ser un n√∫mero entero.',
                },
                status=400,
            )

    if coordenadas and (latitud is None or longitud is None):
        partes = [p.strip() for p in coordenadas.split(',') if p.strip()]
        if len(partes) == 2:
            try:
                latitud = float(partes[0])
                longitud = float(partes[1])
            except ValueError:
                latitud = comunidad.latitud
                longitud = comunidad.longitud
        else:
            latitud = comunidad.latitud
            longitud = comunidad.longitud

    if latitud in ('', None):
        latitud = None
    if longitud in ('', None):
        longitud = None

    comunidad.poblacion = poblacion_valor
    comunidad.cocode = cocode or None
    comunidad.telefono_cocode = telefono_cocode or None
    comunidad.latitud = latitud
    comunidad.longitud = longitud
    comunidad.actualizado_en = timezone.now()
    comunidad.save(
        update_fields=[
            'poblacion',
            'cocode',
            'telefono_cocode',
            'latitud',
            'longitud',
            'actualizado_en',
        ]
    )

    tarjetas_existentes = {
        _normalize_label(tarjeta.titulo): tarjeta
        for tarjeta in TarjetaDato.objects.filter(entidad_tipo='comunidad', entidad_id=comunidad.id)
        if _normalize_label(tarjeta.titulo) in _COMUNIDAD_TITULOS_PREDEFINIDOS
    }

    if latitud is not None and longitud is not None:
        coord_texto = f"{latitud}, {longitud}"
    else:
        coord_texto = (coordenadas or '').strip()

    valor_poblacion = ''
    if poblacion_valor is not None:
        valor_poblacion = f"{poblacion_valor:,} habitantes"

    predefined_cards = [
        {
            'titulo': 'Poblaci√≥n',
            'valor': valor_poblacion,
            'icono': 'üë•',
            'orden': 0,
        },
        {
            'titulo': 'Coordenadas',
            'valor': coord_texto,
            'icono': 'üß≠',
            'orden': 1,
        },
        {
            'titulo': 'COCODE',
            'valor': comunidad.cocode or '',
            'icono': 'üë§',
            'orden': 2,
        },
        {
            'titulo': 'Tel√©fono COCODE',
            'valor': comunidad.telefono_cocode or '',
            'icono': 'üìû',
            'orden': 3,
        },
        {
            'titulo': 'Tipo de Comunidad',
            'valor': comunidad.tipo.get_nombre_display() if comunidad.tipo else '',
            'icono': 'üèòÔ∏è',
            'orden': 4,
        },
    ]

    for card in predefined_cards:
        titulo_norm = _normalize_label(card['titulo'])
        tarjeta_existente = tarjetas_existentes.get(titulo_norm)
        valor_raw = card['valor']
        if isinstance(valor_raw, str):
            valor_limpio = valor_raw.strip()
        else:
            valor_limpio = str(valor_raw) if valor_raw is not None else ''

        if tarjeta_existente:
            tarjeta_existente.titulo = card['titulo']
            tarjeta_existente.valor = valor_limpio
            tarjeta_existente.icono = card['icono']
            tarjeta_existente.orden = card['orden']
            tarjeta_existente.save(update_fields=['titulo', 'valor', 'icono', 'orden', 'actualizado_en'])
        else:
            TarjetaDato.objects.create(
                entidad_tipo='comunidad',
                entidad_id=comunidad.id,
                titulo=card['titulo'],
                valor=valor_limpio,
                icono=card['icono'],
                orden=card['orden'],
            )

    payload_actualizado = _serialize_comunidad_detalle(comunidad, request)

    return JsonResponse(
        {
            'success': True,
            'message': 'Datos actualizados exitosamente.',
            'comunidad': payload_actualizado,
        }
    )

@require_http_methods(["POST"])
@permiso_admin_o_personal_api
def api_actualizar_archivo_region(request, region_id, archivo_id):
    """API: Actualizar la descripci√≥n de un archivo de regi√≥n"""
    try:
        archivo = RegionArchivo.objects.get(id=archivo_id, region_id=region_id)
    except RegionArchivo.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Archivo no encontrado',
            },
            status=404,
        )

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        payload = {}

    descripcion = (payload.get('descripcion') or '').strip()

    archivo.descripcion = descripcion or ''
    archivo.save(update_fields=['descripcion'])

    region = archivo.region
    if region:
        region.actualizado_en = timezone.now()
        region.save(update_fields=['actualizado_en'])

    return JsonResponse(
        {
            'success': True,
            'message': 'Descripci√≥n de archivo actualizada correctamente',
            'archivo': {
                'id': str(archivo.id),
                'name': archivo.nombre_archivo,
                'description': archivo.descripcion or '',
                'type': archivo.archivo_tipo or 'archivo',
                'url': archivo.url_almacenamiento,
                'date': archivo.creado_en.isoformat() if archivo.creado_en else None,
            },
        }
    )
def api_agregar_imagen_comunidad(request, comunidad_id):
    try:
        comunidad = Comunidad.objects.get(id=comunidad_id, activo=True)
    except Comunidad.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Comunidad no encontrada'}, status=404)

    usuario_maga = get_usuario_maga(request.user)
    if not usuario_maga:
        return JsonResponse({'success': False, 'error': 'Usuario no autenticado'}, status=401)

    imagen = request.FILES.get('imagen')
    if not imagen:
        return JsonResponse({'success': False, 'error': 'No se ha enviado ninguna imagen'}, status=400)

    if not imagen.content_type or not imagen.content_type.startswith('image/'):
        return JsonResponse({'success': False, 'error': 'El archivo debe ser una imagen v√°lida'}, status=400)

    # COMPRIMIR IMAGEN autom√°ticamente
    if getattr(settings, 'COMPRESS_IMAGES', True):
        from .image_compression import compress_if_needed, get_compression_settings
        compression_settings = get_compression_settings('gallery')
        imagen = compress_if_needed(imagen, **compression_settings)

    descripcion = (request.POST.get('descripcion') or '').strip()

    galeria_dir = os.path.join(str(settings.MEDIA_ROOT), 'comunidades_galeria')
    os.makedirs(galeria_dir, exist_ok=True)
    
    print(f"üì§ Subiendo imagen a comunidad {comunidad_id}")
    print(f"üìÅ Directorio: {galeria_dir}")
    print(f"‚úÖ Archivo recibido: {imagen.name}, tama√±o: {imagen.size}, tipo: {imagen.content_type}")
    
    # Verificar permisos de escritura
    if not os.access(galeria_dir, os.W_OK):
        print(f"‚ùå No hay permisos de escritura en {galeria_dir}")
        return JsonResponse({'success': False, 'error': f'No se tienen permisos de escritura en {galeria_dir}'}, status=500)

    fs = FileSystemStorage(location=galeria_dir)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
    extension = os.path.splitext(imagen.name)[1]
    filename = f"{timestamp}_{comunidad_id}{extension}"
    print(f"üíæ Intentando guardar archivo: {filename}")
    try:
        saved_name = fs.save(filename, imagen)
        print(f"‚úÖ Archivo guardado exitosamente: {saved_name}")
    except Exception as e:
        import traceback
        print(f"‚ùå Error al guardar archivo: {str(e)}")
        print(f"üìã Traceback:\n{traceback.format_exc()}")
        raise
    file_url = f"/media/comunidades_galeria/{saved_name}"

    foto = ComunidadGaleria.objects.create(
        comunidad=comunidad,
        archivo_nombre=imagen.name,
        url_almacenamiento=file_url,
        descripcion=descripcion or None,
        creado_por=usuario_maga,
    )

    comunidad.actualizado_en = timezone.now()
    comunidad.save(update_fields=['actualizado_en'])

    updated = (
        Comunidad.objects.select_related('region', 'tipo')
        .prefetch_related(
            Prefetch(
                'galeria', queryset=ComunidadGaleria.objects.order_by('-creado_en'), to_attr='galeria_api'
            ),
            Prefetch(
                'archivos', queryset=ComunidadArchivo.objects.order_by('-creado_en'), to_attr='archivos_api'
            ),
            Prefetch(
                'autoridades', queryset=ComunidadAutoridad.objects.filter(activo=True).order_by('nombre'), to_attr='autoridades_api'
            ),
        )
        .get(id=comunidad_id, activo=True)
    )

    payload = _serialize_comunidad_detalle(updated, request)

    return JsonResponse(
        {
            'success': True,
            'message': 'Imagen agregada exitosamente',
            'foto': {
                'id': str(foto.id),
                'url': foto.url_almacenamiento,
                'description': foto.descripcion or 'Imagen de la comunidad',
            },
            'comunidad': payload,
        }
    )


@require_http_methods(["DELETE"])
@permiso_admin_o_personal_api
def api_eliminar_imagen_comunidad(request, comunidad_id, imagen_id):
    try:
        foto = ComunidadGaleria.objects.get(id=imagen_id, comunidad_id=comunidad_id)
    except ComunidadGaleria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Imagen no encontrada'}, status=404)

    ruta = foto.url_almacenamiento or ''
    ruta_media = (settings.MEDIA_URL or '/media/').rstrip('/')
    if ruta.startswith(ruta_media):
        relative_path = ruta[len(ruta_media):].lstrip('/')
    elif ruta.startswith('/'):
        ruta_sin_slash = ruta.lstrip('/')
        media_prefix = ruta_media.lstrip('/')
        if ruta_sin_slash.startswith(media_prefix):
            relative_path = ruta_sin_slash[len(media_prefix):].lstrip('/')
        else:
            relative_path = ruta_sin_slash
    else:
        relative_path = ruta

    if relative_path:
        file_path = os.path.join(str(settings.MEDIA_ROOT), relative_path.replace('/', os.sep))
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                pass

    foto.delete()

    updated = (
        Comunidad.objects.select_related('region', 'tipo')
        .prefetch_related(
            Prefetch(
                'galeria', queryset=ComunidadGaleria.objects.order_by('-creado_en'), to_attr='galeria_api'
            ),
            Prefetch(
                'archivos', queryset=ComunidadArchivo.objects.order_by('-creado_en'), to_attr='archivos_api'
            ),
            Prefetch(
                'autoridades', queryset=ComunidadAutoridad.objects.filter(activo=True).order_by('nombre'), to_attr='autoridades_api'
            ),
        )
        .get(id=comunidad_id, activo=True)
    )

    payload = _serialize_comunidad_detalle(updated, request)

    return JsonResponse(
        {
            'success': True,
            'message': 'Imagen eliminada correctamente',
            'comunidad': payload,
        }
    )


@require_http_methods(["POST"])
@permiso_admin_o_personal_api
def api_actualizar_comunidad_descripcion(request, comunidad_id):
    try:
        comunidad = Comunidad.objects.get(id=comunidad_id, activo=True)
    except Comunidad.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Comunidad no encontrada',
            },
            status=404,
        )

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse(
            {
                'success': False,
                'error': 'Datos inv√°lidos',
            },
            status=400,
        )

    descripcion = (payload.get('descripcion') or '').strip()

    comunidad.descripcion = descripcion or ''
    comunidad.actualizado_en = timezone.now()
    comunidad.save(update_fields=['descripcion', 'actualizado_en'])

    comunidad_actualizada = (
        Comunidad.objects.select_related('region', 'tipo')
        .prefetch_related(
            Prefetch('galeria', queryset=ComunidadGaleria.objects.order_by('-creado_en'), to_attr='galeria_api'),
            Prefetch('archivos', queryset=ComunidadArchivo.objects.order_by('-creado_en'), to_attr='archivos_api'),
            Prefetch('autoridades', queryset=ComunidadAutoridad.objects.filter(activo=True).order_by('nombre'), to_attr='autoridades_api'),
        )
        .get(id=comunidad_id, activo=True)
    )

    payload_actualizado = _serialize_comunidad_detalle(comunidad_actualizada, request)

    return JsonResponse(
        {
            'success': True,
            'message': 'Descripci√≥n actualizada exitosamente.',
            'comunidad': payload_actualizado,
        }
    )


def api_agregar_archivo_comunidad(request, comunidad_id):
    try:
        comunidad = Comunidad.objects.get(id=comunidad_id, activo=True)
    except Comunidad.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Comunidad no encontrada'}, status=404)

    usuario_maga = get_usuario_maga(request.user)
    if not usuario_maga:
        return JsonResponse({'success': False, 'error': 'Usuario no autenticado'}, status=401)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'No se ha enviado ning√∫n archivo'}, status=400)

    nombre = (request.POST.get('nombre') or '').strip()
    descripcion = (request.POST.get('descripcion') or '').strip()

    extension = os.path.splitext(archivo.name)[1].lower()
    if not nombre:
        nombre = os.path.splitext(archivo.name)[0]
    archivo_tipo = extension.replace('.', '') if extension else (archivo.content_type or 'archivo')

    archivos_dir = os.path.join(str(settings.MEDIA_ROOT), 'comunidades_archivos')
    os.makedirs(archivos_dir, exist_ok=True)

    fs = FileSystemStorage(location=archivos_dir)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
    filename = f"{timestamp}_{comunidad_id}{extension}"
    saved_name = fs.save(filename, archivo)

    relative_path = f"comunidades_archivos/{saved_name}"
    media_url = (settings.MEDIA_URL or '/media/').rstrip('/')
    file_url = f"{media_url}/{relative_path}"
    if not file_url.startswith('/'):
        file_url = f"/{file_url}"

    ComunidadArchivo.objects.create(
        comunidad=comunidad,
        nombre_archivo=nombre,
        archivo_tipo=archivo_tipo,
        url_almacenamiento=file_url,
        descripcion=descripcion,
        creado_por=usuario_maga,
    )

    comunidad.actualizado_en = timezone.now()
    comunidad.save(update_fields=['actualizado_en'])

    comunidad_actualizada = (
        Comunidad.objects.select_related('region', 'tipo')
        .prefetch_related(
            Prefetch('galeria', queryset=ComunidadGaleria.objects.order_by('-creado_en'), to_attr='galeria_api'),
            Prefetch('archivos', queryset=ComunidadArchivo.objects.order_by('-creado_en'), to_attr='archivos_api'),
            Prefetch('autoridades', queryset=ComunidadAutoridad.objects.filter(activo=True).order_by('nombre'), to_attr='autoridades_api'),
        )
        .get(id=comunidad_id, activo=True)
    )

    payload_actualizado = _serialize_comunidad_detalle(comunidad_actualizada, request)

    return JsonResponse(
        {
            'success': True,
            'message': 'Archivo agregado exitosamente',
            'comunidad': payload_actualizado,
        }
    )


@require_http_methods(["DELETE"])
@permiso_admin_o_personal_api
def api_eliminar_archivo_comunidad(request, comunidad_id, archivo_id):
    try:
        archivo = ComunidadArchivo.objects.get(id=archivo_id, comunidad_id=comunidad_id)
    except ComunidadArchivo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Archivo no encontrado'}, status=404)

    comunidad = archivo.comunidad

    ruta_url = archivo.url_almacenamiento or ''
    media_url = (settings.MEDIA_URL or '/media/').rstrip('/')
    relative_path = ''

    if ruta_url.startswith(media_url):
        relative_path = ruta_url[len(media_url):].lstrip('/')
    elif ruta_url.startswith('/'):
        ruta_sin_slash = ruta_url.lstrip('/')
        media_prefix = media_url.lstrip('/')
        if ruta_sin_slash.startswith(media_prefix):
            relative_path = ruta_sin_slash[len(media_prefix):].lstrip('/')
        else:
            relative_path = ruta_sin_slash
    else:
        relative_path = ruta_url

    if relative_path:
        file_path = os.path.join(str(settings.MEDIA_ROOT), relative_path.replace('/', os.sep))
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                pass

    archivo.delete()

    if comunidad:
        comunidad.actualizado_en = timezone.now()
        comunidad.save(update_fields=['actualizado_en'])

    comunidad_actualizada = (
        Comunidad.objects.select_related('region', 'tipo')
        .prefetch_related(
            Prefetch('galeria', queryset=ComunidadGaleria.objects.order_by('-creado_en'), to_attr='galeria_api'),
            Prefetch('archivos', queryset=ComunidadArchivo.objects.order_by('-creado_en'), to_attr='archivos_api'),
            Prefetch('autoridades', queryset=ComunidadAutoridad.objects.filter(activo=True).order_by('nombre'), to_attr='autoridades_api'),
        )
        .get(id=comunidad_id, activo=True)
    )

    payload_actualizado = _serialize_comunidad_detalle(comunidad_actualizada, request)

    return JsonResponse(
        {
            'success': True,
            'message': 'Archivo eliminado correctamente',
            'comunidad': payload_actualizado,
        }
    )


@require_http_methods(["POST"])
@permiso_admin_o_personal_api
def api_actualizar_archivo_comunidad(request, comunidad_id, archivo_id):
    try:
        archivo = ComunidadArchivo.objects.get(id=archivo_id, comunidad_id=comunidad_id)
    except ComunidadArchivo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Archivo no encontrado'}, status=404)

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        payload = {}

    descripcion = (payload.get('descripcion') or '').strip()

    archivo.descripcion = descripcion or ''
    archivo.save(update_fields=['descripcion'])

    comunidad = archivo.comunidad
    if comunidad:
        comunidad.actualizado_en = timezone.now()
        comunidad.save(update_fields=['actualizado_en'])

    comunidad_actualizada = (
        Comunidad.objects.select_related('region', 'tipo')
        .prefetch_related(
            Prefetch('galeria', queryset=ComunidadGaleria.objects.order_by('-creado_en'), to_attr='galeria_api'),
            Prefetch('archivos', queryset=ComunidadArchivo.objects.order_by('-creado_en'), to_attr='archivos_api'),
            Prefetch('autoridades', queryset=ComunidadAutoridad.objects.filter(activo=True).order_by('nombre'), to_attr='autoridades_api'),
        )
        .get(id=comunidad_id, activo=True)
    )

    payload_actualizado = _serialize_comunidad_detalle(comunidad_actualizada, request)

    return JsonResponse(
        {
            'success': True,
            'message': 'Descripci√≥n de archivo actualizada correctamente',
            'comunidad': payload_actualizado,
        }
    )

DEFAULT_COMUNIDAD_IMAGE_SMALL = (
    'https://images.unsplash.com/photo-1523978591478-c753949ff840'
    '?ixlib=rb-4.0.3&auto=format&fit=crop&w=400&q=80'
)
DEFAULT_COMUNIDAD_IMAGE_LARGE = (
    'https://images.unsplash.com/photo-1523978591478-c753949ff840'
    '?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80'
)


# =====================================================
# API: LOGIN (usado para modo online y generaci√≥n de credenciales offline)
# =====================================================

@require_http_methods(["POST"])
@api_ratelimit_login_smart(rate_per_user='10/3m', rate_per_ip='20/3m')
def api_login(request):
    """
    Endpoint API para login que devuelve JSON con informaci√≥n del usuario.
    Se usa tanto para login online como para generar credenciales offline.
    
    Rate limit inteligente:
    - 20 intentos por IP cada 3 minutos (permite m√∫ltiples usuarios en la misma red)
    - 10 intentos por usuario cada 3 minutos (protege contra fuerza bruta en cuentas espec√≠ficas)
    """
    try:
        data = json.loads(request.body)
        username_or_email = data.get('username', '').strip()
        password = data.get('password', '')
        remember_me = data.get('remember_me', False)
        device_id = data.get('device_id', '')
        
        if not username_or_email or not password:
            return JsonResponse({
                'success': False,
                'error': 'Usuario y contrase√±a son requeridos'
            }, status=400)
        
        # Intentar autenticar
        user = authenticate(request, username=username_or_email, password=password)
        
        # Si falla con username, intentar con email
        if not user:
            try:
                usuario_maga = Usuario.objects.get(email=username_or_email)
                user = authenticate(request, username=usuario_maga.username, password=password)
            except Usuario.DoesNotExist:
                pass
        
        if not user:
            return JsonResponse({
                'success': False,
                'error': 'Usuario o contrase√±a incorrectos'
            }, status=401)
        
        if not user.is_active:
            return JsonResponse({
                'success': False,
                'error': 'Esta cuenta est√° desactivada'
            }, status=403)
        
        # Iniciar sesi√≥n
        auth_login(request, user)
        request.session.set_expiry(604800)  # 7 d√≠as
        request.session.save()
        
        # Obtener informaci√≥n del usuario
        try:
            usuario_maga = Usuario.objects.select_related('puesto').get(username=user.username)
        except Usuario.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no encontrado en el sistema'
            }, status=404)
        
        # Preparar datos del usuario
        user_data = {
            'id': str(usuario_maga.id),
            'username': usuario_maga.username,
            'nombre': usuario_maga.nombre or '',
            'email': usuario_maga.email,
            'rol': usuario_maga.rol,
            'isAdmin': usuario_maga.rol == 'admin',
            'isPersonal': usuario_maga.rol == 'personal',
            'puesto': usuario_maga.puesto.nombre if usuario_maga.puesto else None,
            'permisos': {
                'es_admin': usuario_maga.rol == 'admin',
                'es_personal': usuario_maga.rol == 'personal',
                'puede_gestionar_eventos': usuario_maga.rol in ['admin', 'personal'],
                'puede_generar_reportes': True,
            }
        }
        
        # Si el usuario quiere recordar credenciales, generar hash para offline
        offline_data = None
        if remember_me:
            import secrets
            # Generar salt y hash para almacenamiento offline
            salt = secrets.token_hex(16)
            password_with_salt = f"{password}{salt}"
            credential_hash = hashlib.sha256(password_with_salt.encode()).hexdigest()
            
            # Guardar en la base de datos
            if not device_id:
                device_id = request.META.get('HTTP_USER_AGENT', '')[:100]
            
            SesionOffline.objects.update_or_create(
                usuario=usuario_maga,
                dispositivo_id=device_id,
                defaults={
                    'token_hash': credential_hash,
                    'permisos_offline': user_data['permisos'],
                    'datos_cache': user_data,
                    'sesion_activa': True,
                    'expira_en': timezone.now() + timedelta(days=30),
                    'navegador': request.META.get('HTTP_USER_AGENT', '')[:100],
                    'sistema_operativo': request.META.get('HTTP_SEC_CH_UA_PLATFORM', '')[:100],
                    'ip_address': request.META.get('REMOTE_ADDR', '')[:50],
                }
            )
            
            offline_data = {
                'salt': salt,
                'hash': credential_hash,
                'expiresAt': (timezone.now() + timedelta(days=30)).isoformat(),
            }
        
        return JsonResponse({
            'success': True,
            'message': 'Inicio de sesi√≥n exitoso',
            'user': user_data,
            'offline': offline_data,
            'redirectUrl': '/'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inv√°lidos'
        }, status=400)
    except Exception as e:
        logger.error(f"Error en api_login: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Error interno del servidor'
        }, status=500)